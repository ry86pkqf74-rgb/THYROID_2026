"""06_verify.py — Cross-check the 4 deliverables against the validation report's headline cells.

Reads the parquet directly (single source of truth) AND the deliverables' Excel files,
prints a reconciliation table.
"""
from pathlib import Path
import pandas as pd
import openpyxl
from datetime import date

HERE = Path(__file__).parent
PKG = HERE.parent
PARQUET = HERE / "m038_per_patient_v2.parquet"

D1 = PKG / "M038_GOITER_patient_level_dataset.xlsx"
D2 = PKG / "M038_GOITER_analysis_workbook.xlsx"
D3 = PKG / "M038_GOITER_tables_figures.xlsx"
D4 = PKG / "M038_GOITER_eMethods.docx"


def main():
    df = pd.read_parquet(PARQUET)
    print("=" * 70)
    print(f"M038 verification — date {date.today().isoformat()}")
    print(f"Source parquet: {PARQUET.name}  ({len(df):,} × {len(df.columns)} cols)")
    print("=" * 70)
    print()

    # File presence + sizes
    print("Deliverable files:")
    for f in [D1, D2, D3, D4]:
        if f.exists():
            print(f"  ✓ {f.name}  ({f.stat().st_size/1024:.1f} KB)")
        else:
            print(f"  ✗ {f.name}  MISSING")
    print()

    # Parquet-derived ground truth
    n = len(df)
    n_m = int(df["is_massive"].sum())
    n_w = int(df["comp_weight_ge100"].sum())
    n_s = int(df["comp_substernal_any"].sum())
    n_a = int(df["comp_airway_any"].sum())
    massive = df[df["is_massive"]]; non_massive = df[~df["is_massive"]]
    n_nm = len(non_massive)
    anycomp_m = int(massive["any_confirmed_complication_flag"].sum())
    anycomp_nm = int(non_massive["any_confirmed_complication_flag"].sum())
    hpt_t_m = int((massive["comp_hypoparathyroidism_confirmed"] & massive["comp_hypoparathyroidism_transient"]).sum())
    hpt_t_nm = int((non_massive["comp_hypoparathyroidism_confirmed"] & non_massive["comp_hypoparathyroidism_transient"]).sum())
    hpt_p_m = int((massive["comp_hypoparathyroidism_confirmed"] & massive["comp_hypoparathyroidism_permanent"]).sum())
    hpt_p_nm = int((non_massive["comp_hypoparathyroidism_confirmed"] & non_massive["comp_hypoparathyroidism_permanent"]).sum())
    rln_m = int(massive["comp_rln_injury_confirmed"].sum())
    rln_nm = int(non_massive["comp_rln_injury_confirmed"].sum())
    # Total thy in massive arm
    tt_m = int(massive["surg_total_thyroidectomy"].sum())
    tt_nm = int(non_massive["surg_total_thyroidectomy"].sum())

    # Era 3-bucket
    era_pre = df[df["era_bucket_3"] == "pre-2015"]
    era_15 = df[df["era_bucket_3"] == "2015-2019"]
    era_20 = df[df["era_bucket_3"] == "2020-2025"]

    cells = [
        ("Cohort total", n, "10,871 (validation)"),
        ("n massive", n_m, "2,501 (validation)"),
        ("Massive %", f"{100*n_m/n:.1f}%", "23.0% (validation)"),
        ("Component W (≥100g)", n_w, "1,429 (validation)"),
        ("Component S (substernal)", n_s, "1,047 (validation)"),
        ("Component A (airway)", n_a, "1,440 (validation)"),
        ("Weight only (W∧¬S∧¬A)", int(((df['comp_weight_ge100']) & (~df['comp_substernal_any']) & (~df['comp_airway_any'])).sum()), "898 (validation)"),
        ("Substernal only", int(((df['comp_substernal_any']) & (~df['comp_weight_ge100']) & (~df['comp_airway_any'])).sum()), "145 (validation, post-v2.1)"),
        ("Airway only", int(((df['comp_airway_any']) & (~df['comp_weight_ge100']) & (~df['comp_substernal_any'])).sum()), "429 (validation, post-v2.1)"),
        ("All three", int((df['comp_weight_ge100'] & df['comp_substernal_any'] & df['comp_airway_any']).sum()), "386 (validation)"),
        ("Any-comp massive", anycomp_m, "132 (validation)"),
        ("Any-comp non-massive", anycomp_nm, "268 (validation)"),
        ("Any-comp massive %", f"{100*anycomp_m/n_m:.2f}%", "5.28% (validation)"),
        ("Any-comp non-massive %", f"{100*anycomp_nm/n_nm:.2f}%", "3.20% (validation)"),
        ("HypoPT transient massive", hpt_t_m, "83 (validation)"),
        ("HypoPT transient non-massive", hpt_t_nm, "197 (validation)"),
        ("HypoPT permanent massive", hpt_p_m, "4 (validation)"),
        ("HypoPT permanent non-massive", hpt_p_nm, "12 (validation)"),
        ("Total thyroidectomy massive %", f"{100*tt_m/n_m:.1f}%", "66.9% (validation)"),
        ("Total thyroidectomy non-massive %", f"{100*tt_nm/n_nm:.1f}%", "51.7% (validation)"),
        ("Era pre-2015 massive %", f"{100*era_pre['is_massive'].mean():.1f}%" if len(era_pre) else "—", "12.0% (validation)"),
        ("Era 2015-2019 massive %", f"{100*era_15['is_massive'].mean():.1f}%" if len(era_15) else "—", "24.9% (validation)"),
        ("Era 2020-2025 massive %", f"{100*era_20['is_massive'].mean():.1f}%" if len(era_20) else "—", "28.5% (validation; mig_254 backfill may shift)"),
    ]

    print(f"{'Cell':<40} {'Live':<14} Validation reference")
    print("-" * 90)
    for k, live, ref in cells:
        print(f"{k:<40} {str(live):<14} {ref}")
    print()

    # Spot-check D1 patient data sheet actually has 10,871 rows
    print("Deliverable spot-checks:")
    wb1 = openpyxl.load_workbook(D1, read_only=True)
    pd_ws = wb1["Patient Data"]
    n_rows_d1 = pd_ws.max_row - 1  # subtract header
    print(f"  D1 Patient Data sheet rows: {n_rows_d1:,}  (expected 10,871)  {'✓' if n_rows_d1 == 10871 else '✗'}")
    n_cols_d1 = pd_ws.max_column
    print(f"  D1 Patient Data sheet cols: {n_cols_d1}  (parquet has {len(df.columns)})  {'✓' if n_cols_d1 == len(df.columns) else '✗'}")
    wb1.close()

    wb2 = openpyxl.load_workbook(D2, read_only=True)
    print(f"  D2 Sheet count: {len(wb2.sheetnames)}  (expected 9)  {'✓' if len(wb2.sheetnames) == 9 else '✗'}")
    print(f"  D2 sheets: {wb2.sheetnames}")
    wb2.close()

    wb3 = openpyxl.load_workbook(D3, read_only=True)
    print(f"  D3 Sheet count: {len(wb3.sheetnames)}  (expected 15: T1-T5 + Fig1-4 + Supp S1-S6)  {'✓' if len(wb3.sheetnames) == 15 else '✗'}")
    print(f"  D3 sheets: {wb3.sheetnames}")
    wb3.close()

    print()
    print(f"  D4 eMethods size: {D4.stat().st_size/1024:.1f} KB")
    print()
    print("Verification complete.")


if __name__ == "__main__":
    main()
