"""02_build_d1_patient_dataset.py — Deliverable 1
M038_GOITER_patient_level_dataset.xlsx (Cover + Patient Data + Data Dictionary).
"""
from pathlib import Path
import pandas as pd
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
PKG = HERE.parent
PARQUET = HERE / "m038_per_patient_v2.parquet"
OUT = PKG / "M038_GOITER_patient_level_dataset.xlsx"

HF = PatternFill("solid", fgColor="1F4E78")
HFONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
TITLE = Font(name="Arial", size=14, bold=True, color="1F4E78")
BD = Border(*[Side(style="thin", color="CCCCCC")] * 4)
AL = Alignment(horizontal="left", vertical="center")
AR = Alignment(horizontal="right", vertical="center")
AC = Alignment(horizontal="center", vertical="center")
AW = Alignment(horizontal="left", vertical="top", wrap_text=True)


def write_df(wb, name, df, freeze_to_col=2, autosize_first=20):
    ws = wb.create_sheet(name)
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = HFONT; c.fill = HF; c.alignment = AC; c.border = BD
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, v in enumerate(row, 1):
            if pd.isna(v):
                v = None
            elif hasattr(v, "to_pydatetime"):
                pyd = v.to_pydatetime()
                v = pyd.date() if hasattr(pyd, "date") else pyd
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = BODY; c.border = BD
            c.alignment = AR if isinstance(v, (int, float)) else AL
    ws.freeze_panes = ws.cell(row=2, column=freeze_to_col)
    for i, col in enumerate(df.columns[:autosize_first], 1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(str(col)) + 2, 14), 36)
    return ws


# Data dictionary (column -> {type, definition, source, allowed_values})
DD = {
    "research_id": ("VARCHAR", "Patient identifier (joins all sources)",
                    "manuscript_workspace.cohort_m038_massive_goiter_v1.research_id", "string"),
    "is_massive": ("BOOLEAN", "PRIMARY EXPOSURE — composite-massive flag (OR of weight/substernal/airway components)",
                   "DERIVED — Methods §2.3", "TRUE/FALSE"),
    "comp_weight_ge100": ("BOOLEAN", "Component flag — gland weight >=100g (synoptic pathology)",
                          "DERIVED — gland_weight_final_g >= 100", "TRUE/FALSE"),
    "comp_substernal_any": ("BOOLEAN", "Component flag — substernal extension on CT or MRI",
                            "DERIVED — ct_substernal_extension_any OR mri_substernal_any", "TRUE/FALSE"),
    "comp_airway_any": ("BOOLEAN", "Component flag — tracheal deviation, narrowing, or airway compromise on CT",
                        "DERIVED — ct_tracheal_deviation_any OR ct_tracheal_narrowing_any OR ct_airway_compromise_any",
                        "TRUE/FALSE"),
    "era_bucket_6": ("VARCHAR", "Six-bucket era based on surg_first_date (incl. 'unknown'); upper-bound rule sweeps pre-1999 into 1999-2004",
                     "DERIVED — surg_first_date CASE WHEN", "1999-2004 / 2005-2009 / 2010-2014 / 2015-2019 / 2020-2025 / unknown"),
    "era_bucket_3": ("VARCHAR", "Three-bucket era used in manuscript headline trend",
                     "DERIVED — surg_first_date CASE WHEN", "pre-2015 / 2015-2019 / 2020-2025 / unknown"),
    "hypopara_postop_class": ("VARCHAR", "Postop hypopara class per standing rule (transient<6mo / permanent>6mo)",
                              "DERIVED — comp_hypoparathyroidism_confirmed × _transient/_permanent",
                              "transient_lt_6mo / permanent_gt_6mo / unclassified / none"),
    "hca_preop_flag": ("BOOLEAN", "Preexisting hypocalcemia flag (timing window OR clinical_preexisting)",
                       "DERIVED — comp_hypocalcemia_timing_window OR _clinical_preexisting", "TRUE/FALSE"),
}

# Column-group documentation (used when explicit DD entry not available)
GROUP_DD = [
    ("age_at_surgery", "BIGINT", "Age at first surgery (years)",
     "manuscript_workspace.cohort_m038_massive_goiter_v1", "integer years"),
    ("sex", "VARCHAR", "Patient sex",
     "manuscript_workspace.cohort_m038_massive_goiter_v1", "Female/Male/Unknown"),
    ("race", "VARCHAR", "Patient race (9-bucket)",
     "manuscript_workspace.cohort_m038_massive_goiter_v1",
     "White / Black or AA / Asian / NHPI / AIAN / Other / Multiple / Unknown / Declined"),
    ("bmi_combined", "DOUBLE", "Combined BMI (kg/m^2) using NSQIP first, fallback to vitals/derived",
     "manuscript_workspace.cohort_m038_massive_goiter_v1", "kg/m^2"),
    ("bmi_source", "VARCHAR", "Source of BMI value",
     "manuscript_workspace.cohort_m038_massive_goiter_v1", "nsqip / vitals / derived / null"),
    ("nsqip_asa_class", "VARCHAR", "ASA physical status class (NSQIP-linked subset only)",
     "manuscript_workspace.cohort_m038_massive_goiter_v1", "1/2/3/4/5"),
]


def build_data_dict(df):
    """Combine explicit DD with auto-inferred entries for remaining cols."""
    rows = []
    seen = set()
    for col, (typ, defn, src, allowed) in DD.items():
        if col in df.columns:
            rows.append({"column": col, "dtype": typ, "definition": defn,
                         "source": src, "allowed_values": allowed})
            seen.add(col)
    for col, typ, defn, src, allowed in GROUP_DD:
        if col in df.columns and col not in seen:
            rows.append({"column": col, "dtype": typ, "definition": defn,
                         "source": src, "allowed_values": allowed})
            seen.add(col)
    # Auto-fill remaining
    for col in df.columns:
        if col in seen:
            continue
        dtype = str(df[col].dtype)
        # Guess definition from prefix
        if col.startswith("nsqip_"):
            defn = "NSQIP-linked field"; src = "manuscript_workspace.cohort_m038_massive_goiter_v1 (NSQIP join)"
        elif col.startswith("pmhx_nlp_"):
            defn = "NLP-extracted past medical history flag"; src = "manuscript_workspace.cohort_m038_massive_goiter_v1 (NLP pmhx)"
        elif col.startswith("pshx_nlp_"):
            defn = "NLP-extracted past surgical history flag"; src = "manuscript_workspace.cohort_m038_massive_goiter_v1 (NLP pshx)"
        elif col.startswith("syn_"):
            defn = "Synoptic pathology field"; src = "manuscript_workspace.cohort_m038_massive_goiter_v1 (synoptic)"
        elif col.startswith("comp_") and ("transient" in col or "permanent" in col or "timing_window" in col
                                          or "preexisting" in col or "new_postop" in col or "limitation_note" in col
                                          or "clinical_preexisting" in col):
            defn = "Complication temporality field (post-mig_255)"; src = "manuscript_workspace.cohort_m038_massive_goiter_v1 (post-mig_255)"
        elif col.startswith("comp_"):
            defn = "Strict-definition complication flag (post-mig_252 rollup)"; src = "manuscript_workspace.cohort_m038_massive_goiter_v1"
        elif col.startswith(("ct_", "mri_", "nlp_airway_")):
            defn = "Imaging-derived flag (composite-massive component)"; src = "manuscript_workspace.cohort_m038_massive_goiter_v1 (imaging)"
        elif col.startswith("surg_"):
            defn = "Surgical episode field (post-mig_253/254)"; src = "manuscript_workspace.cohort_m038_massive_goiter_v1 (surg episode)"
        elif col.startswith("ops_"):
            defn = "Operative-note NLP field"; src = "manuscript_workspace.cohort_m038_massive_goiter_v1 (ops NLP)"
        elif col.startswith("proc_nlp_"):
            defn = "NLP-extracted procedure field"; src = "manuscript_workspace.cohort_m038_massive_goiter_v1 (procedure NLP)"
        elif col.startswith("gland_weight_"):
            defn = "Synoptic pathology gland weight (grams)"; src = "manuscript_workspace.cohort_m038_massive_goiter_v1 (synoptic)"
        else:
            defn = "Reference column"; src = "manuscript_workspace.cohort_m038_massive_goiter_v1"
        rows.append({"column": col, "dtype": dtype, "definition": defn,
                     "source": src, "allowed_values": ""})
    return pd.DataFrame(rows)


def main():
    df = pd.read_parquet(PARQUET)
    print(f"→ Loaded {len(df):,} rows × {len(df.columns)} cols")

    wb = Workbook(); wb.remove(wb.active)

    # Cover
    ws = wb.create_sheet("Cover")
    ws.cell(row=1, column=1, value="M038 — Massive Goiter Descriptive Cohort").font = TITLE
    ws.cell(row=2, column=1, value="Per-patient analytic dataset (Deliverable 1)").font = BOLD
    cover_rows = [
        ("Manuscript ID", "M038"),
        ("Title", "Massive Goiter at a Tertiary Referral Center: A Composite-Definition Descriptive Cohort"),
        ("Date assembled", date.today().isoformat()),
        ("Database", "thyroid_canonical_publication_v1_0 (release pub_v1_1_20260504)"),
        ("Cohort view", "manuscript_workspace.cohort_m038_massive_goiter_v1 (post-mig_255)"),
        ("Cohort definition (composite massive flag)",
         "gland_weight_final_g >= 100g  OR  ct_substernal_extension_any OR mri_substernal_any  "
         "OR  ct_tracheal_deviation_any OR ct_tracheal_narrowing_any OR ct_airway_compromise_any"),
        ("n total", f"{len(df):,}"),
        ("n massive", f"{int(df['is_massive'].sum()):,} ({100*df['is_massive'].mean():.1f}%)"),
        ("n columns", f"{len(df.columns)}"),
        ("Sheets", "Cover (this sheet) → Patient Data (1 row per research_id) → Data Dictionary"),
        ("Standing rule reference", "memory/feedback_complications_transient_vs_permanent.md"),
        ("Validation reference", "09_validation_report.md (156-cell audit; 153 PASS / 3 patched / 0 FAIL)"),
    ]
    for i, (k, v) in enumerate(cover_rows, 4):
        a = ws.cell(row=i, column=1, value=k); a.font = BOLD; a.alignment = AW
        b = ws.cell(row=i, column=2, value=v); b.font = BODY; b.alignment = AW
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 110
    for r in range(4, 4 + len(cover_rows)):
        ws.row_dimensions[r].height = 30

    # Patient Data
    write_df(wb, "Patient Data", df, freeze_to_col=2, autosize_first=10)

    # Data Dictionary
    dd = build_data_dict(df)
    write_df(wb, "Data Dictionary", dd, freeze_to_col=2, autosize_first=5)
    # Wider columns for the DD readability
    dd_ws = wb["Data Dictionary"]
    dd_ws.column_dimensions["A"].width = 42
    dd_ws.column_dimensions["B"].width = 14
    dd_ws.column_dimensions["C"].width = 70
    dd_ws.column_dimensions["D"].width = 60
    dd_ws.column_dimensions["E"].width = 40

    wb.save(OUT)
    print(f"→ Wrote {OUT}  ({OUT.stat().st_size/1024:.1f} KB)")


if __name__ == "__main__":
    main()
