"""07_build_d5_size_symptoms_by_demo.py — Deliverable 5 (additive)
M038_GOITER_size_symptoms_by_demographics.xlsx

Two distinct views:
  A. GOITER SIZE (gland weight) stratified by demographics
  B. GOITER-RELATED SYMPTOMS (preop imaging + perioperative airway findings) by demographics

Demographics cross-cut: sex, age bin, race, BMI bin, era, smoking, comorbidity burden.

Important methodological note (also in Cover sheet):
  • Imaging findings (ct_*, mri_*, nlp_airway_*) are PREOPERATIVE — they describe goiter
    effects on adjacent structures BEFORE surgery and represent goiter symptomatology.
  • ops_difficult_airway is PERIOPERATIVE (captured at intubation).
  • comp_vc_paresis_confirmed / comp_vc_paralysis_confirmed in this cohort view are
    POSTOPERATIVE complications — the cohort view does NOT separately encode preop
    voice hoarseness or preop VC findings as goiter symptoms. We report them as
    "postop VC injury" with that limitation flagged.
  • proc_nlp_tracheostomy may be peri/postop; specific timing in
    proc_nlp_tracheostomy_days_from_surg.
"""
from pathlib import Path
import pandas as pd
import numpy as np
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from _stats import (cont_summary, fmt_mean_sd, fmt_median_iqr, chi2_or_fisher, chi2_table,
                    mannwhitney, ttest, fmt_p, fmt_pct, truthy_count)

HERE = Path(__file__).parent
PKG = HERE.parent
PARQUET = HERE / "m038_per_patient_v2.parquet"
OUT = PKG / "M038_GOITER_size_symptoms_by_demographics.xlsx"

HF = PatternFill("solid", fgColor="1F4E78")
SUBHF = PatternFill("solid", fgColor="D9E1F2")
HFONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHFONT = Font(name="Arial", bold=True, color="1F4E78", size=11)
BODY = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
TITLE = Font(name="Arial", size=14, bold=True, color="1F4E78")
BD = Border(*[Side(style="thin", color="CCCCCC")] * 4)
AL = Alignment(horizontal="left", vertical="center")
AR = Alignment(horizontal="right", vertical="center")
AC = Alignment(horizontal="center", vertical="center")
AW = Alignment(horizontal="left", vertical="top", wrap_text=True)


def write_header(ws, headers, row=1, fill=HF, font=HFONT):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = font; c.fill = fill; c.alignment = AC; c.border = BD


def write_row(ws, row, vals, font=BODY, bold_first=False):
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = BOLD if (bold_first and ci == 1) else font
        c.border = BD
        c.alignment = AR if isinstance(v, (int, float)) and not isinstance(v, bool) else AL


def section_row(ws, row, label, n_cols=10, fill=SUBHF, font=SUBHFONT):
    c = ws.cell(row=row, column=1, value=label)
    c.font = font; c.fill = fill; c.alignment = AL; c.border = BD
    for ci in range(2, n_cols + 1):
        cc = ws.cell(row=row, column=ci, value=""); cc.fill = fill; cc.border = BD


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------- demographic binning helpers ----------
def add_demo_bins(df):
    df = df.copy()
    # Normalize sex (cohort view stores lowercase)
    df["sex_norm"] = df["sex"].astype("string").str.strip().str.lower().map(
        {"female": "Female", "male": "Male"}).fillna("Unknown")
    # Derived: max lobe height (greatest of left/right when both available; otherwise the available one)
    lh = pd.to_numeric(df["syn_left_lobe_height_cm"], errors="coerce")
    rh = pd.to_numeric(df["syn_right_lobe_height_cm"], errors="coerce")
    df["max_lobe_height_cm"] = pd.concat([lh, rh], axis=1).max(axis=1)
    df["sum_lobe_height_cm"] = pd.concat([lh, rh], axis=1).sum(axis=1, min_count=1)

    df["age_bin"] = pd.cut(df["age_at_surgery"],
                           bins=[0, 30, 40, 50, 60, 70, 80, 200],
                           labels=["<30", "30-39", "40-49", "50-59", "60-69", "70-79", "80+"],
                           right=False)
    df["bmi_bin"] = pd.cut(df["bmi_combined"],
                           bins=[0, 18.5, 25, 30, 35, 40, 200],
                           labels=["underweight (<18.5)", "normal (18.5-<25)", "overweight (25-<30)",
                                   "obese I (30-<35)", "obese II (35-<40)", "obese III (≥40)"],
                           right=False)
    # Race buckets — collapse small categories
    def race_bucket(r):
        if pd.isna(r) or r == "None": return "Unknown / Declined"
        r = str(r).strip()
        if r == "White": return "White"
        if r in ("Black or African American", "Black", "African American"): return "Black or African American"
        if r == "Asian": return "Asian"
        if r in ("Hispanic or Latino", "Hispanic"): return "Hispanic or Latino"
        if r in ("American Indian or Alaska Native", "AIAN"): return "AIAN"
        if r in ("Native Hawaiian or Other Pacific Islander", "NHPI"): return "NHPI"
        if r in ("Other", "Multiple", "Mixed"): return "Other / Multiple"
        if r in ("Unknown or Not Reported", "Unknown", "Not Reported", "Declined"): return "Unknown / Declined"
        return "Other / Multiple"
    df["race_bucket"] = df["race"].apply(race_bucket)
    # Smoking
    def smoke(s):
        if pd.isna(s): return "(missing)"
        s = str(s).strip().lower()
        if s in ("current", "yes", "active"): return "current"
        if s in ("former", "ex-smoker", "past"): return "former"
        if s in ("never",): return "never"
        return "(other)"
    df["smoking_bin"] = df["pmhx_nlp_smoking_status"].apply(smoke)
    # Comorbidity burden
    df["comorb_bin"] = pd.cut(df["pmhx_nlp_n_comorbidities"].fillna(0),
                              bins=[-0.1, 0.5, 1.5, 3.5, 100],
                              labels=["0", "1", "2-3", "4+"])
    return df


# ---------- Cover ----------
def build_cover(wb, df):
    ws = wb.create_sheet("Cover")
    ws.cell(row=1, column=1, value="M038 — Goiter Size & Symptoms by Demographics").font = TITLE
    ws.cell(row=2, column=1, value="Additive analysis view (Deliverable 5)").font = BOLD

    rows = [
        ("Manuscript ID", "M038"),
        ("Date assembled", date.today().isoformat()),
        ("Database / cohort view",
         "thyroid_canonical_publication_v1_0 / manuscript_workspace.cohort_m038_massive_goiter_v1 (post-mig_255)"),
        ("n total / n massive", f"{len(df):,} / {int(df['is_massive'].sum()):,}"),
        ("",""),
        ("View A — THYROID SIZE & WEIGHT by demographics",
         "All available size dimensions summarized by sex, age bin, race, BMI bin, era (3-bucket and 5-yr), "
         "smoking status, and comorbidity burden. Each demographic-stratum tab stacks 7 size measures: "
         "(1) gland_weight_final_g, (2) gland_weight_total_reported_g, "
         "(3) max_lobe_height_cm (greater of L/R, derived), (4) sum_lobe_height_cm (L+R, derived), "
         "(5) syn_left_lobe_height_cm, (6) syn_right_lobe_height_cm, (7) syn_isthmus_height_cm. "
         "Cells per measure: n with measure; median (IQR); mean ± SD; three threshold % bands."),
        ("",""),
        ("View B — GOITER SYMPTOMS by demographics",
         "PREOP imaging-derived findings (substernal extension on CT/MRI; tracheal deviation/narrowing/"
         "compromise on CT) and PERIOPERATIVE airway findings (op-note NLP difficult airway; tracheostomy) "
         "by the same demographic strata. POSTOP VC paresis/paralysis included with explicit timing-window "
         "limitation note."),
        ("",""),
        ("Methodological note — preop vs postop", ""),
        ("  Preoperative imaging",
         "ct_substernal_extension_any, mri_substernal_any, ct_tracheal_deviation_any, "
         "ct_tracheal_narrowing_any, ct_airway_compromise_any — describe goiter effect on adjacent structures BEFORE surgery."),
        ("  Perioperative",
         "ops_difficult_airway (op-note NLP at intubation); proc_nlp_tracheostomy "
         "(peri/postop — specific timing in proc_nlp_tracheostomy_days_from_surg)."),
        ("  Postoperative",
         "comp_vc_paresis_confirmed / comp_vc_paralysis_confirmed / comp_rln_injury_confirmed — these are "
         "POSTOP complications. Cohort view does NOT separately encode preop voice hoarseness or preop VC paralysis "
         "as goiter symptoms (carry-forwards CF-VC-PARALYSIS-PREOP-FLAG / CF-RLN-PREOP-FLAG remain open)."),
        ("",""),
        ("Statistical approach", "Continuous (gland weight): Kruskal-Wallis across strata + Mann-Whitney pairwise where applicable. "
                                 "Categorical (symptom prevalence): chi-squared (Fisher when expected<5)."),
        ("Standing rules", "Era binning: upper-bound rule sweeps pre-1999 → 1999-2004. "
                           "Composite massive: weight≥100g OR substernal(CT/MRI) OR airway(CT)."),
    ]
    for i, (k, v) in enumerate(rows, 4):
        a = ws.cell(row=i, column=1, value=k); a.font = BOLD; a.alignment = AW
        b = ws.cell(row=i, column=2, value=v); b.font = BODY; b.alignment = AW
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 110
    for r in range(4, 4 + len(rows)):
        ws.row_dimensions[r].height = max(18, min(60, 18 * (1 + len(str(rows[r-4][1])) // 90)))


# ---------- VIEW A: SIZE by demographics (multi-measure) ----------
# Each stratum tab shows ALL these size dimensions stacked as section blocks.
SIZE_MEASURES = [
    # (col, label, unit, threshold_list, threshold_labels)
    ("gland_weight_final_g", "Gland weight (final, synoptic path)", "g",
     [100, 200, 500], ["%≥100g", "%≥200g", "%≥500g"]),
    ("gland_weight_total_reported_g", "Gland weight (total reported)", "g",
     [100, 200, 500], ["%≥100g", "%≥200g", "%≥500g"]),
    ("max_lobe_height_cm", "Max lobe height (greater of L/R)", "cm",
     [5, 7, 10], ["%≥5cm", "%≥7cm", "%≥10cm"]),
    ("sum_lobe_height_cm", "Sum lobe height (L+R)", "cm",
     [10, 14, 20], ["%≥10cm", "%≥14cm", "%≥20cm"]),
    ("syn_left_lobe_height_cm", "Left lobe height", "cm",
     [5, 7, 10], ["%≥5cm", "%≥7cm", "%≥10cm"]),
    ("syn_right_lobe_height_cm", "Right lobe height", "cm",
     [5, 7, 10], ["%≥5cm", "%≥7cm", "%≥10cm"]),
    ("syn_isthmus_height_cm", "Isthmus height", "cm",
     [1, 2, 3], ["%≥1cm", "%≥2cm", "%≥3cm"]),
]


def size_by_stratum(ws, df, stratum_col, stratum_label, ordered_cats=None):
    """Generic SIZE-by-stratum builder, looping over all SIZE_MEASURES."""
    ws.cell(row=1, column=1, value=f"Thyroid size & weight by {stratum_label}").font = TITLE
    ws.cell(row=2, column=1,
            value=f"All available thyroid size dimensions summarized within each {stratum_label} stratum. "
                  f"Test p-value: Kruskal-Wallis across all non-missing strata. "
                  f"Threshold % cells use stratum n as denominator (not n-with-measure)."
            ).font = BODY

    cats = list(ordered_cats) if ordered_cats is not None else sorted(df[stratum_col].dropna().astype(str).unique().tolist())
    if "(missing)" not in cats and df[stratum_col].isna().any():
        cats = cats + ["(missing)"]

    from scipy import stats as scipy_stats
    n_total = len(df)
    r = 4

    for col, label, unit, thresholds, thr_labels in SIZE_MEASURES:
        if col not in df.columns:
            continue
        # Section header
        section_row(ws, r, f"{label} ({unit})  [column: {col}]", n_cols=10); r += 1
        write_header(ws, ["Stratum", "n total", f"n with {col}",
                          f"Median (IQR), {unit}", f"Mean ± SD, {unit}",
                          thr_labels[0], thr_labels[1], thr_labels[2],
                          "Test", "p-value"], row=r); r += 1

        # KW across cats
        groups = []
        for cat in cats:
            sub = df[df[stratum_col].isna()] if cat == "(missing)" else df[df[stratum_col].astype(str) == str(cat)]
            x = pd.to_numeric(sub[col], errors="coerce").dropna().values
            if len(x) >= 2:
                groups.append(x)
        if len(groups) >= 2:
            try:
                kw_h, kw_p = scipy_stats.kruskal(*groups)
            except Exception:
                kw_h, kw_p = None, float("nan")
        else:
            kw_h, kw_p = None, float("nan")

        for i, cat in enumerate(cats):
            sub = df[df[stratum_col].isna()] if cat == "(missing)" else df[df[stratum_col].astype(str) == str(cat)]
            n = len(sub)
            s = cont_summary(sub[col])
            x_num = pd.to_numeric(sub[col], errors="coerce")
            ns = [int((x_num >= t).sum()) for t in thresholds]
            write_row(ws, r, [
                str(cat), n, s["n"],
                fmt_median_iqr(s), fmt_mean_sd(s),
                fmt_pct(ns[0], n) if n else "—",
                fmt_pct(ns[1], n) if n else "—",
                fmt_pct(ns[2], n) if n else "—",
                "" if i > 0 else "Kruskal-Wallis",
                "" if i > 0 else fmt_p(kw_p),
            ])
            r += 1
        # OVERALL row for this measure
        s_all = cont_summary(df[col])
        x_all = pd.to_numeric(df[col], errors="coerce")
        write_row(ws, r, ["OVERALL", n_total, s_all["n"],
                          fmt_median_iqr(s_all), fmt_mean_sd(s_all),
                          fmt_pct(int((x_all >= thresholds[0]).sum()), n_total),
                          fmt_pct(int((x_all >= thresholds[1]).sum()), n_total),
                          fmt_pct(int((x_all >= thresholds[2]).sum()), n_total),
                          "", ""], bold_first=True)
        r += 1
        # Spacer
        r += 1

    autosize(ws, [28, 12, 18, 22, 22, 12, 12, 12, 18, 14])


def build_size_views(wb, df):
    AGE_ORDER = ["<30", "30-39", "40-49", "50-59", "60-69", "70-79", "80+"]
    BMI_ORDER = ["underweight (<18.5)", "normal (18.5-<25)", "overweight (25-<30)",
                 "obese I (30-<35)", "obese II (35-<40)", "obese III (≥40)"]
    RACE_ORDER = ["White", "Black or African American", "Asian", "Hispanic or Latino",
                  "AIAN", "NHPI", "Other / Multiple", "Unknown / Declined"]
    ERA3_ORDER = ["pre-2015", "2015-2019", "2020-2025"]
    ERA6_ORDER = ["1999-2004", "2005-2009", "2010-2014", "2015-2019", "2020-2025"]
    SMOKE_ORDER = ["never", "former", "current", "(other)", "(missing)"]
    COMORB_ORDER = ["0", "1", "2-3", "4+"]

    size_by_stratum(wb.create_sheet("Size+Wt by Sex"), df, "sex_norm", "sex",
                    ordered_cats=["Female", "Male", "Unknown"])
    size_by_stratum(wb.create_sheet("Size+Wt by Age"), df, "age_bin", "age bin",
                    ordered_cats=AGE_ORDER)
    size_by_stratum(wb.create_sheet("Size+Wt by Race"), df, "race_bucket", "race",
                    ordered_cats=RACE_ORDER)
    size_by_stratum(wb.create_sheet("Size+Wt by BMI"), df, "bmi_bin", "BMI bin",
                    ordered_cats=BMI_ORDER)
    size_by_stratum(wb.create_sheet("Size+Wt by Era 3"), df, "era_bucket_3", "era (3-bucket)",
                    ordered_cats=ERA3_ORDER)
    size_by_stratum(wb.create_sheet("Size+Wt by Era 5yr"), df, "era_bucket_6", "era (5-yr)",
                    ordered_cats=ERA6_ORDER)
    size_by_stratum(wb.create_sheet("Size+Wt by Smoking"), df, "smoking_bin", "smoking status",
                    ordered_cats=SMOKE_ORDER)
    size_by_stratum(wb.create_sheet("Size+Wt by Comorbid"), df, "comorb_bin", "comorbidity burden",
                    ordered_cats=COMORB_ORDER)


# ---------- VIEW B: SYMPTOMS by demographics ----------
SYMPTOMS = [
    # (column, display_label, kind)  kind: 'imaging_pre' / 'periop' / 'postop'
    ("ct_substernal_extension_any", "Substernal extension (CT)", "imaging_pre"),
    ("mri_substernal_any", "Substernal extension (MRI)", "imaging_pre"),
    ("comp_substernal_any", "Substernal — any (CT or MRI)", "imaging_pre_derived"),
    ("ct_tracheal_deviation_any", "Tracheal deviation (CT)", "imaging_pre"),
    ("ct_tracheal_narrowing_any", "Tracheal narrowing (CT)", "imaging_pre"),
    ("ct_airway_compromise_any", "Airway compromise (CT)", "imaging_pre"),
    ("comp_airway_any", "Airway — any (CT)", "imaging_pre_derived"),
    ("nlp_airway_has_data", "Airway documented (NLP)", "imaging_pre"),
    ("ops_difficult_airway_yes", "Difficult airway at intubation (op-note)", "periop"),
    ("proc_nlp_tracheostomy", "Tracheostomy (peri/postop, NLP)", "periop"),
    ("comp_vc_paresis_confirmed", "VC paresis confirmed (postop)*", "postop"),
    ("comp_vc_paralysis_confirmed", "VC paralysis confirmed (postop)*", "postop"),
    ("comp_rln_injury_confirmed", "RLN injury confirmed (postop)*", "postop"),
]


def get_symptom_series(df, col):
    """Return a boolean Series for each symptom column, handling string-encoded ones."""
    if col == "ops_difficult_airway_yes":
        s = df["ops_difficult_airway"].astype("string").str.strip().str.lower()
        return s.str.startswith("yes", na=False) & ~s.isna()
    if col in df.columns:
        s = df[col]
        if s.dtype == bool:
            return s.fillna(False)
        return s.astype("boolean").fillna(False)
    return pd.Series([False] * len(df), index=df.index)


def symptoms_by_stratum(ws, df, stratum_col, stratum_label, ordered_cats=None, n_min=10):
    """Generic symptom-prevalence-by-stratum builder."""
    ws.cell(row=1, column=1, value=f"Goiter symptoms by {stratum_label}").font = TITLE
    ws.cell(row=2, column=1,
            value=f"Prevalence (%) of each symptom signal within each {stratum_label} stratum. "
                  "Tests: chi-squared (or Fisher exact when expected<5) across non-missing strata. "
                  "Asterisked rows (*) are POSTOP complications, not preop symptoms — included for context."
            ).font = BODY

    headers = ["Symptom (kind)"] + [str(c) for c in (ordered_cats or [])] + ["Overall", "Test", "p-value"]
    write_header(ws, headers, row=4)

    cats = ordered_cats
    # Subset frames by stratum
    cat_dfs = []
    for cat in cats:
        if cat == "(missing)":
            cat_dfs.append(df[df[stratum_col].isna()])
        else:
            cat_dfs.append(df[df[stratum_col].astype(str) == str(cat)])
    cat_ns = [len(d) for d in cat_dfs]

    r = 5
    # Denominator row
    write_row(ws, r, ["n (denominator)"] + cat_ns + [len(df), "", ""], bold_first=True); r += 1

    for col, label, kind in SYMPTOMS:
        sym_full = get_symptom_series(df, col)
        cells = []
        obs_table = []  # rows = strata, cols = [pos, neg]
        for d, n_strat in zip(cat_dfs, cat_ns):
            sym = get_symptom_series(d, col)
            n_pos = int(sym.sum())
            cells.append(fmt_pct(n_pos, n_strat) if n_strat else "—")
            obs_table.append([n_pos, max(0, n_strat - n_pos)])
        n_overall = len(df)
        n_pos_overall = int(sym_full.sum())
        # Test only if we have ≥2 strata with n>=n_min
        non_missing_obs = [row for row, n_strat in zip(obs_table, cat_ns)
                           if n_strat >= n_min]
        if len(non_missing_obs) >= 2:
            chi2v, p, dof = chi2_table(non_missing_obs)
            test_str = f"Chi2 (df={dof})" if chi2v is not None else "Chi2"
        else:
            p = float("nan"); test_str = "—"
        kind_str = {
            "imaging_pre": "preop imaging",
            "imaging_pre_derived": "preop derived",
            "periop": "perioperative",
            "postop": "postop complication",
        }[kind]
        write_row(ws, r,
                  [f"{label}  [{kind_str}]"] + cells +
                  [fmt_pct(n_pos_overall, n_overall), test_str, fmt_p(p)])
        r += 1

    autosize(ws, [44] + [16] * len(cats) + [18, 18, 14])


def build_symptom_views(wb, df):
    AGE_ORDER = ["<30", "30-39", "40-49", "50-59", "60-69", "70-79", "80+"]
    BMI_ORDER = ["underweight (<18.5)", "normal (18.5-<25)", "overweight (25-<30)",
                 "obese I (30-<35)", "obese II (35-<40)", "obese III (≥40)"]
    RACE_ORDER = ["White", "Black or African American", "Asian", "Hispanic or Latino",
                  "Other / Multiple", "Unknown / Declined"]
    ERA3_ORDER = ["pre-2015", "2015-2019", "2020-2025"]
    ERA6_ORDER = ["1999-2004", "2005-2009", "2010-2014", "2015-2019", "2020-2025"]
    SMOKE_ORDER = ["never", "former", "current"]

    symptoms_by_stratum(wb.create_sheet("Symptoms by Sex"), df, "sex_norm", "sex",
                        ordered_cats=["Female", "Male"])
    symptoms_by_stratum(wb.create_sheet("Symptoms by Age"), df, "age_bin", "age bin",
                        ordered_cats=AGE_ORDER)
    symptoms_by_stratum(wb.create_sheet("Symptoms by Race"), df, "race_bucket", "race",
                        ordered_cats=RACE_ORDER)
    symptoms_by_stratum(wb.create_sheet("Symptoms by BMI"), df, "bmi_bin", "BMI bin",
                        ordered_cats=BMI_ORDER)
    symptoms_by_stratum(wb.create_sheet("Symptoms by Era 3-bucket"), df, "era_bucket_3", "era (3-bucket)",
                        ordered_cats=ERA3_ORDER)
    symptoms_by_stratum(wb.create_sheet("Symptoms by Era 5-yr"), df, "era_bucket_6", "era (5-yr)",
                        ordered_cats=ERA6_ORDER)
    symptoms_by_stratum(wb.create_sheet("Symptoms by Smoking"), df, "smoking_bin", "smoking status",
                        ordered_cats=SMOKE_ORDER)


# ---------- View C: Composite-massive prevalence by demographic ----------
def build_massive_prevalence_by_demo(wb, df):
    ws = wb.create_sheet("Massive prevalence by demo")
    ws.cell(row=1, column=1, value="Composite-massive prevalence by demographic stratum").font = TITLE
    ws.cell(row=2, column=1,
            value="Single-page rollup: % of patients in each stratum meeting the composite-massive flag. "
                  "Useful for spotting demographic enrichment in the massive phenotype."
            ).font = BODY
    write_header(ws, ["Stratum group", "Category", "n", "n massive", "% massive", "Test", "p-value"], row=4)
    r = 5
    for label, col, order in [
        ("Sex", "sex_norm", ["Female", "Male"]),
        ("Age bin", "age_bin", ["<30", "30-39", "40-49", "50-59", "60-69", "70-79", "80+"]),
        ("Race", "race_bucket", ["White", "Black or African American", "Asian", "Hispanic or Latino",
                                  "Other / Multiple", "Unknown / Declined"]),
        ("BMI bin", "bmi_bin", ["underweight (<18.5)", "normal (18.5-<25)", "overweight (25-<30)",
                                "obese I (30-<35)", "obese II (35-<40)", "obese III (≥40)"]),
        ("Era (3-bucket)", "era_bucket_3", ["pre-2015", "2015-2019", "2020-2025"]),
        ("Smoking", "smoking_bin", ["never", "former", "current"]),
        ("Comorbidity burden", "comorb_bin", ["0", "1", "2-3", "4+"]),
    ]:
        section_row(ws, r, label, n_cols=7); r += 1
        # Build observation table for chi2 across strata
        obs = []; cell_results = []
        for cat in order:
            sub = df[df[col].astype(str) == str(cat)]
            n = len(sub); n_mas = int(sub["is_massive"].sum())
            obs.append([n_mas, max(0, n - n_mas)])
            cell_results.append((cat, n, n_mas))
        # Strip empty rows for chi2
        obs_nz = [row for row in obs if sum(row) >= 10]
        if len(obs_nz) >= 2:
            chi2v, p, dof = chi2_table(obs_nz)
            p_first = fmt_p(p); test = f"Chi2 (df={dof})"
        else:
            p_first = "—"; test = "—"
        for i, (cat, n, n_mas) in enumerate(cell_results):
            write_row(ws, r, [
                "", cat, n, n_mas,
                f"{100*n_mas/n:.1f}%" if n else "—",
                test if i == 0 else "",
                p_first if i == 0 else "",
            ]); r += 1
    autosize(ws, [22, 28, 12, 14, 14, 18, 14])


# ---------- View D: Symptom co-occurrence within massive arm ----------
def build_symptom_cooccurrence(wb, df):
    ws = wb.create_sheet("Symptom co-occurrence")
    ws.cell(row=1, column=1, value="Symptom co-occurrence within MASSIVE arm").font = TITLE
    ws.cell(row=2, column=1,
            value="Within the massive arm (n=2,501): pairwise overlap of preop imaging symptoms "
                  "and perioperative airway findings. Cell = patients with BOTH (row & column). "
                  "Diagonal = solo prevalence."
            ).font = BODY
    items = [
        ("Substernal CT", "ct_substernal_extension_any"),
        ("Substernal MRI", "mri_substernal_any"),
        ("Tracheal deviation CT", "ct_tracheal_deviation_any"),
        ("Tracheal narrowing CT", "ct_tracheal_narrowing_any"),
        ("Airway compromise CT", "ct_airway_compromise_any"),
        ("Difficult airway (op)", "ops_difficult_airway_yes"),
        ("Tracheostomy", "proc_nlp_tracheostomy"),
        ("VC paralysis postop*", "comp_vc_paralysis_confirmed"),
    ]
    massive = df[df["is_massive"]].copy()
    headers = ["Row \\ Col"] + [lbl for lbl, _ in items]
    write_header(ws, headers, row=4)
    series = {col: get_symptom_series(massive, col) for _, col in items}
    r = 5
    for lbl_r, col_r in items:
        cells = [lbl_r]
        sr = series[col_r]
        for lbl_c, col_c in items:
            sc = series[col_c]
            if col_r == col_c:
                cells.append(int(sr.sum()))
            else:
                cells.append(int((sr & sc).sum()))
        write_row(ws, r, cells, bold_first=True)
        r += 1
    r += 1
    write_row(ws, r, ["Note", "Diagonal cell = solo prevalence in massive arm. "
                              "Off-diagonal = overlap (BOTH row & column TRUE). "
                              "* = postop complication, not preop symptom."])
    autosize(ws, [28] + [18] * len(items))


# ---------- main ----------
def main():
    df0 = pd.read_parquet(PARQUET)
    df = add_demo_bins(df0)
    print(f"→ Loaded {len(df):,} rows × {len(df.columns)} cols (with demo bins)")

    wb = Workbook(); wb.remove(wb.active)
    build_cover(wb, df)
    build_size_views(wb, df)
    build_symptom_views(wb, df)
    build_massive_prevalence_by_demo(wb, df)
    build_symptom_cooccurrence(wb, df)

    wb.save(OUT)
    print(f"→ Wrote {OUT}  ({OUT.stat().st_size/1024:.1f} KB)")


if __name__ == "__main__":
    main()
