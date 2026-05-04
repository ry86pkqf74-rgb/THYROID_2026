"""04_build_d3_tables_figures.py — Deliverable 3
M038_GOITER_tables_figures.xlsx — Pub-formatted tables + figure data.

Tabs:
  1-5: Pub Table 1-5 (Demographics / Pathology / Surgical / Complications / Era)
  6-9: Figure 1-4 underlying data (Venn / era trend / complications bar / component coverage)
  10-15: Supplement S1-S6
"""
from pathlib import Path
import pandas as pd
import numpy as np
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from _stats import (cont_summary, fmt_mean_sd, fmt_median_iqr, chi2_or_fisher, chi2_table,
                    mannwhitney, ttest, rr_ci, fmt_p, fmt_rr_ci, fmt_pct, count_pct, truthy_count)

HERE = Path(__file__).parent
PKG = HERE.parent
PARQUET = HERE / "m038_per_patient_v2.parquet"
OUT = PKG / "M038_GOITER_tables_figures.xlsx"

HF = PatternFill("solid", fgColor="1F4E78")
SUBHF = PatternFill("solid", fgColor="D9E1F2")
HFONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHFONT = Font(name="Arial", bold=True, color="1F4E78", size=11)
BODY = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
TITLE = Font(name="Arial", size=14, bold=True, color="1F4E78")
BD = Border(*[Side(style="thin", color="CCCCCC")] * 4)
AL = Alignment(horizontal="left", vertical="center")
AR = Alignment(horizontal="right", vertical="center")
AC = Alignment(horizontal="center", vertical="center")
AW = Alignment(horizontal="left", vertical="top", wrap_text=True)


def write_header(ws, headers, row=1, fill=HF, font=HFONT):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = font; c.fill = fill; c.alignment = AC; c.border = BD


def write_row(ws, row, vals, font=BODY, bold_first=False):
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = BOLD if (bold_first and ci == 1) else font
        c.border = BD
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            c.alignment = AR
        else:
            c.alignment = AL


def section_row(ws, row, label, n_cols=6, fill=SUBHF, font=SUBHFONT):
    c = ws.cell(row=row, column=1, value=label)
    c.font = font; c.fill = fill; c.alignment = AL; c.border = BD
    for ci in range(2, n_cols + 1):
        cc = ws.cell(row=row, column=ci, value=""); cc.fill = fill; cc.border = BD


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------- Pub-formatted tables ----------
def pub_table1(wb, df):
    ws = wb.create_sheet("Table 1")
    ws.cell(row=1, column=1, value="Table 1. Patient Demographics and Comorbidities").font = TITLE
    ws.cell(row=2, column=1,
            value="Stratified by composite-massive status. Continuous variables: median (IQR). "
                  "Categorical: n (%). Tests: Mann-Whitney U for continuous; chi-squared (Fisher when expected<5)."
            ).font = BODY
    write_header(ws, ["Variable", "Massive (n=2,501)", "Non-massive (n=8,370)", "p-value"], row=4)
    r = 5
    n_m = int(df["is_massive"].sum()); n_nm = int((~df["is_massive"]).sum())
    massive = df[df["is_massive"]]; non_massive = df[~df["is_massive"]]

    # Age
    sm = cont_summary(massive["age_at_surgery"]); sn = cont_summary(non_massive["age_at_surgery"])
    _, p = mannwhitney(massive["age_at_surgery"], non_massive["age_at_surgery"])
    write_row(ws, r, ["Age at surgery, y, median (IQR)", fmt_median_iqr(sm), fmt_median_iqr(sn), fmt_p(p)]); r += 1

    # Sex
    a = int((massive["sex"].str.lower() == "female").sum())
    b = int((non_massive["sex"].str.lower() == "female").sum())
    _, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
    write_row(ws, r, ["Female sex, n (%)", fmt_pct(a, n_m), fmt_pct(b, n_nm), fmt_p(p)]); r += 1

    # Race summary rows (top buckets)
    section_row(ws, r, "Race, n (%)", n_cols=4); r += 1
    races_top = ["White", "Black or African American", "Asian", "Other / Multiple", "Unknown / Declined"]
    for race_label in races_top:
        if race_label == "Other / Multiple":
            mask_m = massive["race"].astype(str).isin(["Other", "Multiple", "Mixed"])
            mask_nm = non_massive["race"].astype(str).isin(["Other", "Multiple", "Mixed"])
        elif race_label == "Unknown / Declined":
            mask_m = massive["race"].astype(str).isin(["Unknown", "Declined", "Not Reported"])
            mask_nm = non_massive["race"].astype(str).isin(["Unknown", "Declined", "Not Reported"])
        else:
            mask_m = massive["race"].astype(str) == race_label
            mask_nm = non_massive["race"].astype(str) == race_label
        a = int(mask_m.sum()); b = int(mask_nm.sum())
        _, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
        write_row(ws, r, [f"  {race_label}", fmt_pct(a, n_m), fmt_pct(b, n_nm), fmt_p(p)]); r += 1

    # BMI
    sm = cont_summary(massive["bmi_combined"]); sn = cont_summary(non_massive["bmi_combined"])
    _, p = mannwhitney(massive["bmi_combined"], non_massive["bmi_combined"])
    write_row(ws, r, ["BMI, kg/m², median (IQR)", fmt_median_iqr(sm), fmt_median_iqr(sn), fmt_p(p)]); r += 1

    # Comorbidities
    section_row(ws, r, "Comorbidities, NLP-extracted, n (%)", n_cols=4); r += 1
    for label, col in [("Hypertension", "pmhx_nlp_hypertension"),
                       ("Diabetes", "pmhx_nlp_diabetes"),
                       ("Coronary artery disease", "pmhx_nlp_cad"),
                       ("Chronic kidney disease", "pmhx_nlp_ckd"),
                       ("COPD", "pmhx_nlp_copd")]:
        a = int(massive[col].sum()); b = int(non_massive[col].sum())
        _, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
        write_row(ws, r, [f"  {label}", fmt_pct(a, n_m), fmt_pct(b, n_nm), fmt_p(p)]); r += 1
    sm = cont_summary(massive["pmhx_nlp_n_comorbidities"])
    sn = cont_summary(non_massive["pmhx_nlp_n_comorbidities"])
    _, p = mannwhitney(massive["pmhx_nlp_n_comorbidities"], non_massive["pmhx_nlp_n_comorbidities"])
    write_row(ws, r, ["  Mean # comorbidities", f"{sm['mean']:.2f}", f"{sn['mean']:.2f}", fmt_p(p)]); r += 1

    # Thyroid history
    section_row(ws, r, "Thyroid / surgical history, n (%)", n_cols=4); r += 1
    for label, col in [("Graves disease", "syn_graves"),
                       ("Hashimoto thyroiditis", "syn_hashimoto"),
                       ("Prior thyroidectomy", "pshx_nlp_prior_thyroidectomy"),
                       ("Prior neck surgery", "pshx_nlp_prior_neck_surgery")]:
        a = int(massive[col].sum()); b = int(non_massive[col].sum())
        _, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
        write_row(ws, r, [f"  {label}", fmt_pct(a, n_m), fmt_pct(b, n_nm), fmt_p(p)]); r += 1

    # ASA
    section_row(ws, r, "ASA class (NSQIP-linked subset), n (%)", n_cols=4); r += 1
    massive_ns = massive[massive["nsqip_asa_class"].notna()]
    non_massive_ns = non_massive[non_massive["nsqip_asa_class"].notna()]
    n_ns_m = len(massive_ns); n_ns_nm = len(non_massive_ns)
    write_row(ws, r, [f"  NSQIP-linked denominator", n_ns_m, n_ns_nm, ""]); r += 1
    a = int(massive_ns["nsqip_asa_class"].astype(str).isin(["3", "4", "5"]).sum())
    b = int(non_massive_ns["nsqip_asa_class"].astype(str).isin(["3", "4", "5"]).sum())
    _, p = chi2_or_fisher(a, n_ns_m - a, b, n_ns_nm - b)
    write_row(ws, r, ["  ASA III-V", fmt_pct(a, n_ns_m), fmt_pct(b, n_ns_nm), fmt_p(p)]); r += 1

    # Follow-up
    sm = cont_summary(massive["followup_years"]); sn = cont_summary(non_massive["followup_years"])
    _, p = mannwhitney(massive["followup_years"], non_massive["followup_years"])
    write_row(ws, r, ["Follow-up, y, median (IQR)", fmt_median_iqr(sm), fmt_median_iqr(sn), fmt_p(p)]); r += 1

    autosize(ws, [44, 22, 24, 14])


def pub_table2(wb, df):
    ws = wb.create_sheet("Table 2")
    ws.cell(row=1, column=1, value="Table 2. Pathology").font = TITLE
    write_header(ws, ["Variable", "Massive (n=2,501)", "Non-massive (n=8,370)", "p-value"], row=3)
    r = 4
    n_m = int(df["is_massive"].sum()); n_nm = int((~df["is_massive"]).sum())
    massive = df[df["is_massive"]]; non_massive = df[~df["is_massive"]]

    a = int(massive["is_malignant"].sum()); b = int(non_massive["is_malignant"].sum())
    _, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
    write_row(ws, r, ["Malignant, n (%)", fmt_pct(a, n_m), fmt_pct(b, n_nm), fmt_p(p)]); r += 1
    a = int(massive["bilateral_path_flag"].sum()); b = int(non_massive["bilateral_path_flag"].sum())
    _, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
    write_row(ws, r, ["Bilateral disease (pathology), n (%)", fmt_pct(a, n_m), fmt_pct(b, n_nm), fmt_p(p)]); r += 1

    section_row(ws, r, "Histology distribution, n (%)", n_cols=4); r += 1
    histos = df["histology_final"].fillna("(missing)").value_counts().head(10).index.tolist()
    for h in histos:
        a = int((massive["histology_final"].fillna("(missing)") == h).sum())
        b = int((non_massive["histology_final"].fillna("(missing)") == h).sum())
        _, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
        write_row(ws, r, [f"  {h}", fmt_pct(a, n_m), fmt_pct(b, n_nm), fmt_p(p)]); r += 1

    sm = cont_summary(massive["gland_weight_final_g"])
    sn = cont_summary(non_massive["gland_weight_final_g"])
    _, p = mannwhitney(massive["gland_weight_final_g"], non_massive["gland_weight_final_g"])
    write_row(ws, r, ["Gland weight, g, median (IQR)", fmt_median_iqr(sm), fmt_median_iqr(sn), fmt_p(p)]); r += 1

    autosize(ws, [44, 22, 24, 14])


def pub_table3(wb, df):
    ws = wb.create_sheet("Table 3")
    ws.cell(row=1, column=1, value="Table 3. Surgical and Operative Context").font = TITLE
    write_header(ws, ["Variable", "Massive (n=2,501)", "Non-massive (n=8,370)", "p-value"], row=3)
    r = 4
    n_m = int(df["is_massive"].sum()); n_nm = int((~df["is_massive"]).sum())
    massive = df[df["is_massive"]]; non_massive = df[~df["is_massive"]]

    section_row(ws, r, "Procedure type, n (%)", n_cols=4); r += 1
    a = int((massive["surg_procedure_type"].fillna("") == "total_thyroidectomy").sum())
    b = int((non_massive["surg_procedure_type"].fillna("") == "total_thyroidectomy").sum())
    _, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
    write_row(ws, r, ["  Total thyroidectomy", fmt_pct(a, n_m), fmt_pct(b, n_nm), fmt_p(p)]); r += 1
    a = int((massive["surg_procedure_type"].fillna("") == "hemithyroidectomy").sum())
    b = int((non_massive["surg_procedure_type"].fillna("") == "hemithyroidectomy").sum())
    _, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
    write_row(ws, r, ["  Hemithyroidectomy", fmt_pct(a, n_m), fmt_pct(b, n_nm), fmt_p(p)]); r += 1
    a = int((~massive["surg_procedure_type"].astype(str).isin(["total_thyroidectomy", "hemithyroidectomy", "nan", "None", ""])).sum())
    b = int((~non_massive["surg_procedure_type"].astype(str).isin(["total_thyroidectomy", "hemithyroidectomy", "nan", "None", ""])).sum())
    _, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
    write_row(ws, r, ["  Other / completion / unspecified", fmt_pct(a, n_m), fmt_pct(b, n_nm), fmt_p(p)]); r += 1

    section_row(ws, r, "Operative context (NSQIP-linked subset)", n_cols=4); r += 1
    massive_ns = massive[massive["nsqip_inpatient_outpatient"].notna()]
    non_massive_ns = non_massive[non_massive["nsqip_inpatient_outpatient"].notna()]
    n_ns_m = len(massive_ns); n_ns_nm = len(non_massive_ns)
    write_row(ws, r, ["  NSQIP-linked n", n_ns_m, n_ns_nm, ""]); r += 1

    sm = cont_summary(massive_ns["nsqip_operative_duration_min"])
    sn = cont_summary(non_massive_ns["nsqip_operative_duration_min"])
    _, p = mannwhitney(massive_ns["nsqip_operative_duration_min"], non_massive_ns["nsqip_operative_duration_min"])
    write_row(ws, r, ["  Operative duration, min, median (IQR)", fmt_median_iqr(sm), fmt_median_iqr(sn), fmt_p(p)]); r += 1
    sm = cont_summary(massive_ns["nsqip_length_of_stay_days"])
    sn = cont_summary(non_massive_ns["nsqip_length_of_stay_days"])
    _, p = mannwhitney(massive_ns["nsqip_length_of_stay_days"], non_massive_ns["nsqip_length_of_stay_days"])
    write_row(ws, r, ["  Length of stay, d, median (IQR)", fmt_median_iqr(sm), fmt_median_iqr(sn), fmt_p(p)]); r += 1

    a = truthy_count(massive_ns["nsqip_central_neck_dissection"])
    b = truthy_count(non_massive_ns["nsqip_central_neck_dissection"])
    _, p = chi2_or_fisher(a, n_ns_m - a, b, n_ns_nm - b)
    write_row(ws, r, ["  Central neck dissection, n (%)", fmt_pct(a, n_ns_m), fmt_pct(b, n_ns_nm), fmt_p(p)]); r += 1
    a = truthy_count(massive_ns["nsqip_lateral_neck_dissection"])
    b = truthy_count(non_massive_ns["nsqip_lateral_neck_dissection"])
    _, p = chi2_or_fisher(a, n_ns_m - a, b, n_ns_nm - b)
    write_row(ws, r, ["  Lateral neck dissection, n (%)", fmt_pct(a, n_ns_m), fmt_pct(b, n_ns_nm), fmt_p(p)]); r += 1
    a = truthy_count(massive_ns["nsqip_drain_usage"])
    b = truthy_count(non_massive_ns["nsqip_drain_usage"])
    _, p = chi2_or_fisher(a, n_ns_m - a, b, n_ns_nm - b)
    write_row(ws, r, ["  Drain placed, n (%)", fmt_pct(a, n_ns_m), fmt_pct(b, n_ns_nm), fmt_p(p)]); r += 1

    a = int((massive_ns["nsqip_inpatient_outpatient"].astype(str).str.lower() == "inpatient").sum())
    b = int((non_massive_ns["nsqip_inpatient_outpatient"].astype(str).str.lower() == "inpatient").sum())
    _, p = chi2_or_fisher(a, n_ns_m - a, b, n_ns_nm - b)
    write_row(ws, r, ["  Inpatient, n (%)", fmt_pct(a, n_ns_m), fmt_pct(b, n_ns_nm), fmt_p(p)]); r += 1
    a = int((massive_ns["nsqip_same_day_discharge_flag"] == 1).sum())
    b = int((non_massive_ns["nsqip_same_day_discharge_flag"] == 1).sum())
    _, p = chi2_or_fisher(a, n_ns_m - a, b, n_ns_nm - b)
    write_row(ws, r, ["  Same-day discharge, n (%)", fmt_pct(a, n_ns_m), fmt_pct(b, n_ns_nm), fmt_p(p)]); r += 1

    a = int(massive["proc_nlp_tracheostomy"].sum())
    b = int(non_massive["proc_nlp_tracheostomy"].sum())
    _, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
    write_row(ws, r, ["Tracheostomy (NLP), n (%)", fmt_pct(a, n_m), fmt_pct(b, n_nm), fmt_p(p)]); r += 1

    autosize(ws, [44, 22, 24, 14])


def pub_table4(wb, df):
    ws = wb.create_sheet("Table 4")
    ws.cell(row=1, column=1, value="Table 4. Strict-Definition Postoperative Complications").font = TITLE
    ws.cell(row=2, column=1,
            value="Per standing rule: hypoparathyroidism split into transient (<6 mo) and permanent (>6 mo); "
                  "preexisting (preop) reported separately. Hypocalcemia preexisting reported separately."
            ).font = BODY
    write_header(ws, ["Complication", "Massive n (%)", "Non-massive n (%)", "RR (95% CI)", "p-value"], row=4)
    r = 5
    n_m = int(df["is_massive"].sum()); n_nm = int((~df["is_massive"]).sum())
    massive = df[df["is_massive"]]; non_massive = df[~df["is_massive"]]

    def comp_row(label, a, b):
        nonlocal r
        rr, lo, hi = rr_ci(a, n_m, b, n_nm)
        _, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
        write_row(ws, r, [label, fmt_pct(a, n_m), fmt_pct(b, n_nm),
                          fmt_rr_ci(rr, lo, hi), fmt_p(p)])
        r += 1

    comp_row("Any confirmed complication", int(massive["any_confirmed_complication_flag"].sum()),
             int(non_massive["any_confirmed_complication_flag"].sum()))
    comp_row("Hematoma", int(massive["comp_hematoma_confirmed"].sum()),
             int(non_massive["comp_hematoma_confirmed"].sum()))
    comp_row("Seroma", int(massive["comp_seroma_confirmed"].sum()),
             int(non_massive["comp_seroma_confirmed"].sum()))
    comp_row("Chyle leak", int(massive["comp_chyle_leak_confirmed"].sum()),
             int(non_massive["comp_chyle_leak_confirmed"].sum()))
    comp_row("RLN injury", int(massive["comp_rln_injury_confirmed"].sum()),
             int(non_massive["comp_rln_injury_confirmed"].sum()))
    comp_row("Vocal cord paralysis", int(massive["comp_vc_paralysis_confirmed"].sum()),
             int(non_massive["comp_vc_paralysis_confirmed"].sum()))
    comp_row("Hypocalcemia (postop confirmed)", int(massive["comp_hypocalcemia_confirmed"].sum()),
             int(non_massive["comp_hypocalcemia_confirmed"].sum()))
    comp_row("Hypocalcemia, preexisting (preop)", int(massive["hca_preop_flag"].sum()),
             int(non_massive["hca_preop_flag"].sum()))
    comp_row("Hypoparathyroidism, transient (<6 mo)",
             int((massive["comp_hypoparathyroidism_confirmed"] & massive["comp_hypoparathyroidism_transient"]).sum()),
             int((non_massive["comp_hypoparathyroidism_confirmed"] & non_massive["comp_hypoparathyroidism_transient"]).sum()))
    comp_row("Hypoparathyroidism, permanent (>6 mo)",
             int((massive["comp_hypoparathyroidism_confirmed"] & massive["comp_hypoparathyroidism_permanent"]).sum()),
             int((non_massive["comp_hypoparathyroidism_confirmed"] & non_massive["comp_hypoparathyroidism_permanent"]).sum()))
    comp_row("Hypoparathyroidism, preexisting (preop)",
             int(massive["comp_hypoparathyroidism_preexisting"].sum()),
             int(non_massive["comp_hypoparathyroidism_preexisting"].sum()))
    comp_row("Tracheostomy", int(massive["proc_nlp_tracheostomy"].sum()),
             int(non_massive["proc_nlp_tracheostomy"].sum()))
    comp_row("Mortality (definitive)", int(massive["comp_mortality_definitive"].sum()),
             int(non_massive["comp_mortality_definitive"].sum()))

    autosize(ws, [44, 18, 22, 22, 12])


def pub_table5(wb, df):
    ws = wb.create_sheet("Table 5")
    ws.cell(row=1, column=1, value="Table 5. Composite-Massive Prevalence by Era").font = TITLE
    ws.cell(row=2, column=1,
            value="Era binning: upper-bound rule sweeps pre-1999 surgery dates into 1999-2004 bucket. "
                  "Trend test: chi-squared across 3-bucket eras (pre-2015 / 2015-2019 / 2020-2025)."
            ).font = BODY
    write_header(ws, ["Era", "n total", "n massive", "% massive (95% CI)",
                      "n weight≥100g (%)", "n substernal (%)", "n airway (%)"], row=4)
    r = 5
    section_row(ws, r, "5-year buckets", n_cols=7); r += 1
    for era in ["1999-2004", "2005-2009", "2010-2014", "2015-2019", "2020-2025"]:
        sub = df[df["era_bucket_6"] == era]
        n = len(sub); n_mas = int(sub["is_massive"].sum())
        # Wilson 95% CI on prevalence
        from math import sqrt
        if n > 0:
            phat = n_mas / n
            z = 1.96
            denom = 1 + z**2/n
            centre = (phat + z**2/(2*n)) / denom
            half = z * sqrt((phat*(1-phat) + z**2/(4*n))/n) / denom
            ci_lo = max(0, centre - half) * 100
            ci_hi = min(1, centre + half) * 100
            ci_str = f"{100*phat:.1f}% ({ci_lo:.1f}–{ci_hi:.1f})"
        else:
            ci_str = "—"
        n_w = int(sub["comp_weight_ge100"].sum())
        n_s = int(sub["comp_substernal_any"].sum())
        n_a = int(sub["comp_airway_any"].sum())
        write_row(ws, r, [era, n, n_mas, ci_str,
                          f"{n_w} ({100*n_w/n:.1f}%)" if n else "—",
                          f"{n_s} ({100*n_s/n:.1f}%)" if n else "—",
                          f"{n_a} ({100*n_a/n:.1f}%)" if n else "—"])
        r += 1
    sub = df[df["era_bucket_6"] == "unknown"]
    if len(sub):
        n = len(sub); n_mas = int(sub["is_massive"].sum())
        write_row(ws, r, ["unknown (no surg date)", n, n_mas, f"{100*n_mas/n:.1f}%",
                          int(sub["comp_weight_ge100"].sum()),
                          int(sub["comp_substernal_any"].sum()),
                          int(sub["comp_airway_any"].sum())]); r += 1

    section_row(ws, r, "3-bucket headline (manuscript)", n_cols=7); r += 1
    for era in ["pre-2015", "2015-2019", "2020-2025"]:
        sub = df[df["era_bucket_3"] == era]
        n = len(sub); n_mas = int(sub["is_massive"].sum())
        write_row(ws, r, [era, n, n_mas, f"{100*n_mas/n:.1f}%",
                          int(sub["comp_weight_ge100"].sum()),
                          int(sub["comp_substernal_any"].sum()),
                          int(sub["comp_airway_any"].sum())])
        r += 1
    obs = []
    for era in ["pre-2015", "2015-2019", "2020-2025"]:
        sub = df[df["era_bucket_3"] == era]
        obs.append([int(sub["is_massive"].sum()), int((~sub["is_massive"]).sum())])
    chi2v, p_overall, dof = chi2_table(obs)
    write_row(ws, r, ["Chi-squared trend across 3 eras", "", "",
                      f"Chi2={chi2v:.1f}, df={dof}, p={fmt_p(p_overall)}", "", "", ""])
    autosize(ws, [22, 12, 14, 22, 20, 22, 18])


# ---------- Figure underlying data tabs ----------
def fig1_venn(wb, df):
    ws = wb.create_sheet("Fig 1 Venn")
    ws.cell(row=1, column=1, value="Figure 1 — Composite-Flag Venn Diagram (massive arm)").font = TITLE
    ws.cell(row=2, column=1, value="Counts of patients meeting weight, substernal, and airway component flags within massive arm (n=2,501)."
            ).font = BODY
    write_header(ws, ["Region", "n", "% of massive", "Definition"], row=4)
    massive = df[df["is_massive"]]
    n_m = len(massive)
    w = massive["comp_weight_ge100"]; s = massive["comp_substernal_any"]; a = massive["comp_airway_any"]
    rows = [
        ("Weight only", int((w & ~s & ~a).sum()), "Weight≥100g; no substernal; no airway"),
        ("Substernal only", int((s & ~w & ~a).sum()), "Substernal (CT/MRI); no weight; no airway"),
        ("Airway only", int((a & ~w & ~s).sum()), "Airway compromise (CT); no weight; no substernal"),
        ("Weight ∧ Substernal", int((w & s & ~a).sum()), "Weight + substernal; no airway"),
        ("Weight ∧ Airway", int((w & a & ~s).sum()), "Weight + airway; no substernal"),
        ("Substernal ∧ Airway", int((s & a & ~w).sum()), "Substernal + airway; no weight"),
        ("All three (W∧S∧A)", int((w & s & a).sum()), "Weight + substernal + airway"),
    ]
    r = 5
    for region, n, defn in rows:
        write_row(ws, r, [region, n, f"{100*n/n_m:.1f}%", defn]); r += 1
    write_row(ws, r, ["Massive total", n_m, "100.0%", "Sum of all 7 mutually-exclusive regions"]); r += 1
    autosize(ws, [22, 10, 16, 60])


def fig2_era(wb, df):
    ws = wb.create_sheet("Fig 2 Era Trend")
    ws.cell(row=1, column=1, value="Figure 2 — Composite-Massive Prevalence by Era").font = TITLE
    write_header(ws, ["Era", "n total", "n massive", "% massive", "% massive 95% CI lo", "% massive 95% CI hi"], row=3)
    r = 4
    from math import sqrt
    for era in ["1999-2004", "2005-2009", "2010-2014", "2015-2019", "2020-2025"]:
        sub = df[df["era_bucket_6"] == era]
        n = len(sub); n_mas = int(sub["is_massive"].sum())
        if n == 0: continue
        phat = n_mas / n; z = 1.96
        denom = 1 + z**2/n
        centre = (phat + z**2/(2*n)) / denom
        half = z * sqrt((phat*(1-phat) + z**2/(4*n))/n) / denom
        lo = max(0, centre - half) * 100; hi = min(1, centre + half) * 100
        write_row(ws, r, [era, n, n_mas, round(100*phat, 2), round(lo, 2), round(hi, 2)]); r += 1
    autosize(ws, [16, 12, 14, 14, 18, 18])


def fig3_complications(wb, df):
    ws = wb.create_sheet("Fig 3 Complications")
    ws.cell(row=1, column=1, value="Figure 3 — Complication Rates: Massive vs Non-massive").font = TITLE
    write_header(ws, ["Complication", "Massive %", "Non-massive %", "RR", "RR 95% CI lo", "RR 95% CI hi"], row=3)
    r = 4
    n_m = int(df["is_massive"].sum()); n_nm = int((~df["is_massive"]).sum())
    massive = df[df["is_massive"]]; non_massive = df[~df["is_massive"]]
    items = [
        ("Any confirmed", massive["any_confirmed_complication_flag"].sum(), non_massive["any_confirmed_complication_flag"].sum()),
        ("Hematoma", massive["comp_hematoma_confirmed"].sum(), non_massive["comp_hematoma_confirmed"].sum()),
        ("RLN injury", massive["comp_rln_injury_confirmed"].sum(), non_massive["comp_rln_injury_confirmed"].sum()),
        ("VC paralysis", massive["comp_vc_paralysis_confirmed"].sum(), non_massive["comp_vc_paralysis_confirmed"].sum()),
        ("HypoPT transient",
         (massive["comp_hypoparathyroidism_confirmed"] & massive["comp_hypoparathyroidism_transient"]).sum(),
         (non_massive["comp_hypoparathyroidism_confirmed"] & non_massive["comp_hypoparathyroidism_transient"]).sum()),
        ("HypoPT permanent",
         (massive["comp_hypoparathyroidism_confirmed"] & massive["comp_hypoparathyroidism_permanent"]).sum(),
         (non_massive["comp_hypoparathyroidism_confirmed"] & non_massive["comp_hypoparathyroidism_permanent"]).sum()),
        ("Tracheostomy", massive["proc_nlp_tracheostomy"].sum(), non_massive["proc_nlp_tracheostomy"].sum()),
        ("Mortality", massive["comp_mortality_definitive"].sum(), non_massive["comp_mortality_definitive"].sum()),
    ]
    for label, a, b in items:
        a = int(a); b = int(b)
        rr, lo, hi = rr_ci(a, n_m, b, n_nm)
        rr_v = round(rr, 3) if rr == rr else None
        lo_v = round(lo, 3) if lo == lo else None
        hi_v = round(hi, 3) if hi == hi else None
        write_row(ws, r, [label, round(100*a/n_m, 2), round(100*b/n_nm, 2),
                          rr_v, lo_v, hi_v]); r += 1
    autosize(ws, [22, 14, 18, 12, 16, 16])


def fig4_component_coverage(wb, df):
    ws = wb.create_sheet("Fig 4 Component Coverage")
    ws.cell(row=1, column=1, value="Figure 4 — Component Coverage Across Eras").font = TITLE
    ws.cell(row=2, column=1, value="Documentation prevalence (cohort-wide) of each massive component by era.").font = BODY
    write_header(ws, ["Era", "n total", "% weight≥100g", "% substernal", "% airway"], row=4)
    r = 5
    for era in ["1999-2004", "2005-2009", "2010-2014", "2015-2019", "2020-2025"]:
        sub = df[df["era_bucket_6"] == era]
        n = len(sub)
        if n == 0: continue
        write_row(ws, r, [era, n,
                          round(100*sub["comp_weight_ge100"].mean(), 2),
                          round(100*sub["comp_substernal_any"].mean(), 2),
                          round(100*sub["comp_airway_any"].mean(), 2)])
        r += 1
    autosize(ws, [16, 14, 18, 18, 18])


# ---------- Supplement S1-S6 ----------
def supp_s1(wb, df):
    ws = wb.create_sheet("Supp S1 Cohort Defn")
    ws.cell(row=1, column=1, value="Supplementary Table S1. Cohort Definition & Composite-Flag Components").font = TITLE
    write_header(ws, ["Component", "Source columns", "Operator", "n positive", "% of cohort"], row=3)
    n = len(df)
    rows = [
        ("Weight ≥100g",
         "gland_weight_final_g",
         "gland_weight_final_g >= 100",
         int(df["comp_weight_ge100"].sum()),
         f"{100*df['comp_weight_ge100'].mean():.1f}%"),
        ("Substernal extension",
         "ct_substernal_extension_any OR mri_substernal_any",
         "any TRUE",
         int(df["comp_substernal_any"].sum()),
         f"{100*df['comp_substernal_any'].mean():.1f}%"),
        ("Airway compromise",
         "ct_tracheal_deviation_any OR ct_tracheal_narrowing_any OR ct_airway_compromise_any",
         "any TRUE",
         int(df["comp_airway_any"].sum()),
         f"{100*df['comp_airway_any'].mean():.1f}%"),
        ("Composite massive (any of above)",
         "Disjunction of above",
         "OR",
         int(df["is_massive"].sum()),
         f"{100*df['is_massive'].mean():.1f}%"),
        ("Cohort total",
         "manuscript_workspace.cohort_m038_massive_goiter_v1",
         "—",
         n, "100.0%"),
    ]
    r = 4
    for row in rows:
        write_row(ws, r, row); r += 1
    autosize(ws, [30, 60, 28, 14, 14])


def supp_s2(wb, df):
    ws = wb.create_sheet("Supp S2 Component Era")
    ws.cell(row=1, column=1, value="Supplementary Table S2. Component Coverage by Era × Arm").font = TITLE
    write_header(ws, ["Era", "Arm", "n", "% weight", "% substernal", "% airway"], row=3)
    r = 4
    for era in ["1999-2004", "2005-2009", "2010-2014", "2015-2019", "2020-2025"]:
        for arm_label, arm_mask in [("Massive", df["is_massive"]), ("Non-massive", ~df["is_massive"])]:
            sub = df[(df["era_bucket_6"] == era) & arm_mask]
            n = len(sub)
            if n == 0: continue
            write_row(ws, r, [era, arm_label, n,
                              round(100*sub["comp_weight_ge100"].mean(), 1),
                              round(100*sub["comp_substernal_any"].mean(), 1),
                              round(100*sub["comp_airway_any"].mean(), 1)])
            r += 1
    autosize(ws, [16, 16, 14, 14, 16, 14])


def supp_s3(wb, df):
    ws = wb.create_sheet("Supp S3 NSQIP Linkage")
    ws.cell(row=1, column=1, value="Supplementary Table S3. NSQIP Linkage Coverage").font = TITLE
    write_header(ws, ["Variable", "Massive: n with data (%)", "Non-massive: n with data (%)"], row=3)
    r = 4
    n_m = int(df["is_massive"].sum()); n_nm = int((~df["is_massive"]).sum())
    massive = df[df["is_massive"]]; non_massive = df[~df["is_massive"]]
    items = [
        ("ASA class", "nsqip_asa_class"),
        ("Inpatient/outpatient (link proxy)", "nsqip_inpatient_outpatient"),
        ("Operative duration", "nsqip_operative_duration_min"),
        ("Length of stay (nsqip_length_of_stay_days)", "nsqip_length_of_stay_days"),
        ("Central neck dissection", "nsqip_central_neck_dissection"),
        ("BMI (nsqip_bmi)", "nsqip_bmi"),
        ("Operative approach", "nsqip_operative_approach"),
        ("Drain usage", "nsqip_drain_usage"),
        ("Vessel sealant", "nsqip_vessel_sealant"),
        ("RLN monitoring", "nsqip_rln_monitoring"),
    ]
    for label, col in items:
        a = int(massive[col].notna().sum())
        b = int(non_massive[col].notna().sum())
        write_row(ws, r, [label, fmt_pct(a, n_m), fmt_pct(b, n_nm)]); r += 1
    autosize(ws, [44, 26, 26])


def supp_s4(wb, df):
    ws = wb.create_sheet("Supp S4 Sensitivity 200g")
    ws.cell(row=1, column=1, value="Supplementary Table S4. Sensitivity — Weight ≥200g Subcohort").font = TITLE
    write_header(ws, ["Outcome", "Weight≥200g n (%)", "Weight 100–<200g n (%)", "Non-massive n (%)"], row=3)
    r = 4
    weight = df["gland_weight_final_g"].fillna(-1)
    g200 = df[weight >= 200]
    g100 = df[(weight >= 100) & (weight < 200)]
    nm = df[~df["is_massive"]]
    n_g200 = len(g200); n_g100 = len(g100); n_nm = len(nm)
    items = [
        ("Cohort n (denominator)", n_g200, n_g100, n_nm),
        ("Any complication", int(g200["any_confirmed_complication_flag"].sum()),
         int(g100["any_confirmed_complication_flag"].sum()),
         int(nm["any_confirmed_complication_flag"].sum())),
        ("HypoPT transient",
         int((g200["comp_hypoparathyroidism_confirmed"] & g200["comp_hypoparathyroidism_transient"]).sum()),
         int((g100["comp_hypoparathyroidism_confirmed"] & g100["comp_hypoparathyroidism_transient"]).sum()),
         int((nm["comp_hypoparathyroidism_confirmed"] & nm["comp_hypoparathyroidism_transient"]).sum())),
        ("HypoPT permanent",
         int((g200["comp_hypoparathyroidism_confirmed"] & g200["comp_hypoparathyroidism_permanent"]).sum()),
         int((g100["comp_hypoparathyroidism_confirmed"] & g100["comp_hypoparathyroidism_permanent"]).sum()),
         int((nm["comp_hypoparathyroidism_confirmed"] & nm["comp_hypoparathyroidism_permanent"]).sum())),
        ("RLN injury", int(g200["comp_rln_injury_confirmed"].sum()),
         int(g100["comp_rln_injury_confirmed"].sum()),
         int(nm["comp_rln_injury_confirmed"].sum())),
        ("Tracheostomy", int(g200["proc_nlp_tracheostomy"].sum()),
         int(g100["proc_nlp_tracheostomy"].sum()),
         int(nm["proc_nlp_tracheostomy"].sum())),
        ("Mortality", int(g200["comp_mortality_definitive"].sum()),
         int(g100["comp_mortality_definitive"].sum()),
         int(nm["comp_mortality_definitive"].sum())),
    ]
    for label, a, b, c in items:
        if label == "Cohort n (denominator)":
            write_row(ws, r, [label, a, b, c])
        else:
            write_row(ws, r, [label, fmt_pct(a, n_g200), fmt_pct(b, n_g100), fmt_pct(c, n_nm)])
        r += 1
    autosize(ws, [32, 22, 24, 22])


def supp_s5(wb, df):
    ws = wb.create_sheet("Supp S5 Missingness")
    ws.cell(row=1, column=1, value="Supplementary Table S5. Missing-Data Coverage").font = TITLE
    write_header(ws, ["Variable", "n missing (massive)", "% missing (massive)",
                      "n missing (non-massive)", "% missing (non-massive)"], row=3)
    r = 4
    n_m = int(df["is_massive"].sum()); n_nm = int((~df["is_massive"]).sum())
    massive = df[df["is_massive"]]; non_massive = df[~df["is_massive"]]
    items = ["age_at_surgery", "sex", "race", "bmi_combined", "gland_weight_final_g",
             "histology_final", "surg_first_date", "surg_procedure_type",
             "nsqip_asa_class", "nsqip_length_of_stay_days", "nsqip_operative_duration_min",
             "followup_years"]
    for col in items:
        m_miss = int(massive[col].isna().sum())
        nm_miss = int(non_massive[col].isna().sum())
        write_row(ws, r, [col, m_miss, f"{100*m_miss/n_m:.1f}%",
                          nm_miss, f"{100*nm_miss/n_nm:.1f}%"]); r += 1
    autosize(ws, [36, 22, 22, 26, 26])


def supp_s6(wb, df):
    ws = wb.create_sheet("Supp S6 Era Comp Rates")
    ws.cell(row=1, column=1, value="Supplementary Table S6. Complication Rates by Era").font = TITLE
    write_header(ws, ["Era", "Arm", "n", "Any-comp n (%)",
                      "HypoPT transient n (%)", "RLN n (%)", "Tracheostomy n (%)"], row=3)
    r = 4
    for era in ["pre-2015", "2015-2019", "2020-2025"]:
        for arm_label, arm_mask in [("Massive", df["is_massive"]), ("Non-massive", ~df["is_massive"])]:
            sub = df[(df["era_bucket_3"] == era) & arm_mask]
            n = len(sub)
            if n == 0: continue
            anycomp = int(sub["any_confirmed_complication_flag"].sum())
            hpt_t = int((sub["comp_hypoparathyroidism_confirmed"] & sub["comp_hypoparathyroidism_transient"]).sum())
            rln = int(sub["comp_rln_injury_confirmed"].sum())
            tr = int(sub["proc_nlp_tracheostomy"].sum())
            write_row(ws, r, [era, arm_label, n,
                              fmt_pct(anycomp, n), fmt_pct(hpt_t, n),
                              fmt_pct(rln, n), fmt_pct(tr, n)])
            r += 1
    autosize(ws, [16, 14, 12, 22, 22, 18, 22])


def main():
    df = pd.read_parquet(PARQUET)
    print(f"→ Loaded {len(df):,} rows × {len(df.columns)} cols")
    wb = Workbook(); wb.remove(wb.active)
    pub_table1(wb, df)
    pub_table2(wb, df)
    pub_table3(wb, df)
    pub_table4(wb, df)
    pub_table5(wb, df)
    fig1_venn(wb, df)
    fig2_era(wb, df)
    fig3_complications(wb, df)
    fig4_component_coverage(wb, df)
    supp_s1(wb, df)
    supp_s2(wb, df)
    supp_s3(wb, df)
    supp_s4(wb, df)
    supp_s5(wb, df)
    supp_s6(wb, df)
    wb.save(OUT)
    print(f"→ Wrote {OUT}  ({OUT.stat().st_size/1024:.1f} KB)")


if __name__ == "__main__":
    main()
