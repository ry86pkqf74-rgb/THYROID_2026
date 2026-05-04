"""03_build_d2_analysis_workbook.py — Deliverable 2
M038_GOITER_analysis_workbook.xlsx — 9 tabs of descriptive analyses.

Tabs:
  1. Cohort Overview
  2. Table 1 Demographics
  3. Table 2 Pathology
  4. Table 3 Surgical
  5. Table 4 Complications
  6. Table 5 Era Stratification
  7. NSQIP Complications
  8. Component Subgroup
  9. Exploratory
"""
from pathlib import Path
import pandas as pd
import numpy as np
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from _stats import (cont_summary, fmt_mean_sd, fmt_median_iqr, chi2_or_fisher, chi2_table,
                    mannwhitney, ttest, rr_ci, fmt_p, fmt_rr_ci, fmt_pct, count_pct, truthy_count)

HERE = Path(__file__).parent
PKG = HERE.parent
PARQUET = HERE / "m038_per_patient_v2.parquet"
OUT = PKG / "M038_GOITER_analysis_workbook.xlsx"

HF = PatternFill("solid", fgColor="1F4E78")
SUBHF = PatternFill("solid", fgColor="D9E1F2")
HFONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHFONT = Font(name="Arial", bold=True, color="1F4E78", size=11)
BODY = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
TITLE = Font(name="Arial", size=14, bold=True, color="1F4E78")
BD = Border(*[Side(style="thin", color="CCCCCC")] * 4)
AL = Alignment(horizontal="left", vertical="center")
AR = Alignment(horizontal="right", vertical="center")
AC = Alignment(horizontal="center", vertical="center")
AW = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ---------- helpers ----------
def write_header(ws, headers, row=1, fill=HF, font=HFONT):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = font; c.fill = fill; c.alignment = AC; c.border = BD


def write_row(ws, row, vals, font=BODY, bold_first=False):
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = BOLD if (bold_first and ci == 1) else font
        c.border = BD
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            c.alignment = AR
        else:
            c.alignment = AL


def section_row(ws, row, label, n_cols=6, fill=SUBHF, font=SUBHFONT):
    c = ws.cell(row=row, column=1, value=label)
    c.font = font; c.fill = fill; c.alignment = AL; c.border = BD
    for ci in range(2, n_cols + 1):
        cc = ws.cell(row=row, column=ci, value=""); cc.fill = fill; cc.border = BD


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------- Tab builders ----------
def build_cohort_overview(wb, df):
    ws = wb.create_sheet("Cohort Overview")
    ws.cell(row=1, column=1, value="M038 Cohort Overview").font = TITLE
    ws.cell(row=2, column=1,
            value="thyroid_canonical_publication_v1_0.manuscript_workspace.cohort_m038_massive_goiter_v1 (post-mig_255)"
            ).font = BODY
    write_header(ws, ["Metric", "Value", "% of total"], row=4)
    n = len(df)
    n_m = int(df["is_massive"].sum())
    rows = [
        ("Total cohort", n, "100.0%"),
        ("Massive (composite-flag positive)", n_m, f"{100*n_m/n:.1f}%"),
        ("Non-massive", n - n_m, f"{100*(n-n_m)/n:.1f}%"),
        ("", "", ""),
        ("— Component composition —", "", ""),
        ("Weight ≥100g (any)", int(df["comp_weight_ge100"].sum()), f"{100*df['comp_weight_ge100'].mean():.1f}%"),
        ("Substernal (CT or MRI)", int(df["comp_substernal_any"].sum()), f"{100*df['comp_substernal_any'].mean():.1f}%"),
        ("Airway (CT deviation/narrowing/compromise)", int(df["comp_airway_any"].sum()), f"{100*df['comp_airway_any'].mean():.1f}%"),
        ("", "", ""),
        ("— Component overlap (Venn) —", "", ""),
    ]
    w = df["comp_weight_ge100"]; s = df["comp_substernal_any"]; a = df["comp_airway_any"]
    rows += [
        ("Weight only (W∧¬S∧¬A)", int(((w) & (~s) & (~a)).sum()), ""),
        ("Substernal only (S∧¬W∧¬A)", int(((s) & (~w) & (~a)).sum()), ""),
        ("Airway only (A∧¬W∧¬S)", int(((a) & (~w) & (~s)).sum()), ""),
        ("Weight ∧ Substernal", int((w & s).sum()), ""),
        ("Weight ∧ Airway", int((w & a).sum()), ""),
        ("Substernal ∧ Airway", int((s & a).sum()), ""),
        ("All three (W∧S∧A)", int((w & s & a).sum()), ""),
        ("Inclusion-exclusion check (= n_massive)",
         int(w.sum()) + int(s.sum()) + int(a.sum())
         - int((w&s).sum()) - int((w&a).sum()) - int((s&a).sum()) + int((w&s&a).sum()),
         f"{n_m} expected"),
        ("", "", ""),
        ("— Era distribution (3-bucket) —", "", ""),
    ]
    for era, sub in df.groupby("era_bucket_3"):
        rows.append((era, len(sub), f"{100*len(sub)/n:.1f}%"))
    rows.append(("", "", ""))
    rows.append(("Date assembled", date.today().isoformat(), ""))
    for i, r in enumerate(rows, 5):
        write_row(ws, i, r, bold_first=True)
    autosize(ws, [44, 18, 14])


def chi2_value_p(df, col, true_val=True):
    """Compute count(true) in massive vs non-massive and chi-squared p."""
    s = df[col]
    if s.dtype == bool:
        a = int((df["is_massive"] & s).sum())
        b = int((~df["is_massive"] & s).sum())
    else:
        a = int((df["is_massive"] & (s == true_val)).sum())
        b = int((~df["is_massive"] & (s == true_val)).sum())
    n1 = int(df["is_massive"].sum())
    n2 = int((~df["is_massive"]).sum())
    test, p = chi2_or_fisher(a, n1 - a, b, n2 - b)
    return a, n1, b, n2, test, p


def build_table1_demographics(wb, df):
    ws = wb.create_sheet("Table 1 Demographics")
    ws.cell(row=1, column=1, value="Table 1 — Demographics & Comorbidities").font = TITLE
    write_header(ws, ["Variable", "Massive (n=2,501)", "Non-massive (n=8,370)", "Test", "p-value"], row=3)
    r = 4
    n_m = int(df["is_massive"].sum())
    n_nm = int((~df["is_massive"]).sum())
    massive = df[df["is_massive"]]
    non_massive = df[~df["is_massive"]]

    section_row(ws, r, "Demographics", n_cols=5); r += 1

    # Age (median IQR + Mann-Whitney)
    sm = cont_summary(massive["age_at_surgery"])
    sn = cont_summary(non_massive["age_at_surgery"])
    test, p = mannwhitney(massive["age_at_surgery"], non_massive["age_at_surgery"])
    write_row(ws, r, ["Age at surgery, median (IQR)", fmt_median_iqr(sm), fmt_median_iqr(sn), test, fmt_p(p)]); r += 1
    test2, p2 = ttest(massive["age_at_surgery"], non_massive["age_at_surgery"])
    write_row(ws, r, ["  mean ± SD", fmt_mean_sd(sm), fmt_mean_sd(sn), test2, fmt_p(p2)]); r += 1

    # Sex
    section_row(ws, r, "Sex", n_cols=5); r += 1
    for label, key in [("Female", "female"), ("Male", "male")]:
        a = int((massive["sex"].str.lower() == key).sum())
        b = int((non_massive["sex"].str.lower() == key).sum())
        test, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
        write_row(ws, r, [f"  {label}", fmt_pct(a, n_m), fmt_pct(b, n_nm), test, fmt_p(p)]); r += 1

    # Race (full breakdown + chi2 across all)
    section_row(ws, r, "Race", n_cols=5); r += 1
    races = sorted(df["race"].dropna().unique().tolist(), key=lambda x: -int((df["race"] == x).sum()))
    obs = []
    for race in races:
        a = int((massive["race"] == race).sum())
        b = int((non_massive["race"] == race).sum())
        obs.append([a, b])
    chi2v, p_overall, dof = chi2_table(obs)
    for race, (a, b) in zip(races, obs):
        write_row(ws, r, [f"  {race}", fmt_pct(a, n_m), fmt_pct(b, n_nm), "", ""]); r += 1
    write_row(ws, r, ["  Race overall (chi-squared)", "", "", f"Chi2 (df={dof})", fmt_p(p_overall)]); r += 1

    # BMI
    section_row(ws, r, "BMI", n_cols=5); r += 1
    sm = cont_summary(massive["bmi_combined"])
    sn = cont_summary(non_massive["bmi_combined"])
    test, p = mannwhitney(massive["bmi_combined"], non_massive["bmi_combined"])
    write_row(ws, r, ["BMI (kg/m^2), median (IQR)", fmt_median_iqr(sm), fmt_median_iqr(sn), test, fmt_p(p)]); r += 1
    write_row(ws, r, [f"  n with BMI", sm["n"], sn["n"], "", ""]); r += 1

    # NLP comorbidities
    section_row(ws, r, "Comorbidities (NLP-extracted)", n_cols=5); r += 1
    for label, col in [("Hypertension", "pmhx_nlp_hypertension"),
                       ("Diabetes", "pmhx_nlp_diabetes"),
                       ("CAD", "pmhx_nlp_cad"),
                       ("CKD", "pmhx_nlp_ckd"),
                       ("COPD", "pmhx_nlp_copd")]:
        a = int(massive[col].sum())
        b = int(non_massive[col].sum())
        test, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
        write_row(ws, r, [f"  {label}", fmt_pct(a, n_m), fmt_pct(b, n_nm), test, fmt_p(p)]); r += 1
    sm = cont_summary(massive["pmhx_nlp_n_comorbidities"])
    sn = cont_summary(non_massive["pmhx_nlp_n_comorbidities"])
    test, p = mannwhitney(massive["pmhx_nlp_n_comorbidities"], non_massive["pmhx_nlp_n_comorbidities"])
    write_row(ws, r, ["  Mean # comorbidities", f"{sm['mean']:.2f}", f"{sn['mean']:.2f}", test, fmt_p(p)]); r += 1

    # Thyroid history
    section_row(ws, r, "Thyroid / surgical history", n_cols=5); r += 1
    for label, col in [("Graves disease", "syn_graves"),
                       ("Hashimoto thyroiditis", "syn_hashimoto"),
                       ("Prior thyroidectomy", "pshx_nlp_prior_thyroidectomy"),
                       ("Prior neck surgery", "pshx_nlp_prior_neck_surgery"),
                       ("Autoimmune thyroid hx", "pmhx_nlp_autoimmune_thyroid_hx")]:
        a = int(massive[col].sum())
        b = int(non_massive[col].sum())
        test, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
        write_row(ws, r, [f"  {label}", fmt_pct(a, n_m), fmt_pct(b, n_nm), test, fmt_p(p)]); r += 1

    # Anticoagulation (truthy-string)
    a = truthy_count(massive["ops_anticoagulation_meds"])
    b = truthy_count(non_massive["ops_anticoagulation_meds"])
    test, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
    write_row(ws, r, ["  Anticoagulation meds (any)", fmt_pct(a, n_m), fmt_pct(b, n_nm), test, fmt_p(p)]); r += 1

    # ASA (NSQIP-linked subset)
    section_row(ws, r, "ASA class (NSQIP-linked subset only)", n_cols=5); r += 1
    massive_ns = massive[massive["nsqip_asa_class"].notna()]
    non_massive_ns = non_massive[non_massive["nsqip_asa_class"].notna()]
    n_ns_m = len(massive_ns); n_ns_nm = len(non_massive_ns)
    write_row(ws, r, ["  NSQIP-linked n", n_ns_m, n_ns_nm, "", ""]); r += 1
    for asa in ["1", "2", "3", "4", "5"]:
        a = int((massive_ns["nsqip_asa_class"].astype(str) == asa).sum())
        b = int((non_massive_ns["nsqip_asa_class"].astype(str) == asa).sum())
        test, p = chi2_or_fisher(a, n_ns_m - a, b, n_ns_nm - b)
        write_row(ws, r, [f"  ASA {asa}", fmt_pct(a, n_ns_m), fmt_pct(b, n_ns_nm), test, fmt_p(p)]); r += 1
    # ASA III-IV combined
    a = int(massive_ns["nsqip_asa_class"].astype(str).isin(["3", "4"]).sum())
    b = int(non_massive_ns["nsqip_asa_class"].astype(str).isin(["3", "4"]).sum())
    test, p = chi2_or_fisher(a, n_ns_m - a, b, n_ns_nm - b)
    write_row(ws, r, ["  ASA III-IV combined", fmt_pct(a, n_ns_m), fmt_pct(b, n_ns_nm), test, fmt_p(p)]); r += 1

    # Smoking
    section_row(ws, r, "Smoking", n_cols=5); r += 1
    a = truthy_count(massive["nsqip_smoker"])
    b = truthy_count(non_massive["nsqip_smoker"])
    test, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
    write_row(ws, r, ["  NSQIP smoker (yes)", fmt_pct(a, n_m), fmt_pct(b, n_nm), test, fmt_p(p)]); r += 1

    # Follow-up
    section_row(ws, r, "Follow-up", n_cols=5); r += 1
    sm = cont_summary(massive["followup_years"])
    sn = cont_summary(non_massive["followup_years"])
    test, p = mannwhitney(massive["followup_years"], non_massive["followup_years"])
    write_row(ws, r, ["Follow-up (years), median (IQR)", fmt_median_iqr(sm), fmt_median_iqr(sn), test, fmt_p(p)]); r += 1
    write_row(ws, r, ["  mean ± SD", fmt_mean_sd(sm), fmt_mean_sd(sn), "t", "—"]); r += 1

    autosize(ws, [44, 22, 22, 14, 12])


def build_table2_pathology(wb, df):
    ws = wb.create_sheet("Table 2 Pathology")
    ws.cell(row=1, column=1, value="Table 2 — Pathology").font = TITLE
    write_header(ws, ["Variable", "Massive (n=2,501)", "Non-massive (n=8,370)", "Test", "p-value"], row=3)
    r = 4
    n_m = int(df["is_massive"].sum())
    n_nm = int((~df["is_massive"]).sum())
    massive = df[df["is_massive"]]; non_massive = df[~df["is_massive"]]

    # Malignancy
    section_row(ws, r, "Malignancy", n_cols=5); r += 1
    a = int(massive["is_malignant"].sum()); b = int(non_massive["is_malignant"].sum())
    test, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
    write_row(ws, r, ["  Malignant (any histology)", fmt_pct(a, n_m), fmt_pct(b, n_nm), test, fmt_p(p)]); r += 1
    a = int(massive["bilateral_disease_flag"].sum()); b = int(non_massive["bilateral_disease_flag"].sum())
    test, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
    write_row(ws, r, ["  Bilateral disease (clinical/imaging)", fmt_pct(a, n_m), fmt_pct(b, n_nm), test, fmt_p(p)]); r += 1
    a = int(massive["bilateral_path_flag"].sum()); b = int(non_massive["bilateral_path_flag"].sum())
    test, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
    write_row(ws, r, ["  Bilateral disease (pathology)", fmt_pct(a, n_m), fmt_pct(b, n_nm), test, fmt_p(p)]); r += 1

    # Histology distribution
    section_row(ws, r, "Histology (final, all patients)", n_cols=5); r += 1
    histos = df["histology_final"].fillna("(missing)").value_counts().index.tolist()
    obs = []
    for h in histos:
        a = int((massive["histology_final"].fillna("(missing)") == h).sum())
        b = int((non_massive["histology_final"].fillna("(missing)") == h).sum())
        obs.append([a, b])
    chi2v, p_overall, dof = chi2_table(obs)
    for h, (a, b) in zip(histos, obs):
        write_row(ws, r, [f"  {h}", fmt_pct(a, n_m), fmt_pct(b, n_nm), "", ""]); r += 1
    write_row(ws, r, ["  Histology overall (chi-squared)", "", "", f"Chi2 (df={dof})", fmt_p(p_overall)]); r += 1

    # Margin (continuous)
    section_row(ws, r, "Closest margin (mm) — malignant only", n_cols=5); r += 1
    sm = cont_summary(massive[massive["is_malignant"]]["closest_margin_mm"])
    sn = cont_summary(non_massive[non_massive["is_malignant"]]["closest_margin_mm"])
    test, p = mannwhitney(massive[massive["is_malignant"]]["closest_margin_mm"],
                          non_massive[non_massive["is_malignant"]]["closest_margin_mm"])
    write_row(ws, r, ["  median (IQR)", fmt_median_iqr(sm), fmt_median_iqr(sn), test, fmt_p(p)]); r += 1
    write_row(ws, r, ["  n with margin", sm["n"], sn["n"], "", ""]); r += 1

    # Gland weight
    section_row(ws, r, "Gland weight (g) — synoptic pathology", n_cols=5); r += 1
    sm = cont_summary(massive["gland_weight_final_g"])
    sn = cont_summary(non_massive["gland_weight_final_g"])
    test, p = mannwhitney(massive["gland_weight_final_g"], non_massive["gland_weight_final_g"])
    write_row(ws, r, ["  median (IQR)", fmt_median_iqr(sm), fmt_median_iqr(sn), test, fmt_p(p)]); r += 1
    write_row(ws, r, ["  mean ± SD", fmt_mean_sd(sm), fmt_mean_sd(sn), "t", "—"]); r += 1
    write_row(ws, r, ["  n with weight", sm["n"], sn["n"], "", ""]); r += 1

    autosize(ws, [44, 22, 22, 14, 12])



def build_table3_surgical(wb, df):
    ws = wb.create_sheet("Table 3 Surgical")
    ws.cell(row=1, column=1, value="Table 3 — Surgical & Operative Context").font = TITLE
    write_header(ws, ["Variable", "Massive (n=2,501)", "Non-massive (n=8,370)", "Test", "p-value"], row=3)
    r = 4
    n_m = int(df["is_massive"].sum())
    n_nm = int((~df["is_massive"]).sum())
    massive = df[df["is_massive"]]; non_massive = df[~df["is_massive"]]

    section_row(ws, r, "Procedure type", n_cols=5); r += 1
    procs = df["surg_procedure_type"].fillna("(missing)").value_counts().index.tolist()
    obs = []
    for p_ in procs:
        a = int((massive["surg_procedure_type"].fillna("(missing)") == p_).sum())
        b = int((non_massive["surg_procedure_type"].fillna("(missing)") == p_).sum())
        obs.append([a, b])
    chi2v, p_overall, dof = chi2_table(obs)
    for p_, (a, b) in zip(procs, obs):
        write_row(ws, r, [f"  {p_}", fmt_pct(a, n_m), fmt_pct(b, n_nm), "", ""]); r += 1
    write_row(ws, r, ["  Procedure type overall", "", "", f"Chi2 (df={dof})", fmt_p(p_overall)]); r += 1

    a = int(massive["surg_total_thyroidectomy"].sum()); b = int(non_massive["surg_total_thyroidectomy"].sum())
    test, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
    write_row(ws, r, ["  Total thyroidectomy (any-procedure flag)", fmt_pct(a, n_m), fmt_pct(b, n_nm), test, fmt_p(p)]); r += 1
    a = int(massive["surg_hemithyroidectomy"].sum()); b = int(non_massive["surg_hemithyroidectomy"].sum())
    test, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
    write_row(ws, r, ["  Hemithyroidectomy (any-procedure flag)", fmt_pct(a, n_m), fmt_pct(b, n_nm), test, fmt_p(p)]); r += 1

    sm = cont_summary(massive["surg_n_procedures"])
    sn = cont_summary(non_massive["surg_n_procedures"])
    test, p = mannwhitney(massive["surg_n_procedures"], non_massive["surg_n_procedures"])
    write_row(ws, r, ["  # surgical procedures, median (IQR)", fmt_median_iqr(sm), fmt_median_iqr(sn), test, fmt_p(p)]); r += 1

    section_row(ws, r, "Neck dissection (NSQIP-linked)", n_cols=5); r += 1
    massive_ns = massive[massive["nsqip_central_neck_dissection"].notna() | massive["nsqip_lateral_neck_dissection"].notna()]
    non_massive_ns = non_massive[non_massive["nsqip_central_neck_dissection"].notna() | non_massive["nsqip_lateral_neck_dissection"].notna()]
    n_ns_m = len(massive_ns); n_ns_nm = len(non_massive_ns)
    a = truthy_count(massive["nsqip_central_neck_dissection"]); b = truthy_count(non_massive["nsqip_central_neck_dissection"])
    test, p = chi2_or_fisher(a, n_ns_m - a, b, n_ns_nm - b)
    write_row(ws, r, ["  Central neck dissection", fmt_pct(a, n_ns_m), fmt_pct(b, n_ns_nm), test, fmt_p(p)]); r += 1
    a = truthy_count(massive["nsqip_lateral_neck_dissection"]); b = truthy_count(non_massive["nsqip_lateral_neck_dissection"])
    test, p = chi2_or_fisher(a, n_ns_m - a, b, n_ns_nm - b)
    write_row(ws, r, ["  Lateral neck dissection", fmt_pct(a, n_ns_m), fmt_pct(b, n_ns_nm), test, fmt_p(p)]); r += 1

    section_row(ws, r, "Operative duration (NSQIP-linked, minutes)", n_cols=5); r += 1
    sm = cont_summary(massive["nsqip_operative_duration_min"])
    sn = cont_summary(non_massive["nsqip_operative_duration_min"])
    test, p = mannwhitney(massive["nsqip_operative_duration_min"], non_massive["nsqip_operative_duration_min"])
    write_row(ws, r, ["  median (IQR)", fmt_median_iqr(sm), fmt_median_iqr(sn), test, fmt_p(p)]); r += 1
    write_row(ws, r, ["  mean ± SD", fmt_mean_sd(sm), fmt_mean_sd(sn), "t", "—"]); r += 1
    write_row(ws, r, ["  n", sm["n"], sn["n"], "", ""]); r += 1

    section_row(ws, r, "Length of stay (NSQIP nsqip_length_of_stay_days)", n_cols=5); r += 1
    sm = cont_summary(massive["nsqip_length_of_stay_days"])
    sn = cont_summary(non_massive["nsqip_length_of_stay_days"])
    test, p = mannwhitney(massive["nsqip_length_of_stay_days"], non_massive["nsqip_length_of_stay_days"])
    write_row(ws, r, ["  median (IQR)", fmt_median_iqr(sm), fmt_median_iqr(sn), test, fmt_p(p)]); r += 1
    write_row(ws, r, ["  mean ± SD", fmt_mean_sd(sm), fmt_mean_sd(sn), "t", "—"]); r += 1
    write_row(ws, r, ["  n", sm["n"], sn["n"], "", ""]); r += 1

    section_row(ws, r, "Operative adjuncts (NSQIP-linked)", n_cols=5); r += 1
    for label, col in [("Drain usage", "nsqip_drain_usage"),
                       ("Vessel sealant", "nsqip_vessel_sealant"),
                       ("RLN monitoring", "nsqip_rln_monitoring")]:
        ns_m = massive[massive[col].notna()]; ns_nm = non_massive[non_massive[col].notna()]
        a = truthy_count(massive[col]); b = truthy_count(non_massive[col])
        test, p = chi2_or_fisher(a, len(ns_m) - a, b, len(ns_nm) - b)
        write_row(ws, r, [f"  {label}", fmt_pct(a, len(ns_m)), fmt_pct(b, len(ns_nm)), test, fmt_p(p)]); r += 1

    section_row(ws, r, "Inpatient/outpatient (NSQIP-linked)", n_cols=5); r += 1
    for val in ["Inpatient", "Outpatient"]:
        ns_m = massive[massive["nsqip_inpatient_outpatient"].notna()]
        ns_nm = non_massive[non_massive["nsqip_inpatient_outpatient"].notna()]
        a = int((ns_m["nsqip_inpatient_outpatient"].astype(str).str.lower() == val.lower()).sum())
        b = int((ns_nm["nsqip_inpatient_outpatient"].astype(str).str.lower() == val.lower()).sum())
        test, p = chi2_or_fisher(a, len(ns_m) - a, b, len(ns_nm) - b)
        write_row(ws, r, [f"  {val}", fmt_pct(a, len(ns_m)), fmt_pct(b, len(ns_nm)), test, fmt_p(p)]); r += 1
    a = int((massive["nsqip_same_day_discharge_flag"] == 1).sum())
    b = int((non_massive["nsqip_same_day_discharge_flag"] == 1).sum())
    n_ns_m = int(massive["nsqip_same_day_discharge_flag"].notna().sum())
    n_ns_nm = int(non_massive["nsqip_same_day_discharge_flag"].notna().sum())
    test, p = chi2_or_fisher(a, n_ns_m - a, b, n_ns_nm - b)
    write_row(ws, r, ["  Same-day discharge flag", fmt_pct(a, n_ns_m), fmt_pct(b, n_ns_nm), test, fmt_p(p)]); r += 1

    section_row(ws, r, "Difficult airway (op-note NLP)", n_cols=5); r += 1
    a = truthy_count(massive["ops_difficult_airway"])
    b = truthy_count(non_massive["ops_difficult_airway"])
    test, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
    write_row(ws, r, ["  Yes", fmt_pct(a, n_m), fmt_pct(b, n_nm), test, fmt_p(p)]); r += 1

    section_row(ws, r, "Tracheostomy (NLP-extracted)", n_cols=5); r += 1
    a = int(massive["proc_nlp_tracheostomy"].sum())
    b = int(non_massive["proc_nlp_tracheostomy"].sum())
    test, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
    write_row(ws, r, ["  Tracheostomy (any)", fmt_pct(a, n_m), fmt_pct(b, n_nm), test, fmt_p(p)]); r += 1

    autosize(ws, [44, 22, 22, 14, 12])


def build_table4_complications(wb, df):
    ws = wb.create_sheet("Table 4 Complications")
    ws.cell(row=1, column=1, value="Table 4 — Strict-Definition Complications").font = TITLE
    ws.cell(row=2, column=1,
            value="Per standing rule (memory/feedback_complications_transient_vs_permanent.md): "
                  "hypopara split into transient<6mo / permanent>6mo; hypocalcemia preexisting reported separately."
            ).font = BODY
    write_header(ws, ["Complication", "Massive n (%)", "Non-massive n (%)", "RR (95% CI)", "Test", "p-value"], row=4)
    r = 5
    n_m = int(df["is_massive"].sum()); n_nm = int((~df["is_massive"]).sum())
    massive = df[df["is_massive"]]; non_massive = df[~df["is_massive"]]

    def write_comp(label, a, b):
        nonlocal r
        rr, lo, hi = rr_ci(a, n_m, b, n_nm)
        test, p = chi2_or_fisher(a, n_m - a, b, n_nm - b)
        write_row(ws, r, [label, fmt_pct(a, n_m), fmt_pct(b, n_nm),
                          fmt_rr_ci(rr, lo, hi), test, fmt_p(p)])
        r += 1

    section_row(ws, r, "Composite + structural", n_cols=6); r += 1
    write_comp("Any confirmed complication", int(massive["any_confirmed_complication_flag"].sum()),
               int(non_massive["any_confirmed_complication_flag"].sum()))
    write_comp("Hematoma", int(massive["comp_hematoma_confirmed"].sum()),
               int(non_massive["comp_hematoma_confirmed"].sum()))
    write_comp("Seroma", int(massive["comp_seroma_confirmed"].sum()),
               int(non_massive["comp_seroma_confirmed"].sum()))
    write_comp("Chyle leak", int(massive["comp_chyle_leak_confirmed"].sum()),
               int(non_massive["comp_chyle_leak_confirmed"].sum()))

    section_row(ws, r, "Recurrent laryngeal nerve / vocal cord", n_cols=6); r += 1
    write_comp("RLN injury (any)", int(massive["comp_rln_injury_confirmed"].sum()),
               int(non_massive["comp_rln_injury_confirmed"].sum()))
    a = int((massive["comp_rln_injury_confirmed"] & massive["comp_rln_injury_transient"]).sum())
    b = int((non_massive["comp_rln_injury_confirmed"] & non_massive["comp_rln_injury_transient"]).sum())
    write_comp("  RLN injury — transient", a, b)
    write_comp("VC paresis (any)", int(massive["comp_vc_paresis_confirmed"].sum()),
               int(non_massive["comp_vc_paresis_confirmed"].sum()))
    a = int((massive["comp_vc_paresis_confirmed"] & massive["comp_vc_paresis_permanent"]).sum())
    b = int((non_massive["comp_vc_paresis_confirmed"] & non_massive["comp_vc_paresis_permanent"]).sum())
    write_comp("  VC paresis — permanent", a, b)
    write_comp("VC paralysis (any)", int(massive["comp_vc_paralysis_confirmed"].sum()),
               int(non_massive["comp_vc_paralysis_confirmed"].sum()))
    a = int((massive["comp_vc_paralysis_confirmed"] & massive["comp_vc_paralysis_permanent"]).sum())
    b = int((non_massive["comp_vc_paralysis_confirmed"] & non_massive["comp_vc_paralysis_permanent"]).sum())
    write_comp("  VC paralysis — permanent", a, b)

    section_row(ws, r, "Hypocalcemia (postop confirmed + preexisting reported separately)", n_cols=6); r += 1
    write_comp("Hypocalcemia (any confirmed)", int(massive["comp_hypocalcemia_confirmed"].sum()),
               int(non_massive["comp_hypocalcemia_confirmed"].sum()))
    a = int((massive["comp_hypocalcemia_confirmed"] & massive["comp_hypocalcemia_transient"]).sum())
    b = int((non_massive["comp_hypocalcemia_confirmed"] & non_massive["comp_hypocalcemia_transient"]).sum())
    write_comp("  Hypocalcemia — transient", a, b)
    a = int((massive["comp_hypocalcemia_confirmed"] & massive["comp_hypocalcemia_permanent"]).sum())
    b = int((non_massive["comp_hypocalcemia_confirmed"] & non_massive["comp_hypocalcemia_permanent"]).sum())
    write_comp("  Hypocalcemia — permanent", a, b)
    a = int(massive["hca_preop_flag"].sum())
    b = int(non_massive["hca_preop_flag"].sum())
    write_comp("  Hypocalcemia — preexisting (preop / pre-surgery window)", a, b)

    section_row(ws, r, "Hypoparathyroidism (standing rule split)", n_cols=6); r += 1
    write_comp("Hypopara (any confirmed)", int(massive["comp_hypoparathyroidism_confirmed"].sum()),
               int(non_massive["comp_hypoparathyroidism_confirmed"].sum()))
    a = int((massive["comp_hypoparathyroidism_confirmed"] & massive["comp_hypoparathyroidism_transient"]).sum())
    b = int((non_massive["comp_hypoparathyroidism_confirmed"] & non_massive["comp_hypoparathyroidism_transient"]).sum())
    write_comp("  Hypopara — transient (<6mo)", a, b)
    a = int((massive["comp_hypoparathyroidism_confirmed"] & massive["comp_hypoparathyroidism_permanent"]).sum())
    b = int((non_massive["comp_hypoparathyroidism_confirmed"] & non_massive["comp_hypoparathyroidism_permanent"]).sum())
    write_comp("  Hypopara — permanent (>6mo)", a, b)
    a = int((massive["comp_hypoparathyroidism_confirmed"]
             & ~massive["comp_hypoparathyroidism_transient"]
             & ~massive["comp_hypoparathyroidism_permanent"]).sum())
    b = int((non_massive["comp_hypoparathyroidism_confirmed"]
             & ~non_massive["comp_hypoparathyroidism_transient"]
             & ~non_massive["comp_hypoparathyroidism_permanent"]).sum())
    write_comp("  Hypopara — unclassified timing", a, b)
    write_comp("  Hypopara — preexisting (preop)", int(massive["comp_hypoparathyroidism_preexisting"].sum()),
               int(non_massive["comp_hypoparathyroidism_preexisting"].sum()))
    write_comp("  Hypopara — new postop (after surgery)", int(massive["comp_hypoparathyroidism_new_postop"].sum()),
               int(non_massive["comp_hypoparathyroidism_new_postop"].sum()))

    section_row(ws, r, "Airway / pulmonary / mortality", n_cols=6); r += 1
    write_comp("Airway complication (definitive)", int(massive["comp_airway_complication_definitive"].sum()),
               int(non_massive["comp_airway_complication_definitive"].sum()))
    write_comp("Pneumothorax (definitive)", int(massive["comp_pneumothorax_definitive"].sum()),
               int(non_massive["comp_pneumothorax_definitive"].sum()))
    write_comp("Mortality (definitive, any time)", int(massive["comp_mortality_definitive"].sum()),
               int(non_massive["comp_mortality_definitive"].sum()))
    write_comp("Death occurred (survival rollup)", int(massive["death_occurred"].sum()),
               int(non_massive["death_occurred"].sum()))

    section_row(ws, r, "Tracheostomy (procedure NLP)", n_cols=6); r += 1
    write_comp("Tracheostomy (any)", int(massive["proc_nlp_tracheostomy"].sum()),
               int(non_massive["proc_nlp_tracheostomy"].sum()))

    autosize(ws, [44, 18, 20, 20, 12, 10])


def build_table5_era(wb, df):
    ws = wb.create_sheet("Table 5 Era Stratification")
    ws.cell(row=1, column=1, value="Table 5 — Era Stratification (massive prevalence + components)").font = TITLE
    ws.cell(row=2, column=1,
            value="Era binning rule: upper-bound — pre-1999 surgery dates sweep into 1999-2004 bucket. "
                  "3-bucket headline trend uses pre-2015 / 2015-2019 / 2020-2025."
            ).font = BODY
    write_header(ws, ["Era bucket", "n total", "n massive", "% massive",
                      "n weight≥100g", "n substernal (CT/MRI)", "n airway (CT)"], row=4)
    r = 5
    # 6-bucket
    section_row(ws, r, "6-bucket (5-yr)", n_cols=7); r += 1
    for era, sub in df.groupby("era_bucket_6"):
        if era == "unknown": continue
        write_row(ws, r, [era, len(sub), int(sub["is_massive"].sum()),
                          f"{100*sub['is_massive'].mean():.1f}%",
                          int(sub["comp_weight_ge100"].sum()),
                          int(sub["comp_substernal_any"].sum()),
                          int(sub["comp_airway_any"].sum())])
        r += 1
    # unknown line
    sub = df[df["era_bucket_6"] == "unknown"]
    if len(sub):
        write_row(ws, r, ["unknown (no surg date)", len(sub), int(sub["is_massive"].sum()),
                          f"{100*sub['is_massive'].mean():.1f}%",
                          int(sub["comp_weight_ge100"].sum()),
                          int(sub["comp_substernal_any"].sum()),
                          int(sub["comp_airway_any"].sum())])
        r += 1
    section_row(ws, r, "3-bucket headline (manuscript abstract)", n_cols=7); r += 1
    for era in ["pre-2015", "2015-2019", "2020-2025"]:
        sub = df[df["era_bucket_3"] == era]
        write_row(ws, r, [era, len(sub), int(sub["is_massive"].sum()),
                          f"{100*sub['is_massive'].mean():.1f}%",
                          int(sub["comp_weight_ge100"].sum()),
                          int(sub["comp_substernal_any"].sum()),
                          int(sub["comp_airway_any"].sum())])
        r += 1
    sub = df[df["era_bucket_3"] == "unknown"]
    write_row(ws, r, ["unknown (no surg date)", len(sub), int(sub["is_massive"].sum()),
                      f"{100*sub['is_massive'].mean():.1f}%",
                      int(sub["comp_weight_ge100"].sum()),
                      int(sub["comp_substernal_any"].sum()),
                      int(sub["comp_airway_any"].sum())])
    r += 1
    # Chi-squared trend across the 3 known buckets
    obs = []
    for era in ["pre-2015", "2015-2019", "2020-2025"]:
        sub = df[df["era_bucket_3"] == era]
        obs.append([int(sub["is_massive"].sum()), int((~sub["is_massive"]).sum())])
    chi2v, p_overall, dof = chi2_table(obs)
    r += 1
    write_row(ws, r, ["Chi-squared across 3-bucket eras", "", "",
                      f"Chi2={chi2v:.2f}, df={dof}, p={fmt_p(p_overall)}", "", "", ""])
    autosize(ws, [22, 12, 14, 14, 18, 22, 16])



def build_nsqip_complications(wb, df):
    ws = wb.create_sheet("NSQIP Complications")
    ws.cell(row=1, column=1, value="NSQIP-Linked 30-Day Complications").font = TITLE
    ws.cell(row=2, column=1,
            value="NSQIP-linked subset only (denominator = patients with NSQIP record). "
                  "These are NSQIP-flagged events, distinct from the strict-definition canonical complications in Table 4."
            ).font = BODY
    write_header(ws, ["Complication (NSQIP)", "Massive (n_NSQIP)", "Non-massive (n_NSQIP)",
                      "RR (95% CI)", "Test", "p-value"], row=4)
    r = 5
    massive = df[df["is_massive"]]
    non_massive = df[~df["is_massive"]]

    # Build a "has any NSQIP record" denominator (using any non-null NSQIP field as proxy)
    # We use nsqip_inpatient_outpatient as the proxy for NSQIP linkage status.
    massive_ns = massive[massive["nsqip_inpatient_outpatient"].notna()]
    non_massive_ns = non_massive[non_massive["nsqip_inpatient_outpatient"].notna()]
    n_ns_m = len(massive_ns); n_ns_nm = len(non_massive_ns)
    write_row(ws, r, ["NSQIP-linked denominator", n_ns_m, n_ns_nm, "", "", ""]); r += 1

    def write_ns(label, ns_m_count, ns_nm_count):
        nonlocal r
        rr, lo, hi = rr_ci(ns_m_count, n_ns_m, ns_nm_count, n_ns_nm)
        test, p = chi2_or_fisher(ns_m_count, n_ns_m - ns_m_count, ns_nm_count, n_ns_nm - ns_nm_count)
        write_row(ws, r, [label, fmt_pct(ns_m_count, n_ns_m), fmt_pct(ns_nm_count, n_ns_nm),
                          fmt_rr_ci(rr, lo, hi), test, fmt_p(p)])
        r += 1

    section_row(ws, r, "Bleeding / hematoma / RLN / hypocalcemia (NSQIP flags)", n_cols=6); r += 1
    write_ns("Transfusion (any unit)", int((massive_ns["nsqip_transfusion"].fillna(0) >= 1).sum()),
             int((non_massive_ns["nsqip_transfusion"].fillna(0) >= 1).sum()))
    write_ns("Neck hematoma (NSQIP truthy)", truthy_count(massive_ns["nsqip_neck_hematoma"]),
             truthy_count(non_massive_ns["nsqip_neck_hematoma"]))
    write_ns("Hematoma flag (=1)", int((massive_ns["nsqip_hematoma_flag"] == 1).sum()),
             int((non_massive_ns["nsqip_hematoma_flag"] == 1).sum()))
    write_ns("RLN injury flag (=1)", int((massive_ns["nsqip_rln_injury_flag"] == 1).sum()),
             int((non_massive_ns["nsqip_rln_injury_flag"] == 1).sum()))
    write_ns("Hypocalcemia flag (=1)", int((massive_ns["nsqip_hypocalcemia_flag"] == 1).sum()),
             int((non_massive_ns["nsqip_hypocalcemia_flag"] == 1).sum()))

    section_row(ws, r, "Airway / unplanned events", n_cols=6); r += 1
    write_ns("Unplanned intubation (>=1)", int((massive_ns["nsqip_unplanned_intubation"].fillna(0) >= 1).sum()),
             int((non_massive_ns["nsqip_unplanned_intubation"].fillna(0) >= 1).sum()))
    write_ns("Unplanned return to OR (=1)", int((massive_ns["nsqip_unplanned_return_or"] == 1).sum()),
             int((non_massive_ns["nsqip_unplanned_return_or"] == 1).sum()))
    write_ns("30-day readmission flag (=1)", int((massive_ns["nsqip_readmission_30d_flag"] == 1).sum()),
             int((non_massive_ns["nsqip_readmission_30d_flag"] == 1).sum()))
    write_ns("30-day death", truthy_count(massive_ns["nsqip_death_30d"]),
             truthy_count(non_massive_ns["nsqip_death_30d"]))

    section_row(ws, r, "Other 30-day NSQIP events", n_cols=6); r += 1
    for label, col in [("Pneumonia (>=1)", "nsqip_pneumonia"),
                       ("DVT (>=1)", "nsqip_dvt"),
                       ("PE (>=1)", "nsqip_pe"),
                       ("Sepsis (>=1)", "nsqip_sepsis"),
                       ("Superficial SSI (>=1)", "nsqip_superficial_ssi"),
                       ("Deep SSI (>=1)", "nsqip_deep_ssi"),
                       ("Organ-space SSI (>=1)", "nsqip_organ_space_ssi")]:
        write_ns(label, int((massive_ns[col].fillna(0) >= 1).sum()),
                 int((non_massive_ns[col].fillna(0) >= 1).sum()))

    autosize(ws, [44, 22, 22, 20, 12, 10])


def build_component_subgroup(wb, df):
    ws = wb.create_sheet("Component Subgroup")
    ws.cell(row=1, column=1, value="Component Subgroup — Outcomes by Massive Component(s)").font = TITLE
    ws.cell(row=2, column=1,
            value="Within massive arm only (n=2,501): which component(s) the patient met. "
                  "Subgroups: weight-only (n=898), substernal-only (n=145), airway-only (n=429), all-three (n=386)."
            ).font = BODY
    write_header(ws, ["Subgroup", "n", "Any complication n (%)",
                      "HypoPT transient n (%)", "HypoPT permanent n (%)",
                      "Tracheostomy n (%)", "Mortality n (%)"], row=4)
    r = 5
    massive = df[df["is_massive"]].copy()
    w = massive["comp_weight_ge100"]
    s = massive["comp_substernal_any"]
    a = massive["comp_airway_any"]

    def line(name, mask):
        nonlocal r
        sub = massive[mask]
        n = len(sub)
        if n == 0:
            write_row(ws, r, [name, 0, "—", "—", "—", "—", "—"]); r += 1; return
        anycomp = int(sub["any_confirmed_complication_flag"].sum())
        hpt_t = int((sub["comp_hypoparathyroidism_confirmed"] & sub["comp_hypoparathyroidism_transient"]).sum())
        hpt_p = int((sub["comp_hypoparathyroidism_confirmed"] & sub["comp_hypoparathyroidism_permanent"]).sum())
        tr = int(sub["proc_nlp_tracheostomy"].sum())
        mort = int(sub["comp_mortality_definitive"].sum())
        write_row(ws, r, [name, n,
                          fmt_pct(anycomp, n), fmt_pct(hpt_t, n), fmt_pct(hpt_p, n),
                          fmt_pct(tr, n), fmt_pct(mort, n)])
        r += 1

    line("Weight only (W∧¬S∧¬A)", w & ~s & ~a)
    line("Substernal only (S∧¬W∧¬A)", s & ~w & ~a)
    line("Airway only (A∧¬W∧¬S)", a & ~w & ~s)
    line("Weight ∧ Substernal", w & s)
    line("Weight ∧ Airway", w & a)
    line("Substernal ∧ Airway", s & a)
    line("All three (W∧S∧A)", w & s & a)
    line("Any massive (overall)", pd.Series([True] * len(massive), index=massive.index))

    autosize(ws, [32, 10, 22, 22, 22, 20, 18])


def build_exploratory(wb, df):
    ws = wb.create_sheet("Exploratory")
    ws.cell(row=1, column=1, value="Exploratory Cross-tabs").font = TITLE

    # 1) Complication rate by procedure type within massive arm
    write_header(ws, ["Procedure type (massive arm only)", "n", "Any-comp n (%)",
                      "HypoPT transient n (%)", "RLN n (%)"], row=3)
    r = 4
    massive = df[df["is_massive"]]
    procs = massive["surg_procedure_type"].fillna("(missing)").value_counts().index.tolist()
    for p_ in procs:
        sub = massive[massive["surg_procedure_type"].fillna("(missing)") == p_]
        n = len(sub)
        anycomp = int(sub["any_confirmed_complication_flag"].sum())
        hpt_t = int((sub["comp_hypoparathyroidism_confirmed"] & sub["comp_hypoparathyroidism_transient"]).sum())
        rln = int(sub["comp_rln_injury_confirmed"].sum())
        write_row(ws, r, [p_, n, fmt_pct(anycomp, n), fmt_pct(hpt_t, n), fmt_pct(rln, n)])
        r += 1

    r += 2
    write_header(ws, ["Subgroup", "n", "n_malignant (%)", "PTC n (%)", "MTC/ATC n (%)"], row=r); r += 1
    # 2) Malignancy rate by massive component
    for label, mask_fn in [
        ("Massive — weight only", lambda d: d["comp_weight_ge100"] & ~d["comp_substernal_any"] & ~d["comp_airway_any"]),
        ("Massive — substernal only", lambda d: d["comp_substernal_any"] & ~d["comp_weight_ge100"] & ~d["comp_airway_any"]),
        ("Massive — airway only", lambda d: d["comp_airway_any"] & ~d["comp_weight_ge100"] & ~d["comp_substernal_any"]),
        ("Massive — all three", lambda d: d["comp_weight_ge100"] & d["comp_substernal_any"] & d["comp_airway_any"]),
        ("Massive — any", lambda d: d["is_massive"]),
        ("Non-massive", lambda d: ~d["is_massive"]),
    ]:
        sub = df[mask_fn(df)]
        n = len(sub)
        n_mal = int(sub["is_malignant"].sum())
        ptc = int(sub["histology_final"].astype(str).str.contains("apillary", na=False).sum())
        bad = int(sub["histology_final"].astype(str).str.contains("edullary|naplas", na=False).sum())
        write_row(ws, r, [label, n, fmt_pct(n_mal, n), fmt_pct(ptc, n), fmt_pct(bad, n)])
        r += 1

    r += 2
    write_header(ws, ["Gland weight bin (g)", "n total", "n massive", "% massive"], row=r); r += 1
    # 3) Gland weight distribution
    bins = [0, 50, 100, 150, 200, 300, 500, 1e9]
    labels_ = ["<50", "50–<100", "100–<150", "150–<200", "200–<300", "300–<500", ">=500"]
    df_w = df.dropna(subset=["gland_weight_final_g"]).copy()
    df_w["bin"] = pd.cut(df_w["gland_weight_final_g"], bins=bins, labels=labels_, right=False)
    for b in labels_:
        sub = df_w[df_w["bin"] == b]
        n = len(sub); n_m = int(sub["is_massive"].sum())
        write_row(ws, r, [b, n, n_m, f"{100*n_m/n:.1f}%" if n else "—"])
        r += 1
    # Patients with NULL gland weight
    null_w = df[df["gland_weight_final_g"].isna()]
    write_row(ws, r, ["NULL gland weight", len(null_w), int(null_w["is_massive"].sum()),
                      f"{100*null_w['is_massive'].mean():.1f}%" if len(null_w) else "—"])

    autosize(ws, [38, 12, 22, 22, 18])


def main():
    df = pd.read_parquet(PARQUET)
    print(f"→ Loaded {len(df):,} rows × {len(df.columns)} cols")

    wb = Workbook(); wb.remove(wb.active)

    build_cohort_overview(wb, df)
    build_table1_demographics(wb, df)
    build_table2_pathology(wb, df)
    build_table3_surgical(wb, df)
    build_table4_complications(wb, df)
    build_table5_era(wb, df)
    build_nsqip_complications(wb, df)
    build_component_subgroup(wb, df)
    build_exploratory(wb, df)

    wb.save(OUT)
    print(f"→ Wrote {OUT}  ({OUT.stat().st_size/1024:.1f} KB)")


if __name__ == "__main__":
    main()
