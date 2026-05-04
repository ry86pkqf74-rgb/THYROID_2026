"""Build M038 per-patient master Excel with column-source map.

Mirrors M044's 05b_per_patient_with_sources.xlsx structure:
- README sheet
- Per-patient analytic sheet (n=10,871 × ~80 cols)
- Source map sheet (per-column DB.schema.table.column + aggregation rule + dtype + sections used)
- Audit (gaps) sheet — surfacing patients with missing key fields

Prereq: local duckdb CLI must be authed as logan.glosser.eras@gmail.com
        (the master account that owns the canonical_publication share).
        See README §"What is parked" for instructions.

Usage:
  python3 build_m038_per_patient.py
  → writes <repo>/M038_submission_package_v1_0/05b_per_patient_with_sources.xlsx
"""
import os, subprocess, sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# mig_299: portable path — script lives at <PKG>/08_analysis_code/<this>.py
PKG = Path(__file__).resolve().parents[1]
OUT = PKG / "05b_per_patient_with_sources.xlsx"
TMP_PARQUET = Path("/tmp/m038_per_patient_v1_0.parquet")

EXTRACT_SQL = """
ATTACH 'md:thyroid_canonical_publication_v1_0';
USE thyroid_canonical_publication_v1_0;

COPY (
WITH base AS (
  SELECT * FROM manuscript_workspace.cohort_m038_massive_goiter_v1
)
SELECT
  research_id,
  -- Derived
  (COALESCE(b.gland_weight_final_g >= 100, FALSE)
   OR COALESCE(b.ct_substernal_extension_any, FALSE) OR COALESCE(b.mri_substernal_any, FALSE)
   OR COALESCE(b.ct_tracheal_deviation_any, FALSE) OR COALESCE(b.ct_tracheal_narrowing_any, FALSE) OR COALESCE(b.ct_airway_compromise_any, FALSE)) AS is_massive,
  COALESCE(b.gland_weight_final_g >= 100, FALSE) AS comp_weight_ge100,
  (COALESCE(b.ct_substernal_extension_any, FALSE) OR COALESCE(b.mri_substernal_any, FALSE)) AS comp_substernal_any,
  (COALESCE(b.ct_tracheal_deviation_any, FALSE) OR COALESCE(b.ct_tracheal_narrowing_any, FALSE) OR COALESCE(b.ct_airway_compromise_any, FALSE)) AS comp_airway_any,
  CASE WHEN b.surg_first_date IS NULL THEN 'unknown'
       WHEN b.surg_first_date <= '2004-12-31' THEN '1999-2004'
       WHEN b.surg_first_date <= '2009-12-31' THEN '2005-2009'
       WHEN b.surg_first_date <= '2014-12-31' THEN '2010-2014'
       WHEN b.surg_first_date <= '2019-12-31' THEN '2015-2019'
       ELSE '2020-2025' END AS era_bucket,
  CASE WHEN b.comp_hypoparathyroidism_confirmed AND b.comp_hypoparathyroidism_transient THEN 'transient_lt_6mo'
       WHEN b.comp_hypoparathyroidism_confirmed AND b.comp_hypoparathyroidism_permanent THEN 'permanent_gt_6mo'
       WHEN b.comp_hypoparathyroidism_confirmed THEN 'unclassified'
       ELSE 'none' END AS hypopara_postop_class,
  (COALESCE(b.comp_hypocalcemia_timing_window = 'pre_surgery', FALSE)
   OR COALESCE(b.comp_hypocalcemia_clinical_preexisting, FALSE)) AS hca_preop_flag,
  -- Demographics
  age_at_surgery, sex, race, bmi_combined, bmi_source, nsqip_bmi, nsqip_asa_class,
  nsqip_smoker, nsqip_tobacco_use, pmhx_nlp_smoking_status,
  -- Comorbidities NLP
  pmhx_nlp_diabetes, pmhx_nlp_hypertension, pmhx_nlp_cad, pmhx_nlp_ckd, pmhx_nlp_copd,
  pmhx_nlp_n_comorbidities, pmhx_nlp_autoimmune_thyroid_hx,
  -- Comorbidities NSQIP
  nsqip_diabetes, nsqip_hypertension, nsqip_copd, nsqip_heart_failure,
  nsqip_bleeding_disorder, nsqip_disseminated_cancer, nsqip_functional_status,
  -- Thyroid-specific
  syn_graves, syn_hashimoto, pshx_nlp_prior_thyroidectomy, pshx_nlp_prior_neck_surgery,
  -- Surgical
  surg_first_date, surg_procedure_type, surg_total_thyroidectomy, surg_hemithyroidectomy,
  surg_n_procedures, nsqip_central_neck_dissection, nsqip_lateral_neck_dissection,
  nsqip_operative_approach, nsqip_operative_duration_min, nsqip_drain_usage, nsqip_vessel_sealant,
  nsqip_rln_monitoring, ops_difficult_airway, ops_surgeon, ops_surg_date,
  nsqip_inpatient_outpatient, nsqip_primary_indication,
  -- Anatomy / exposure
  gland_weight_final_g, gland_weight_total_reported_g,
  ct_substernal_extension_any, mri_substernal_any,
  ct_tracheal_deviation_any, ct_tracheal_narrowing_any, ct_airway_compromise_any,
  ct_goiter_present_any, nlp_airway_has_data,
  bilateral_disease_flag, bilateral_path_flag, closest_margin_mm,
  -- Pathology
  histology_final, is_malignant,
  -- LOS
  nsqip_length_of_stay_days, nsqip_hospital_los_days,
  -- Complications strict
  any_confirmed_complication_flag, comp_hematoma_confirmed, comp_seroma_confirmed,
  comp_chyle_leak_confirmed, comp_rln_injury_confirmed, comp_vc_paresis_confirmed,
  comp_vc_paralysis_confirmed, comp_hypocalcemia_confirmed, comp_hypoparathyroidism_confirmed,
  comp_mortality_definitive, comp_airway_complication_definitive, comp_pneumothorax_definitive,
  -- Complications temporality (post-mig_255)
  comp_hypoparathyroidism_transient, comp_hypoparathyroidism_permanent,
  comp_hypoparathyroidism_timing_window, comp_hypoparathyroidism_preexisting,
  comp_hypoparathyroidism_new_postop, comp_hypopara_permanent_limitation_note,
  comp_hypocalcemia_transient, comp_hypocalcemia_permanent,
  comp_hypocalcemia_timing_window, comp_hypocalcemia_clinical_preexisting,
  comp_rln_injury_transient, comp_rln_injury_timing_window,
  comp_vc_paresis_timing_window, comp_vc_paralysis_timing_window,
  comp_vc_paresis_permanent, comp_vc_paralysis_permanent,
  -- NSQIP perioperative
  nsqip_transfusion, nsqip_unplanned_intubation, nsqip_readmission_30d_flag,
  nsqip_readmission_count, nsqip_death_30d,
  -- Tracheostomy & follow-up
  proc_nlp_tracheostomy, proc_nlp_tracheostomy_days_from_surg,
  followup_years, death_occurred
FROM base b
ORDER BY research_id
) TO '%s' (FORMAT 'parquet');
""" % str(TMP_PARQUET)


def extract():
    """Run the duckdb CLI extract."""
    print("→ Extracting via duckdb CLI (must be authed as logan.glosser.eras)...")
    sql_path = Path("/tmp/m038_extract.sql")
    sql_path.write_text(EXTRACT_SQL)
    r = subprocess.run(["duckdb", "-f", str(sql_path)], capture_output=True, text=True)
    if r.returncode != 0 or "no database/share" in (r.stdout + r.stderr).lower():
        print("EXTRACT FAILED:")
        print(r.stdout)
        print(r.stderr)
        print()
        print("→ Need MotherDuck eras-account auth. Run in a fresh terminal:")
        print("    rm -f ~/.duckdb/credentials*")
        print("    duckdb -c \"ATTACH 'md:thyroid_canonical_publication_v1_0';\"")
        print("    # When the SSO browser opens, choose logan.glosser.eras@gmail.com")
        print("    # Then re-run this script.")
        sys.exit(1)
    print(f"→ Extracted to {TMP_PARQUET}")


def build_excel():
    df = pd.read_parquet(TMP_PARQUET)
    print(f"→ Loaded {len(df):,} rows × {len(df.columns)} cols")

    # Source Map (per-column provenance, mirrors M044 schema)
    source_map = []
    DB = "thyroid_canonical_publication_v1_0"
    COHORT = f"{DB}.manuscript_workspace.cohort_m038_massive_goiter_v1"
    CPM = f"{DB}.main.canonical_patient_master"
    for col in df.columns:
        # Best-effort provenance lookup
        if col == "research_id":
            src = (DB, "manuscript_workspace.cohort_m038_massive_goiter_v1", "research_id", "PASS", "VARCHAR", "ALL (key)", "Patient identifier; joins all sources.")
        elif col == "is_massive":
            src = (DB, "DERIVED", "6-flag disjunction", "DERIVED", "BOOLEAN", "PRIMARY EXPOSURE", "Composite per Methods §2.3: weight≥100g OR substernal(CT|MRI) OR airway(CT)")
        elif col in ("comp_weight_ge100","comp_substernal_any","comp_airway_any"):
            src = (DB, "DERIVED", "(see column name)", "DERIVED", "BOOLEAN", "Composite component", "Per Methods §2.3 component definition")
        elif col == "era_bucket":
            src = (DB, "DERIVED", "surg_first_date (CASE)", "DERIVED", "VARCHAR", "Era stratification §3.6", "Upper-bound binning rule sweeps pre-1999 dates into 1999–2004")
        elif col == "hypopara_postop_class":
            src = (DB, "DERIVED", "comp_hypoparathyroidism_confirmed × _transient/_permanent", "DERIVED", "VARCHAR", "Standing rule §3.5b", "transient_lt_6mo / permanent_gt_6mo / unclassified / none")
        elif col == "hca_preop_flag":
            src = (DB, "DERIVED", "comp_hypocalcemia_timing_window OR _clinical_preexisting", "DERIVED", "BOOLEAN", "Standing rule §3.5b", "TRUE if pre-surgery hypocalcemia by either signal")
        elif col.startswith("comp_") and col.endswith(("_transient","_permanent","_timing_window","_preexisting","_new_postop","_limitation_note","_clinical_preexisting")):
            src = (DB, COHORT.split(".",1)[1], col, "PASS", str(df[col].dtype), "Standing rule (§3.5b)", "Post-mig_255 cohort view passthrough from canonical_patient_master")
        elif col.startswith("comp_"):
            src = (DB, COHORT.split(".",1)[1], col, "PASS", str(df[col].dtype), "§3.5 Table 4", "Strict rollup post-mig_252 (present + def/probable)")
        elif col.startswith("nsqip_"):
            src = (DB, COHORT.split(".",1)[1], col, "PASS", str(df[col].dtype), "§3.4 Table 3 (NSQIP-linked subset)", "NSQIP perioperative; NULL = not in NSQIP linkage")
        elif col.startswith("pmhx_nlp_"):
            src = (DB, COHORT.split(".",1)[1], col, "PASS", str(df[col].dtype), "§3.2 Table 1 (NLP comorbidities)", "NLP-extracted from PMHx narrative")
        elif col.startswith("syn_"):
            src = (DB, COHORT.split(".",1)[1], col, "PASS", str(df[col].dtype), "§3.2 Table 1 (thyroid-specific)", "Synoptic pathology field")
        elif col.startswith("pshx_nlp_"):
            src = (DB, COHORT.split(".",1)[1], col, "PASS", str(df[col].dtype), "§3.2 Table 1 (surgical history)", "NLP-extracted from PSHx narrative")
        elif col.startswith(("ct_", "mri_", "nlp_airway_")):
            src = (DB, COHORT.split(".",1)[1], col, "PASS", str(df[col].dtype), "Composite-flag component (§2.3)", "Imaging-derived; NULL = not documented")
        elif col == "histology_final":
            src = (DB, COHORT.split(".",1)[1], col, "PASS", "VARCHAR", "§3.3 Table 2", "Resolved histologic dx (single category per patient)")
        elif col == "is_malignant":
            src = (DB, COHORT.split(".",1)[1], col, "PASS", "BOOLEAN", "§3.2/§3.3", "Malignancy flag")
        elif col == "gland_weight_final_g":
            src = (DB, COHORT.split(".",1)[1], col, "PASS", "DOUBLE", "Composite-flag component / §3.4", "Synoptic pathology gland weight (grams)")
        elif col in ("surg_procedure_type","surg_total_thyroidectomy","surg_hemithyroidectomy"):
            src = (DB, COHORT.split(".",1)[1], col, "PASS", str(df[col].dtype), "§3.4 Table 3", "Post-mig_253 procedure-type fill")
        elif col == "surg_first_date":
            src = (DB, COHORT.split(".",1)[1], col, "PASS", "DATE", "§3.6 Era / §3.2 Table 1", "Post-mig_254 backfill from first_surgery_date_v2")
        elif col in ("followup_years","death_occurred"):
            src = (DB, COHORT.split(".",1)[1], col, "PASS", str(df[col].dtype), "§3.5 narrative / Table 4 mortality", "Survival rollup")
        elif col in ("bilateral_disease_flag","bilateral_path_flag"):
            src = (DB, COHORT.split(".",1)[1], col, "PASS", "BOOLEAN", "§3.2 Table 1 (pathology)", "")
        elif col.startswith("proc_nlp_tracheostomy"):
            src = (DB, COHORT.split(".",1)[1], col, "PASS", str(df[col].dtype), "§3.4 Table 3 (tracheostomy)", "NLP-extracted")
        else:
            src = (DB, COHORT.split(".",1)[1], col, "PASS", str(df[col].dtype), "Reference column", "Available for downstream use")
        source_map.append((col,) + src)

    # QA gaps
    qa = []
    for _, r in df.iterrows():
        flags = []
        if pd.isna(r.get("surg_first_date")): flags.append("surg_first_date_missing")
        if pd.isna(r.get("bmi_combined")): flags.append("bmi_missing")
        if r.get("followup_years") == 0: flags.append("zero_followup")
        if pd.isna(r.get("histology_final")) and r.get("is_malignant"): flags.append("malignant_no_histology")
        if pd.isna(r.get("nsqip_asa_class")): flags.append("asa_class_null")
        if flags:
            qa.append({"research_id": r["research_id"], "is_massive": r.get("is_massive"),
                       "n_flags": len(flags), "flags": "; ".join(flags)})

    # ---------- Build workbook ----------
    wb = Workbook(); wb.remove(wb.active)
    HF = PatternFill("solid", fgColor="1F4E78"); HFONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    BODY = Font(name="Arial", size=10); BD = Border(left=Side(style="thin", color="CCCCCC"),
                                                     right=Side(style="thin", color="CCCCCC"),
                                                     top=Side(style="thin", color="CCCCCC"),
                                                     bottom=Side(style="thin", color="CCCCCC"))
    AL = Alignment(horizontal="left", vertical="center")
    AR = Alignment(horizontal="right", vertical="center")
    AC = Alignment(horizontal="center", vertical="center")

    def write_df(name, df, freeze_to_col=2):
        ws = wb.create_sheet(name)
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(row=1, column=ci, value=col); c.font = HFONT; c.fill = HF; c.alignment = AC; c.border = BD
        for ri, row in enumerate(df.itertuples(index=False), 2):
            for ci, v in enumerate(row, 1):
                if pd.isna(v): v = None
                # convert pandas Timestamp/Timedelta to native types
                if hasattr(v, "to_pydatetime"):
                    v = v.to_pydatetime().date() if hasattr(v.to_pydatetime(), "date") else v.to_pydatetime()
                c = ws.cell(row=ri, column=ci, value=v); c.font = BODY; c.border = BD
                c.alignment = AR if isinstance(v, (int, float)) else AL
        ws.freeze_panes = ws.cell(row=2, column=freeze_to_col)
        # autosize first ~20 cols
        for i, col in enumerate(df.columns[:20], 1):
            try:
                ws.column_dimensions[get_column_letter(i)].width = min(max(len(str(col)) + 2, 14), 36)
            except Exception:
                pass
        return ws

    # README
    ws = wb.create_sheet("README")
    readme = [
        ("Workbook", "M038_per_patient_with_sources.xlsx"),
        ("Manuscript", "M038 v2 — Massive Goiter Composite-Definition Descriptive Cohort"),
        ("Database", "thyroid_canonical_publication_v1_0 (release pub_v1_1_20260504)"),
        ("Cohort view", "manuscript_workspace.cohort_m038_massive_goiter_v1 (post-mig_255)"),
        ("Rows", f"{len(df):,} (1 row per research_id)"),
        ("Columns", f"{len(df.columns)}"),
        ("",""),
        ("Sheets", ""),
        ("  Per-patient analytic", "Patient-grain wide table"),
        ("  Source map", "Per-column DB.schema.table.column + aggregation rule + dtype + manuscript section"),
        ("  Audit (gaps)", "Patients with missing key fields"),
        ("",""),
        ("Standing rule", "memory/feedback_complications_transient_vs_permanent.md"),
        ("Audit doc", "manuscript_outputs/v1_0_20260501/M038_v2_DATA_VALIDITY_AUDIT_20260501.md"),
    ]
    for i, (k, v) in enumerate(readme, 1):
        ws.cell(row=i, column=1, value=k).font = Font(name="Arial", size=11, bold=True)
        ws.cell(row=i, column=2, value=v).font = BODY
    ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 80

    write_df("Per-patient analytic", df)
    write_df("Source map", pd.DataFrame(source_map,
        columns=["column_name","source_database","source_object","source_column","aggregation_rule","data_type","used_in_sections","notes"]))
    write_df("Audit (gaps)", pd.DataFrame(qa) if qa else pd.DataFrame([{"note":"no gaps flagged"}]))

    wb.save(OUT)
    print(f"→ Saved {OUT}")


if __name__ == "__main__":
    extract()
    build_excel()
