"""Build M038_submission_package_v1_0/04_tables.xlsx.

Mirrors M044 04_tables.xlsx structure: Cover + Table 1-5 (manuscript) +
Supp S1-S5 + Data dictionary + QA. Values hard-coded from live SQL re-derivation
performed during the 2026-05-01 audit (see audit doc + Excel build trail).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

OUT = "/Users/loganglosser/THYROID_2026/M038_submission_package_v1_0/04_tables.xlsx"

wb = Workbook(); wb.remove(wb.active)

# Styling
HF = PatternFill("solid", fgColor="1F4E78"); HFONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TF = Font(name="Arial", bold=True, size=14, color="1F4E78"); SF = Font(name="Arial", bold=True, size=11, color="404040")
BODY = Font(name="Arial", size=10); DESC = Font(name="Arial", size=10, italic=True, color="595959")
SECT = Font(name="Arial", bold=True, size=10, color="1F4E78")
THIN = Side(style="thin", color="CCCCCC"); BD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
AL = Alignment(horizontal="left", vertical="center"); AR = Alignment(horizontal="right", vertical="center")
AC = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
PCT = "0.00%"; PCT1 = "0.0%"; INT = "#,##0"; RR = "0.00"; DEC2 = "0.00"

def write(name, desc, headers, rows, widths, fmts=None, footer=None, freeze=True):
    ws = wb.create_sheet(name)
    ws["A1"] = desc; ws["A1"].font = DESC; ws["A1"].alignment = WRAP
    if len(desc) > 100:
        ws.row_dimensions[1].height = 45
    # blank row 2
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h); c.font = HFONT; c.fill = HF; c.alignment = AC; c.border = BD
    for ri, row in enumerate(rows, 4):
        if isinstance(row, str):
            # section subheader
            c = ws.cell(row=ri, column=1, value=row); c.font = SECT
            c.fill = PatternFill("solid", fgColor="DDEBF7")
            ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=len(headers))
            continue
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=v); c.font = BODY; c.border = BD
            c.alignment = AR if isinstance(v, (int, float)) else WRAP
            if fmts and ci in fmts: c.number_format = fmts[ci]
    if footer:
        fr = 4 + len(rows)
        c = ws.cell(row=fr, column=1, value=footer); c.font = DESC; c.alignment = WRAP
        ws.merge_cells(start_row=fr, start_column=1, end_row=fr, end_column=len(headers))
        ws.row_dimensions[fr].height = 80
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if freeze: ws.freeze_panes = "A4"
    return ws

# ============================================================
# Cover
# ============================================================
ws = wb.create_sheet("Cover")
ws["A1"] = "M038 — Massive Goiter at a Tertiary Referral Center"
ws["A1"].font = TF; ws.merge_cells("A1:D1")
ws["A2"] = "A Composite-Definition Descriptive Cohort of 2,501 Patients (Emory University, 1999-2025)"
ws["A2"].font = SF; ws.merge_cells("A2:D2")
cover = [
    ("Manuscript", "M038 v2 — Massive Goiter Composite-Definition Descriptive Cohort"),
    ("Source draft", "manuscript_outputs/v1_0_20260501/M038_massive_goiter_DRAFT_v2_post_mig_252_253.md"),
    ("Status", "Submission package v1.0 (2026-05-01)"),
    ("Target journal", "Surgery / Annals of Surgical Oncology / Thyroid (TBD)"),
    ("",""),
    ("Database", "thyroid_canonical_publication_v1_0 (MotherDuck)"),
    ("Release ID", "pub_v1_0_20260430"),
    ("Cohort view", "manuscript_workspace.cohort_m038_massive_goiter_v1"),
    ("Most-recent migration", "mig_255 (cohort_m038 complication temporality columns) — applied 2026-05-02"),
    ("Upstream fixes", "mig_252 (strict comp_*_confirmed rollup) + mig_253 (surg_procedure_type fill) + mig_254 (surg_first_date backfill) + mig_255 (temporality columns)"),
    ("",""),
    ("Tabs in this workbook", ""),
    ("  Table 1 — Baseline", "Demographics & comorbidities, massive vs non-massive"),
    ("  Table 2 — Histology", "Malignant histology distribution within massive arm"),
    ("  Table 3 — Procedure", "Procedure type + operative context (NSQIP)"),
    ("  Table 4 — Complications", "Strict-definition complications (per standing rule)"),
    ("  Table 5 — Era stratification", "5-year buckets, 1999-2025"),
    ("  Supp S1 — Cohort assembly", "Composite-flag composition + inclusion-exclusion check"),
    ("  Supp S2 — Component coverage", "Per-era × per-arm coverage of 6 component-flag source columns"),
    ("  Supp S3 — Race detail", "Full 9-bucket race breakdown"),
    ("  Supp S4 — ASA detail", "Full ASA-class breakdown (NSQIP-linked subset)"),
    ("  Supp S5 — Hypopara limitation notes", "Permanence-classification quality notes"),
    ("  Supp S6 — Sensitivity ≥200g", "Weight-only focal cohort (n=475)"),
    ("  Data dictionary", "Per-column type / definition / source / range"),
    ("  QA reconciliation", "All audit cells reconciled to live SQL"),
    ("",""),
    ("Audit reference", "manuscript_outputs/v1_0_20260501/M038_v2_DATA_VALIDITY_AUDIT_20260501.md"),
    ("Standing rule applied", "memory/feedback_complications_transient_vs_permanent.md"),
    ("Generated (UTC)", datetime.utcnow().isoformat(timespec="seconds") + "Z"),
    ("Cowork commits", "b232007 + c3203b0 + b6cc14d (mig_255)"),
]
for i, (k, v) in enumerate(cover, 4):
    ws.cell(row=i, column=1, value=k).font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=i, column=2, value=v).font = BODY
    ws.cell(row=i, column=2).alignment = WRAP
ws.column_dimensions["A"].width = 38; ws.column_dimensions["B"].width = 90

# ============================================================
# Table 1 — Demographics & Baseline
# ============================================================
write(
    "Table 1 — Baseline",
    "Demographics & baseline characteristics, composite-massive vs non-massive arms. "
    "Whole-cohort denominators (massive 2,501; non-massive 8,370) unless 'Subset n' noted. "
    "Footnote: BMI subset = bmi_combined non-NULL; ASA subset = NSQIP-linked. Two pre-1999 dates "
    "(1945-07-13, 1993-04-01) are swept into the 1999–2004 era bucket via upper-bound binning.",
    ["Characteristic","Massive (n=2,501)","Massive %","Non-massive (n=8,370)","Non-massive %","Subset n (M)","Subset n (NM)","Notes"],
    [
        "Age at surgery",
        ("  Mean (years)", 55.4, None, 50.5, None, None, None, ""),
        ("  Median (years)", 56, None, 50, None, None, None, ""),
        ("  IQR Q25–Q75", "45–66", None, "39–62", None, None, None, ""),
        "Sex",
        ("  Female", 1771, 1771/2501, 6688, 6688/8370, None, None, "sex='female'"),
        ("  Male", 730, 730/2501, 1682, 1682/8370, None, None, "sex='male'"),
        "Race",
        ("  Black or African American", 1555, 1555/2501, 2613, 2613/8370, None, None, ""),
        ("  White", 714, 714/2501, 4552, 4552/8370, None, None, ""),
        ("  Asian", 57, 57/2501, 419, 419/8370, None, None, ""),
        ("  Other / AIAN / NH-PI / Hispanic", 44, 44/2501, 187, 187/8370, None, None, "Composite of 4 buckets (see Supp S3)"),
        ("  Unknown or Not Reported", 130, 130/2501, 591, 591/8370, None, None, "+1/+8 NULL counted separately"),
        "BMI (kg/m²)",
        ("  Mean", 33.5, None, 29.8, None, 417, 1668, "bmi_combined non-NULL"),
        ("  Median", 32.1, None, 28.5, None, 417, 1668, ""),
        ("  IQR Q25–Q75", "27.7–37.5", None, "24.4–33.6", None, 417, 1668, ""),
        "NLP-extracted comorbidities",
        ("  Hypertension", 696, 696/2501, 1079, 1079/8370, None, None, ""),
        ("  Diabetes mellitus", 500, 500/2501, 966, 966/8370, None, None, ""),
        ("  Coronary artery disease", 84, 84/2501, 140, 140/8370, None, None, ""),
        ("  Chronic kidney disease", 85, 85/2501, 136, 136/8370, None, None, ""),
        ("  COPD", 47, 47/2501, 60, 60/8370, None, None, ""),
        ("  Mean N comorbidities (NLP)", 2.78, None, 2.38, None, None, None, ""),
        "Thyroid-specific history",
        ("  Graves disease (synoptic)", 108, 108/2501, 466, 466/8370, None, None, ""),
        ("  Hashimoto thyroiditis (synoptic)", 39, 39/2501, 209, 209/8370, None, None, ""),
        ("  Prior thyroidectomy (NLP)", 209, 209/2501, 650, 650/8370, None, None, ""),
        ("  Prior neck surgery (NLP)", 38, 38/2501, 102, 102/8370, None, None, ""),
        "ASA class (NSQIP-linked subset)",
        ("  ASA I — Normal/Healthy", 6, 6/246, 84, 84/1164, 246, 1164, ""),
        ("  ASA II — Mild systemic", 80, 80/246, 583, 583/1164, 246, 1164, ""),
        ("  ASA III — Severe systemic", 144, 144/246, 473, 473/1164, 246, 1164, ""),
        ("  ASA IV — Severe / threat", 16, 16/246, 24, 24/1164, 246, 1164, ""),
        "Surgical era",
        ("  1999–2004", 110, 110/2501, 793, 793/8370, None, None, "Includes 2 pre-1999 cases"),
        ("  2005–2009", 142, 142/2501, 1049, 1049/8370, None, None, ""),
        ("  2010–2014", 240, 240/2501, 1645, 1645/8370, None, None, ""),
        ("  2015–2019", 731, 731/2501, 2204, 2204/8370, None, None, ""),
        ("  2020–2025", 517, 517/2501, 1300, 1300/8370, None, None, ""),
        ("  Surgical date unknown", 761, 761/2501, 1379, 1379/8370, None, None, ""),
        "Pathology",
        ("  Malignant histology", 646, 646/2501, 3491, 3491/8370, None, None, "is_malignant"),
        ("  Bilateral disease", 749, 749/2501, 1393, 1393/8370, None, None, ""),
        "Follow-up",
        ("  Mean (years, all)", 1.22, None, 1.84, None, None, None, ""),
        ("  Patients with FU>0", 997, 997/2501, 3174, 3174/8370, None, None, ""),
        ("  Mean (years, FU>0 subset)", 3.06, None, 4.85, None, 997, 3174, ""),
    ],
    [42,18,14,18,14,12,12,32],
    {2: DEC2, 3: PCT, 4: DEC2, 5: PCT, 6: INT, 7: INT},
)

# ============================================================
# Table 2 — Histology
# ============================================================
write(
    "Table 2 — Histology",
    "Histology distribution within the malignant subset of the massive cohort (n=646). "
    "Cross-reference: M032 broader malignant cohort n=4,022, PTC 80.9%.",
    ["Histology","n","% of 646","Notes"],
    [
        ("PTC", 417, 417/646, ""),
        ("Follicular carcinoma", 97, 97/646, ""),
        ("MTC (medullary)", 32, 32/646, ""),
        ("NIFTP", 25, 25/646, ""),
        ("Poorly differentiated thyroid carcinoma", 22, 22/646, ""),
        ("Anaplastic carcinoma", 13, 13/646, ""),
        ("FTUMP", 9, 9/646, ""),
        ("Metastatic / rare variants (long tail, 14 categories)", 31, 31/646, "Includes NUT n=1, thymus-like n=1"),
        ("TOTAL malignant in massive arm", 646, 1.0, ""),
    ],
    [50, 12, 14, 40],
    {2: INT, 3: PCT},
)

# ============================================================
# Table 3 — Procedure type + op context
# ============================================================
write(
    "Table 3 — Procedure",
    "Surgical procedure type (whole-arm denominator) + operative context (NSQIP-derived; "
    "denominators heterogeneous; LOS uses nsqip_length_of_stay_days, n=246/1,164).",
    ["Variable","Massive value","Massive %","Non-massive value","Non-massive %","Notes"],
    [
        "Procedure type",
        ("  Total thyroidectomy", 1672, 1672/2501, 4327, 4327/8370, ""),
        ("  Hemithyroidectomy", 792, 792/2501, 3640, 3640/8370, ""),
        ("  Other", 36, 36/2501, 386, 386/8370, ""),
        ("  Isthmusectomy", 1, 1/2501, 6, 6/8370, ""),
        ("  Unknown / NULL", 0, 0/2501, 11, 11/8370, "9 'unknown' string + 2 NULL"),
        ("  Procedure-type completeness", None, 1.0, None, 8368/8370, "Post-mig_253"),
        "Operative context (NSQIP, where coverage permits)",
        ("  Central neck dissection (n)", 55, None, 193, None, ""),
        ("  Lateral neck dissection (n)", 19, None, 20, None, ""),
        ("  Mean operative duration (min)", 130.83, None, 121.33, None, ""),
        ("  Median operative duration (min)", 113.5, None, 107, None, ""),
        ("  Mean hospital LOS (days)", 1.264, None, 1.067, None, "nsqip_length_of_stay_days"),
        ("  Median hospital LOS (days)", 1, None, 1, None, ""),
        ("  Transfusion (NSQIP, ≥1 unit) (n)", 2, None, 2, None, ""),
        ("  Unplanned reintubation (n)", 5, None, 7, None, ""),
        ("  30-day readmission (n)", 11, None, 18, None, ""),
        ("  NLP tracheostomy (any timing)", 121, 121/2501, 263, 263/8370, ""),
    ],
    [44, 18, 14, 18, 14, 32],
    {2: DEC2, 3: PCT, 4: DEC2, 5: PCT},
)

# ============================================================
# Table 4 — Complications (per standing rule)
# ============================================================
write(
    "Table 4 — Complications",
    "Strict-definition perioperative complications (post-mig_252). Strict postop definition: "
    "comp_*_confirmed = TRUE ⇔ underlying finding_status='present' AND evidence_strength IN "
    "('definitive','probable'). Per standing rule (memory/feedback_complications_transient_vs_permanent.md): "
    "hypoparathyroidism is split into postop transient (<6mo) vs postop permanent (>6mo); "
    "hypocalcemia adds present-preop flag (timing_window='pre_surgery'); RLN injury and VC "
    "paralysis preop status not currently encoded (carry-forwards open).",
    ["Outcome","Massive n","Massive %","Non-massive n","Non-massive %","RR (M/NM)","Notes"],
    [
        "Postop confirmed (whole-arm denominator)",
        ("Any confirmed complication flag", 132, 132/2501, 268, 268/8370, (132/2501)/(268/8370), ""),
        ("Confirmed hematoma", 23, 23/2501, 45, 45/8370, (23/2501)/(45/8370), ""),
        ("Confirmed seroma", 12, 12/2501, 27, 27/8370, (12/2501)/(27/8370), ""),
        ("Confirmed chyle leak", 2, 2/2501, 1, 1/8370, (2/2501)/(1/8370), "Small counts; CI wide"),
        ("Confirmed RLN injury (postop) †", 14, 14/2501, 7, 7/8370, (14/2501)/(7/8370), ""),
        ("Confirmed VC paresis (postop) †", 0, 0, 0, 0, None, ""),
        ("Confirmed VC paralysis (postop) †", 19, 19/2501, 4, 4/8370, (19/2501)/(4/8370), ""),
        ("Confirmed hypocalcemia (postop)", 1, 1/2501, 8, 8/8370, (1/2501)/(8/8370), "Excludes preop"),
        ("All-cause in-record mortality", 59, 59/2501, 133, 133/8370, (59/2501)/(133/8370), ""),
        "Hypoparathyroidism — temporality split (standing rule)",
        ("Hypoparathyroidism — postop transient (<6mo)", 83, 83/2501, 197, 197/8370, (83/2501)/(197/8370), ""),
        ("Hypoparathyroidism — postop permanent (>6mo)", 4, 4/2501, 12, 12/8370, (4/2501)/(12/8370), ""),
        ("Hypoparathyroidism — postop unclassified", 0, 0, 0, 0, None, "Zero in M038"),
        ("  Sum check (transient + permanent)", 87, 87/2501, 209, 209/8370, (87/2501)/(209/8370), "83+4=87 ✓; 197+12=209 ✓"),
        "Preop yes/no flag (standing rule)",
        ("Hypocalcemia — present preop", 7, 7/2501, 46, 46/8370, None, "timing_window='pre_surgery'"),
        ("Hypoparathyroidism — preexisting (FYI)", 17, 17/2501, 42, 42/8370, None, "Not used in trans/perm split"),
        ("RLN injury — present preop †", None, None, None, None, None, "Not encoded; CF-RLN-PREOP-FLAG"),
        ("VC paralysis — present preop †", None, None, None, None, None, "Not encoded; CF-VC-PARALYSIS-PREOP-FLAG"),
        ("VC paresis — present preop †", None, None, None, None, None, "Not encoded; same CF as VC paralysis"),
    ],
    [50, 12, 14, 16, 16, 14, 32],
    {2: INT, 3: PCT, 4: INT, 5: PCT, 6: RR},
    footer="† Preop status for RLN injury, VC paresis, and VC paralysis is not currently encoded in canonical_patient_master; postop confirmed counts only. Carry-forwards: CF-RLN-PREOP-FLAG, CF-VC-PARALYSIS-PREOP-FLAG. Hypopara permanence-classification limitation notes for 14 cohort-wide cases — see Supp S5.",
)

# ============================================================
# Table 5 — Era stratification
# ============================================================
write(
    "Table 5 — Era",
    "Era stratification of composite-massive flag prevalence. Era binning: surg_first_date "
    "<= '2004-12-31' first bucket (sweeps 2 pre-1999 dates); 5-year buckets; NULL → unknown.",
    ["Era","Total n","Massive n","Non-massive n","% Massive","Notes"],
    [
        ("1999–2004", 903, 110, 793, 110/903, "Includes 1945-07-13, 1993-04-01"),
        ("2005–2009", 1191, 142, 1049, 142/1191, ""),
        ("2010–2014", 1885, 240, 1645, 240/1885, ""),
        ("2015–2019", 2935, 731, 2204, 731/2935, ""),
        ("2020–2025", 1817, 517, 1300, 517/1817, ""),
        ("Surgical date unknown", 2140, 761, 1379, 761/2140, "surg_first_date IS NULL"),
        ("TOTAL", 10871, 2501, 8370, 2501/10871, ""),
    ],
    [22, 12, 12, 16, 14, 38],
    {2: INT, 3: INT, 4: INT, 5: PCT},
)

# ============================================================
# Supp S1 — Cohort assembly + inclusion-exclusion
# ============================================================
write(
    "Supp S1 — Cohort assembly",
    "Composite-flag composition (per Methods §2.3) + inclusion-exclusion check.",
    ["Component / overlap","n","% of cohort (10,871)","% of massive (2,501)","Notes"],
    [
        ("N total cohort", 10871, 1.0, None, ""),
        ("N massive (composite)", 2501, 2501/10871, 1.0, ""),
        ("Weight ≥100 g", 1429, 1429/10871, 1429/2501, ""),
        ("Substernal (CT or MRI)", 1047, 1047/10871, 1047/2501, ""),
        ("Airway compromise (CT)", 1440, 1440/10871, 1440/2501, ""),
        ("Weight ∩ Substernal", 404, 404/10871, 404/2501, ""),
        ("Weight ∩ Airway", 513, 513/10871, 513/2501, ""),
        ("Substernal ∩ Airway", 884, 884/10871, 884/2501, ""),
        ("All three components", 386, 386/10871, 386/2501, ""),
        ("Weight only", 898, 898/10871, 898/2501, ""),
        ("Substernal only", 145, 145/10871, 145/2501, "v2.1 corrected from 114"),
        ("Airway only", 429, 429/10871, 429/2501, "v2.1 corrected from 309"),
        ("Inclusion-exclusion sum", 2501, None, None, "1429+1047+1440 - 404-513-884 + 386 = 2501 ✓"),
    ],
    [42, 12, 22, 22, 36],
    {2: INT, 3: PCT, 4: PCT},
)

# ============================================================
# Supp S2 — Component coverage by era × arm
# ============================================================
cov = [
    ("1999–2004","Massive",110,110,1,0,1,1,1),
    ("1999–2004","Non-massive",793,673,0,0,0,0,0),
    ("2005–2009","Massive",142,141,0,0,1,1,1),
    ("2005–2009","Non-massive",1049,918,1,1,0,0,2),
    ("2010–2014","Massive",240,234,25,8,31,31,39),
    ("2010–2014","Non-massive",1645,1505,17,6,35,23,71),
    ("2015–2019","Massive",731,657,406,72,473,457,497),
    ("2015–2019","Non-massive",2204,1886,153,46,224,186,388),
    ("2020–2025","Massive",517,421,331,65,380,373,404),
    ("2020–2025","Non-massive",1300,988,128,59,215,179,378),
    ("Unknown","Massive",761,595,566,114,634,617,683),
    ("Unknown","Non-massive",1379,1002,176,91,335,274,570),
]
write(
    "Supp S2 — Coverage",
    "Per-era × per-arm coverage of the six composite-flag source columns. 'Coverage' = "
    "fraction with non-NULL value (TRUE or FALSE). Documents the era-rise narrative: "
    "pre-2010 CT/MRI documentation was essentially absent.",
    ["Era","Arm","Arm n","gland_weight n","gland_weight cov","ct_subst n","ct_subst cov","mri_subst n","mri_subst cov","ct_trach_dev n","ct_trach_dev cov","ct_trach_narr n","ct_trach_narr cov","ct_airway n","ct_airway cov"],
    [(era, arm, n,
      gw, gw/n, cts, cts/n, mrs, mrs/n, ctd, ctd/n, ctn, ctn/n, cta, cta/n) for (era, arm, n, gw, cts, mrs, ctd, ctn, cta) in cov],
    [14, 14, 10] + [12, 14] * 6,
    {3: INT, 4: INT, 5: PCT1, 6: INT, 7: PCT1, 8: INT, 9: PCT1, 10: INT, 11: PCT1, 12: INT, 13: PCT1, 14: INT, 15: PCT1},
)

# ============================================================
# Supp S3 — Race detail
# ============================================================
write(
    "Supp S3 — Race",
    "Full per-arm race breakdown (9 buckets + NULL). Manuscript Table 1 collapses 4 buckets "
    "(Other, AIAN, NH-PI, Hispanic) into one row.",
    ["Race","Massive n","Massive %","Non-massive n","Non-massive %"],
    [
        ("Black or African American", 1555, 1555/2501, 2613, 2613/8370),
        ("White", 714, 714/2501, 4552, 4552/8370),
        ("Unknown or Not Reported", 130, 130/2501, 591, 591/8370),
        ("Asian", 57, 57/2501, 419, 419/8370),
        ("Other", 38, 38/2501, 105, 105/8370),
        ("American Indian or Alaska Native", 4, 4/2501, 35, 35/8370),
        ("Native Hawaiian or Other Pacific Islander", 1, 1/2501, 26, 26/8370),
        ("Hispanic or Latino", 1, 1/2501, 21, 21/8370),
        ("NULL (race not coded)", 1, 1/2501, 8, 8/8370),
        ("TOTAL", 2501, 1.0, 8370, 1.0),
    ],
    [44, 14, 14, 16, 16],
    {2: INT, 3: PCT, 4: INT, 5: PCT},
)

# ============================================================
# Supp S4 — ASA detail
# ============================================================
write(
    "Supp S4 — ASA",
    "Full per-arm ASA-class breakdown (NSQIP-linked subset only).",
    ["ASA class (verbatim)","Massive n","Massive %","Non-massive n","Non-massive %"],
    [
        ("ASA  I - Normal/Healthy", 6, 6/246, 84, 84/1164),
        ("ASA II - Mild systemic disease", 80, 80/246, 583, 583/1164),
        ("ASA III - Severe systemic disease", 144, 144/246, 473, 473/1164),
        ("ASA IV - Severe / threat to life", 16, 16/246, 24, 24/1164),
        ("TOTAL (NSQIP-linked)", 246, 1.0, 1164, 1.0),
    ],
    [50, 14, 14, 16, 16],
    {2: INT, 3: PCT, 4: INT, 5: PCT},
)

# ============================================================
# Supp S5 — Hypopara permanence limitation notes
# ============================================================
write(
    "Supp S5 — Hypopara limitations",
    "Hypoparathyroidism permanence-classification limitation notes (per standing rule §4). "
    "These notes signal where the trans/perm assignment is constrained by short follow-up, "
    "post-reset data uncertainty, or absent persistent biochem evidence.",
    ["Limitation note","Massive n","Non-massive n","Total","Interpretation"],
    [
        ("(none — clean classification)", 84, 198, 84+198, "Trans/perm assignment uses full evidence"),
        ("followup_too_short_for_permanence_classification", 1, 5, 6, "Follow-up < 180 days after surgery — permanence cannot be determined"),
        ("reset_20260417:confirmed_duration_unknown", 1, 4, 5, "Post-reset case; pre-reset duration data was lost"),
        ("confirmed_hypopara_no_persistent_biochem_evidence_followup_gt_6mo", 1, 2, 3, "Followup >6mo but no persistent biochem evidence — leans transient"),
        ("TOTAL confirmed hypoparathyroidism cases", 87, 209, 87+209, ""),
    ],
    [60, 14, 16, 12, 60],
    {2: INT, 3: INT, 4: INT},
)

# ============================================================
# Supp S6 — Sensitivity ≥200g focal cohort
# ============================================================
write(
    "Supp S6 — Sensitivity ≥200g",
    "Sensitivity analysis: weight-only ≥200g focal cohort (per author-input gap #5). "
    "n=475 patients with gland_weight_final_g ≥ 200; comparison arm n=10,396.",
    ["Outcome","≥200g (n=475)","≥200g %","<200g or NULL (n=10,396)","<200g %","Notes"],
    [
        ("Cohort n", 475, 475/10871, 10396, 10396/10871, ""),
        ("Any confirmed complication", 10, 10/475, 390, 390/10396, "RR ≈ 0.56 — small numerator"),
        ("Hypoparathyroidism — postop transient", 7, 7/475, None, None, ""),
        ("Hypoparathyroidism — postop permanent", 0, 0, None, None, ""),
        ("Confirmed RLN injury (postop)", 2, 2/475, None, None, ""),
        ("Confirmed VC paralysis (postop)", 2, 2/475, None, None, ""),
    ],
    [44, 16, 14, 28, 14, 36],
    {2: INT, 3: PCT, 4: INT, 5: PCT},
)

# ============================================================
# Data dictionary
# ============================================================
dd = [
    # (column, type, definition, allowed_values_or_range, source)
    ("research_id", "VARCHAR", "Patient identifier (joins all sources)", "Free-text key", "canonical_patient_master.research_id"),
    ("is_massive", "BOOLEAN", "Composite massive-goiter flag", "TRUE/FALSE", "DERIVED: 6-flag disjunction per Methods §2.3"),
    ("comp_weight_ge100", "BOOLEAN", "Gland weight ≥100 g component", "TRUE/FALSE", "DERIVED: gland_weight_final_g >= 100"),
    ("comp_substernal_any", "BOOLEAN", "Substernal extension (CT or MRI)", "TRUE/FALSE", "DERIVED: ct_substernal_extension_any OR mri_substernal_any"),
    ("comp_airway_any", "BOOLEAN", "Airway compromise (CT)", "TRUE/FALSE", "DERIVED: ct_tracheal_deviation_any OR ct_tracheal_narrowing_any OR ct_airway_compromise_any"),
    ("era_bucket", "VARCHAR", "Surgical-era 5-year bucket", "1999-2004 / 2005-2009 / 2010-2014 / 2015-2019 / 2020-2025 / unknown", "DERIVED: surg_first_date upper-bound binning"),
    ("hypopara_postop_class", "VARCHAR", "Hypopara temporality class", "transient_lt_6mo / permanent_gt_6mo / unclassified / none", "DERIVED: comp_hypoparathyroidism_confirmed × _transient/_permanent"),
    ("hca_preop_flag", "BOOLEAN", "Hypocalcemia present preop", "TRUE/FALSE", "DERIVED: comp_hypocalcemia_timing_window='pre_surgery' OR _clinical_preexisting"),
    ("age_at_surgery", "BIGINT", "Patient age at index surgery (years)", "0–110", "canonical_patient_master.age_at_surgery"),
    ("sex", "VARCHAR", "Patient sex", "'female' / 'male'", "canonical_patient_master.sex"),
    ("race", "VARCHAR", "Patient race (self-reported)", "9 buckets (see Supp S3)", "canonical_patient_master.race"),
    ("bmi_combined", "DOUBLE", "Harmonized BMI", "10–80 kg/m²", "canonical_patient_master.bmi_combined"),
    ("nsqip_asa_class", "VARCHAR", "ASA class", "'ASA I' to 'ASA IV' (verbatim)", "canonical_patient_master.nsqip_asa_class (NSQIP-linked subset only)"),
    ("pmhx_nlp_*", "BOOLEAN", "NLP-extracted comorbidity flags", "TRUE/FALSE/NULL", "canonical_patient_master.pmhx_nlp_* (NLP pipeline)"),
    ("syn_graves / syn_hashimoto", "BOOLEAN", "Synoptic thyroid-specific dx", "TRUE/FALSE/NULL", "canonical_patient_master.syn_*"),
    ("pshx_nlp_prior_thyroidectomy", "BOOLEAN", "Prior thyroidectomy (NLP)", "TRUE/FALSE/NULL", "canonical_patient_master.pshx_nlp_*"),
    ("surg_first_date", "DATE", "Index thyroid surgery date", "1945-2025 (2 pre-1999 outliers)", "canonical_patient_master.surg_first_date (post-mig_254 backfill)"),
    ("surg_procedure_type", "VARCHAR", "Procedure type (post-mig_253)", "total_thyroidectomy / hemithyroidectomy / other / isthmusectomy / unknown / NULL", "canonical_patient_master.surg_procedure_type"),
    ("nsqip_central_neck_dissection", "VARCHAR", "Central neck dissection", "'Yes' / 'No' / NULL (NSQIP-linked)", "canonical_patient_master.nsqip_central_neck_dissection"),
    ("nsqip_lateral_neck_dissection", "VARCHAR", "Lateral neck dissection", "'Yes' / 'No' / NULL", "canonical_patient_master.nsqip_lateral_neck_dissection"),
    ("nsqip_operative_duration_min", "BIGINT", "Operative duration in minutes", "0–600", "canonical_patient_master.nsqip_operative_duration_min"),
    ("nsqip_length_of_stay_days", "BIGINT", "Length of stay (days)", "0–60+", "canonical_patient_master.nsqip_length_of_stay_days"),
    ("nsqip_transfusion", "BIGINT", "Transfusion units", "0+", "canonical_patient_master.nsqip_transfusion"),
    ("nsqip_unplanned_intubation", "BIGINT", "Unplanned reintubation", "0/1+", "canonical_patient_master.nsqip_unplanned_intubation"),
    ("nsqip_readmission_30d_flag", "BIGINT", "30-day readmission flag", "0/1", "canonical_patient_master.nsqip_readmission_30d_flag"),
    ("proc_nlp_tracheostomy", "BOOLEAN", "Tracheostomy mentioned (NLP)", "TRUE/FALSE/NULL", "canonical_patient_master.proc_nlp_tracheostomy"),
    ("gland_weight_final_g", "DOUBLE", "Gland weight (grams; harmonized)", "0–2000+", "canonical_patient_master.gland_weight_final_g"),
    ("ct_substernal_extension_any", "BOOLEAN", "CT-documented substernal extension", "TRUE/FALSE/NULL", "canonical_patient_master.ct_substernal_extension_any"),
    ("mri_substernal_any", "BOOLEAN", "MRI-documented substernal extension", "TRUE/FALSE/NULL", "canonical_patient_master.mri_substernal_any"),
    ("ct_tracheal_deviation_any / narrowing_any / airway_compromise_any", "BOOLEAN", "Airway-component flags from CT", "TRUE/FALSE/NULL", "canonical_patient_master.ct_*"),
    ("histology_final", "VARCHAR", "Resolved histologic dx", "PTC / follicular carcinoma / MTC / NIFTP / poorly differentiated / anaplastic / FTUMP / + 14 metastatic/rare", "canonical_patient_master.histology_final"),
    ("is_malignant", "BOOLEAN", "Malignancy flag", "TRUE/FALSE", "canonical_patient_master.is_malignant"),
    ("bilateral_disease_flag", "BOOLEAN", "Bilateral disease (path + imaging)", "TRUE/FALSE/NULL", "canonical_patient_master.bilateral_disease_flag"),
    ("any_confirmed_complication_flag", "BOOLEAN", "Any strict-confirmed complication", "TRUE/FALSE", "canonical_patient_master.any_confirmed_complication_flag (post-mig_252)"),
    ("comp_*_confirmed (hematoma/seroma/chyle_leak/rln_injury/vc_paresis/vc_paralysis/hypocalcemia/hypoparathyroidism)", "BOOLEAN", "Strict postop confirmation per family", "TRUE/FALSE", "canonical_patient_master.comp_*_confirmed (post-mig_252 strict definition)"),
    ("comp_hypoparathyroidism_transient", "BOOLEAN", "Hypopara resolves <6mo", "TRUE/FALSE/NULL", "canonical_patient_master.comp_hypoparathyroidism_transient (post-mig_255 view passthrough)"),
    ("comp_hypoparathyroidism_permanent", "BOOLEAN", "Hypopara persists >6mo", "TRUE/FALSE/NULL", "canonical_patient_master.comp_hypoparathyroidism_permanent"),
    ("comp_hypoparathyroidism_timing_window", "VARCHAR", "Underlying timing bucket", "pre_surgery/0_30d/31_180d/181_365d/gt_365d/unknown", "canonical_patient_master.comp_hypoparathyroidism_timing_window"),
    ("comp_hypopara_permanent_limitation_note", "VARCHAR", "Permanence-classification caveat", "(see Supp S5)", "canonical_patient_master.comp_hypopara_permanent_limitation_note"),
    ("comp_hypocalcemia_timing_window", "VARCHAR", "Hypocalcemia timing bucket", "pre_surgery/0_30d/etc.", "canonical_patient_master.comp_hypocalcemia_timing_window"),
    ("comp_hypocalcemia_clinical_preexisting", "BOOLEAN", "Hypocalcemia preop (clinical evidence)", "TRUE/FALSE/NULL", "canonical_patient_master.comp_hypocalcemia_clinical_preexisting"),
    ("comp_mortality_definitive", "BOOLEAN", "Definitive mortality flag", "TRUE/FALSE", "canonical_patient_master.comp_mortality_definitive"),
    ("death_occurred", "BOOLEAN", "All-cause in-record mortality", "TRUE/FALSE", "canonical_patient_master.death_occurred"),
    ("followup_years", "DOUBLE", "Follow-up duration (years)", "0–20+", "canonical_patient_master.followup_years"),
]
write(
    "Data dictionary",
    "Per-column type / definition / allowed-values / source. Keep aligned with 05b Source Map.",
    ["column", "type", "definition", "allowed_values_or_range", "source_object"],
    [(c, t, d, av, s) for (c, t, d, av, s) in dd],
    [42, 12, 60, 50, 80],
    {},
)

# ============================================================
# QA reconciliation
# ============================================================
write(
    "QA",
    "Audit reconciliation: 156 numeric cells in M038 v2 manuscript re-derived against live SQL "
    "on thyroid_canonical_publication_v1_0. Result: 153 PASS, 3 DIFF (v2.1 patch landed), 0 FAIL.",
    ["Section","Cell description","Manuscript value","Live result","Status"],
    [
        ("Cohort", "N total", "10,871", "10,871", "PASS"),
        ("Cohort", "N massive (composite)", "2,501 (23.0%)", "2,501 (23.00%)", "PASS"),
        ("Cohort", "Inclusion-exclusion sum", "2,501", "1429+1047+1440-404-513-884+386 = 2,501", "PASS"),
        ("§3.1", "Substernal only (v2.1 corrected)", "145 (orig 114)", "145", "PASS (post-patch)"),
        ("§3.1", "Airway only (v2.1 corrected)", "429 (orig 309)", "429", "PASS (post-patch)"),
        ("§3.2 Table 1", "All 30+ rows (age/sex/race/BMI/comorbidities/era/path/FU)", "(see manuscript)", "All match within rounding", "PASS"),
        ("§3.3 Table 2", "11 histology rows + 4,022 cross-ref to M032", "(see manuscript)", "All match", "PASS"),
        ("§3.4 Table 3", "5 procedure-type rows + 11 op-context rows", "(see manuscript)", "All match", "PASS"),
        ("§3.5 Table 4", "10 strict complication rows × counts × % × RR", "(see manuscript)", "All match (RR computed within rounding)", "PASS"),
        ("§3.5 Table 4", "Hypopara split (per standing rule)", "Massive 83+4=87; Non-massive 197+12=209", "Cohort-wide 280+16=296; M038 partition clean", "PASS (post-patch)"),
        ("§3.6 Era table", "6 era rows × total / massive / %", "(see manuscript)", "All match (post-mig_254 sweep)", "PASS"),
        ("§5 Limitations (v2.1)", "surgical date 69.6% (massive) / 80.3% (cohort-wide)", "1740/2501 = 69.57%; 8731/10871 = 80.31%", "PASS (post-patch)"),
        ("Strict-definition rollup", "comp_*_confirmed = present + def/probable", "Cohort any_confirmed 2,490 → 400", "PASS (mig_252)"),
        ("Procedure-type completeness", "100% massive / 99.98% non-massive", "0 NULL / 2 NULL out of 8,370", "PASS (mig_253)"),
        ("Strict-definition complications applied", "All 10 outcomes use mig_252 strict rollup", "Confirmed", "PASS"),
        ("Standing rule applied", "Hypopara split + hypocalcemia preop flag", "post-mig_255 view passthrough", "PASS"),
    ],
    [22, 60, 36, 50, 22],
    {},
)

wb.save(OUT)
print(f"Saved {OUT}")
print(f"Sheets: {wb.sheetnames}")
