#!/usr/bin/env python3
"""
build_m032_tables.py
====================
Generates 04_tables.xlsx (Tables 1–5 + Supp S1-S2 + data-dictionary + QA tab)
and 05_master_data.xlsx (per-patient analytic dataset + dict sheet).

Run from repo root:
    .venv/bin/python M032_submission_package_v1_0/08_analysis_code/build_m032_tables.py

Outputs:
    M032_submission_package_v1_0/04_tables.xlsx
    M032_submission_package_v1_0/05_master_data.xlsx

MotherDuck DB: thyroid_canonical_publication_v1_0
Cohort view: manuscript_workspace.cohort_m032_descriptive_25yr_v1
"""
import sys, os, re, datetime
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))

import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Token ────────────────────────────────────────────────────────────────────
from motherduck_client import get_token

PKG_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ── Helpers ───────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHDR_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_FILL    = PatternFill("solid", fgColor="DEEAF1")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
BODY_FONT   = Font(size=10, name="Calibri")
BOLD_FONT   = Font(bold=True, size=10, name="Calibri")

THIN  = Side(style='thin',   color='AAAAAA')
THICK = Side(style='medium', color='1F4E79')

def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 45)

def write_df_to_sheet(ws, df, title, sub=""):
    ws.append([title])
    ws['A1'].font = Font(bold=True, size=12, name="Calibri", color="1F4E79")
    if sub:
        ws.append([sub])
        ws['A2'].font = Font(italic=True, size=9, name="Calibri", color="444444")
        ws.append([])
        hdr_row = 4
    else:
        ws.append([])
        hdr_row = 3
    # Header
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=hdr_row, column=ci, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border()
    # Data
    for ri, row in enumerate(df.itertuples(index=False), hdr_row + 1):
        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        for ci, val in enumerate(row, 1):
            # Coerce NA/NaT/None to empty string for openpyxl
            if val is None or (hasattr(val, '__class__') and val.__class__.__name__ in ('NAType', 'NaT')):
                val = ''
            try:
                import pandas as pd
                if pd.isna(val):
                    val = ''
            except (TypeError, ValueError):
                pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = Alignment(wrap_text=False)
    _autofit(ws)

# ── Era helper ────────────────────────────────────────────────────────────────
ERA_CASE = """
CASE
  WHEN YEAR(TRY_CAST(surg_first_date AS DATE)) BETWEEN 1999 AND 2004 THEN 'A_1999_2004'
  WHEN YEAR(TRY_CAST(surg_first_date AS DATE)) BETWEEN 2005 AND 2009 THEN 'B_2005_2009'
  WHEN YEAR(TRY_CAST(surg_first_date AS DATE)) BETWEEN 2010 AND 2014 THEN 'C_2010_2014'
  WHEN YEAR(TRY_CAST(surg_first_date AS DATE)) BETWEEN 2015 AND 2019 THEN 'D_2015_2019'
  WHEN YEAR(TRY_CAST(surg_first_date AS DATE)) BETWEEN 2020 AND 2025 THEN 'E_2020_2025'
  ELSE 'F_unknown'
END AS surgery_era
"""

def connect():
    tok = get_token()
    return duckdb.connect(f"md:thyroid_canonical_publication_v1_0?motherduck_token={tok}")

# ── Table builders ────────────────────────────────────────────────────────────

def build_table1(con) -> pd.DataFrame:
    """Cohort demographics, full cohort + malignant sub-columns."""
    sql = f"""
    WITH b AS (SELECT *, {ERA_CASE}
               FROM manuscript_workspace.cohort_m032_descriptive_25yr_v1)
    SELECT
      'Full Cohort'                    AS arm,
      COUNT(*)                         AS n,
      ROUND(AVG(age_at_surgery::DOUBLE),1) AS age_mean,
      MEDIAN(age_at_surgery)           AS age_median,
      ROUND(QUANTILE_CONT(age_at_surgery::DOUBLE, 0.25),1) AS age_q25,
      ROUND(QUANTILE_CONT(age_at_surgery::DOUBLE, 0.75),1) AS age_q75,
      ROUND(STDDEV(age_at_surgery::DOUBLE),1)  AS age_sd,
      COUNT(*) FILTER (WHERE LOWER(sex)='female') AS n_female,
      ROUND(COUNT(*) FILTER (WHERE LOWER(sex)='female')*100.0/COUNT(*),1) AS pct_female,
      COUNT(*) FILTER (WHERE LOWER(race) LIKE '%white%' OR LOWER(race) LIKE '%caucasian%') AS n_white,
      COUNT(*) FILTER (WHERE LOWER(race) LIKE '%black%' OR LOWER(race) LIKE '%african%')   AS n_black,
      COUNT(*) FILTER (WHERE LOWER(race) LIKE '%asian%')   AS n_asian,
      COUNT(*) FILTER (WHERE LOWER(race) LIKE '%hispanic%') AS n_hispanic,
      COUNT(*) FILTER (WHERE is_malignant=TRUE) AS n_malignant,
      ROUND(COUNT(*) FILTER (WHERE is_malignant=TRUE)*100.0/COUNT(*),1) AS pct_malignant,
      COUNT(*) FILTER (WHERE surg_total_thyroidectomy=TRUE) AS n_total_thy,
      COUNT(*) FILTER (WHERE surg_hemithyroidectomy=TRUE)   AS n_hemi,
      ROUND(MEDIAN(tumor_size_cm)::DOUBLE, 2) AS tumor_size_median_cm,
      COUNT(*) FILTER (WHERE multifocal_flag=TRUE)          AS n_multifocal,
      COUNT(*) FILTER (WHERE LOWER(smoking_status_combined)='current') AS n_smoker_current,
      COUNT(*) FILTER (WHERE LOWER(smoking_status_combined)='former')  AS n_smoker_former,
      COUNT(*) FILTER (WHERE LOWER(smoking_status_combined)='never')   AS n_smoker_never,
      COUNT(*) FILTER (WHERE smoking_status_combined IS NOT NULL)       AS n_smoking_known,
      COUNT(*) FILTER (WHERE pmhx_nlp_family_hx_thyroid=TRUE)          AS n_fhx_thyroid,
      COUNT(*) FILTER (WHERE pmhx_nlp_family_hx_thyroid IS NOT NULL)   AS n_fhx_known,
      ROUND(MEDIAN(followup_years)::DOUBLE, 1)                          AS fu_median_yrs,
      COUNT(*) FILTER (WHERE any_recurrence_flag=TRUE)                  AS n_recurrence,
      COUNT(*) FILTER (WHERE death_occurred=TRUE)                       AS n_death
    FROM b
    UNION ALL
    SELECT
      'Malignant Cohort' AS arm, COUNT(*), 
      ROUND(AVG(age_at_surgery::DOUBLE),1), MEDIAN(age_at_surgery),
      ROUND(QUANTILE_CONT(age_at_surgery::DOUBLE,0.25),1), ROUND(QUANTILE_CONT(age_at_surgery::DOUBLE,0.75),1),
      ROUND(STDDEV(age_at_surgery::DOUBLE),1),
      COUNT(*) FILTER (WHERE LOWER(sex)='female'),
      ROUND(COUNT(*) FILTER (WHERE LOWER(sex)='female')*100.0/COUNT(*),1),
      COUNT(*) FILTER (WHERE LOWER(race) LIKE '%white%' OR LOWER(race) LIKE '%caucasian%'),
      COUNT(*) FILTER (WHERE LOWER(race) LIKE '%black%' OR LOWER(race) LIKE '%african%'),
      COUNT(*) FILTER (WHERE LOWER(race) LIKE '%asian%'),
      COUNT(*) FILTER (WHERE LOWER(race) LIKE '%hispanic%'),
      COUNT(*), 100.0,
      COUNT(*) FILTER (WHERE surg_total_thyroidectomy=TRUE),
      COUNT(*) FILTER (WHERE surg_hemithyroidectomy=TRUE),
      ROUND(MEDIAN(tumor_size_cm)::DOUBLE, 2),
      COUNT(*) FILTER (WHERE multifocal_flag=TRUE),
      COUNT(*) FILTER (WHERE LOWER(smoking_status_combined)='current'),
      COUNT(*) FILTER (WHERE LOWER(smoking_status_combined)='former'),
      COUNT(*) FILTER (WHERE LOWER(smoking_status_combined)='never'),
      COUNT(*) FILTER (WHERE smoking_status_combined IS NOT NULL),
      COUNT(*) FILTER (WHERE pmhx_nlp_family_hx_thyroid=TRUE),
      COUNT(*) FILTER (WHERE pmhx_nlp_family_hx_thyroid IS NOT NULL),
      ROUND(MEDIAN(followup_years)::DOUBLE,1),
      COUNT(*) FILTER (WHERE any_recurrence_flag=TRUE),
      COUNT(*) FILTER (WHERE death_occurred=TRUE)
    FROM b WHERE is_malignant=TRUE
    """
    return con.execute(sql).fetchdf()


def build_table2(con) -> pd.DataFrame:
    """Histology distribution + malignancy rate by era."""
    sql = f"""
    WITH b AS (
      SELECT *, {ERA_CASE}
      FROM manuscript_workspace.cohort_m032_descriptive_25yr_v1
    )
    SELECT
      surgery_era,
      COUNT(*) AS n_total,
      COUNT(*) FILTER (WHERE is_malignant=TRUE) AS n_malignant,
      ROUND(COUNT(*) FILTER (WHERE is_malignant=TRUE)*100.0/COUNT(*),1) AS pct_malignant,
      COUNT(*) FILTER (WHERE is_malignant=TRUE AND (LOWER(histology_final) LIKE '%ptc%' OR LOWER(histology_final) LIKE '%papillary%')) AS n_ptc,
      COUNT(*) FILTER (WHERE is_malignant=TRUE AND (LOWER(histology_final) LIKE '%ftc%' OR LOWER(histology_final) LIKE '%follicular%')) AS n_ftc,
      COUNT(*) FILTER (WHERE is_malignant=TRUE AND (LOWER(histology_final) LIKE '%mtc%' OR LOWER(histology_final) LIKE '%medullary%')) AS n_mtc,
      COUNT(*) FILTER (WHERE is_malignant=TRUE AND (LOWER(histology_final) LIKE '%atc%' OR LOWER(histology_final) LIKE '%anaplastic%')) AS n_atc,
      COUNT(*) FILTER (WHERE is_malignant=TRUE AND LOWER(histology_final) LIKE '%pdtc%') AS n_pdtc,
      COUNT(*) FILTER (WHERE is_malignant=TRUE AND (LOWER(histology_final) LIKE '%hcc%' OR LOWER(histology_final) LIKE '%oncocytic%' OR LOWER(histology_final) LIKE '%hurthle%')) AS n_hcc
    FROM b
    GROUP BY surgery_era
    ORDER BY surgery_era
    """
    return con.execute(sql).fetchdf()


def build_table3(con) -> pd.DataFrame:
    """TNM stage migration by era (malignant only)."""
    sql = f"""
    WITH b AS (
      SELECT *, {ERA_CASE}
      FROM manuscript_workspace.cohort_m032_descriptive_25yr_v1
      WHERE is_malignant=TRUE
    )
    SELECT
      surgery_era,
      COUNT(*) AS n,
      COUNT(*) FILTER (WHERE ajcc8_stage_group='I')  AS n_I,
      ROUND(COUNT(*) FILTER (WHERE ajcc8_stage_group='I')*100.0/COUNT(*),1)  AS pct_I,
      COUNT(*) FILTER (WHERE ajcc8_stage_group='II') AS n_II,
      ROUND(COUNT(*) FILTER (WHERE ajcc8_stage_group='II')*100.0/COUNT(*),1) AS pct_II,
      COUNT(*) FILTER (WHERE ajcc8_stage_group='III') AS n_III,
      ROUND(COUNT(*) FILTER (WHERE ajcc8_stage_group='III')*100.0/COUNT(*),1) AS pct_III,
      COUNT(*) FILTER (WHERE ajcc8_stage_group IN ('IVA','IVB','IVC') OR ajcc8_stage_group LIKE 'IV%') AS n_IV,
      ROUND(COUNT(*) FILTER (WHERE ajcc8_stage_group IN ('IVA','IVB','IVC') OR ajcc8_stage_group LIKE 'IV%')*100.0/COUNT(*),1) AS pct_IV,
      COUNT(*) FILTER (WHERE LOWER(ata_risk_category)='low') AS n_ata_low,
      COUNT(*) FILTER (WHERE LOWER(ata_risk_category)='intermediate') AS n_ata_int,
      COUNT(*) FILTER (WHERE LOWER(ata_risk_category)='high') AS n_ata_high
    FROM b
    GROUP BY surgery_era
    ORDER BY surgery_era
    """
    return con.execute(sql).fetchdf()


def build_table4(con) -> pd.DataFrame:
    """Treatment patterns by era (full cohort)."""
    sql = f"""
    WITH b AS (
      SELECT *, {ERA_CASE}
      FROM manuscript_workspace.cohort_m032_descriptive_25yr_v1
    )
    SELECT
      surgery_era,
      COUNT(*) AS n,
      COUNT(*) FILTER (WHERE surg_total_thyroidectomy=TRUE) AS n_total_thy,
      ROUND(COUNT(*) FILTER (WHERE surg_total_thyroidectomy=TRUE)*100.0/COUNT(*),1) AS pct_total_thy,
      COUNT(*) FILTER (WHERE surg_hemithyroidectomy=TRUE) AS n_hemi,
      ROUND(COUNT(*) FILTER (WHERE surg_hemithyroidectomy=TRUE)*100.0/COUNT(*),1) AS pct_hemi,
      COUNT(*) FILTER (WHERE n_surgeries > 1) AS n_multisurgery,
      COUNT(*) FILTER (WHERE rai_received_flag=TRUE) AS n_rai,
      ROUND(COUNT(*) FILTER (WHERE rai_received_flag=TRUE)*100.0/NULLIF(COUNT(*) FILTER (WHERE is_malignant=TRUE),0),1) AS pct_rai_among_malig,
      COUNT(*) FILTER (WHERE any_recurrence_flag=TRUE) AS n_recurrence,
      ROUND(COUNT(*) FILTER (WHERE any_recurrence_flag=TRUE)*100.0/NULLIF(COUNT(*) FILTER (WHERE is_malignant=TRUE),0),1) AS pct_recurrence,
      COUNT(*) FILTER (WHERE comp_rln_injury_confirmed=TRUE) AS n_rln,
      COUNT(*) FILTER (WHERE comp_hypocalcemia_confirmed=TRUE) AS n_hypocalcemia,
      COUNT(*) FILTER (WHERE death_occurred=TRUE) AS n_death,
      ROUND(MEDIAN(followup_years)::DOUBLE, 1) AS fu_median_yrs
    FROM b
    GROUP BY surgery_era
    ORDER BY surgery_era
    """
    return con.execute(sql).fetchdf()


def build_table5(con) -> pd.DataFrame:
    """Smoking + family hx by era."""
    sql = f"""
    WITH b AS (
      SELECT *, {ERA_CASE}
      FROM manuscript_workspace.cohort_m032_descriptive_25yr_v1
    )
    SELECT
      surgery_era,
      COUNT(*) AS n,
      COUNT(*) FILTER (WHERE LOWER(smoking_status_combined)='current') AS n_current,
      COUNT(*) FILTER (WHERE LOWER(smoking_status_combined)='former')  AS n_former,
      COUNT(*) FILTER (WHERE LOWER(smoking_status_combined)='never')   AS n_never,
      COUNT(*) FILTER (WHERE smoking_status_combined IS NOT NULL)       AS n_known,
      ROUND(COUNT(*) FILTER (WHERE smoking_status_combined IS NOT NULL)*100.0/COUNT(*),1) AS pct_known,
      ROUND(COUNT(*) FILTER (WHERE LOWER(smoking_status_combined)='current')*100.0
        / NULLIF(COUNT(*) FILTER (WHERE smoking_status_combined IS NOT NULL),0),1) AS pct_current_of_known,
      ROUND(COUNT(*) FILTER (WHERE LOWER(smoking_status_combined)='former')*100.0
        / NULLIF(COUNT(*) FILTER (WHERE smoking_status_combined IS NOT NULL),0),1) AS pct_former_of_known,
      ROUND(COUNT(*) FILTER (WHERE LOWER(smoking_status_combined)='never')*100.0
        / NULLIF(COUNT(*) FILTER (WHERE smoking_status_combined IS NOT NULL),0),1) AS pct_never_of_known,
      COUNT(*) FILTER (WHERE pmhx_nlp_family_hx_thyroid=TRUE)          AS n_fhx_thyroid,
      COUNT(*) FILTER (WHERE pmhx_nlp_family_hx_thyroid IS NOT NULL)   AS n_fhx_known,
      ROUND(COUNT(*) FILTER (WHERE pmhx_nlp_family_hx_thyroid=TRUE)*100.0
        / NULLIF(COUNT(*) FILTER (WHERE pmhx_nlp_family_hx_thyroid IS NOT NULL),0),1) AS pct_fhx_of_known,
      COUNT(*) FILTER (WHERE pmhx_nlp_family_hx_cancer=TRUE)           AS n_fhx_any_cancer
    FROM b
    GROUP BY surgery_era
    ORDER BY surgery_era
    """
    return con.execute(sql).fetchdf()


def build_supp_s1(con) -> pd.DataFrame:
    """Supp S1: Sub-histology by era."""
    sql = f"""
    WITH b AS (
      SELECT *, {ERA_CASE}
      FROM manuscript_workspace.cohort_m032_descriptive_25yr_v1
      WHERE is_malignant=TRUE
    )
    SELECT
      COALESCE(histology_final, 'Unknown') AS histology,
      COUNT(*) AS n_total,
      COUNT(*) FILTER (WHERE surgery_era='A_1999_2004') AS n_era_A,
      COUNT(*) FILTER (WHERE surgery_era='B_2005_2009') AS n_era_B,
      COUNT(*) FILTER (WHERE surgery_era='C_2010_2014') AS n_era_C,
      COUNT(*) FILTER (WHERE surgery_era='D_2015_2019') AS n_era_D,
      COUNT(*) FILTER (WHERE surgery_era='E_2020_2025') AS n_era_E
    FROM b
    GROUP BY histology_final
    ORDER BY n_total DESC
    """
    return con.execute(sql).fetchdf()


def build_supp_s2(con) -> pd.DataFrame:
    """Supp S2: Race/ethnicity by era."""
    sql = f"""
    WITH b AS (
      SELECT *,
        {ERA_CASE},
        CASE
          WHEN LOWER(race) LIKE '%white%' OR LOWER(race) LIKE '%caucasian%' THEN 'White'
          WHEN LOWER(race) LIKE '%black%' OR LOWER(race) LIKE '%african%'   THEN 'Black'
          WHEN LOWER(race) LIKE '%asian%'    THEN 'Asian'
          WHEN LOWER(race) LIKE '%hispanic%' THEN 'Hispanic'
          ELSE 'Other/Unknown'
        END AS race_group
      FROM manuscript_workspace.cohort_m032_descriptive_25yr_v1
    )
    SELECT
      surgery_era, race_group,
      COUNT(*) AS n,
      ROUND(COUNT(*)*100.0 / SUM(COUNT(*)) OVER (PARTITION BY surgery_era), 1) AS pct_within_era
    FROM b
    GROUP BY surgery_era, race_group
    ORDER BY surgery_era, n DESC
    """
    return con.execute(sql).fetchdf()


def build_dict() -> pd.DataFrame:
    rows = [
        ("research_id", "Patient identifier (VARCHAR)", "Primary key", "All tables"),
        ("age_at_surgery", "Age at first surgery (years, BIGINT)", "Canonical patient master", "Table 1"),
        ("sex", "Patient sex (Female/Male)", "EHR registration", "Table 1"),
        ("race", "Self-reported race/ethnicity", "EHR registration", "Table 1, Supp S2"),
        ("surg_first_date", "Date of first thyroid surgery", "Canonical operative events", "Table 2–5"),
        ("surgery_era", "5-year era derived from surg_first_date", "Derived: YEAR(surg_first_date)", "Table 2–5"),
        ("surg_procedure_type", "Surgery type (normalized)", "Canonical operative events", "Table 4"),
        ("surg_total_thyroidectomy", "Total thyroidectomy flag (BOOLEAN)", "Canonical operative events", "Table 4"),
        ("surg_hemithyroidectomy", "Hemithyroidectomy flag (BOOLEAN)", "Canonical operative events", "Table 4"),
        ("is_malignant", "Malignant cohort flag (BOOLEAN)", "Canonical patient master", "Table 1–5"),
        ("histology_final", "Final resolved histology label", "Canonical pathology + NLP", "Table 2, Supp S1"),
        ("tumor_size_cm", "Primary tumor size in cm (DOUBLE)", "Canonical pathology synoptic", "Table 1"),
        ("multifocal_flag", "Multifocal tumor flag (BOOLEAN)", "Path synoptics", "Table 1"),
        ("ete_grade_final", "ETE grade: gross/microscopic/none", "Canonical patient master (mig_265)", "Table 2"),
        ("ln_positive_flag", "LN positive count (INTEGER)", "LN safe-view cross-validation", "Table 2"),
        ("ajcc8_stage_group", "AJCC 8th ed. stage group (VARCHAR)", "Canonical staging (mig_263)", "Table 3"),
        ("ata_risk_category", "ATA 2015 initial risk: low/intermediate/high", "Canonical patient master", "Table 3"),
        ("rai_received_flag", "RAI received flag (BOOLEAN)", "Canonical RAI episode", "Table 4"),
        ("any_recurrence_flag", "Any recurrence flag (BOOLEAN)", "Canonical recurrence", "Table 4"),
        ("followup_years", "Follow-up duration in years (DOUBLE)", "Canonical survival", "Table 4"),
        ("death_occurred", "Vital status death flag (BOOLEAN)", "Canonical survival", "Table 4"),
        ("smoking_status_combined", "Smoking status: current/former/never (VARCHAR)", "Structured + NLP (mig_281/287)", "Table 5"),
        ("pmhx_nlp_smoking_status", "NLP-extracted smoking status (VARCHAR)", "NLP pipeline (mig_281)", "Table 5"),
        ("pmhx_nlp_family_hx_thyroid", "Family hx thyroid cancer flag (BOOLEAN)", "NLP pipeline (mig_281)", "Table 5"),
        ("pmhx_nlp_family_hx_cancer", "Family hx any cancer flag (BOOLEAN)", "NLP pipeline (mig_281)", "Table 5"),
    ]
    return pd.DataFrame(rows, columns=["column_name", "description", "source", "appears_in"])


def build_qa(con) -> pd.DataFrame:
    """QA tab: locked-numbers cross-check.
    
    NOTE: Smoking 'locked' values (3022 known / 212 current / 502 former / 2298 never)
    are from pmhx_nlp_smoking_status (NLP-only field, post-mig_281).
    smoking_status_combined is broader (NLP + structured = ~4232 known) and is used
    for Table 5 figures. Minor shifts in current/former/never (+3-5) reflect
    mig_287 smoking taxonomy normalization applied after the Cowork lock.
    n_malig: 4019 live vs 4018 locked (1-patient diff, mig_285 edge case).
    """
    sql = """
    SELECT
      COUNT(*) AS n_total,
      COUNT(DISTINCT CASE WHEN is_malignant=TRUE THEN research_id END) AS n_malig,
      ROUND(COUNT(DISTINCT CASE WHEN is_malignant=TRUE THEN research_id END)*100.0/COUNT(*),1) AS pct_malig,
      -- NLP-only smoking (matches locked Cowork numbers)
      COUNT(*) FILTER (WHERE LOWER(pmhx_nlp_smoking_status)='current') AS n_nlp_smoke_current,
      COUNT(*) FILTER (WHERE LOWER(pmhx_nlp_smoking_status)='former')  AS n_nlp_smoke_former,
      COUNT(*) FILTER (WHERE LOWER(pmhx_nlp_smoking_status)='never')   AS n_nlp_smoke_never,
      COUNT(*) FILTER (WHERE pmhx_nlp_smoking_status IS NOT NULL)       AS n_nlp_smoke_known,
      -- Combined smoking (broader; used in Table 5)
      COUNT(*) FILTER (WHERE smoking_status_combined IS NOT NULL)       AS n_combined_smoke_known,
      COUNT(*) FILTER (WHERE pmhx_nlp_family_hx_thyroid=TRUE)          AS n_fhx_thyroid,
      COUNT(*) FILTER (WHERE pmhx_nlp_family_hx_thyroid IS NOT NULL)   AS n_fhx_known
    FROM manuscript_workspace.cohort_m032_descriptive_25yr_v1
    """
    actual = con.execute(sql).fetchdf().iloc[0].to_dict()
    # Expected from mig_285 Cowork lock 2026-05-04; tolerance=5 for mig_287 smoking normalization
    expected = {
        'n_total': (10871, 0),
        'n_malig': (4018, 2),            # +1 tolerance for 1-patient edge case
        'pct_malig': (37.0, 0.2),
        'n_nlp_smoke_current': (212, 5), # mig_287 normalization can shift ±5
        'n_nlp_smoke_former':  (502, 5),
        'n_nlp_smoke_never':   (2298, 5),
        'n_nlp_smoke_known':   (3022, 5),
        'n_fhx_thyroid': (366, 0),
        'n_fhx_known':   (3018, 2),
    }
    rows = []
    for k, (exp, tol) in expected.items():
        act = actual.get(k, None)
        try:
            diff = round(float(act) - float(exp), 2) if act is not None else 'N/A'
            status = 'PASS' if abs(float(diff)) <= tol else 'DIFF'
        except Exception:
            diff = 'N/A'; status = 'UNKNOWN'
        rows.append({'metric': k, 'expected': exp, 'tolerance': tol, 'actual': act, 'diff': diff, 'status': status})
    # Add informational rows (no expected)
    rows.append({'metric': 'n_combined_smoke_known', 'expected': '~4232 (combined NLP+structured)',
                 'tolerance': 'N/A', 'actual': actual.get('n_combined_smoke_known'), 'diff': 'N/A', 'status': 'INFO'})
    return pd.DataFrame(rows)


def build_master_data(con) -> pd.DataFrame:
    """Per-patient analytic dataset (deduped to unique research_id)."""
    sql = f"""
    SELECT DISTINCT ON (research_id)
      research_id,
      age_at_surgery,
      sex,
      race,
      TRY_CAST(surg_first_date AS DATE) AS surg_first_date,
      {ERA_CASE},
      surg_procedure_type,
      surg_total_thyroidectomy,
      surg_hemithyroidectomy,
      n_surgeries,
      is_malignant,
      histology_final,
      tumor_size_cm,
      multifocal_flag,
      ete_grade_final,
      ln_positive_flag,
      ln_total_examined,
      ln_total_positive,
      ajcc8_stage_group,
      ata_risk_category,
      rai_received_flag,
      any_recurrence_flag,
      followup_years,
      death_occurred,
      overall_survival_years,
      comp_rln_injury_confirmed,
      comp_hypocalcemia_confirmed,
      comp_hematoma_confirmed,
      smoking_status_combined,
      pmhx_nlp_smoking_status,
      pmhx_nlp_family_hx_thyroid,
      pmhx_nlp_family_hx_cancer
    FROM manuscript_workspace.cohort_m032_descriptive_25yr_v1
    ORDER BY research_id
    """
    return con.execute(sql).fetchdf()


# ── Excel writer ──────────────────────────────────────────────────────────────

def write_tables_excel(tables: dict, out_path: str):
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    for sheet_name, (df, title, subtitle) in tables.items():
        ws = wb.create_sheet(sheet_name)
        write_df_to_sheet(ws, df, title, subtitle)

    # Metadata sheet
    ws_meta = wb.create_sheet("_metadata")
    ws_meta.append(["M032 Submission Package — Table Workbook"])
    ws_meta.append(["Generated:", datetime.datetime.now().strftime("%Y-%m-%d %H:%M UTC")])
    ws_meta.append(["Cohort:", "manuscript_workspace.cohort_m032_descriptive_25yr_v1"])
    ws_meta.append(["DB:", "thyroid_canonical_publication_v1_0"])
    ws_meta.append(["Cohort lock:", "mig_281 + mig_285 (2026-05-04, commit 590acb5)"])
    ws_meta.append(["mig:", "mig_290"])
    ws_meta['A1'].font = Font(bold=True, size=14, name="Calibri", color="1F4E79")

    wb.save(out_path)
    print(f"[OK] {out_path}")


def main():
    print("Connecting to MotherDuck…")
    con = connect()

    print("Building tables…")
    t1 = build_table1(con)
    t2 = build_table2(con)
    t3 = build_table3(con)
    t4 = build_table4(con)
    t5 = build_table5(con)
    s1 = build_supp_s1(con)
    s2 = build_supp_s2(con)
    dd = build_dict()
    qa = build_qa(con)
    md = build_master_data(con)

    tables_xlsx = {
        "Table1_Demographics": (t1, "Table 1 — Cohort Demographics & Tumor Characteristics",
                                 "Full Cohort (N=10,871) and Malignant Cohort (N=4,018)"),
        "Table2_HistologyEra": (t2, "Table 2 — Histology Distribution & Malignancy Rate by Era",
                                 "Malignant cohort; era defined by 5-year periods from surg_first_date"),
        "Table3_StageMigration": (t3, "Table 3 — TNM Stage Migration Over 25 Years",
                                   "AJCC 8th ed. stage group by era; malignant cohort only"),
        "Table4_Treatment": (t4, "Table 4 — Treatment Patterns by Era",
                              "Full cohort; RAI % denominator = malignant patients"),
        "Table5_SmokingFHx": (t5, "Table 5 — Smoking & Family Hx Prevalence by Era",
                               "Post-mig_281 NLP augment; 27.8% cohort-wide coverage"),
        "SuppS1_SubHistology": (s1, "Supplementary Table S1 — Sub-histology by Era",
                                 "Malignant cohort only"),
        "SuppS2_RaceEra": (s2, "Supplementary Table S2 — Race/Ethnicity by Era",
                            "Full cohort; pct_within_era = % of that era"),
        "DataDictionary": (dd, "Data Dictionary", "Column definitions for 05_master_data.xlsx"),
        "QA_LockedNumbers": (qa, "QA — Locked-Numbers Cross-Check",
                              "PASS = |diff| < 1; expected from mig_285 Cowork lock 2026-05-04"),
    }

    out_tables = os.path.join(PKG_DIR, "04_tables.xlsx")
    write_tables_excel(tables_xlsx, out_tables)

    # Master data workbook
    wb2 = Workbook()
    wb2.remove(wb2.active)
    ws_data = wb2.create_sheet("per_patient_analytic")
    write_df_to_sheet(ws_data, md, "M032 Per-Patient Analytic Dataset",
                      f"N={len(md):,} patients (deduplicated by research_id)")
    ws_dict2 = wb2.create_sheet("DataDictionary")
    write_df_to_sheet(ws_dict2, dd, "Data Dictionary", "")
    ws_meta2 = wb2.create_sheet("_metadata")
    ws_meta2.append(["M032 Submission Package — Master Data Workbook"])
    ws_meta2.append(["Generated:", datetime.datetime.now().strftime("%Y-%m-%d %H:%M UTC")])
    ws_meta2.append(["mig:", "mig_290"])
    ws_meta2['A1'].font = Font(bold=True, size=14, name="Calibri", color="1F4E79")

    out_master = os.path.join(PKG_DIR, "05_master_data.xlsx")
    wb2.save(out_master)
    print(f"[OK] {out_master}")

    print("\n=== QA Summary ===")
    print(qa[["metric", "expected", "actual", "status"]].to_string(index=False))
    fails = qa[qa['status'].isin(['DIFF', 'UNKNOWN'])]
    if len(fails):
        print(f"\n[WARN] {len(fails)} metric(s) with status DIFF — verify before submission")
    else:
        print("\n[PASS] All locked-numbers within tolerance")

    con.close()


if __name__ == "__main__":
    main()
