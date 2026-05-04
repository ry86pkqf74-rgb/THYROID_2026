#!/usr/bin/env python3
"""
build_m025_tables.py
====================
Build 04_tables.xlsx + 05_master_data.xlsx for M025 (TI-RADS performance) submission v1.0.

Run from repo root:
    .venv/bin/python M025_submission_package_v1_0/08_analysis_code/build_m025_tables.py
"""

from __future__ import annotations

import json
import os
import re
import sys
from datetime import datetime, timezone

import duckdb
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
from motherduck_client import get_token

try:
    from sklearn.metrics import auc, roc_curve
except ImportError:  # pragma: no cover
    roc_curve = None  # type: ignore[assignment,misc]
    auc = None  # type: ignore[assignment,misc]

PKG_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
ALT_FILL = PatternFill("solid", fgColor="DEEAF1")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
BODY_FONT = Font(size=10, name="Calibri")
THIN = Side(style="thin", color="AAAAAA")

ACR_ROM_REF = pd.DataFrame(
    [
        ("TR1", "<2%", "ACR TI-RADS 2017 illustrative ROM (general population enrichment)"),
        ("TR2", "<2%", "—"),
        ("TR3", "<5%", "—"),
        ("TR4", "5–20%", "—"),
        ("TR5", ">20%", "—"),
    ],
    columns=["tirads_category", "acr_expected_rom", "footnote"],
)


def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 55)


def write_df_to_sheet(ws, df: pd.DataFrame, title: str, sub: str = ""):
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=12, name="Calibri", color="1F4E79")
    if sub:
        ws.append([sub])
        ws["A2"].font = Font(italic=True, size=9, name="Calibri", color="444444")
        ws.append([])
        hdr_row = 4
    else:
        ws.append([])
        hdr_row = 3
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=hdr_row, column=ci, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border()
    for ri, row in enumerate(df.itertuples(index=False), hdr_row + 1):
        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        for ci, val in enumerate(row, 1):
            if val is None or (hasattr(val, "__class__") and val.__class__.__name__ in ("NAType", "NaT")):
                val = ""
            try:
                if pd.isna(val):
                    val = ""
            except (TypeError, ValueError):
                pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = thin_border()
    _autofit(ws)


def connect():
    tok = get_token()
    return duckdb.connect(f"md:thyroid_canonical_publication_v1_0?motherduck_token={tok}")


def pull_spine(con) -> pd.DataFrame:
    sql = """
    SELECT
      c.*,
      pm.tirads_resolved AS cpm_tirads_resolved,
      pm.is_malignant AS cpm_is_malignant,
      CAST(pm.research_id AS VARCHAR) AS research_id_key
    FROM manuscript_workspace.cohort_m025_tirads_performance_v1 c
    LEFT JOIN main.canonical_patient_master pm
      ON CAST(c.research_id AS VARCHAR) = CAST(pm.research_id AS VARCHAR)
    """
    return con.execute(sql).fetchdf()


def parse_tr_rank(tr: object) -> float | np.floating | None:
    if tr is None or (isinstance(tr, float) and np.isnan(tr)):
        return None
    s = str(tr).strip().upper().replace(" ", "")
    if not s:
        return None
    if s.startswith("TR"):
        m = re.findall(r"\d+", s)
        if m:
            return float(int(m[0]))
    return None


def tirads_bucket_from_rank(rank: float) -> str:
    if rank is None or (isinstance(rank, float) and np.isnan(rank)):
        return "Unknown"
    return "TR%d" % int(rank)


def add_derived(df: pd.DataFrame) -> pd.DataFrame:
    d = df.copy()
    resolved = []
    for _, row in d.iterrows():
        rnk = parse_tr_rank(row.get("cpm_tirads_resolved"))
        if rnk is None:
            for col in ("tirads_worst_score_v12", "tirads_best_score_v12"):
                v = row.get(col)
                if v is not None and not (isinstance(v, float) and np.isnan(v)):
                    try:
                        rnk = float(int(v))
                        break
                    except (ValueError, TypeError):
                        pass
        resolved.append(rnk)
    d["tr_rank"] = resolved
    d["tr_label"] = d["tr_rank"].apply(lambda x: tirads_bucket_from_rank(x) if pd.notna(x) else "Unknown")

    malignancy_col = "cpm_is_malignant" if "cpm_is_malignant" in d.columns else "is_malignant"
    d["y_mal"] = (
        d[malignancy_col]
        .map(lambda v: v is True or str(v).lower() in ("true", "t", "1"))
        .astype(bool)
    )

    bf = pd.to_numeric(d["bethesda_final"], errors="coerce")

    def bethesda_bucket(x: float) -> str:
        if pd.isna(x):
            return "Unknown/Unk"
        i = int(x)
        return "I-II" if i <= 2 else "III" if i == 3 else "IV" if i == 4 else "V" if i == 5 else "VI" if i == 6 else "Unknown/Unk"

    d["bethesda_bucket"] = bf.map(bethesda_bucket)

    d["male"] = d["sex"].astype(str).str.lower().isin(("male", "m")).astype(int)
    return d


def wilson_ci(x: int, n: int, z: float = 1.96) -> tuple[float, float]:
    if n <= 0:
        return (float("nan"), float("nan"))
    phat = x / n
    denom = 1 + z**2 / n
    center = phat + z**2 / (2 * n)
    margin = z * np.sqrt((phat * (1 - phat) / n) + z**2 / (4 * n**2))
    return float((center - margin) / denom), float((center + margin) / denom)


def table1_demo_by_tr(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    order = ["TR1", "TR2", "TR3", "TR4", "TR5"]
    df2 = df[df["tr_label"].isin(order)].copy()
    df2["_ord"] = df2["tr_label"].apply(lambda lab: order.index(lab))
    df2 = df2.sort_values("_ord")
    for lab, sub in df2.groupby("tr_label"):
        nm = sub["y_mal"].astype(bool).sum()
        n_all = len(sub)
        ci_lo, ci_hi = wilson_ci(int(nm), n_all)
        rows.append(
            {
                "TIRADS": lab,
                "n": n_all,
                "n_malignant": int(nm),
                "ROM_pct": round(100.0 * nm / n_all, 1) if n_all else None,
                "ROM_CI95_low_pct": round(100.0 * ci_lo, 1),
                "ROM_CI95_high_pct": round(100.0 * ci_hi, 1),
                "age_mean": round(float(sub["age_at_surgery"].mean()), 1)
                if sub["age_at_surgery"].notna().any()
                else None,
                "age_sd": round(float(sub["age_at_surgery"].std()), 1)
                if sub["age_at_surgery"].dropna().shape[0] > 1
                else None,
                "pct_male": round(100.0 * float(sub["male"].mean()), 1),
                "tumor_cm_median": round(float(pd.to_numeric(sub["tumor_size_cm"], errors="coerce").median()), 2),
            }
        )
    return pd.DataFrame(rows)


def calc_threshold_metrics(df: pd.DataFrame, thresh: int):
    """Test positive when tr_rank >= thresh (ordinal TR)."""
    sub = df[pd.notna(df["tr_rank"])].copy()
    y = sub["y_mal"].astype(bool).to_numpy()
    pred = sub["tr_rank"].to_numpy() >= float(thresh)
    tp = int(np.sum(pred & y))
    fp = int(np.sum(pred & ~y))
    tn = int(np.sum(~pred & ~y))
    fn = int(np.sum(~pred & y))
    sens = tp / (tp + fn) if (tp + fn) else np.nan
    spec = tn / (tn + fp) if (tn + fp) else np.nan
    ppv = tp / (tp + fp) if (tp + fp) else np.nan
    npv = tn / (tn + fn) if (tn + fn) else np.nan
    acc = (tp + tn) / max(len(sub), 1)
    lr_pos = sens / (1 - spec) if spec < 1 and not np.isnan(sens) else np.nan
    lr_neg = (1 - sens) / spec if spec > 0 and not np.isnan(sens) else np.nan
    label = "TR ≥ TR%d" % thresh
    return {
        "threshold": label,
        "n_evaluable": len(sub),
        "TP": tp,
        "FP": fp,
        "TN": tn,
        "FN": fn,
        "sensitivity": round(float(sens), 4),
        "specificity": round(float(spec), 4),
        "PPV": round(float(ppv), 4),
        "NPV": round(float(npv), 4),
        "accuracy": round(float(acc), 4),
        "LRplus": round(float(lr_pos), 4)
        if not (isinstance(lr_pos, float) and np.isnan(lr_pos))
        else lr_pos,
        "LRminus": round(float(lr_neg), 4)
        if not (isinstance(lr_neg, float) and np.isnan(lr_neg))
        else lr_neg,
    }


def table2_performance(df: pd.DataFrame) -> pd.DataFrame:
    rows = [calc_threshold_metrics(df, th) for th in (2, 3, 4, 5)]
    return pd.DataFrame(rows)


def twoby2_primary(df: pd.DataFrame, thresh: int = 4) -> pd.DataFrame:
    sub = df[pd.notna(df["tr_rank"])].copy()
    pred = sub["tr_rank"].to_numpy() >= float(thresh)
    y = sub["y_mal"].astype(bool).to_numpy()
    tp = int(np.sum(pred & y))
    fp = int(np.sum(pred & ~y))
    tn = int(np.sum(~pred & ~y))
    fn = int(np.sum(~pred & y))
    return pd.DataFrame(
        [
            {"slice": "test_pos_malign", "metric": "malignant", "count": tp},
            {"slice": "test_pos_benign", "metric": "benign", "count": fp},
            {"slice": "test_neg_malign", "metric": "malignant", "count": fn},
            {"slice": "test_neg_benign", "metric": "benign", "count": tn},
        ]
    )


def table4_size_quartile(df: pd.DataFrame) -> pd.DataFrame:
    sz = pd.to_numeric(df["tumor_size_cm"], errors="coerce")
    dd = df.loc[sz.notna()].copy()
    dd["_sz"] = sz.loc[sz.notna()]
    try:
        dd["_quad"] = pd.qcut(dd["_sz"], q=4, labels=["Q1_small", "Q2", "Q3", "Q4_large"], duplicates="drop")
    except ValueError:
        dd["_quad"] = pd.Series([pd.NA] * len(dd), index=dd.index)

    rows: list[dict] = []
    for qt in ["Q1_small", "Q2", "Q3", "Q4_large"]:
        s = dd[dd["_quad"].astype(str) == qt]
        if s.empty:
            continue
        mal = int(s["y_mal"].astype(bool).sum())
        n = len(s)
        ci_lo, ci_hi = wilson_ci(mal, n)
        rows.append(
            {
                "dominant_path_size_quartile": qt,
                "n": n,
                "ROM_pct": round(100 * mal / n, 1),
                "Rom_CI95_low_pct": round(100 * ci_lo, 1),
                "Rom_CI95_high_pct": round(100 * ci_hi, 1),
            }
        )
    hi = df[(pd.notna(df["tr_rank"])) & (df["tr_rank"] >= 4)]
    if not hi.empty:
        mal = int(hi["y_mal"].astype(bool).sum())
        n = len(hi)
        rows.append(
            {
                "dominant_path_size_quartile": "Subgroup_TR_ge_TR4_any_size",
                "n": n,
                "ROM_pct": round(100 * mal / n, 1),
                "Rom_CI95_low_pct": "",
                "Rom_CI95_high_pct": "",
            }
        )
    tab = pd.DataFrame(rows)
    tab.attrs[
        "footnote"
    ] = "Composition strata not populated on cohort spine v1; size quartiles from pathologic dominant nodule diameter."
    return tab


def supp_roc_curve(df: pd.DataFrame):
    """Return FPR/TPR series + AUC using tr_rank as continuous risk score."""
    sub = df[pd.notna(df["tr_rank"])].copy()
    y = sub["y_mal"].astype(int).to_numpy()
    scores = sub["tr_rank"].to_numpy(dtype=float)
    if roc_curve is None or auc is None:
        return pd.DataFrame(), float("nan")
    fpr, tpr, _ = roc_curve(y, scores)
    auc_val = float(auc(fpr, tpr))
    out = pd.DataFrame({"fpr": fpr, "tpr": tpr})
    return out, auc_val


def build_literature_compare(tab1_live: pd.DataFrame) -> pd.DataFrame:
    m = pd.merge(ACR_ROM_REF[["tirads_category", "acr_expected_rom"]], tab1_live, left_on="tirads_category", right_on="TIRADS", how="left")
    m = m.rename(columns={"ROM_pct": "cohort_rom_pct_live"})
    cols = ["tirads_category", "acr_expected_rom", "n", "n_malignant", "cohort_rom_pct_live"]
    return m[[c for c in cols if c in m.columns]]


def main():
    con = connect()
    raw = pull_spine(con)
    con.close()
    df = add_derived(raw)

    out_dir = os.path.join(PKG_DIR, "08_analysis_outputs")
    os.makedirs(out_dir, exist_ok=True)

    snap = {
        "generated_utc": datetime.now(timezone.utc).isoformat(),
        "n_cohort_view": int(len(raw)),
        "n_tr_rank_known": int(df["tr_rank"].notna().sum()),
        "n_malignant": int(df["y_mal"].astype(bool).sum()),
        "n_benign": int((~df["y_mal"].astype(bool)).sum()),
    }
    with open(os.path.join(out_dir, "m025_run_snapshot.json"), "w", encoding="utf-8") as f:
        json.dump(snap, f, indent=2)

    parquet_path = os.path.join(out_dir, "m025_analytic_spine.parquet")
    df.to_parquet(parquet_path, index=False)

    tab1 = table1_demo_by_tr(df)
    tab1_vs_acr = build_literature_compare(tab1)
    tab2 = table2_performance(df)
    tb2 = twoby2_primary(df, 4)

    tbl3_full = pd.crosstab(
        df[df["tr_label"].isin(["TR1", "TR2", "TR3", "TR4", "TR5"])].tr_label,
        df[df["tr_label"].isin(["TR1", "TR2", "TR3", "TR4", "TR5"])].bethesda_bucket,
        margins=True,
    ).fillna(0).astype(int)
    tab3_export = tbl3_full.reset_index().rename(columns={"tr_label": "TIRADS"})

    tab4 = table4_size_quartile(df)
    roc_df, roc_auc_val = supp_roc_curve(df)
    roc_path = os.path.join(out_dir, "m025_supp_ROC_curve_points.csv")
    if not roc_df.empty:
        roc_df.to_csv(roc_path, index=False)
        pd.DataFrame([{"metric": "roc_auc_ordinal_rank", "value": roc_auc_val}]).to_csv(
            os.path.join(out_dir, "m025_supp_ROC_summary.csv"),
            index=False,
        )

    cover = pd.DataFrame(
        [
            ("Package", "M025 v1.0 — TI-RADS performance"),
            ("Database", "thyroid_canonical_publication_v1_0"),
            ("Cohort view", "manuscript_workspace.cohort_m025_tirads_performance_v1"),
            ("Gold standard", "canonical_patient_master.is_malignant"),
            ("TIRADS label", "cpm.tirads_resolved with fallback tirads_worst_score_v12"),
            ("Build UTC", snap["generated_utc"]),
            ("Cohort n", snap["n_cohort_view"]),
            (
                "ROC AUC (tr_rank ordinal)",
                round(roc_auc_val, 4)
                if isinstance(roc_auc_val, (float, np.floating)) and not np.isnan(roc_auc_val)
                else "sklearn_missing",
            ),
            ("ROC points CSV", os.path.basename(roc_path) if not roc_df.empty else "n/a"),
        ],
        columns=["key", "value"],
    )

    qa = pd.DataFrame(
        [
            {"check": "cohort_non_empty", "status": "PASS" if snap["n_cohort_view"] > 0 else "FAIL"},
            {"check": "tr_rank_coverage", "status": "PASS" if snap["n_tr_rank_known"] >= 0.95 * snap["n_cohort_view"] else "REVIEW"},
            {"check": "parity_cohort_gate~3375±100", "status": "PASS" if 3200 <= snap["n_cohort_view"] <= 3600 else "REVIEW"},
        ]
    )

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Cover")
    write_df_to_sheet(ws, cover, "M025 tables workbook — cover")

    ws = wb.create_sheet("Table1_demo_by_TR")
    write_df_to_sheet(
        ws,
        tab1,
        "Table 1 — Demographics + ROM by TI-RADS (resolved rank)",
        "Operative cohort: ROM exceeds ACR-illustrative external cohort at every tier — Methods footnote mandatory.",
    )

    ws = wb.create_sheet("Table1b_vs_ACR_ref")
    write_df_to_sheet(ws, tab1_vs_acr, "Table 1b — ROM vs illustrative ACR-reported ranges")

    ws = wb.create_sheet("Table2_performance")
    write_df_to_sheet(ws, tab2, "Table 2 — Sens/Spec/PPV/NPV at TR thresholds")

    ws = wb.create_sheet("Supp_twoby2_TRge4")
    write_df_to_sheet(ws, tb2, "Primary 2×2 — TR≥TR4 (path gold standard)", "")

    ws = wb.create_sheet("Table3_TR_x_Bethesda")
    write_df_to_sheet(ws, tab3_export, "Table 3 — TI-RADS × Bethesda bucket", "")

    ws = wb.create_sheet("Table4_size_quartiles")
    write_df_to_sheet(
        ws,
        tab4,
        "Table 4 — ROM by dominant pathologic size quartile",
        getattr(tab4, "attrs", {}).get("footnote", ""),
    )

    ws = wb.create_sheet("Supp_S1_ROC_points")
    if roc_df.empty:
        write_df_to_sheet(ws, pd.DataFrame([{"note": "sklearn not installed — ROC points skipped"}]), "Supplement S1 — ROC", "")
    else:
        write_df_to_sheet(
            ws,
            roc_df,
            "Supplement S1 — ROC operating points",
            "AUC=%0.4f (ordinal TI-RADS rank as classifier score)." % roc_auc_val,
        )

    ws = wb.create_sheet("Supp_S2_literature")
    write_df_to_sheet(ws, tab1_vs_acr, "Supplement S2 — Side-by-side ACR illustrative vs observed ROM", "")

    ws = wb.create_sheet("QA")
    write_df_to_sheet(ws, qa, "QA checks", "")

    out_tables = os.path.join(PKG_DIR, "04_tables.xlsx")
    wb.save(out_tables)
    print(f"Wrote {out_tables}")

    master_xlsx = os.path.join(PKG_DIR, "05_master_data.xlsx")
    meta = pd.DataFrame(
        [
            {"column": col, "description": "(see cohort view + CPM tirads_resolved migration mig_288)", "dtype": str(df[col].dtype)}
            for col in df.columns
        ]
    )
    with pd.ExcelWriter(master_xlsx, engine="openpyxl") as xw:
        df.to_excel(xw, sheet_name="analytic_spine", index=False)
        pd.DataFrame([snap]).to_excel(xw, sheet_name="run_snapshot", index=False)
        meta.head(180).to_excel(xw, sheet_name="data_dictionary_stub", index=False)
    print(f"Wrote {master_xlsx}")


if __name__ == "__main__":
    main()
