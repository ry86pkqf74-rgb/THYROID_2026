#!/usr/bin/env python3
"""
build_m025_tables.py — M025 v2.0 (nodule-level)
================================================
Build 04_tables.xlsx + 05_master_data.xlsx from MotherDuck using
``manuscript_workspace.cohort_m025_nodule_level_v1`` (mig_306).

Primary analytic subset: ``analytic_eligible_strict_acr_pernodule = TRUE``.
Gold standard at nodule grain: ``nodule_path_proven_malignant``.

Patient-level comparator ROM uses ``cohort_m025_tirads_performance_v1`` (v1.0 spine).

Run from repo root:
    .venv/bin/python M025_submission_package_v2_0/08_analysis_code/build_m025_tables.py
"""

from __future__ import annotations

import json
import os
import re
import sys
from datetime import datetime, timezone

import duckdb
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
from motherduck_client import get_token

try:
    from sklearn.metrics import auc, roc_curve
except ImportError:  # pragma: no cover
    roc_curve = None  # type: ignore[assignment,misc]
    auc = None  # type: ignore[assignment,misc]

PKG_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
ALT_FILL = PatternFill("solid", fgColor="DEEAF1")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
BODY_FONT = Font(size=10, name="Calibri")
THIN = Side(style="thin", color="AAAAAA")

ACR_ROM_REF = {
    "TR1": "<2%",
    "TR2": "<2%",
    "TR3": "<5%",
    "TR4": "5–20%",
    "TR5": ">20%",
}


def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 55)


def write_df_to_sheet(ws, df: pd.DataFrame, title: str, sub: str = ""):
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=12, name="Calibri", color="1F4E79")
    if sub:
        ws.append([sub])
        ws["A2"].font = Font(italic=True, size=9, name="Calibri", color="444444")
        ws.append([])
        hdr_row = 4
    else:
        ws.append([])
        hdr_row = 3
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=hdr_row, column=ci, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border()
    for ri, row in enumerate(df.itertuples(index=False), hdr_row + 1):
        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        for ci, val in enumerate(row, 1):
            if val is None or (hasattr(val, "__class__") and val.__class__.__name__ in ("NAType", "NaT")):
                val = ""
            try:
                if pd.isna(val):
                    val = ""
            except (TypeError, ValueError):
                pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = thin_border()
    _autofit(ws)


def connect():
    tok = get_token()
    return duckdb.connect(f"md:thyroid_canonical_publication_v1_0?motherduck_token={tok}")


def _excel_safe_df(d: pd.DataFrame) -> pd.DataFrame:
    """Strip timezone from datetimes — openpyxl cannot write tz-aware."""
    o = d.copy()
    for c in o.columns:
        ser = o[c]
        if not pd.api.types.is_datetime64_any_dtype(ser):
            continue
        try:
            if getattr(ser.dtype, "tz", None) is not None:
                o[c] = ser.dt.tz_convert("UTC").dt.tz_localize(None)
        except (TypeError, AttributeError, ValueError):
            pass
    return o


def pull_nodule_spine(con) -> pd.DataFrame:
    return con.execute(
        """
        SELECT
          v.*,
          pm.age_at_surgery,
          pm.sex,
          pm.tumor_size_cm_max AS tumor_size_cm
        FROM manuscript_workspace.cohort_m025_nodule_level_v1 v
        LEFT JOIN main.canonical_patient_master pm
          ON CAST(v.research_id AS VARCHAR) = CAST(pm.research_id AS VARCHAR)
        """
    ).fetchdf()


def pull_patient_spine(con) -> pd.DataFrame:
    return con.execute(
        """
        SELECT
          c.*,
          pm.tirads_resolved AS cpm_tirads_resolved,
          pm.is_malignant AS cpm_is_malignant,
          CAST(pm.research_id AS VARCHAR) AS research_id_key
        FROM manuscript_workspace.cohort_m025_tirads_performance_v1 c
        LEFT JOIN main.canonical_patient_master pm
          ON CAST(c.research_id AS VARCHAR) = CAST(pm.research_id AS VARCHAR)
        """
    ).fetchdf()


def parse_tr_rank(tr: object) -> float | np.floating | None:
    if tr is None or (isinstance(tr, float) and np.isnan(tr)):
        return None
    s = str(tr).strip().upper().replace(" ", "")
    if not s:
        return None
    if s.startswith("TR"):
        m = re.findall(r"\d+", s)
        if m:
            return float(int(m[0]))
    return None


def tirads_bucket_from_rank(rank: float) -> str:
    if rank is None or (isinstance(rank, float) and np.isnan(rank)):
        return "Unknown"
    return "TR%d" % int(rank)


def _strict_eligible_mask(s: pd.Series) -> pd.Series:
    return s.map(lambda v: v is True or str(v).lower() in ("true", "t", "1", "yes"))


def add_derived_patient(df: pd.DataFrame) -> pd.DataFrame:
    """Match v1.0 patient-level spine derivation (MAX patient TR + CPM malignancy)."""
    d = df.copy()
    resolved = []
    for _, row in d.iterrows():
        rnk = parse_tr_rank(row.get("cpm_tirads_resolved"))
        if rnk is None:
            for col in ("tirads_worst_score_v12", "tirads_best_score_v12"):
                v = row.get(col)
                if v is not None and not (isinstance(v, float) and np.isnan(v)):
                    try:
                        rnk = float(int(v))
                        break
                    except (ValueError, TypeError):
                        pass
        resolved.append(rnk)
    d["tr_rank"] = resolved
    d["tr_label"] = d["tr_rank"].apply(lambda x: tirads_bucket_from_rank(x) if pd.notna(x) else "Unknown")
    col = "cpm_is_malignant" if "cpm_is_malignant" in d.columns else "is_malignant"
    d["y_mal"] = d[col].map(lambda v: v is True or str(v).lower() in ("true", "t", "1")).astype(bool)

    def bethesda_bucket(x: float) -> str:
        if pd.isna(x):
            return "Unknown/Unk"
        i = int(x)
        return (
            "I-II"
            if i <= 2
            else "III"
            if i == 3
            else "IV"
            if i == 4
            else "V"
            if i == 5
            else "VI"
            if i == 6
            else "Unknown/Unk"
        )

    if "bethesda_final" in d.columns:
        bf = pd.to_numeric(d["bethesda_final"], errors="coerce")
        d["bethesda_bucket"] = bf.map(bethesda_bucket)
    else:
        d["bethesda_bucket"] = pd.Series(["Unknown/Unk"] * len(d), index=d.index)
    d["male"] = d["sex"].astype(str).str.lower().isin(("male", "m")).astype(int)
    return d


def add_derived_nodule(df: pd.DataFrame) -> pd.DataFrame:
    d = df.copy()
    cat_col = "acr2017_tirads_category"
    if cat_col not in d.columns:
        raise SystemExit(f"Missing {cat_col} on cohort_m025_nodule_level_v1 — cannot build v2.0 tables.")
    d["tr_rank"] = d[cat_col].map(parse_tr_rank)
    d["tr_label"] = d["tr_rank"].apply(lambda x: tirads_bucket_from_rank(x) if pd.notna(x) else "Unknown")
    if "nodule_path_proven_malignant" not in d.columns:
        raise SystemExit("Missing nodule_path_proven_malignant on nodule spine.")
    d["y_mal"] = d["nodule_path_proven_malignant"].map(
        lambda v: v is True or str(v).lower() in ("true", "t", "1")
    ).astype(bool)
    bcol = None
    for c in ("bethesda_2023_num", "bethesda_final_num", "bethesda_final"):
        if c in d.columns:
            bcol = c
            break
    bf = pd.to_numeric(d[bcol], errors="coerce") if bcol else pd.Series([np.nan] * len(d))

    def bethesda_bucket_bethesda2023(x: float) -> str:
        if pd.isna(x):
            return "missing"
        i = int(x)
        if i <= 2:
            return "I-II"
        if i == 3:
            return "III"
        if i == 4:
            return "IV"
        if i == 5:
            return "V"
        if i == 6:
            return "VI"
        return "missing"

    d["bethesda_bucket"] = bf.map(bethesda_bucket_bethesda2023)
    d["male"] = d["sex"].astype(str).str.lower().isin(("male", "m")).astype(int)
    return d


def wilson_ci(x: int, n: int, z: float = 1.96) -> tuple[float, float]:
    if n <= 0:
        return (float("nan"), float("nan"))
    phat = x / n
    denom = 1 + z**2 / n
    center = phat + z**2 / (2 * n)
    margin = z * np.sqrt((phat * (1 - phat) / n) + z**2 / (4 * n**2))
    return float((center - margin) / denom), float((center + margin) / denom)


def table1_demo_by_tr(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    order = ["TR1", "TR2", "TR3", "TR4", "TR5"]
    df2 = df[df["tr_label"].isin(order)].copy()
    df2["_ord"] = df2["tr_label"].apply(lambda lab: order.index(lab))
    df2 = df2.sort_values("_ord")
    for lab, sub in df2.groupby("tr_label"):
        nm = int(sub["y_mal"].astype(bool).sum())
        n_all = len(sub)
        ci_lo, ci_hi = wilson_ci(nm, n_all)
        age_mean = None
        age_sd = None
        if "age_at_surgery" in sub.columns and sub["age_at_surgery"].notna().any():
            age_mean = round(float(pd.to_numeric(sub["age_at_surgery"], errors="coerce").mean()), 1)
            ad = sub["age_at_surgery"].dropna()
            if len(ad) > 1:
                age_sd = round(float(pd.to_numeric(ad, errors="coerce").std()), 1)
        tum_med = None
        if "tumor_size_cm" in sub.columns:
            tum_med = round(float(pd.to_numeric(sub["tumor_size_cm"], errors="coerce").median()), 2)
        rows.append(
            {
                "TIRADS": lab,
                "n": n_all,
                "n_malignant": nm,
                "ROM_pct": round(100.0 * nm / n_all, 1) if n_all else None,
                "ROM_CI95_low_pct": round(100.0 * ci_lo, 1),
                "ROM_CI95_high_pct": round(100.0 * ci_hi, 1),
                "age_mean": age_mean,
                "age_sd": age_sd,
                "pct_male": round(100.0 * float(sub["male"].mean()), 1) if "male" in sub.columns else None,
                "tumor_cm_median": tum_med,
            }
        )
    return pd.DataFrame(rows)


def calc_threshold_metrics_ci(df: pd.DataFrame, thresh: int) -> dict:
    sub = df[pd.notna(df["tr_rank"])].copy()
    y = sub["y_mal"].astype(bool).to_numpy()
    pred = sub["tr_rank"].to_numpy() >= float(thresh)
    tp = int(np.sum(pred & y))
    fp = int(np.sum(pred & ~y))
    tn = int(np.sum(~pred & ~y))
    fn = int(np.sum(~pred & y))
    sens = tp / (tp + fn) if (tp + fn) else np.nan
    spec = tn / (tn + fp) if (tn + fp) else np.nan
    ppv = tp / (tp + fp) if (tp + fp) else np.nan
    npv = tn / (tn + fn) if (tn + fn) else np.nan
    s_lo, s_hi = wilson_ci(tp, tp + fn) if (tp + fn) > 0 else (float("nan"), float("nan"))
    p_lo, p_hi = wilson_ci(tn, tn + fp) if (tn + fp) > 0 else (float("nan"), float("nan"))
    pv_lo, pv_hi = wilson_ci(tp, tp + fp) if (tp + fp) > 0 else (float("nan"), float("nan"))
    nv_lo, nv_hi = wilson_ci(tn, tn + fn) if (tn + fn) > 0 else (float("nan"), float("nan"))
    label = "TR>=" + "TR%d" % thresh
    return {
        "threshold": label,
        "tp": tp,
        "fp": fp,
        "fn": fn,
        "tn": tn,
        "sens_pct": round(100 * sens, 1),
        "sens_lo_95": round(100 * s_lo, 1),
        "sens_hi_95": round(100 * s_hi, 1),
        "spec_pct": round(100 * spec, 1),
        "spec_lo_95": round(100 * p_lo, 1),
        "spec_hi_95": round(100 * p_hi, 1),
        "ppv_pct": round(100 * ppv, 1),
        "ppv_lo_95": round(100 * pv_lo, 1),
        "ppv_hi_95": round(100 * pv_hi, 1),
        "npv_pct": round(100 * npv, 1),
        "npv_lo_95": round(100 * nv_lo, 1),
        "npv_hi_95": round(100 * nv_hi, 1),
    }


def table_threshold_export_rows(df_strict: pd.DataFrame) -> pd.DataFrame:
    rows = [calc_threshold_metrics_ci(df_strict, th) for th in (3, 4, 5)]
    return pd.DataFrame(rows)


def table_patient_vs_nodule_rom(df_p: pd.DataFrame, df_n_strict: pd.DataFrame) -> pd.DataFrame:
    p = table1_demo_by_tr(df_p).set_index("TIRADS")
    n = table1_demo_by_tr(df_n_strict).set_index("TIRADS")
    order = ["TR1", "TR2", "TR3", "TR4", "TR5"]
    out_rows = []
    for lab in order:
        prow = p.loc[lab] if lab in p.index else None
        nrow = n.loc[lab] if lab in n.index else None
        inflation = None
        if prow is not None and nrow is not None:
            inflation = round(float(prow["ROM_pct"]) - float(nrow["ROM_pct"]), 1)
        band = ACR_ROM_REF.get(lab, "")
        in_band = ""
        if nrow is not None:
            rom = float(nrow["ROM_pct"])
            if lab == "TR4":
                in_band = "YES" if 5 <= rom <= 20 else "no"
            elif lab == "TR5":
                in_band = "YES" if rom > 20 else "no"
            elif lab in ("TR1", "TR2"):
                in_band = "YES" if rom < 2 else "no"
            elif lab == "TR3":
                in_band = "YES" if rom < 5 else "no"
            else:
                in_band = ""
        out_rows.append(
            {
                "tirads": lab,
                "patient_n": int(prow["n"]) if prow is not None else None,
                "patient_k": int(prow["n_malignant"]) if prow is not None else None,
                "patient_rom_pct": prow["ROM_pct"] if prow is not None else None,
                "patient_lo_95": prow["ROM_CI95_low_pct"] if prow is not None else None,
                "patient_hi_95": prow["ROM_CI95_high_pct"] if prow is not None else None,
                "nodule_n": int(nrow["n"]) if nrow is not None else None,
                "nodule_k": int(nrow["n_malignant"]) if nrow is not None else None,
                "nodule_rom_pct": nrow["ROM_pct"] if nrow is not None else None,
                "nodule_lo_95": nrow["ROM_CI95_low_pct"] if nrow is not None else None,
                "nodule_hi_95": nrow["ROM_CI95_high_pct"] if nrow is not None else None,
                "inflation_pp": inflation,
                "acr_expected_band": band,
                "nodule_in_acr_band": in_band,
            }
        )
    return pd.DataFrame(out_rows)


def table_bethesda_by_tr_cross(df_strict: pd.DataFrame) -> pd.DataFrame:
    use = df_strict[df_strict["tr_label"].isin(["TR2", "TR3", "TR4", "TR5"])].copy()
    tbl = pd.crosstab(use["bethesda_bucket"], use["tr_label"], margins=True)
    return tbl.fillna(0).astype(int)


def supp_roc_curve(df: pd.DataFrame):
    sub = df[pd.notna(df["tr_rank"])].copy()
    y = sub["y_mal"].astype(int).to_numpy()
    scores = sub["tr_rank"].to_numpy(dtype=float)
    if roc_curve is None or auc is None or len(sub) < 5:
        return pd.DataFrame(), float("nan")
    fpr, tpr, _ = roc_curve(y, scores)
    auc_val = float(auc(fpr, tpr))
    return pd.DataFrame({"fpr": fpr, "tpr": tpr}), auc_val


def main():
    con = connect()
    raw_n = pull_nodule_spine(con)
    raw_p = pull_patient_spine(con)
    con.close()

    if "analytic_eligible_strict_acr_pernodule" not in raw_n.columns:
        raise SystemExit("cohort view missing analytic_eligible_strict_acr_pernodule")

    df_n = add_derived_nodule(raw_n)
    df_p = add_derived_patient(raw_p)
    strict_mask = _strict_eligible_mask(raw_n["analytic_eligible_strict_acr_pernodule"])
    df_ns = df_n.loc[strict_mask].copy()

    out_dir = os.path.join(PKG_DIR, "08_analysis_outputs")
    os.makedirs(out_dir, exist_ok=True)

    roc_df, roc_auc = supp_roc_curve(df_ns)

    snap = {
        "generated_utc": datetime.now(timezone.utc).isoformat(),
        "view": "manuscript_workspace.cohort_m025_nodule_level_v1",
        "build_mig": "mig_306",
        "n_total_rows": int(len(raw_n)),
        "n_strict_acr_analytic_eligible": int(strict_mask.sum()),
        "n_distinct_patients_all": int(raw_n["research_id"].nunique()),
        "patient_cohort_n": int(len(raw_p)),
        "n_tr_rank_known_strict": int(df_ns["tr_rank"].notna().sum()),
        "n_malignant_nodules_strict": int(df_ns["y_mal"].sum()),
        "roc_auc_ordinal_rank_strict": round(roc_auc, 4) if not np.isnan(roc_auc) else None,
    }
    with open(os.path.join(out_dir, "m025v2_run_snapshot.json"), "w", encoding="utf-8") as f:
        json.dump(snap, f, indent=2)

    # Parquet exports
    df_n.to_parquet(os.path.join(out_dir, "m025_analytic_spine.parquet"), index=False)
    df_p.to_parquet(os.path.join(out_dir, "m025v2_patient_level_compare.parquet"), index=False)

    tab1_n = table1_demo_by_tr(df_ns)
    tab_compare = table_patient_vs_nodule_rom(df_p, df_ns)
    tab_thresh = table_threshold_export_rows(df_ns)
    tab_beth = table_bethesda_by_tr_cross(df_ns)

    tab_compare.to_csv(os.path.join(out_dir, "m025v2_per_tr_rom_with_ci.csv"), index=False)
    tab_thresh.to_csv(os.path.join(out_dir, "m025v2_threshold_metrics_per_nodule.csv"), index=False)
    tab_beth.to_csv(os.path.join(out_dir, "m025v2_bethesda_x_tirads_counts.csv"))

    roc_path = os.path.join(out_dir, "m025v2_supp_ROC_curve_points.csv")
    if not roc_df.empty:
        roc_df.to_csv(roc_path, index=False)
        pd.DataFrame([{"metric": "roc_auc_ordinal_rank_nodule_strict", "value": roc_auc}]).to_csv(
            os.path.join(out_dir, "m025v2_supp_ROC_summary.csv"), index=False
        )

    roc_p_df, roc_p_auc = supp_roc_curve(df_p)
    if not roc_p_df.empty:
        roc_p_df.to_csv(os.path.join(out_dir, "m025v2_supp_ROC_patient_curve_points.csv"), index=False)
        pd.DataFrame([{"metric": "roc_auc_ordinal_rank_patient_level", "value": roc_p_auc}]).to_csv(
            os.path.join(out_dir, "m025v2_supp_ROC_patient_summary.csv"), index=False
        )

    cover = pd.DataFrame(
        [
            ("Package", "M025 v2.0 — nodule-level TI-RADS (mig_307)"),
            ("Database", "thyroid_canonical_publication_v1_0"),
            ("Nodule cohort view", "manuscript_workspace.cohort_m025_nodule_level_v1"),
            ("Patient comparator", "manuscript_workspace.cohort_m025_tirads_performance_v1"),
            ("Strict filter", "analytic_eligible_strict_acr_pernodule"),
            ("Gold (nodule)", "nodule_path_proven_malignant"),
            ("Predictor", "acr2017_tirads_category"),
            ("Build UTC", snap["generated_utc"]),
            ("Nodule rows (all)", snap["n_total_rows"]),
            ("Strict nodules", snap["n_strict_acr_analytic_eligible"]),
            (
                "ROC AUC (strict)",
                snap["roc_auc_ordinal_rank_strict"] if snap["roc_auc_ordinal_rank_strict"] is not None else "sklearn_missing",
            ),
        ],
        columns=["key", "value"],
    )

    qa = pd.DataFrame(
        [
            {"check": "nodule_spine_non_empty", "status": "PASS" if snap["n_total_rows"] > 0 else "FAIL"},
            {"check": "strict_cohort_~3687", "status": "PASS" if 3500 <= snap["n_strict_acr_analytic_eligible"] <= 3900 else "REVIEW"},
            {"check": "patient_comparator_nonempty", "status": "PASS" if snap["patient_cohort_n"] > 0 else "FAIL"},
        ]
    )

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Cover")
    write_df_to_sheet(ws, cover, "M025 v2.0 — tables workbook cover")

    ws = wb.create_sheet("Table1_nodule_ROM_by_TR")
    write_df_to_sheet(ws, tab1_n, "Table 1 — Nodule-level ROM by TR (strict ACR cohort)", "")

    ws = wb.create_sheet("Table2_thresholds_TR3_4_5")
    write_df_to_sheet(ws, tab_thresh, "Table 2 — Sens/Spec/PPV/NPV at TR≥TR3, TR4, TR5 (Wilson 95% CIs)", "")

    ws = wb.create_sheet("Table3_patient_vs_nodule_ROM")
    write_df_to_sheet(
        ws,
        tab_compare,
        "Table 3 — Patient- vs nodule-level ROM (headline comparison)",
        "Patient grain: cohort_m025_tirads_performance_v1 + CPM malignancy. Nodule grain: strict per-nodule path bridge.",
    )

    ws = wb.create_sheet("Table4_Bethesda_x_TR")
    write_df_to_sheet(ws, tab_beth.reset_index(), "Table 4 — Bethesda × TI-RADS (nodule grain, counts)", "")

    ws = wb.create_sheet("Supp_ROC_points")
    if roc_df.empty:
        write_df_to_sheet(ws, pd.DataFrame([{"note": "ROC skipped — sklearn missing or insufficient data"}]), "ROC", "")
    else:
        write_df_to_sheet(
            ws,
            roc_df,
            "Supplement — ROC operating points (ordinal TR rank, strict nodule cohort)",
            "AUC=%0.4f" % roc_auc,
        )

    ws = wb.create_sheet("QA")
    write_df_to_sheet(ws, qa, "QA gates", "")

    out_tables = os.path.join(PKG_DIR, "04_tables.xlsx")
    wb.save(out_tables)
    print(f"Wrote {out_tables}")

    meta = pd.DataFrame([{"column": col, "dtype": str(df_n[col].dtype)} for col in df_n.columns])

    preferred_nodule = [
        "research_id",
        "exam_date",
        "nodule_master_id",
        "acr2017_tirads_category",
        "acr2017_tirads_points",
        "analytic_eligible_strict_acr_pernodule",
        "nodule_path_proven_malignant",
        "bethesda_final_num",
        "bethesda_2023_num",
        "laterality_norm",
        "age_at_surgery",
        "sex",
        "tumor_size_cm",
        "tr_rank",
        "tr_label",
        "y_mal",
        "bethesda_bucket",
    ]
    have_pref = [c for c in preferred_nodule if c in df_ns.columns]
    rest_ns = [c for c in df_ns.columns if c not in have_pref]
    ns_export_cols = have_pref + rest_ns[: max(0, 80 - len(have_pref))]

    master_xlsx = os.path.join(PKG_DIR, "05_master_data.xlsx")
    with pd.ExcelWriter(master_xlsx, engine="openpyxl") as xw:
        # Full-rowcount spine lives in m025_analytic_spine.parquet (avoid 10+ MB xlsx).
        _excel_safe_df(df_ns[ns_export_cols]).to_excel(xw, sheet_name="nodule_spine_strict", index=False)
        dp = _excel_safe_df(df_p)
        _cap = min(len(dp.columns), 55)
        dp.iloc[:, :_cap].to_excel(xw, sheet_name="patient_comparator", index=False)
        pd.DataFrame([snap]).to_excel(xw, sheet_name="run_snapshot", index=False)
        meta.head(200).to_excel(xw, sheet_name="nodule_columns", index=False)
    print(f"Wrote {master_xlsx}")


if __name__ == "__main__":
    main()
