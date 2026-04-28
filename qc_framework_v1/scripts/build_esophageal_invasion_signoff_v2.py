"""
qc_framework_v1/scripts/build_esophageal_invasion_signoff_v2.py
===============================================================

Esophageal sign-off review v2 (.xlsx, 2 sheets).

After the 32-row bulk reclass under Logan's compression / adjacency-only /
non-cancer pathology rules (mig_93a-row-1), the canonical now sits at
124 present / 64 negated. This builder produces a single workbook with:

  Sheet 1 -- "review_124_present"
    All 124 remaining 'present' rows. ONE column to set per row:
    your_decision  in {ACCEPT, FLIP, REJECT}.

  Sheet 2 -- "audit_32_auto_reclassed"
    The 32 rows that were auto-flipped present -> negated by the
    bulk-reclass rule. Read-only audit so Logan can spot-check that
    the rule caught the right cases. Each row carries
    auto_reclass_reason  in {compression, adjacency, noncancer}
    (or multiple, joined by '+'). No decision column here.

Format: .xlsx, evidence_text in dedicated wide wrap-text column,
header row bold + filled, decision col yellow-tinted, frozen panes.
"""
from __future__ import annotations

from pathlib import Path

import duckdb
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

REPO_ROOT = Path(__file__).resolve().parents[2]
OUT = (
    REPO_ROOT
    / "verification_csvs"
    / "canonical_esophageal_invasion_events_v1"
    / "esophageal_signoff__mig_93__v2.xlsx"
)
DB = "thyroid_canonical_publication_v1_0"

POSITIVE_ENTITY_TYPES = (
    "esophageal_invasion_present",
    "esophageal_invasion_extent",
    "esophageal_muscularis_invasion",
    "esophageal_mucosal_invasion",
    "esophageal_invasion_length_cm",
    "esophageal_repair_performed",
)

# Same filter logic as the bulk-reclass UPDATE -- applied here to
# label the 64 'negated' rows so we can split out the 32 auto-reclassed.
RECLASS_REASON_SQL = """
CASE
  WHEN LOWER(COALESCE(evidence_text,'')) LIKE '%compress%'
    OR LOWER(COALESCE(evidence_text,'')) LIKE '%displace%'
    OR LOWER(COALESCE(evidence_text,'')) LIKE '%deviat%'
    OR LOWER(COALESCE(evidence_text,'')) LIKE '%effac%'
    OR LOWER(COALESCE(evidence_text,'')) LIKE '%mass effect%'
  THEN 'compression'
  WHEN (
        LOWER(COALESCE(evidence_text,'')) LIKE '%behind the esophagus%'
     OR LOWER(COALESCE(evidence_text,'')) LIKE '%posterior to the esophagus%'
     OR LOWER(COALESCE(evidence_text,'')) LIKE '%retroesophageal%'
     OR LOWER(COALESCE(evidence_text,'')) LIKE '%peeled%'
     OR LOWER(COALESCE(evidence_text,'')) LIKE '%separated from%'
     OR LOWER(COALESCE(evidence_text,'')) LIKE '%adjacent to%'
     OR LOWER(COALESCE(evidence_text,'')) LIKE '%along the%esophagus%')
   AND NOT (
        LOWER(COALESCE(evidence_text,'')) LIKE '%invad%'
     OR LOWER(COALESCE(evidence_text,'')) LIKE '%invasion%'
     OR LOWER(COALESCE(evidence_text,'')) LIKE '%involve%'
     OR LOWER(COALESCE(evidence_text,'')) LIKE '%infiltrat%'
     OR LOWER(COALESCE(evidence_text,'')) LIKE '%inseparable%'
     OR LOWER(COALESCE(evidence_text,'')) LIKE '%defect%'
     OR LOWER(COALESCE(evidence_text,'')) LIKE '%fistul%')
  THEN 'adjacency'
  WHEN LOWER(COALESCE(evidence_text,'')) LIKE '%goiter%'
    OR LOWER(COALESCE(evidence_text,'')) LIKE '%benign%'
    OR LOWER(COALESCE(evidence_text,'')) LIKE '%hyperplas%'
    OR LOWER(COALESCE(evidence_text,'')) LIKE '%parathyroid adenoma%'
    OR LOWER(COALESCE(evidence_text,'')) LIKE '%nodular thyroid%'
    OR LOWER(COALESCE(evidence_text,'')) LIKE '%multinodular%'
  THEN 'noncancer'
  ELSE NULL
END
"""


def _format_header(ws, headers, header_row):
    fill = PatternFill(start_color="FFE6E6E6",
                       end_color="FFE6E6E6", fill_type="solid")
    for col_ix in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_ix)
        cell.font = Font(bold=True)
        cell.fill = fill


def _set_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def main() -> None:
    con = duckdb.connect("md:")
    con.execute(f'USE "{DB}"')

    # ---------- Sheet 1: 124 remaining 'present' rows ----------
    present_rows = con.execute(
        f"""
        SELECT
          research_id, note_type, note_date, source_column,
          entity_type, entity_value, present_or_negated,
          confidence, source_line, entity_date, evidence_text,
          CASE WHEN entity_type IN
               ({", ".join(f"'{t}'" for t in POSITIVE_ENTITY_TYPES)})
               THEN 1 ELSE 0 END AS is_invasion_entity
        FROM main.canonical_esophageal_invasion_events_v1
        WHERE present_or_negated = 'present'
        ORDER BY is_invasion_entity DESC, entity_type, research_id
        """
    ).fetchall()
    print(f"sheet1 present rows: {len(present_rows)}")

    # ---------- Sheet 2: 32 auto-reclassed (now 'negated') ----------
    audit_rows = con.execute(
        f"""
        SELECT
          research_id, note_type, note_date,
          entity_type, entity_value, present_or_negated,
          confidence, evidence_text,
          {RECLASS_REASON_SQL} AS auto_reclass_reason
        FROM main.canonical_esophageal_invasion_events_v1
        WHERE present_or_negated = 'negated'
          AND {RECLASS_REASON_SQL} IS NOT NULL
        ORDER BY auto_reclass_reason, research_id, entity_type
        """
    ).fetchall()
    print(f"sheet2 audit rows: {len(audit_rows)}")

    wb = openpyxl.Workbook()

    # ---------- SHEET 1 ----------
    ws1 = wb.active
    ws1.title = "review_124_present"
    instructions = [
        "Esophageal invasion sign-off -- canonical_esophageal_invasion_events_v1",
        "Sheet 1 of 2: 124 'present' rows remaining after the 32-row auto-reclass.",
        "",
        "Decision rule: ONE column per row -> your_decision",
        "  ACCEPT  (default; blank == ACCEPT) -- real esophageal invasion event",
        "  FLIP    -- LLM got the sense wrong; should be 'negated'",
        "  REJECT  -- not a real esophageal invasion event (drop the row)",
        "",
        "Cancer-only and compression-vs-invasion rules already applied as",
        "the bulk-reclass UPDATE (32 rows -> 'negated'). See Sheet 2 for audit.",
        "",
    ]
    for line in instructions:
        ws1.append([line])
    headers1 = [
        "research_id", "note_type", "note_date", "source_column",
        "entity_type", "entity_value", "present_or_negated",
        "confidence", "source_line", "entity_date", "evidence_text",
        "is_invasion_entity",
        "your_decision", "your_note",
    ]
    ws1.append(headers1)
    header_row1 = ws1.max_row
    _format_header(ws1, headers1, header_row1)
    for r in present_rows:
        ws1.append(list(r) + ["", ""])
    _set_widths(ws1, {
        "A": 12, "B": 16, "C": 12, "D": 14,
        "E": 32, "F": 22, "G": 16,
        "H": 10, "I": 10, "J": 12,
        "K": 80, "L": 16,
        "M": 14, "N": 30,
    })
    ev_col = headers1.index("evidence_text") + 1
    dec_col = headers1.index("your_decision") + 1
    yellow = PatternFill(start_color="FFFFFFCC",
                         end_color="FFFFFFCC", fill_type="solid")
    for row_ix in range(header_row1 + 1, ws1.max_row + 1):
        ws1.cell(row=row_ix, column=ev_col).alignment = Alignment(
            wrap_text=True, vertical="top")
        ws1.cell(row=row_ix, column=dec_col).fill = yellow
    ws1.freeze_panes = ws1.cell(row=header_row1 + 1, column=1)

    # ---------- SHEET 2 ----------
    ws2 = wb.create_sheet("audit_32_auto_reclassed")
    instructions2 = [
        "AUDIT: 32 rows auto-flipped present -> negated by the bulk-reclass rule.",
        "READ-ONLY. Spot-check the auto_reclass_reason against the evidence_text.",
        "",
        "auto_reclass_reason values:",
        "  compression -- evidence has compress/displace/deviate/efface/mass-effect",
        "  adjacency   -- evidence has behind/posterior/retroesophageal/peeled/etc.",
        "                 WITHOUT any invasion verb (invad/invasion/involve/etc.)",
        "  noncancer   -- evidence mentions goiter/benign/hyperplasia/parathyroid",
        "                 adenoma/nodular thyroid/multinodular",
        "",
        "If you spot a row that should NOT have been auto-reclassed, leave a note",
        "in the spot_check column at the bottom of the table. I'll surface those",
        "in the next round.",
        "",
    ]
    for line in instructions2:
        ws2.append([line])
    headers2 = [
        "research_id", "note_type", "note_date",
        "entity_type", "entity_value", "present_or_negated",
        "confidence", "evidence_text",
        "auto_reclass_reason",
        "spot_check_note",
    ]
    ws2.append(headers2)
    header_row2 = ws2.max_row
    _format_header(ws2, headers2, header_row2)
    for r in audit_rows:
        ws2.append(list(r) + [""])
    _set_widths(ws2, {
        "A": 12, "B": 16, "C": 12,
        "D": 32, "E": 22, "F": 16,
        "G": 10, "H": 80,
        "I": 16, "J": 30,
    })
    ev_col2 = headers2.index("evidence_text") + 1
    for row_ix in range(header_row2 + 1, ws2.max_row + 1):
        ws2.cell(row=row_ix, column=ev_col2).alignment = Alignment(
            wrap_text=True, vertical="top")
    ws2.freeze_panes = ws2.cell(row=header_row2 + 1, column=1)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"-> wrote {OUT}")


if __name__ == "__main__":
    main()
