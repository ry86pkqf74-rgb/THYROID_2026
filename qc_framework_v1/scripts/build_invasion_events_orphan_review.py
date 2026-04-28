"""
qc_framework_v1/scripts/build_invasion_events_orphan_review.py
================================================================

Builds the orphan-review .xlsx for canonical_invasion_events_v1 mig_91 under
Protocol v2.

Discovery (Cowork session, 2026-04-28): the published canonical at
'thyroid_canonical_publication_v1_0' (logan.glosser.eras account) and the
archive snapshot

    "Thyroid 2026 UPdated".archive_pub_v1_0
        .canonical_invasion_events_v1_pre363v3_20260422_032942

share an identical row count (51,773) and identical row-identifying keys
(invasion_event_id, research_id, invasion_type, source_table, source_row_id,
ROW_NUMBER() OVER (...) for collision-prone mention-grain rows). This admits
a CTC-equivalence verification pattern even though canonical_invasion_events_v1
was assembled by Script 363's UNION pipeline rather than a SELECT*+filter+UPDATE
chain on a single source.

Mass-equivalence on the 11 not_started cols against the snapshot:

  ZERO diffs:    invasion_type, finding_date, source_modality, source_kind,
                 linkage_method, n_candidate_episodes,
                 linkage_ambiguous_multi_episode  (7 cols → CTC-pass)

  Localized:     finding_status (1,353 rows / 100% LLM / 100% downgrades)
                 evidence_qualifier (70 rows — mostly row-pair swaps)
                 evidence_span_hash (4 rows — pair swap on 2 patients)
                 confidence (2 rows — pair swap on 1 patient)

Of the 1,353 finding_status downgrades:
  ~312  rule-library matches (cannot be assessed / equivocal / not applicable
        / compression / adjacency / explicit-negative / adherent-only)
  ~828  patient has structured 'present' covering same/related invasion type
        (Script 363 correctly de-duplicated)
  ~213  remaining defensible (compression/adjacency/explicit-negative orphans)
  101   ORPHAN downgrades, split by Rule #1 (cancer-only):
          54 BENIGN orphans (canonical_path_benign_events_v1 entry, NO
                             malignant path) -- per Rule #1 Script 363
                             correctly downgraded; massive goiter / MNG
                             substernal extension being mis-extracted as
                             malignant ETE. AUDIT-confirm only.
          47 CANCER orphans (canonical_path_malignant_events_v1 entry).
                             TRUE review queue.

Output:
  verification_csvs/canonical_invasion_events_v1/orphan_review__mig_91.xlsx

Sheets (v3 with cancer sub-buckets per Logan's clinical rules):
  1. summary                      -- methodology, statistics, sign-off block
  2. 1_high_pos                   -- HIGH-CONFIDENCE POSITIVES (anatomic +
                                     invasion verb). Default = FLIP_TO_PRESENT.
  3. 2_keyword_needs_context      -- bare 'extrathyroidal extension' rows;
                                     mark NEEDS_CONTEXT and I'll pull source.
  4. 3_ln_ene                     -- lymph node extranodal extension rows;
                                     NOT thyroid-tumor ETE. Default = RECLASS
                                     (probably to lymphatic_microscopic) or
                                     REJECT.
  5. 4_ambig_extracap             -- 'extracapsular extension' without
                                     thyroid-vs-LN qualifier; usually LN in
                                     pathology context. NEEDS_CONTEXT.
  6. 5_vocal_or_not_invasion      -- vocal cord paralysis (different column)
                                     or pure mass effect / incidental.
                                     Default = REJECT or RECLASS.
  7. 6_benign_orphans             -- 54 audit-only rows (benign path; Rule #1
                                     confirms Script 363 correctly downgraded
                                     -- goiter / MNG substernal extension).
  8. hash_conf_swaps              -- 6 informational pair-swap rows.

Decision vocabulary (your_decision col):
  ACCEPT             — downgrade is correct (LLM finding correctly weakened)
  FLIP_TO_PRESENT    — downgrade was wrong; evidence supports 'present'
  FLIP_TO_SUSPECTED  — indeterminate too weak; should be 'suspected'
  REJECT             — drop the row entirely (LLM mis-extracted)
  RECLASS_INVASION_TYPE — row should be a different invasion_type (specify)
  NEEDS_CONTEXT      — pull source-note context before deciding (KEYWORD /
                       AMBIG_EC rows)

Cancer sub-bucket categorization rules (categorize_cancer_orphan):
  HIGH_POS  -- contains anatomic structure + invasion verb (strap muscle
               invasion, cartilage erosion, extends into trachea, encasement,
               mediastinum involvement, etc.)
  LN_ENE    -- contains 'extranodal' OR ('lymph node' + 'invasion'/'extension')
               OR 'pericapsular' + 'lymph node'. Lymph node ENE is NOT
               thyroid-tumor ETE.
  VOCAL_OR_NOT_INVASION -- contains vocal cord / vocal fold / RLN /
               recurrent laryngeal (different column entirely) OR pure
               deviation/displacement without invasion verb (mass effect)
               OR incidental finding (laryngocele).
  AMBIG_EC  -- 'extracapsular extension' without thyroid-vs-LN qualifier
               (usually LN in pathology context but ambiguous w/o source).
  KEYWORD   -- default for bare 'extrathyroidal extension' / 'ETE' without
               surrounding anatomic detail.

Usage:
  Re-pulls the data via duckdb-py with MotherDuck SSO (logan.glosser.eras
  per memory reference_protocol_v2_md_accounts.md). Builds .xlsx via
  openpyxl (csv.QUOTE_ALL fragile in Excel per memory
  feedback_review_csv_formatting.md).

  python3 qc_framework_v1/scripts/build_invasion_events_orphan_review.py
"""
from __future__ import annotations

import os
from pathlib import Path

import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
OUT = (
    REPO_ROOT
    / "verification_csvs"
    / "canonical_invasion_events_v1"
    / "orphan_review__mig_91.xlsx"
)
LIVE_DB = "thyroid_canonical_publication_v1_0"
ARCHIVE_DB = "Thyroid 2026 UPdated"
ARCHIVE_TABLE = (
    'archive_pub_v1_0.canonical_invasion_events_v1_pre363v3_20260422_032942'
)

HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF")
HDR_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(border_style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PRIORITY_FILL = PatternFill("solid", fgColor="FFE6CC")
SAFE_FILL = PatternFill("solid", fgColor="E2EFDA")


def _connect():
    token = os.environ.get("MOTHERDUCK_TOKEN") or os.environ.get("motherduck_token")
    if not token:
        raise RuntimeError(
            "MOTHERDUCK_TOKEN env var required (auth as "
            "logan.glosser.eras@gmail.com per memory "
            "reference_protocol_v2_md_accounts.md)"
        )
    con = duckdb.connect(f"md:?motherduck_token={token}")
    return con


def _diff_cte(con):
    """Return the SQL CTE that produces the (live, arc) diff key-pair join."""
    return f"""
    WITH live AS (
      SELECT *,
        ROW_NUMBER() OVER (
          PARTITION BY invasion_event_id, research_id, invasion_type, source_table, source_row_id
          ORDER BY finding_date, finding_status
        ) AS rn
      FROM {LIVE_DB}.main.canonical_invasion_events_v1
    ),
    arc AS (
      SELECT *,
        ROW_NUMBER() OVER (
          PARTITION BY invasion_event_id, research_id, invasion_type, source_table, source_row_id
          ORDER BY finding_date, finding_status
        ) AS rn
      FROM "{ARCHIVE_DB}".{ARCHIVE_TABLE}
    ),
    paired AS (
      SELECT
        live.* EXCLUDE (rn),
        arc.finding_status      AS arc_finding_status,
        arc.evidence_qualifier  AS arc_evidence_qualifier,
        arc.evidence_span_hash  AS arc_evidence_span_hash,
        arc.confidence          AS arc_confidence
      FROM live
      JOIN arc
        ON live.invasion_event_id IS NOT DISTINCT FROM arc.invasion_event_id
        AND live.research_id      IS NOT DISTINCT FROM arc.research_id
        AND live.invasion_type    IS NOT DISTINCT FROM arc.invasion_type
        AND live.source_table     IS NOT DISTINCT FROM arc.source_table
        AND live.source_row_id    IS NOT DISTINCT FROM arc.source_row_id
        AND live.rn = arc.rn
    )
    """


def categorize_cancer_orphan(evidence_qualifier: str) -> str:
    """Sub-bucket a CANCER orphan row per Logan's clinical rules (2026-04-28).

    Returns one of:
        HIGH_POS               -- anatomic structure + invasion verb
        LN_ENE                 -- lymph node extranodal extension
                                  (NOT thyroid-tumor ETE)
        VOCAL_OR_NOT_INVASION  -- vocal cord / RLN (different column) OR
                                  pure mass effect / displacement / incidental
        AMBIG_EC               -- 'extracapsular extension' without thyroid
                                  vs LN qualifier; usually LN in path context
        KEYWORD                -- bare 'extrathyroidal extension' / 'ETE'
                                  without surrounding anatomic detail;
                                  needs source-note context
    """
    q = (evidence_qualifier or "").lower()

    # LN ENE -- explicit
    if "extranodal" in q:
        return "LN_ENE"
    if "lymph node" in q and ("invasion" in q or "extension" in q):
        return "LN_ENE"
    if "pericapsular" in q and "lymph node" in q:
        return "LN_ENE"

    # Vocal cord / RLN / paralysis -- different column
    if any(t in q for t in ["vocal cord", "vocal fold", "rln",
                              "recurrent laryngeal"]):
        return "VOCAL_OR_NOT_INVASION"

    # Pure mass effect / displacement / incidental finding
    if "laryngocele" in q:
        return "VOCAL_OR_NOT_INVASION"
    if any(t in q for t in ["deviation", "displaced", "displacement",
                              "deviated"]) \
            and not any(v in q for v in ["invasion", "extension", "erosion",
                                          "involvement", "extends",
                                          "encasement"]):
        return "VOCAL_OR_NOT_INVASION"

    # Ambiguous extracapsular (could be LN or thyroid tumor; needs source)
    if "extracapsular" in q and "extrathyroidal" not in q and "thyroid" not in q:
        return "AMBIG_EC"

    # High-confidence positive: anatomic structure + invasion verb
    high_pos_patterns = (
        "strap muscle invasion", "cartilage erosion", "tracheoesophageal",
        "cricothyroid membrane", "cricoid cartilage", "thyroid cartilage",
        "extends into", "encasement", "encased", "circumferential",
        "brachiocephalic", "svc", "trachea", "esophagus", "mediastinum",
        "involvement of",
    )
    if any(t in q for t in high_pos_patterns):
        return "HIGH_POS"

    return "KEYWORD"


def fetch_all_orphans(con):
    """Pull ALL orphan rows (LLM-only, no structured fallback) split into
    Z + Y buckets, with a cancer_status flag (CANCER if patient has
    canonical_path_malignant_events_v1 entry, else BENIGN).

    Per Rule #1 (cancer-only) of the clinical rule library, BENIGN orphans
    are correctly downgraded by Script 363 -- they represent massive goiter /
    MNG / multinodular substernal extension being mis-extracted as malignant
    ETE on patients with NO malignant pathology. CANCER orphans are the true
    review queue.
    """
    sql = f"""
    {_diff_cte(con)}
    , diffs AS (
      SELECT *,
        LOWER(TRIM(evidence_qualifier)) AS qn
      FROM paired
      WHERE source_kind = 'llm'
        AND finding_status IS DISTINCT FROM arc_finding_status
    )
    SELECT
      invasion_event_id,
      research_id,
      invasion_type,
      source_modality,
      finding_date,
      arc_finding_status   AS arc_status,
      finding_status       AS live_status,
      evidence_qualifier,
      confidence,
      CASE WHEN EXISTS (SELECT 1 FROM {LIVE_DB}.main.canonical_path_malignant_events_v1 pm
                        WHERE pm.research_id = d.research_id)
           THEN 'CANCER' ELSE 'BENIGN' END AS cancer_status,
      CASE WHEN qn LIKE '%extrathyroidal%' OR qn LIKE '%ete%'
                OR qn LIKE '%strap muscle%' OR qn LIKE '%muscle invasion%'
                OR qn LIKE '%invasion%' THEN 'Z' ELSE 'Y' END AS rule_bucket
    FROM diffs d
    WHERE NOT EXISTS (
      SELECT 1 FROM {LIVE_DB}.main.canonical_invasion_events_v1 s
      WHERE s.research_id = d.research_id
        AND s.source_kind = 'structured'
        AND s.finding_status = 'present'
    )
    AND (
      -- Z bucket
      ((qn LIKE '%extrathyroidal%' OR qn LIKE '%ete%' OR qn LIKE '%strap muscle%'
        OR qn LIKE '%muscle invasion%' OR qn LIKE '%invasion%')
       AND qn NOT LIKE '%no %invasion%' AND qn NOT LIKE '%not %invasion%'
       AND qn NOT LIKE '%cannot%' AND qn NOT LIKE '%equivoc%'
       AND qn NOT LIKE '%suspect%' AND qn NOT LIKE '%suggest%'
       AND qn NOT LIKE '%possib%' AND qn NOT LIKE '%question%'
       AND qn NOT LIKE '%uncert%')
      OR
      -- Y bucket
      (NOT (qn LIKE '%cannot%' OR qn LIKE '%equivoc%' OR qn LIKE '%uncertain%'
            OR qn LIKE '%probable%' OR qn LIKE '%suspect%' OR qn LIKE '%suggest%'
            OR qn LIKE '%possib%' OR qn LIKE '%question%'
            OR qn LIKE '%not_applicable%' OR qn LIKE '%not applicable%')
       AND NOT (qn LIKE '%compress%' OR qn LIKE '%displace%'
            OR qn LIKE '%mass effect%' OR qn LIKE '%deviate%' OR qn LIKE '%efface%')
       AND NOT (qn LIKE '%adjacent%' OR qn LIKE '%abut%'
            OR qn LIKE '%posterior to%' OR qn LIKE '%along the%')
       AND NOT (qn LIKE '%no %invasion%' OR qn LIKE '%no entrance%'
            OR qn LIKE '%not identified%' OR qn LIKE '%not violated%')
       AND NOT (qn LIKE '%adherent%' OR qn LIKE '%adhesion%')
       AND NOT (qn LIKE '%extrathyroidal%' OR qn LIKE '%ete%'
            OR qn LIKE '%strap muscle%' OR qn LIKE '%muscle invasion%'
            OR qn LIKE '%invasion%'))
    )
    ORDER BY cancer_status, invasion_type, research_id
    """
    return con.execute(sql).fetchall()


def fetch_orphans(con, bucket: str):
    """[DEPRECATED] use fetch_all_orphans -- kept for backward compat."""
    if bucket == "Z":
        clause = (
            "(qn LIKE '%extrathyroidal%' OR qn LIKE '%ete%' "
            "OR qn LIKE '%strap muscle%' OR qn LIKE '%muscle invasion%' "
            "OR qn LIKE '%invasion%') "
            "AND qn NOT LIKE '%no %invasion%' AND qn NOT LIKE '%not %invasion%' "
            "AND qn NOT LIKE '%cannot%' AND qn NOT LIKE '%equivoc%' "
            "AND qn NOT LIKE '%suspect%' AND qn NOT LIKE '%suggest%' "
            "AND qn NOT LIKE '%possib%' AND qn NOT LIKE '%question%' "
            "AND qn NOT LIKE '%uncert%'"
        )
    elif bucket == "Y":
        clause = (
            "NOT (qn LIKE '%cannot%' OR qn LIKE '%equivoc%' OR qn LIKE '%uncertain%' "
            "OR qn LIKE '%probable%' OR qn LIKE '%suspect%' OR qn LIKE '%suggest%' "
            "OR qn LIKE '%possib%' OR qn LIKE '%question%' OR qn LIKE '%not_applicable%' "
            "OR qn LIKE '%not applicable%') "
            "AND NOT (qn LIKE '%compress%' OR qn LIKE '%displace%' "
            "OR qn LIKE '%mass effect%' OR qn LIKE '%deviate%' OR qn LIKE '%efface%') "
            "AND NOT (qn LIKE '%adjacent%' OR qn LIKE '%abut%' "
            "OR qn LIKE '%posterior to%' OR qn LIKE '%along the%') "
            "AND NOT (qn LIKE '%no %invasion%' OR qn LIKE '%no entrance%' "
            "OR qn LIKE '%not identified%' OR qn LIKE '%not violated%') "
            "AND NOT (qn LIKE '%adherent%' OR qn LIKE '%adhesion%') "
            "AND NOT (qn LIKE '%extrathyroidal%' OR qn LIKE '%ete%' "
            "OR qn LIKE '%strap muscle%' OR qn LIKE '%muscle invasion%' "
            "OR qn LIKE '%invasion%')"
        )
    else:
        raise ValueError(bucket)

    sql = f"""
    {_diff_cte(con)}
    , diffs AS (
      SELECT *,
        LOWER(TRIM(evidence_qualifier)) AS qn
      FROM paired
      WHERE source_kind = 'llm'
        AND finding_status IS DISTINCT FROM arc_finding_status
    )
    SELECT
      invasion_event_id,
      research_id,
      invasion_type,
      source_modality,
      finding_date,
      arc_finding_status   AS arc_status,
      finding_status       AS live_status,
      evidence_qualifier,
      confidence
    FROM diffs d
    WHERE {clause}
      AND NOT EXISTS (
        SELECT 1 FROM {LIVE_DB}.main.canonical_invasion_events_v1 s
        WHERE s.research_id = d.research_id
          AND s.source_kind = 'structured'
          AND s.finding_status = 'present'
      )
    ORDER BY invasion_type, research_id
    """
    return con.execute(sql).fetchall()


def fetch_swaps(con):
    sql = f"""
    {_diff_cte(con)}
    SELECT 'hash_diff' AS diff_kind, invasion_event_id, research_id,
           invasion_type, source_modality, source_kind, finding_date,
           arc_evidence_span_hash AS arc_val,
           evidence_span_hash     AS live_val,
           CASE WHEN LEN(evidence_qualifier) > 200
                THEN LEFT(evidence_qualifier, 200) || '…'
                ELSE evidence_qualifier END AS evidence_qualifier
      FROM paired
     WHERE evidence_span_hash IS DISTINCT FROM arc_evidence_span_hash
    UNION ALL
    SELECT 'confidence_diff', invasion_event_id, research_id,
           invasion_type, source_modality, source_kind, finding_date,
           CAST(arc_confidence AS VARCHAR), CAST(confidence AS VARCHAR),
           CASE WHEN LEN(evidence_qualifier) > 200
                THEN LEFT(evidence_qualifier, 200) || '…'
                ELSE evidence_qualifier END
      FROM paired
     WHERE confidence IS DISTINCT FROM arc_confidence
    """
    return con.execute(sql).fetchall()


def _apply_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = HDR_ALIGN
        cell.border = BORDER


def _set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_summary(wb):
    ws = wb.create_sheet("summary", 0)
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, "canonical_invasion_events_v1 — mig_91 orphan review (Protocol v2)").font = Font(size=14, bold=True)

    rows = [
        ("Live state",                  "51,773 rows · 10,871 patients · 20 cols · 6 modality×kind slices"),
        ("Pre-363 archive",             f'"{ARCHIVE_DB}".{ARCHIVE_TABLE}'),
        ("CTC-equivalence pattern",     "Live vs pre-363 snapshot — same row count, same identifying keys, deterministic key-pair JOIN"),
        ("", ""),
        ("RESULT — 7 of 11 not_started cols pass at 0 diffs:", ""),
        ("",  "invasion_type / finding_date / source_modality / source_kind / linkage_method / n_candidate_episodes / linkage_ambiguous_multi_episode"),
        ("", ""),
        ("4 of 11 cols carry localized diffs:", ""),
        ("  finding_status",            "1,353 rows changed (100% on source_kind='llm'; 100% confidence-downgrades)"),
        ("  evidence_qualifier",        "70 rows — mostly row-pair swaps + 1 typo fix (informational)"),
        ("  evidence_span_hash",        "4 rows — pair swap on 2 patients (informational)"),
        ("  confidence",                "2 rows — pair swap on 1 patient (informational)"),
        ("", ""),
        ("Downgrade decomposition (1,353 finding_status changes):", ""),
        ("  rule-library matches",      "~312 rows — DEFENSIBLE"),
        ("  de-duplicated by structured", "~828 rows — DEFENSIBLE"),
        ("  ORPHAN downgrades",         "101 rows split by Rule #1 (cancer-only):"),
        ("    BENIGN orphans (54)",     "patient has canonical_path_benign_events_v1; NO malignant path. Per Rule #1, Script 363 correctly downgraded -- massive goiter / MNG substernal extension mis-extracted as malignant ETE. AUDIT-CONFIRM ONLY."),
        ("    CANCER orphans (47)",     "patient has canonical_path_malignant_events_v1. TRUE review queue."),
        ("  remaining ~113",            "compression/adjacency/explicit-negative/adherent orphans -- defensible"),
        ("", ""),
        ("Linkage cluster",             "All linkage cols show 0 diffs vs pre-363. 759-group ambiguous-linkage CSV unchanged from verified pre-363 — defer as CF-91-LINKAGE-COL-NAME."),
        ("", ""),
        ("Sign-off plan",               "1) Logan reviews cancer_orphans (47), ACCEPT/FLIP/REJECT per row. 2) Spot-check ~5 benign_orphans for Rule #1 audit. 3) FLIPs applied as UPDATE on canonical. 4) All 11 cols flagged 'verified'. 5) table_status='verified' (8/184 tables)."),
        ("", ""),
        ("Decision vocabulary:", ""),
        ("  ACCEPT",                    "downgrade is correct (LLM finding correctly weakened)"),
        ("  FLIP_TO_PRESENT",           "downgrade was wrong; evidence supports 'present'"),
        ("  FLIP_TO_SUSPECTED",         "indeterminate too weak; should be 'suspected'"),
        ("  REJECT",                    "drop the row entirely (LLM mis-extracted)"),
        ("  RECLASS_INVASION_TYPE",     "row should be a different invasion_type (specify in note)"),
    ]
    for r, (label, value) in enumerate(rows, start=2):
        ws.cell(r, 1, label).font = Font(bold=True)
        ws.cell(r, 2, value).alignment = WRAP
    _set_widths(ws, [42, 110])
    for r in range(2, len(rows) + 2):
        if ws.cell(r, 2).value and len(str(ws.cell(r, 2).value)) > 80:
            ws.row_dimensions[r].height = 30


def write_orphan_sheet(wb, name, data, intro, fill=None):
    """Write an orphan-review sheet. Each row is a 11-tuple from
    fetch_all_orphans:
        (eid, rid, itype, smod, fdate, arc_s, live_s, eq, conf,
         cancer_status, rule_bucket)
    """
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, intro).font = Font(size=11, bold=True, color="9C5700")
    ws.cell(1, 1).fill = fill if fill is not None else PRIORITY_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    ws.row_dimensions[1].height = 40

    headers = [
        "invasion_event_id", "research_id", "invasion_type", "source_modality",
        "finding_date", "arc_status", "live_status", "rule_bucket",
        "evidence_qualifier", "your_decision", "your_note",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(2, c, h)
    _apply_header(ws, row=2)
    ws.row_dimensions[2].height = 32

    for r_idx, row in enumerate(data, start=3):
        (eid, rid, itype, smod, fdate, arc_s, live_s, eq, conf,
         _cancer, bucket) = row
        ws.cell(r_idx, 1, eid)
        ws.cell(r_idx, 2, rid)
        ws.cell(r_idx, 3, itype)
        ws.cell(r_idx, 4, smod)
        ws.cell(r_idx, 5, str(fdate) if fdate is not None else "")
        ws.cell(r_idx, 6, arc_s)
        ws.cell(r_idx, 7, live_s)
        ws.cell(r_idx, 8, bucket)
        ws.cell(r_idx, 9, eq).alignment = WRAP
        ws.cell(r_idx, 10, "")
        ws.cell(r_idx, 11, "")
        ws.row_dimensions[r_idx].height = 30
        for c in range(1, 12):
            ws.cell(r_idx, c).border = BORDER
            if c == 9:
                ws.cell(r_idx, c).alignment = WRAP

    _set_widths(ws, [22, 10, 14, 14, 12, 12, 14, 7, 60, 22, 28])
    ws.freeze_panes = "A3"


def write_swap_sheet(wb, data):
    ws = wb.create_sheet("hash_conf_swaps")
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, "Hash + confidence diffs are PAIR SWAPS between near-duplicate rows on 2 patients (5986, 9846). Both rows exist with both qualifiers; just hash/confidence association flipped. Informational only — no decision needed.").font = Font(italic=True, color="595959")
    ws.cell(1, 1).fill = SAFE_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.row_dimensions[1].height = 32

    headers = ["diff_kind", "invasion_event_id", "research_id", "invasion_type",
               "source_modality", "source_kind", "finding_date",
               "arc_val", "live_val", "evidence_qualifier"]
    for c, h in enumerate(headers, start=1):
        ws.cell(2, c, h)
    _apply_header(ws, row=2)

    for r_idx, row in enumerate(data, start=3):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(r_idx, c, value if value is not None else "")
            cell.alignment = WRAP
            cell.border = BORDER
        ws.row_dimensions[r_idx].height = 28

    _set_widths(ws, [14, 22, 10, 14, 14, 10, 12, 36, 36, 36])
    ws.freeze_panes = "A3"


def main():
    con = _connect()
    rows = fetch_all_orphans(con)
    cancer = [r for r in rows if r[9] == "CANCER"]
    benign = [r for r in rows if r[9] == "BENIGN"]
    swaps = fetch_swaps(con)

    # Sub-bucket cancer orphans per Logan's clinical rules (rule_bucket col
    # is overridden with the finer-grained sub-bucket).
    def relabel(t, sub):
        # Replace position 10 (rule_bucket -- index 10) with the new label.
        return tuple(list(t[:10]) + [sub])

    cancer_sub = [relabel(t, categorize_cancer_orphan(t[7])) for t in cancer]

    HIGH_POS = [t for t in cancer_sub if t[10] == "HIGH_POS"]
    KEYWORD  = [t for t in cancer_sub if t[10] == "KEYWORD"]
    LN_ENE   = [t for t in cancer_sub if t[10] == "LN_ENE"]
    AMBIG_EC = [t for t in cancer_sub if t[10] == "AMBIG_EC"]
    VOCAL_OR = [t for t in cancer_sub if t[10] == "VOCAL_OR_NOT_INVASION"]

    wb = Workbook()
    wb.remove(wb.active)

    write_summary(wb)
    write_orphan_sheet(
        wb, "1_high_pos", HIGH_POS,
        f"PRIORITY 1 -- {len(HIGH_POS)} HIGH-CONFIDENCE POSITIVES. Anatomic "
        "structure + invasion verb. Default disposition = FLIP_TO_PRESENT.",
        fill=PRIORITY_FILL,
    )
    write_orphan_sheet(
        wb, "2_keyword_needs_context", KEYWORD,
        f"PRIORITY 2 -- {len(KEYWORD)} KEYWORD-ONLY rows. Bare 'extrathyroidal "
        "extension' or similar without surrounding anatomic detail. Mark rows "
        "with your_decision = NEEDS_CONTEXT and I'll re-pull with the source-"
        "note sentence(s) for those rows.",
        fill=PRIORITY_FILL,
    )
    write_orphan_sheet(
        wb, "3_ln_ene", LN_ENE,
        f"REVIEW -- {len(LN_ENE)} LN_ENE rows. Lymph node extranodal extension "
        "-- NOT thyroid-tumor ETE. Default disposition = RECLASS_INVASION_TYPE "
        "(probably to lymphatic_microscopic) or REJECT. NOT FLIP_TO_PRESENT for "
        "soft_tissue / perineural ETE.",
        fill=SAFE_FILL,
    )
    write_orphan_sheet(
        wb, "4_ambig_extracap", AMBIG_EC,
        f"REVIEW -- {len(AMBIG_EC)} AMBIGUOUS 'extracapsular extension' rows. "
        "Without thyroid-vs-LN qualifier, these are usually LN ENE in pathology "
        "context but ambiguous without source-note context. Mark NEEDS_CONTEXT "
        "and I'll pull surrounding text.",
        fill=PRIORITY_FILL,
    )
    write_orphan_sheet(
        wb, "5_vocal_or_not_invasion", VOCAL_OR,
        f"REVIEW -- {len(VOCAL_OR)} VOCAL CORD / NOT-INVASION rows. Vocal cord "
        "paralysis / RLN involvement go in a different column. Mass effect / "
        "displacement / laryngoceles are not invasion. Default = REJECT or "
        "RECLASS.",
        fill=SAFE_FILL,
    )
    write_orphan_sheet(
        wb, "6_benign_orphans", benign,
        f"AUDIT ONLY -- {len(benign)} BENIGN orphans (patient has "
        "canonical_path_benign_events_v1 only; NO malignant path). Per Rule #1 "
        "(cancer-only), Script 363 correctly downgraded these -- they are "
        "massive goiter / MNG / multinodular substernal extension being mis-"
        "extracted by the LLM as malignant ETE. Spot-check ~5 to audit-confirm "
        "Rule #1; default disposition = ACCEPT all.",
        fill=SAFE_FILL,
    )
    write_swap_sheet(wb, swaps)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(
        f"wrote {OUT} "
        f"(cancer: HIGH_POS={len(HIGH_POS)} KEYWORD={len(KEYWORD)} "
        f"LN_ENE={len(LN_ENE)} AMBIG_EC={len(AMBIG_EC)} "
        f"VOCAL_OR_NOT={len(VOCAL_OR)} | benign={len(benign)} swaps={len(swaps)})"
    )


if __name__ == "__main__":
    main()
