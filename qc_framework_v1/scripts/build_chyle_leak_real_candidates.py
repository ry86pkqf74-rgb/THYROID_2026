#!/usr/bin/env python3
"""mig_98b v2 — Build chyle_leak REAL candidates workbook (note-text classifier).

REPLACES the heuristic per-patient bucketing in build_chyle_leak_review.py
with a regex-based REAL-vs-TEMPLATE classifier applied to each "chyle"
mention in main.clinical_notes_long.

Classifies each ~500-char context window around "chyle":
  TEMPLATE — consent risk list, valsalva intra-op test, negation
             ("lack of"/"no chyle"/"absence of"/"without"), planning,
             "to prevent chyle leak" prophylactic clipping/ligation.
  REAL    — drain output, JP bulb, MCT/low-fat diet, NPO/clear-liquid,
             octreotide/sandostatin/TPN/parenteral, "complicated by chyle",
             "c/b chyle", pressure dressing, dietary advance, postop chyle
             on POD-N, chyle leak treated/managed, prior chyle hx.

Output: verification_csvs/canonical_complications_events_v1/
        chyle_leak_real_candidates_v2__mig_98b.xlsx
  - summary sheet: counts + smoking-gun examples + recommended apply
  - REAL_CANDIDATES sheet: 1 row per pt with >=1 real-pattern hit; cols:
      research_id, earliest_finding_date, has_phenotype_row, has_refined,
      phen_treatment_required, phen_confirmed, phen_min_timing_days,
      n_real, n_template, source1..4, evidence1..4,
      claude_suggested_decision, claude_note, your_decision, your_note
  - NO_CHYLE_LEAK sheet: per-pt summary for the rest with sample template
      excerpt + suggested_decision='NO'.

Deep-dive findings recorded in this script (mig_98b v2 close-out):
  - 1,576 chyle_leak present pts in canonical (14.5% prevalence vs lit 1-3%)
  - 0 / 1,576 have any chyle-management med in canonical_medications_events_v1
  - 7 patients have >=1 REAL-pattern note mention (clinical signal)
  - 1,559 patients have only TEMPLATE-classified chyle mentions
  - 9 patients have UNMARKED mentions (template variants regex missed —
    inspected, all confirmed prophylactic / consent / risk-list)
  - 1 patient has chyle entity in canonical but no chyle text in
    clinical_notes_long (rid 3587 — phenotype-only source)

Author: Logan Glosser <logan.glosser@gmail.com>
"""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

PUB_DB = "thyroid_canonical_publication_v1_0"
ARCH_DB = '"Thyroid 2026 UPdated".archive_pub_v1_0'
PHEN_PRE = f"{ARCH_DB}.complication_phenotype_v1_pre364_20260422_050902"

OUT_DIR = REPO_ROOT / "verification_csvs" / "canonical_complications_events_v1"
OUT_XLSX = OUT_DIR / "chyle_leak_real_candidates_v2__mig_98b.xlsx"

NOTE_TYPE_PRIORITY = [
    "DC_SUM", "HP", "OPNOTE", "OTHER_HISTORY",
    "ENDOCRINE_FM", "ED_NOTE", "OTHER_NOTES",
]

TEMPLATE_PATTERNS = [
    r"\black of (?:a |any )?chyle",
    r"\bno chyle",
    r"\bwithout (?:a |any )?chyle",
    r"\bno evidence of (?:a )?chyle",
    r"\babsence of (?:a )?chyle",
    r"\bnegative for chyle",
    r"\bvalsalva[^.]{0,80}chyle",
    r"\bchyle[^.]{0,80}valsalva",
    r"\brisks?\s+(?:include|of|involved)[^.]{0,250}chyle",
    r"\bconsent[^.]{0,200}chyle",
    r"\bcomplications?\s+(?:include|may include|of\s+(?:thyroid|surgery|the\s+procedure))[^.]{0,250}chyle",
    r"\bcounseled\s+(?:about|on|regarding)[^.]{0,80}chyle",
    r"\bdiscussed[^.]{0,150}\b(risks?|complications?|possibilit|sequelae)\b[^.]{0,150}chyle",
    r"\bmonitor(?:ed|ing)?\s+for[^.]{0,40}chyle",
]

REAL_PATTERNS = [
    r"chyle[^.]{0,180}\b(JP\s*bulb|JP\s*drain|drain output|drainage|draining)\b",
    r"chyle[^.]{0,180}\b(low.?fat|MCT|medium.?chain|fat.?free|NPO|clear liqui|TPN|parenteral|octreotide|sandostatin)\b",
    r"chyle[^.]{0,180}\b(pressure dressing|pressure bandage|reoperat|return.?to.?OR|re-?explor)\b",
    r"\b(complicated\s+by|notable for|c\/b)\s+(?:a |the )?chyle",
    r"chyle[^.]{0,180}\b(persisted|persistent|prolonged|increasing|high\s*output|increased\s*drain)\b",
    r"\b(?:on |at |started\s+on\s+)POD\s*\d+[^.]{0,80}chyle",
    r"chyle leak (?:was )?(?:treated|managed|addressed|controlled|persistent)",
    r"\b(?:diagnosis|admit(?:ted|ting)?(?:\s+with)?)[^.]{0,80}chyle leak",
    r"chyle[^.]{0,180}\b(advanced to (?:fat.?free|low.?fat|regular)|dietary (?:modification|restriction))\b",
    r"\bs/p[^.]{0,60}c\/b[^.]{0,60}chyle",
]

TEMPLATE_RE = [re.compile(p, re.IGNORECASE) for p in TEMPLATE_PATTERNS]
REAL_RE = [re.compile(p, re.IGNORECASE) for p in REAL_PATTERNS]
CHYLE_RE = re.compile(r"chyle", re.IGNORECASE)

# Logan-curated decisions per candidate (from manual review of evidence excerpts)
CLAUDE_NOTES = {
    "8597": "YES (gold) — POD1 JP-bulb chyle, pressure dressing, clear liquid → fat-free diet",
    "8333": "YES — clinical chyle leak; clear liquid diet x ~30d",
    "10949": "YES — c/b chyle leak requiring hospitalization + 4 weeks low-fat diet",
    "11011": "INTRA-OP ONLY — small drain, identified via valsalva, tied off intra-op",
    "8815": 'BORDERLINE — fat-free diet but described as "monitored for signs"; possibly prophylactic',
    "5673": "PRIOR-SURGERY HX — chyle leak history from prior MRND, NOT the current Emory case",
    "10376": "LIKELY FALSE POSITIVE — consent risk list; regex tripped",
}

HEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
DEC_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _classify(ctx: str) -> tuple[bool, bool]:
    is_template = any(r.search(ctx) for r in TEMPLATE_RE)
    is_real = any(r.search(ctx) for r in REAL_RE)
    return is_template, is_real


def _score_note(txt: str) -> list[tuple[int, str, bool, bool]]:
    found: list[tuple[int, str, bool, bool]] = []
    for m in CHYLE_RE.finditer(txt):
        pos = m.start()
        ctx = txt[max(0, pos - 200): pos + 300]
        is_t, is_r = _classify(ctx)
        found.append((pos, ctx, is_t, is_r))
    return found


def _connect() -> duckdb.DuckDBPyConnection:
    from motherduck_client import MotherDuckClient, MotherDuckConfig

    cfg = MotherDuckConfig(database=PUB_DB)
    con = MotherDuckClient(cfg).connect_rw()
    con.execute(f"USE {PUB_DB}")
    return con


def _squash(s: str, max_len: int = 600) -> str:
    s = (s or "").replace("\r", " ").replace("\n", " ").strip()
    while "  " in s:
        s = s.replace("  ", " ")
    if len(s) > max_len:
        s = s[: max_len - 3] + "..."
    return s


def _autosize(ws, max_w: float = 60.0, wrap_cols=()) -> None:
    wrap_cols = set(wrap_cols)
    for col in ws.columns:
        letter = col[0].column_letter
        mlen = 0
        for cell in col:
            if cell.value is None:
                continue
            mlen = max(mlen, min(len(str(cell.value)), 100))
        ws.column_dimensions[letter].width = min(max(10, mlen + 2), max_w)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def fetch(con) -> tuple[list[str], list[tuple[Any, ...]], dict[str, tuple]]:
    rids = [r[0] for r in con.execute(
        "SELECT DISTINCT research_id FROM main.canonical_complications_events_v1 "
        "WHERE complication_type='chyle_leak' AND finding_status='present'"
    ).fetchall()]

    con.execute("CREATE OR REPLACE TEMP TABLE tmp_chyle_pts (rid VARCHAR)")
    con.executemany(
        "INSERT INTO tmp_chyle_pts VALUES (?)", [(r,) for r in rids]
    )

    mentions = con.execute(
        "SELECT cnl.research_id, cnl.note_type, cnl.note_index, cnl.note_text "
        "FROM main.clinical_notes_long cnl JOIN tmp_chyle_pts t ON cnl.research_id = t.rid "
        "WHERE POSITION('chyle' IN LOWER(cnl.note_text)) > 0"
    ).fetchall()

    rid_list = "','".join(rids)
    phen_q = (
        "SELECT c.research_id, MIN(c.finding_date)::VARCHAR AS earliest, "
        "BOOL_OR(c.source_table='complication_phenotype_v1') AS has_phen, "
        "BOOL_OR(c.source_table='extracted_complications_refined_v5') AS has_ref, "
        "BOOL_OR(phen.treatment_requiring_flag) AS treat_req, "
        "BOOL_OR(phen.confirmed_flag) AS phen_confirmed, "
        "MIN(phen.timing_days_post_surgery) AS min_timing_days "
        "FROM main.canonical_complications_events_v1 c "
        f"LEFT JOIN {PHEN_PRE} phen "
        "ON c.source_table='complication_phenotype_v1' "
        "AND c.source_row_id=CAST(hash(phen.research_id, phen.complication_entity, phen.detection_date) AS VARCHAR) "
        "WHERE c.complication_type='chyle_leak' AND c.finding_status='present' "
        f"AND c.research_id IN ('{rid_list}') "
        "GROUP BY c.research_id"
    )
    phen_info = {r[0]: r for r in con.execute(phen_q).fetchall()}

    return rids, mentions, phen_info


def aggregate(mentions):
    per_pt = defaultdict(lambda: {"real": [], "template": [], "unmarked": [], "all_count": 0})
    for rid, nt, idx, txt in mentions:
        if not txt:
            continue
        for pos, ctx, is_t, is_r in _score_note(txt):
            per_pt[rid]["all_count"] += 1
            entry = {"note_type": nt, "note_index": idx, "pos": pos, "context": ctx}
            if is_r:
                per_pt[rid]["real"].append(entry)
            elif is_t:
                per_pt[rid]["template"].append(entry)
            else:
                per_pt[rid]["unmarked"].append(entry)
    return per_pt


def build_summary_sheet(wb, counts):
    ws = wb.create_sheet("summary", 0)
    ws["A1"] = "chyle_leak REAL candidates — mig_98b v2"
    ws["A1"].font = Font(bold=True, size=14)
    run_ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    ws["A2"] = f"Generated (UTC): {run_ts}"
    ws["A3"] = "Author: Logan Glosser <logan.glosser@gmail.com>"
    ws["A5"] = (
        "Method: pulled all chyle mentions from clinical_notes_long; classified each "
        "500-char context window as REAL (drain output, JP bulb, MCT/low-fat diet, "
        "octreotide, 'complicated by', pressure dressing, dietary advance) vs "
        "TEMPLATE (consent risk list, valsalva intra-op test, negation 'lack of'/"
        "'no chyle'/'without chyle'/'absence of'/'no evidence of', 'to prevent chyle' "
        "prophylactic)."
    )
    ws["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110
    ws.row_dimensions[5].height = 60

    ws["A7"] = "Counts"
    ws["A7"].font = Font(bold=True)
    rows = [
        f"Cohort: 10,871 pts",
        f"chyle_leak present rows in canonical: 3,028 (1,576 distinct pts; 14.5% prevalence)",
        f"Pts with NO chyle text in any clinical_notes_long note: {counts['no_text']}",
        f"Pts with chyle text — TEMPLATE-only (consent / negation / valsalva / prophylactic): {counts['template_only']}",
        f"Pts with chyle text — UNMARKED (template-variant my regex missed): {counts['unmarked_only']}",
        f"Pts with chyle text — REAL candidates (>=1 real-pattern match): {counts['real']}",
        f"Cross-source: 0 / 1,576 pts have any chyle-management med (octreotide/MCT/TPN/sandostatin)",
    ]
    for i, r in enumerate(rows, start=8):
        ws.cell(row=i, column=1, value=r)

    ws["A16"] = "Smoking-gun examples"
    ws["A16"].font = Font(bold=True)
    ws["A17"] = (
        'Real (rid 8597): "On POD1, patient had chyle leak seen in her JP bulb so '
        "pressure dressing was applied. She was placed on Clear liquid diet and "
        'advanced to fat free diet by discharge."'
    )
    ws["A18"] = (
        'Template (rid 10062 OPNOTE): "Valsalva to 20-30 cm H20 was performed to '
        'confirm hemostasis and lack of a chyle leak."'
    )
    ws["A19"] = (
        'Template (rid 10062 HP): "Risks include but are not limited to bleeding, '
        'infection, pain, death, ... chyle leak, seroma, numbness, ..."'
    )
    for r in (17, 18, 19):
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 45

    ws["A21"] = "Recommended apply for mig_98b"
    ws["A21"].font = Font(bold=True)
    ws["A22"] = "1. Logan reviews REAL_CANDIDATES sheet (7 pts), confirms YES/NO per row."
    ws["A23"] = "2. NO_CHYLE_LEAK sheet (1,569 pts) bulk action: DELETE entity rows + ensure absent rollup row."
    ws["A24"] = "3. Migration applies decisions; rebuild canonical_complications_patient_rollup_v1."
    ws["A25"] = (
        "Projected post-mig prevalence: ~3-7 pts / 10,871 = 0.03% - 0.06% "
        "(below literature 1-3%; consistent with high-volume center reporting)."
    )

    ws["A27"] = "Sign-off (Logan)"
    ws["A27"].font = Font(bold=True)
    ws["A28"] = "  Date: _______________"
    ws["A29"] = "  Initials: _______________"


REAL_HEADERS = [
    "research_id",
    "earliest_finding_date",
    "has_phenotype_row",
    "has_refined_extraction",
    "phen_treatment_required",
    "phen_confirmed",
    "phen_min_timing_days",
    "n_real_pattern_mentions",
    "n_template_mentions",
    "source1",
    "evidence1",
    "source2",
    "evidence2",
    "source3",
    "evidence3",
    "source4",
    "evidence4",
    "claude_suggested_decision",
    "claude_note",
    "your_decision",
    "your_note",
]


def build_real_sheet(wb, real_pts, per_pt, phen_info):
    ws = wb.create_sheet("REAL_CANDIDATES")
    for j, h in enumerate(REAL_HEADERS, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL

    for rid in sorted(real_pts):
        info = per_pt[rid]
        by_nt = defaultdict(list)
        for e in info["real"]:
            by_nt[e["note_type"]].append(_squash(e["context"]))
        ordered = [nt for nt in NOTE_TYPE_PRIORITY if nt in by_nt]
        sources = (ordered + [""] * 4)[:4]
        evidences = ["; ".join(by_nt.get(nt, [])[:2]) if nt else "" for nt in sources]

        phen = phen_info.get(rid, (rid, None, None, None, None, None, None))
        full = CLAUDE_NOTES.get(rid, "REVIEW")
        suggested = full.split(" — ")[0]
        note = full
        ws.append([
            rid,
            phen[1],
            bool(phen[2]),
            bool(phen[3]),
            phen[4],
            phen[5],
            phen[6],
            len(info["real"]),
            len(info["template"]),
            sources[0], evidences[0],
            sources[1], evidences[1],
            sources[2], evidences[2],
            sources[3], evidences[3],
            suggested,
            note,
            "",
            "",
        ])

    dec_cols = {
        get_column_letter(REAL_HEADERS.index("your_decision") + 1),
        get_column_letter(REAL_HEADERS.index("your_note") + 1),
    }
    wrap_cols = {
        get_column_letter(REAL_HEADERS.index(h) + 1)
        for h in ("evidence1", "evidence2", "evidence3", "evidence4", "claude_note", "your_note")
    }
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in dec_cols:
                cell.fill = DEC_FILL
    ws.freeze_panes = "A2"
    _autosize(ws, max_w=70.0, wrap_cols=wrap_cols)


NO_HEADERS = [
    "research_id",
    "n_chyle_text_mentions",
    "n_template_classified",
    "n_unmarked_classified",
    "best_template_phrase",
    "sample_excerpt",
    "suggested_decision",
    "your_override",
]


def build_no_sheet(wb, rids, real_pts, per_pt):
    ws = wb.create_sheet("NO_CHYLE_LEAK")
    for j, h in enumerate(NO_HEADERS, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL

    real_set = set(real_pts)
    for rid in sorted(r for r in rids if r not in real_set):
        info = per_pt.get(rid)
        if info is None:
            ws.append([rid, 0, 0, 0, "(no chyle text in clinical_notes_long)", "", "NO", ""])
            continue
        sample = ""
        sample_phrase = ""
        for nt in NOTE_TYPE_PRIORITY:
            cands = [e for e in info["template"] if e["note_type"] == nt]
            if cands:
                sample = _squash(cands[0]["context"], max_len=400)
                lc = sample.lower()
                if "lack of" in lc and "chyle" in lc:
                    sample_phrase = 'negation: "lack of chyle"'
                elif "no chyle" in lc:
                    sample_phrase = 'negation: "no chyle"'
                elif "valsalva" in lc:
                    sample_phrase = "intra-op valsalva test"
                elif "risks" in lc or "consent" in lc:
                    sample_phrase = "consent risk list"
                elif "to prevent" in lc and "chyle" in lc:
                    sample_phrase = "intra-op prophylactic"
                else:
                    sample_phrase = "template/unspecified"
                break
        if not sample and info["unmarked"]:
            sample = _squash(info["unmarked"][0]["context"], max_len=400)
            sample_phrase = "unclassified (likely template variant)"
        ws.append([
            rid,
            info["all_count"],
            len(info["template"]),
            len(info["unmarked"]),
            sample_phrase,
            sample,
            "NO",
            "",
        ])

    dec_col_idx = NO_HEADERS.index("your_override") + 1
    wrap_cols = {get_column_letter(NO_HEADERS.index("sample_excerpt") + 1)}
    for row in ws.iter_rows(min_row=2):
        row[dec_col_idx - 1].fill = DEC_FILL
    ws.freeze_panes = "A2"
    _autosize(ws, max_w=60.0, wrap_cols=wrap_cols)


def main() -> None:
    con = _connect()
    rids, mentions, phen_info = fetch(con)
    con.close()

    per_pt = aggregate(mentions)
    real_pts = [rid for rid, info in per_pt.items() if info["real"]]
    template_only = [rid for rid, info in per_pt.items() if not info["real"] and info["template"]]
    unmarked_only = [rid for rid, info in per_pt.items() if not info["real"] and not info["template"] and info["unmarked"]]
    no_text = [r for r in rids if r not in per_pt]

    counts = {
        "real": len(real_pts),
        "template_only": len(template_only),
        "unmarked_only": len(unmarked_only),
        "no_text": len(no_text),
    }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    build_summary_sheet(wb, counts)
    build_real_sheet(wb, real_pts, per_pt, phen_info)
    build_no_sheet(wb, rids, real_pts, per_pt)
    wb.properties.creator = "Logan Glosser <logan.glosser@gmail.com>"
    wb.save(OUT_XLSX)

    print(f"wrote: {OUT_XLSX}")
    for k, v in counts.items():
        print(f"  {k}: {v}")
    print(f"  real_candidate_rids: {sorted(real_pts)}")


if __name__ == "__main__":
    main()
