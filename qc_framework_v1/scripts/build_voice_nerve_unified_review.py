#!/usr/bin/env python3
"""mig_98c v2 — Build unified voice/nerve injury review workbook.

Per Logan's request:
  - RLN_injury and vocal_cord_paralysis (VCP / paresis) are two views of the
    same clinical category (mechanism vs functional outcome). Review jointly.
  - Each patient gets ONE row with RLN + VCP columns side-by-side.
  - Claude derives a date-based attribution (PREEXISTING / OPERATIVE /
    POSTOP_LATE / NOT_OPERATIVE / UNKNOWN) so Logan can verify per-patient.
  - Phenotype source_tier_label (laryngoscopy_confirmed / chart_documented /
    nlp_refined / nlp_single_mention) is shown as evidence — closes the
    "?no evidence?" gap where the structured signal flagged a pt but
    clinical_notes_long didn't carry the source text.

Scope: the 95 candidates from the original mig_98c structured-signal selection
(phen_confirmed OR v2 extractor OR mig_98a VCP overlap). Logan's 58 partial
decisions from the prior workbook are pre-populated under
`logan_prior_decision` / `logan_prior_note` (raw text) +
`logan_prior_mapped` (mapped to vocabulary).

Output: verification_csvs/canonical_complications_events_v1/
        voice_nerve_unified__mig_98c_v2.xlsx
  - summary sheet
  - VOICE_NERVE_REVIEW sheet: 1 row per pt, RLN + VCP grouped, with
    decision col

Decision vocabulary (your_decision):
  OPERATIVE             — RLN/VCP attributable to index surgery
  CANCER_RELATED        — operative subset; nerve sacrificed for cancer
  INTENTIONAL_SACRIFICE — operative subset; planned nerve sacrifice
  POSTOP_LATE           — operative but late detection (31-365d post-op)
  PRIOR_OP              — from a prior (non-index) surgery
  PREEXISTING           — pre-existing or non-surgical finding
  NOT_OPERATIVE         — not from index surgery (idiopathic / other)
  NEEDS_CONTEXT         — needs further chart review / date check
  NO                    — false positive / template noise

Author: Logan Glosser <logan.glosser@gmail.com>
"""
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

PUB_DB = "thyroid_canonical_publication_v1_0"
ARCH_DB = '"Thyroid 2026 UPdated".archive_pub_v1_0'
PHEN_PRE = f"{ARCH_DB}.complication_phenotype_v1_pre364_20260422_050902"

OUT_DIR = REPO_ROOT / "verification_csvs" / "canonical_complications_events_v1"
OUT_XLSX = OUT_DIR / "voice_nerve_unified__mig_98c_v2.xlsx"
PARTIAL_JSON = OUT_DIR / "mig_98c_logan_decisions_partial.json"

NOTE_TYPE_PRIORITY = [
    "DC_SUM", "HP", "OPNOTE", "ENDOCRINE_FM", "ED_NOTE",
    "OTHER_HISTORY", "OTHER_NOTES",
]

_ILLEGAL_CTRL_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
RLN_VCP_RE = re.compile(
    r"\b(rln|recurrent\s+laryngeal|vocal\s+cord\s+paralys|vocal\s+fold\s+paralys|"
    r"\bvcp\b|hoarse|dysphon)",
    re.IGNORECASE,
)

HEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
RLN_FILL = PatternFill(start_color="E8F4DD", end_color="E8F4DD", fill_type="solid")
VCP_FILL = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
DEC_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _connect():
    from motherduck_client import MotherDuckClient, MotherDuckConfig

    cfg = MotherDuckConfig(database=PUB_DB)
    con = MotherDuckClient(cfg).connect_rw()
    con.execute(f"USE {PUB_DB}")
    return con


def _squash(s: str | None, max_len: int = 500) -> str:
    s = (s or "").replace("\r", " ").replace("\n", " ").strip()
    s = _ILLEGAL_CTRL_RE.sub(" ", s)
    while "  " in s:
        s = s.replace("  ", " ")
    if len(s) > max_len:
        s = s[: max_len - 3] + "..."
    return s


def _load_partial_decisions() -> dict[str, dict]:
    if not PARTIAL_JSON.exists():
        return {}
    return json.loads(PARTIAL_JSON.read_text())


def _classify_attribution(timing_days, timing_window):
    """Date-based attribution rule. timing_days is from phenotype timing_days_post_surgery."""
    if timing_window == "pre_surgery":
        return "PREEXISTING", f"phenotype timing_window=pre_surgery"
    if timing_days is None:
        return "UNKNOWN", "no timing_days_post_surgery"
    td = int(timing_days)
    if td < 0:
        return "PREEXISTING", f"finding precedes first surgery by {-td} days"
    if td <= 30:
        return "OPERATIVE", f"timing_days={td} (intra-op or early postop window)"
    if td <= 180:
        return "POSTOP_LATE", f"timing_days={td} (still likely operative-attributable)"
    if td <= 365:
        return "POSTOP_VERY_LATE", f"timing_days={td} (operative attribution unclear)"
    if td <= 365 * 5:
        return "POSSIBLY_PRIOR_OP", f"timing_days={td} ({td // 365}y; likely prior op or unrelated)"
    return "NOT_OPERATIVE", f"timing_days={td} ({td // 365}y; almost certainly not from index surgery)"


def fetch(con):
    """Pull RLN + VCP side-by-side for the 95 mig_98c candidates."""
    # Get the 95 candidate rids: structured signals (phen_confirmed OR v2 OR vcp_overlap)
    rids = [r[0] for r in con.execute(
        """
        WITH rln_pts AS (
          SELECT DISTINCT research_id FROM main.canonical_complications_events_v1
          WHERE complication_type='rln_injury' AND finding_status IN ('present','suspected')
        ),
        v2_pts AS (
          SELECT DISTINCT research_id FROM main.canonical_complications_events_v1
          WHERE complication_type='rln_injury' AND source_table='extracted_rln_injury_refined_v2'
        ),
        phen_conf_pts AS (
          SELECT DISTINCT c.research_id
          FROM main.canonical_complications_events_v1 c
          LEFT JOIN """ + PHEN_PRE + """ phen
            ON c.source_table='complication_phenotype_v1'
           AND c.source_row_id=CAST(hash(phen.research_id, phen.complication_entity, phen.detection_date) AS VARCHAR)
          WHERE c.complication_type='rln_injury'
            AND phen.confirmed_flag = TRUE
        ),
        vcp_keepers AS (
          SELECT DISTINCT research_id FROM main.canonical_complications_events_v1
          WHERE complication_type='vocal_cord_paralysis' AND finding_status='present'
        )
        SELECT research_id FROM (
          SELECT research_id FROM v2_pts
          UNION SELECT research_id FROM phen_conf_pts
          UNION SELECT research_id FROM rln_pts r WHERE EXISTS (SELECT 1 FROM vcp_keepers v WHERE v.research_id = r.research_id)
        )
        ORDER BY research_id
        """
    ).fetchall()]

    # First-surgery date per pt from canonical_operative_events_v1
    rid_list = "','".join(rids)
    fs_q = (
        f"SELECT research_id, MIN(surgery_date_native)::VARCHAR AS first_surgery_date "
        f"FROM main.canonical_operative_events_v1 "
        f"WHERE research_id IN ('{rid_list}') "
        f"GROUP BY research_id"
    )
    first_surgery = {r[0]: r[1] for r in con.execute(fs_q).fetchall()}

    # RLN canonical rows (per-pt aggregates)
    rln_q = (
        f"SELECT c.research_id, "
        f"STRING_AGG(DISTINCT c.finding_status, ',' ORDER BY c.finding_status) AS statuses, "
        f"STRING_AGG(DISTINCT c.onset_class, ',' ORDER BY c.onset_class) AS onsets, "
        f"MIN(c.finding_date)::VARCHAR AS first_dt, "
        f"BOOL_OR(c.evidence_strength='definitive') AS any_definitive, "
        f"BOOL_OR(c.source_table='extracted_rln_injury_refined_v2') AS has_v2 "
        f"FROM main.canonical_complications_events_v1 c "
        f"WHERE c.complication_type='rln_injury' "
        f"AND c.research_id IN ('{rid_list}') "
        f"GROUP BY c.research_id"
    )
    rln_canon = {r[0]: r for r in con.execute(rln_q).fetchall()}

    # VCP canonical rows
    vcp_q = (
        f"SELECT c.research_id, "
        f"STRING_AGG(DISTINCT c.finding_status, ',' ORDER BY c.finding_status) AS statuses, "
        f"STRING_AGG(DISTINCT c.onset_class, ',' ORDER BY c.onset_class) AS onsets, "
        f"MIN(c.finding_date)::VARCHAR AS first_dt, "
        f"BOOL_OR(c.evidence_strength='definitive') AS any_definitive "
        f"FROM main.canonical_complications_events_v1 c "
        f"WHERE c.complication_type='vocal_cord_paralysis' "
        f"AND c.finding_status IN ('present','suspected') "
        f"AND c.research_id IN ('{rid_list}') "
        f"GROUP BY c.research_id"
    )
    vcp_canon = {r[0]: r for r in con.execute(vcp_q).fetchall()}

    # Phenotype data per pt × entity (rln_injury, vocal_cord_paralysis, vocal_cord_paresis)
    phen_q = (
        f"SELECT phen.research_id::VARCHAR, phen.complication_entity, phen.confirmed_flag, "
        f"phen.suspected_flag, phen.transient_flag, phen.permanent_flag, "
        f"phen.treatment_requiring_flag, phen.timing_window, phen.timing_days_post_surgery, "
        f"phen.evidence_tier, phen.source_tier_label, "
        f"phen.detection_date::VARCHAR AS det_dt, phen.first_surgery_date::VARCHAR AS first_sx "
        f"FROM {PHEN_PRE} phen "
        f"WHERE phen.research_id::VARCHAR IN ('{rid_list}') "
        f"AND (phen.complication_entity ILIKE '%rln%' OR phen.complication_entity ILIKE '%vocal%')"
    )
    phen_rows = con.execute(phen_q).fetchall()
    phen_by_pt = defaultdict(lambda: {"rln": None, "vcp": None})
    for r in phen_rows:
        rid, entity = r[0], r[1]
        if "rln" in entity.lower():
            phen_by_pt[rid]["rln"] = r
        elif "vocal_cord" in entity.lower() or "vocal cord" in entity.lower():
            # prefer paralysis over paresis if both exist
            existing = phen_by_pt[rid]["vcp"]
            if existing is None or (
                "paralysis" in entity.lower() and "paralysis" not in existing[1].lower()
            ):
                phen_by_pt[rid]["vcp"] = r

    # Note-text excerpts from clinical_notes_long
    con.execute("CREATE OR REPLACE TEMP TABLE tmp_pts (rid VARCHAR)")
    con.executemany("INSERT INTO tmp_pts VALUES (?)", [(r,) for r in rids])
    notes = con.execute(
        "SELECT cnl.research_id, cnl.note_type, cnl.note_index, cnl.note_text "
        "FROM main.clinical_notes_long cnl JOIN tmp_pts t ON cnl.research_id = t.rid "
        "WHERE regexp_matches(LOWER(cnl.note_text), "
        "'(\\brln\\b|recurrent laryngeal|vocal cord paralys|vocal fold paralys|\\bvcp\\b|hoarse|dysphon)')"
    ).fetchall()

    notes_by_pt = defaultdict(lambda: defaultdict(list))
    for rid, nt, idx, txt in notes:
        if not txt:
            continue
        for m in RLN_VCP_RE.finditer(txt):
            pos = m.start()
            ctx = _squash(txt[max(0, pos - 200): pos + 300], max_len=400)
            notes_by_pt[rid][nt].append(ctx)

    return rids, first_surgery, rln_canon, vcp_canon, phen_by_pt, notes_by_pt


HEADERS = [
    # Identity
    "research_id",
    "first_surgery_date",
    # Claude attribution
    "claude_attribution",
    "claude_attribution_basis",
    # RLN side
    "rln_status",
    "rln_first_finding_date",
    "rln_timing_days",
    "rln_evidence_tier",
    "rln_source_tier_label",
    "rln_phen_confirmed",
    "rln_phen_treat_req",
    "rln_phen_permanent",
    "rln_phen_transient",
    "rln_evidence_excerpt",
    # VCP side
    "vcp_entity",
    "vcp_status",
    "vcp_first_finding_date",
    "vcp_timing_days",
    "vcp_evidence_tier",
    "vcp_source_tier_label",
    "vcp_phen_confirmed",
    "vcp_phen_treat_req",
    "vcp_phen_permanent",
    "vcp_phen_transient",
    "vcp_evidence_excerpt",
    # Note-text
    "notetext_source",
    "notetext_excerpt",
    # Decision
    "logan_prior_decision",
    "logan_prior_note",
    "logan_prior_mapped",
    "your_decision",
    "your_note",
]


def _phen_evidence(phen_row):
    """Format phenotype row as an evidence string when no clinical_notes excerpt is available."""
    if not phen_row:
        return ""
    _, entity, conf, susp, trans, perm, treat, tw, td, ev_tier, src_tier, det_dt, first_sx = phen_row
    parts = []
    parts.append(f"phenotype:{entity}")
    if src_tier:
        parts.append(f"tier={src_tier}")
    if ev_tier:
        parts.append(f"ev={ev_tier}")
    if conf:
        parts.append("confirmed")
    if susp:
        parts.append("suspected")
    if perm:
        parts.append("permanent")
    if trans:
        parts.append("transient")
    if treat:
        parts.append("treat_req")
    if td is not None:
        parts.append(f"+{td}d")
    if tw:
        parts.append(f"window={tw}")
    if det_dt:
        parts.append(f"detected={det_dt}")
    return " | ".join(parts)


def _best_note_excerpt(notes_for_pt):
    """Pick the highest-priority note_type with clinical evidence."""
    for nt in NOTE_TYPE_PRIORITY:
        if nt in notes_for_pt and notes_for_pt[nt]:
            return nt, "; ".join(notes_for_pt[nt][:2])
    return "", ""


def build_summary(wb, counts):
    ws = wb.create_sheet("summary", 0)
    ws["A1"] = "voice_nerve unified review — mig_98c v2"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated (UTC): {datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')}"
    ws["A3"] = "Author: Logan Glosser <logan.glosser@gmail.com>"
    ws["A5"] = (
        "Per Logan's architectural guidance: rln_injury and vocal_cord_paralysis "
        "(RLN nerve damage vs functional paralysis outcome) reviewed jointly. "
        "One row per patient with RLN + VCP columns side-by-side. claude_attribution "
        "is date-based (timing_days_post_surgery from phenotype): PREEXISTING / "
        "OPERATIVE / POSTOP_LATE / POSTOP_VERY_LATE / POSSIBLY_PRIOR_OP / "
        "NOT_OPERATIVE / UNKNOWN. Phenotype source_tier_label "
        "(laryngoscopy_confirmed / chart_documented / nlp_refined / nlp_single_mention) "
        "shown as evidence to close the prior '?no evidence?' gap. Logan's 58 "
        "partial decisions from the previous workbook are pre-populated under "
        "logan_prior_* columns."
    )
    ws["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110
    ws.row_dimensions[5].height = 100

    ws["A7"] = "Counts"
    ws["A7"].font = Font(bold=True)
    rows = [
        f"Cohort: 10,871 pts",
        f"VOICE_NERVE_REVIEW candidates: {counts['total']}",
        f"  -- with RLN canonical row: {counts['has_rln']}",
        f"  -- with VCP canonical row: {counts['has_vcp']}",
        f"  -- with both:               {counts['has_both']}",
        f"  -- Logan prior decisions pre-populated: {counts['prior_decisions']}",
        "",
        "Decision vocabulary (your_decision):",
        "  OPERATIVE             — RLN/VCP attributable to index surgery",
        "  CANCER_RELATED        — operative; nerve sacrificed for cancer",
        "  INTENTIONAL_SACRIFICE — operative; planned nerve sacrifice",
        "  POSTOP_LATE           — operative but late detection (31-365d)",
        "  PRIOR_OP              — from a prior (non-index) surgery",
        "  PREEXISTING           — pre-existing or non-surgical finding",
        "  NOT_OPERATIVE         — not from index surgery (idiopathic / other)",
        "  NEEDS_CONTEXT         — needs further chart review / date check",
        "  NO                    — false positive / template noise",
    ]
    for i, r in enumerate(rows, start=8):
        ws.cell(row=i, column=1, value=r)


def build_review_sheet(wb, rids, first_surgery, rln_canon, vcp_canon, phen_by_pt, notes_by_pt, partial):
    ws = wb.create_sheet("VOICE_NERVE_REVIEW")
    for j, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL

    rln_cols = [h for h in HEADERS if h.startswith("rln_")]
    vcp_cols = [h for h in HEADERS if h.startswith("vcp_")]

    counts = {
        "total": len(rids),
        "has_rln": 0,
        "has_vcp": 0,
        "has_both": 0,
        "prior_decisions": 0,
    }

    for rid in rids:
        rln_row = rln_canon.get(rid)
        vcp_row = vcp_canon.get(rid)
        phen = phen_by_pt.get(rid, {"rln": None, "vcp": None})
        rln_phen = phen["rln"]
        vcp_phen = phen["vcp"]

        if rln_row:
            counts["has_rln"] += 1
        if vcp_row:
            counts["has_vcp"] += 1
        if rln_row and vcp_row:
            counts["has_both"] += 1

        # Pick the most-clinical timing_days for attribution: prefer rln_phen, fall back to vcp_phen
        timing_days = None
        timing_window = None
        if rln_phen and rln_phen[8] is not None:
            timing_days = rln_phen[8]
            timing_window = rln_phen[7]
        elif vcp_phen and vcp_phen[8] is not None:
            timing_days = vcp_phen[8]
            timing_window = vcp_phen[7]
        attr, basis = _classify_attribution(timing_days, timing_window)

        notes_pt = notes_by_pt.get(rid, {})
        notetext_src, notetext_excerpt = _best_note_excerpt(notes_pt)

        rln_evidence = _phen_evidence(rln_phen) if rln_phen else (
            "rln_injury canonical row only" if rln_row else ""
        )
        vcp_entity = vcp_phen[1] if vcp_phen else ("vocal_cord_paralysis" if vcp_row else "")
        vcp_evidence = _phen_evidence(vcp_phen) if vcp_phen else (
            "vocal_cord_paralysis canonical row only" if vcp_row else ""
        )

        prior = partial.get(str(rid), {})
        if prior:
            counts["prior_decisions"] += 1

        ws.append([
            rid,
            first_surgery.get(rid),
            attr,
            basis,
            rln_row[1] if rln_row else "",                  # rln_status
            rln_row[3] if rln_row else "",                  # rln_first_dt
            rln_phen[8] if rln_phen else None,              # timing_days
            rln_phen[9] if rln_phen else None,              # evidence_tier
            rln_phen[10] if rln_phen else "",               # source_tier_label
            rln_phen[2] if rln_phen else None,              # confirmed
            rln_phen[6] if rln_phen else None,              # treat_req
            rln_phen[5] if rln_phen else None,              # permanent
            rln_phen[4] if rln_phen else None,              # transient
            rln_evidence,
            vcp_entity,
            vcp_row[1] if vcp_row else "",
            vcp_row[3] if vcp_row else "",
            vcp_phen[8] if vcp_phen else None,
            vcp_phen[9] if vcp_phen else None,
            vcp_phen[10] if vcp_phen else "",
            vcp_phen[2] if vcp_phen else None,
            vcp_phen[6] if vcp_phen else None,
            vcp_phen[5] if vcp_phen else None,
            vcp_phen[4] if vcp_phen else None,
            vcp_evidence,
            notetext_src,
            notetext_excerpt,
            prior.get("raw_decision", ""),
            prior.get("raw_note", ""),
            prior.get("mapped_decision", ""),
            "",
            "",
        ])

    # Apply group-color fills for header band (RLN green, VCP orange)
    for col_name in rln_cols:
        c = ws.cell(row=1, column=HEADERS.index(col_name) + 1)
        c.fill = RLN_FILL
    for col_name in vcp_cols:
        c = ws.cell(row=1, column=HEADERS.index(col_name) + 1)
        c.fill = VCP_FILL

    # Yellow-highlight decision cols
    dec_cols = {
        get_column_letter(HEADERS.index("your_decision") + 1),
        get_column_letter(HEADERS.index("your_note") + 1),
    }
    wrap_cols = {
        get_column_letter(HEADERS.index(h) + 1)
        for h in (
            "claude_attribution_basis", "rln_evidence_excerpt",
            "vcp_evidence_excerpt", "notetext_excerpt",
            "logan_prior_note", "your_note",
        )
    }
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in dec_cols:
                cell.fill = DEC_FILL
    ws.freeze_panes = "C2"
    # autosize
    for col in ws.columns:
        letter = col[0].column_letter
        mlen = 0
        for cell in col:
            if cell.value is None:
                continue
            mlen = max(mlen, min(len(str(cell.value)), 100))
        ws.column_dimensions[letter].width = min(max(10, mlen + 2), 60.0)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    return counts


def main():
    con = _connect()
    rids, first_surgery, rln_canon, vcp_canon, phen_by_pt, notes_by_pt = fetch(con)
    con.close()

    partial = _load_partial_decisions()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    summary_placeholder = {"total": len(rids), "has_rln": 0, "has_vcp": 0, "has_both": 0, "prior_decisions": 0}
    build_summary(wb, summary_placeholder)
    counts = build_review_sheet(wb, rids, first_surgery, rln_canon, vcp_canon, phen_by_pt, notes_by_pt, partial)
    # Re-write summary with real counts
    wb.remove(wb["summary"])
    build_summary(wb, counts)
    wb.move_sheet("summary", offset=-1)  # move to first

    wb.properties.creator = "Logan Glosser <logan.glosser@gmail.com>"
    wb.save(OUT_XLSX)
    print(f"wrote: {OUT_XLSX}")
    for k, v in counts.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
