#!/usr/bin/env python3
"""mig_98b — Build Logan review workbook for chyle_leak in canonical_complications_events_v1.

Reads MotherDuck `thyroid_canonical_publication_v1_0` + pre-364 archives in
`Thyroid 2026 UPdated`.archive_pub_v1_0 for evidence_span and phenotype fields.

Does NOT mutate canonical tables or registry.

Key design difference from mig_98a (vocal_cord_paralysis):
    chyle_leak NEC `evidence_span` is just the 10-char entity match string
    ("chyle leak"), with NO surrounding context. Per-row evidence review is
    therefore useless; review must be PER-PATIENT on metadata fields
    (source_modality, onset_class, source overlap, phenotype flags).

Bucketing (per-patient on present rows):
    A_OPNOTE_ONLY              957 pts / 1,223 op_note rows
        op_note entity hits with NO clinic/discharge/phenotype/refined corroboration.
        ALL 957 already carry an absent rollup row from phenotype source, so DELETE
        is safe. Default disposition: DELETE present rows.
    B_OPNOTE_WITH_POSTOP        589 pts / 1,771 rows
        op_note + at least one of (clinic / discharge / phenotype / refined).
        Likely real clinical chyle_leak. Default disposition: ACCEPT.
        Logan can FLIP_TO_ABSENT individuals on closer read of timing/sources.
    C_NO_OPNOTE_BACKED_BY_POSTOP  30 pts / 34 rows
        Post-op clinical mentions only (no op_note). Default disposition: ACCEPT.

Plus per-row review of:
    D_PHENOTYPE_ROWS            20 rows
        Includes treatment_requiring_flag, suspected_flag, confirmed_flag,
        timing_window/days. 1 with treatment_required=TRUE (high-confidence ACCEPT);
        16 with timing 0_30d / 0d (intra-op echo, candidate for FLIP); 3 with >365d
        (stale FP candidate).
    E_REFINED_EXTRACTION        20 rows
        Refined NLP extraction; default ACCEPT, surface for spot-check.

Author: Logan Glosser <logan.glosser@gmail.com>
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))
OUT_DIR = REPO_ROOT / "verification_csvs" / "canonical_complications_events_v1"
OUT_XLSX = OUT_DIR / "chyle_leak_review__mig_98b.xlsx"
DECISIONS_JSON = OUT_DIR / "mig_98_decisions.json"

PUB_DB = "thyroid_canonical_publication_v1_0"
ARCH_DB = '"Thyroid 2026 UPdated".archive_pub_v1_0'
PHEN_PRE = f"{ARCH_DB}.complication_phenotype_v1_pre364_20260422_050902"
NEC_PRE = f"{ARCH_DB}.note_entities_complications_pre364_20260422_050902"

# Cohort denominator from canonical_patient_master at time of build
COHORT_DENOM = 10871

# Bucket bulk-action defaults (Logan ratifies/modifies on summary sheet)
DEFAULT_ACTIONS = {
    "A_OPNOTE_ONLY": "DELETE",
    "B_OPNOTE_WITH_POSTOP": "ACCEPT",
    "C_NO_OPNOTE_BACKED_BY_POSTOP": "ACCEPT",
}


def _connect_md() -> duckdb.DuckDBPyConnection:
    from motherduck_client import MotherDuckClient, MotherDuckConfig

    cfg = MotherDuckConfig(database=PUB_DB)
    con = MotherDuckClient(cfg).connect_rw()
    con.execute(f"USE {PUB_DB}")
    return con


def _fetch_per_patient(con: duckdb.DuckDBPyConnection) -> list[dict[str, Any]]:
    q = f"""
    WITH ev AS (
        SELECT
            research_id,
            source_table,
            source_modality,
            source_kind,
            onset_class,
            permanence_class,
            evidence_strength,
            finding_date,
            confidence
        FROM main.canonical_complications_events_v1
        WHERE complication_type = 'chyle_leak'
          AND finding_status = 'present'
    )
    SELECT
        research_id,
        SUM(CASE WHEN source_modality = 'op_note'
              AND source_table = 'note_entities_complications' THEN 1 ELSE 0 END) AS n_op,
        SUM(CASE WHEN source_modality = 'clinic_note'
              AND source_table = 'note_entities_complications' THEN 1 ELSE 0 END) AS n_clinic,
        SUM(CASE WHEN source_modality = 'discharge_summary'
              AND source_table = 'note_entities_complications' THEN 1 ELSE 0 END) AS n_disch_entity,
        SUM(CASE WHEN source_table = 'complication_phenotype_v1' THEN 1 ELSE 0 END) AS n_phen,
        SUM(CASE WHEN source_table = 'extracted_complications_refined_v5' THEN 1 ELSE 0 END) AS n_refined,
        MIN(finding_date) AS min_finding_date,
        MAX(finding_date) AS max_finding_date,
        MIN(CASE WHEN source_modality = 'op_note' THEN finding_date END) AS min_op_date,
        MIN(CASE WHEN source_modality IN ('clinic_note','discharge_summary')
              THEN finding_date END) AS min_postop_date,
        MAX(CASE WHEN source_modality IN ('clinic_note','discharge_summary')
              THEN finding_date END) AS max_postop_date,
        STRING_AGG(DISTINCT onset_class, ',' ORDER BY onset_class) AS onset_classes,
        AVG(confidence) AS avg_confidence
    FROM ev
    GROUP BY research_id
    """
    res = con.execute(q)
    cols = [d[0] for d in res.description]
    rows: list[dict[str, Any]] = []
    for tup in res.fetchall():
        rec = dict(zip(cols, tup))
        n_op = int(rec.get("n_op") or 0)
        n_clinic = int(rec.get("n_clinic") or 0)
        n_disch = int(rec.get("n_disch_entity") or 0)
        n_phen = int(rec.get("n_phen") or 0)
        n_ref = int(rec.get("n_refined") or 0)
        if n_op > 0 and (n_clinic + n_disch + n_phen + n_ref) == 0:
            bucket = "A_OPNOTE_ONLY"
        elif n_op > 0:
            bucket = "B_OPNOTE_WITH_POSTOP"
        else:
            bucket = "C_NO_OPNOTE_BACKED_BY_POSTOP"
        rec["bucket"] = bucket
        rec["bulk_default"] = DEFAULT_ACTIONS[bucket]
        rec["days_op_to_postop"] = None
        if rec.get("min_op_date") and rec.get("min_postop_date"):
            try:
                rec["days_op_to_postop"] = (
                    rec["min_postop_date"] - rec["min_op_date"]
                ).days
            except Exception:
                pass
        rows.append(rec)
    return rows


def _fetch_phenotype_rows(con: duckdb.DuckDBPyConnection) -> list[dict[str, Any]]:
    q = f"""
    SELECT
        c.research_id,
        c.finding_date,
        c.source_modality,
        c.onset_class,
        c.evidence_strength,
        c.confidence,
        phen.suspected_flag,
        phen.confirmed_flag,
        phen.transient_flag,
        phen.permanent_flag,
        phen.surgery_related_flag,
        phen.historical_only_flag,
        phen.timing_window,
        phen.timing_days_post_surgery,
        phen.treatment_requiring_flag,
        phen.final_complication_status,
        phen.evidence_tier,
        phen.source_tier_label,
        phen.n_raw_nlp_mentions,
        phen.n_valid_nlp_mentions,
        phen.detection_date,
        phen.first_surgery_date
    FROM main.canonical_complications_events_v1 c
    LEFT JOIN {PHEN_PRE} phen
      ON CAST(phen.research_id AS VARCHAR) = c.research_id
     AND c.source_table = 'complication_phenotype_v1'
     AND c.source_row_id = CAST(hash(phen.research_id, phen.complication_entity,
                                      phen.detection_date) AS VARCHAR)
    WHERE c.complication_type = 'chyle_leak'
      AND c.finding_status = 'present'
      AND c.source_table = 'complication_phenotype_v1'
    ORDER BY c.research_id, c.finding_date
    """
    res = con.execute(q)
    cols = [d[0] for d in res.description]
    rows: list[dict[str, Any]] = []
    for tup in res.fetchall():
        rec = dict(zip(cols, tup))
        # Suggest action based on phenotype flags
        treat = bool(rec.get("treatment_requiring_flag"))
        confirmed = bool(rec.get("confirmed_flag"))
        timing_days = rec.get("timing_days_post_surgery")
        if treat:
            rec["suggested_action"] = "ACCEPT — treatment_requiring=TRUE"
        elif timing_days is not None and timing_days > 365:
            rec["suggested_action"] = (
                f"FLIP_TO_ABSENT — stale phenotype hit (timing_days={timing_days})"
            )
        elif timing_days is not None and timing_days <= 1:
            rec["suggested_action"] = (
                "FLIP_TO_ABSENT — intra-op echo (timing_days=0/1; not treatment-req)"
            )
        elif confirmed and not treat:
            rec["suggested_action"] = (
                "REVIEW — confirmed phenotype but no treatment_requiring evidence"
            )
        else:
            rec["suggested_action"] = "REVIEW — no strong phenotype flags"
        rows.append(rec)
    return rows


def _fetch_refined_rows(con: duckdb.DuckDBPyConnection) -> list[dict[str, Any]]:
    q = """
    SELECT
        c.research_id,
        c.finding_date,
        c.source_modality,
        c.onset_class,
        c.evidence_strength,
        c.permanence_class,
        c.confidence,
        c.evidence_span_hash,
        c.detection_date_inferred
    FROM main.canonical_complications_events_v1 c
    WHERE c.complication_type = 'chyle_leak'
      AND c.finding_status = 'present'
      AND c.source_table = 'extracted_complications_refined_v5'
    ORDER BY c.research_id, c.finding_date
    """
    res = con.execute(q)
    cols = [d[0] for d in res.description]
    rows: list[dict[str, Any]] = []
    for tup in res.fetchall():
        rec = dict(zip(cols, tup))
        rec["suggested_action"] = "ACCEPT — refined NLP extraction; spot-check"
        rows.append(rec)
    return rows


def _autosize(ws, max_width: float = 50.0, wrap_cols: set[str] | None = None) -> None:
    wrap_cols = wrap_cols or set()
    for col in ws.columns:
        letter = col[0].column_letter
        mlen = 0
        for cell in col:
            if cell.value is None:
                continue
            mlen = max(mlen, min(len(str(cell.value)), 80))
        ws.column_dimensions[letter].width = min(max(10, mlen + 2), max_width)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


HEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
DECISION_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _write_summary(
    wb: Workbook,
    bucket_counts: dict[str, int],
    bucket_row_counts: dict[str, int],
    phen_n: int,
    refined_n: int,
    run_ts: str,
) -> None:
    ws = wb.create_sheet("summary", 0)
    ws["A1"] = "chyle_leak review — mig_98b"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated (UTC): {run_ts}"
    ws["A3"] = "Author: Logan Glosser <logan.glosser@gmail.com>"
    ws["A5"] = (
        f"Cohort: {COHORT_DENOM} pts. Pre-mig prevalence (present/cohort): "
        f"{sum(bucket_counts.values())}/{COHORT_DENOM} = "
        f"{100.0*sum(bucket_counts.values())/COHORT_DENOM:.2f}% "
        f"(literature ~1-3%)."
    )
    ws["A5"].font = Font(italic=True)

    ws["A7"] = "Buckets (per-patient on present rows)"
    ws["A7"].font = Font(bold=True)

    headers = [
        "bucket",
        "n_pts",
        "n_rows_affected",
        "default_action",
        "rule_summary",
        "your_bulk_action",
        "your_note",
    ]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=8, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL

    rules = {
        "A_OPNOTE_ONLY": (
            "op_note entity hits ONLY (no clinic/discharge/phenotype/refined). "
            "evidence_span is 10-char literal; cannot distinguish intra-op observation "
            "from documentation-of-concern. ALL 957 already have absent rollup row "
            "from phenotype source. DELETE is safe."
        ),
        "B_OPNOTE_WITH_POSTOP": (
            "op_note + at least one post-op source (clinic / discharge / phenotype / "
            "refined). Likely real clinical chyle_leak. Default ACCEPT; Logan can "
            "FLIP_TO_ABSENT specific patients on closer read."
        ),
        "C_NO_OPNOTE_BACKED_BY_POSTOP": (
            "Post-op clinical mentions only (no op_note hit). 30 pts. Default ACCEPT."
        ),
    }

    bucket_order = ["A_OPNOTE_ONLY", "B_OPNOTE_WITH_POSTOP", "C_NO_OPNOTE_BACKED_BY_POSTOP"]
    for i, b in enumerate(bucket_order, start=9):
        ws.cell(row=i, column=1, value=b)
        ws.cell(row=i, column=2, value=bucket_counts.get(b, 0))
        ws.cell(row=i, column=3, value=bucket_row_counts.get(b, 0))
        ws.cell(row=i, column=4, value=DEFAULT_ACTIONS.get(b, ""))
        ws.cell(row=i, column=5, value=rules.get(b, ""))
        c_action = ws.cell(row=i, column=6, value="")
        c_action.fill = DECISION_FILL
        c_note = ws.cell(row=i, column=7, value="")
        c_note.fill = DECISION_FILL

    r0 = 13
    ws.cell(row=r0, column=1, value="Per-row sub-buckets (review individually)").font = Font(
        bold=True
    )
    sub_headers = ["sub_bucket", "n_rows", "default_action", "rule_summary"]
    for j, h in enumerate(sub_headers, start=1):
        c = ws.cell(row=r0 + 1, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
    ws.cell(row=r0 + 2, column=1, value="D_PHENOTYPE_ROWS")
    ws.cell(row=r0 + 2, column=2, value=phen_n)
    ws.cell(row=r0 + 2, column=3, value="ROW-BY-ROW")
    ws.cell(
        row=r0 + 2,
        column=4,
        value=(
            "phenotype source rows; surface treatment_requiring_flag, suspected_flag, "
            "timing_window per row. 1 treat=TRUE → ACCEPT; 16 timing=0d → FLIP; "
            "3 timing>365d → FLIP."
        ),
    )
    ws.cell(row=r0 + 3, column=1, value="E_REFINED_EXTRACTION")
    ws.cell(row=r0 + 3, column=2, value=refined_n)
    ws.cell(row=r0 + 3, column=3, value="ACCEPT (default)")
    ws.cell(
        row=r0 + 3,
        column=4,
        value="refined NLP extraction; default ACCEPT, surface for spot-check.",
    )

    r0 += 5
    ws.cell(row=r0, column=1, value="Decision vocabulary").font = Font(bold=True)
    vocab_lines = [
        "ACCEPT — keep row(s) as-is",
        "DELETE — remove row(s) from canonical (patient retains existing absent rollup row)",
        "FLIP_TO_ABSENT — set finding_status='absent' (preserves provenance)",
        "REVIEW — case-by-case (phenotype/refined sheets)",
        "NEEDS_CONTEXT — defer; need note re-pull",
    ]
    for ln in vocab_lines:
        r0 += 1
        ws.cell(row=r0, column=1, value=ln)

    r0 += 2
    ws.cell(row=r0, column=1, value="Sign-off (Logan)").font = Font(bold=True)
    ws.cell(row=r0 + 1, column=1, value="  Date: _______________")
    ws.cell(row=r0 + 2, column=1, value="  Initials: _______________")

    _autosize(ws, max_width=70.0, wrap_cols={"E", "G"})


PT_HEADERS = [
    "research_id",
    "bucket",
    "bulk_default",
    "n_op_rows",
    "n_clinic_rows",
    "n_disch_entity_rows",
    "n_phen_rows",
    "n_refined_rows",
    "min_finding_date",
    "max_finding_date",
    "min_op_date",
    "min_postop_date",
    "max_postop_date",
    "days_op_to_postop",
    "onset_classes",
    "your_decision",
    "your_note",
]


def _write_pt_sheet(
    wb: Workbook,
    title: str,
    rows: list[dict[str, Any]],
) -> None:
    ws = wb.create_sheet(title)
    for j, h in enumerate(PT_HEADERS, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
    for r in rows:
        ws.append(
            [
                str(r.get("research_id", "")),
                r.get("bucket", ""),
                r.get("bulk_default", ""),
                int(r.get("n_op") or 0),
                int(r.get("n_clinic") or 0),
                int(r.get("n_disch_entity") or 0),
                int(r.get("n_phen") or 0),
                int(r.get("n_refined") or 0),
                r.get("min_finding_date"),
                r.get("max_finding_date"),
                r.get("min_op_date"),
                r.get("min_postop_date"),
                r.get("max_postop_date"),
                r.get("days_op_to_postop"),
                r.get("onset_classes", ""),
                "",
                "",
            ]
        )
    # Highlight decision/note cols
    dec_col = get_column_letter(PT_HEADERS.index("your_decision") + 1)
    note_col = get_column_letter(PT_HEADERS.index("your_note") + 1)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in (dec_col, note_col):
                cell.fill = DECISION_FILL
    ws.freeze_panes = "A2"
    _autosize(ws, max_width=22.0, wrap_cols={note_col})


PHEN_HEADERS = [
    "research_id",
    "finding_date",
    "source_modality",
    "onset_class",
    "evidence_strength",
    "confidence",
    "treatment_requiring_flag",
    "suspected_flag",
    "confirmed_flag",
    "transient_flag",
    "permanent_flag",
    "surgery_related_flag",
    "timing_window",
    "timing_days_post_surgery",
    "n_raw_nlp_mentions",
    "n_valid_nlp_mentions",
    "evidence_tier",
    "source_tier_label",
    "final_complication_status",
    "detection_date",
    "first_surgery_date",
    "suggested_action",
    "your_decision",
    "your_note",
]


def _write_phen_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("4_phenotype_rows")
    for j, h in enumerate(PHEN_HEADERS, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
    for r in rows:
        ws.append(
            [
                str(r.get("research_id", "")),
                r.get("finding_date"),
                r.get("source_modality", ""),
                r.get("onset_class", ""),
                r.get("evidence_strength", ""),
                r.get("confidence"),
                r.get("treatment_requiring_flag"),
                r.get("suspected_flag"),
                r.get("confirmed_flag"),
                r.get("transient_flag"),
                r.get("permanent_flag"),
                r.get("surgery_related_flag"),
                r.get("timing_window", ""),
                r.get("timing_days_post_surgery"),
                r.get("n_raw_nlp_mentions"),
                r.get("n_valid_nlp_mentions"),
                r.get("evidence_tier"),
                r.get("source_tier_label", ""),
                r.get("final_complication_status", ""),
                r.get("detection_date"),
                r.get("first_surgery_date"),
                r.get("suggested_action", ""),
                "",
                "",
            ]
        )
    dec_col = get_column_letter(PHEN_HEADERS.index("your_decision") + 1)
    note_col = get_column_letter(PHEN_HEADERS.index("your_note") + 1)
    sug_col = get_column_letter(PHEN_HEADERS.index("suggested_action") + 1)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in (dec_col, note_col):
                cell.fill = DECISION_FILL
    ws.freeze_panes = "A2"
    _autosize(ws, max_width=40.0, wrap_cols={sug_col, note_col})


REFINED_HEADERS = [
    "research_id",
    "finding_date",
    "source_modality",
    "onset_class",
    "permanence_class",
    "evidence_strength",
    "confidence",
    "detection_date_inferred",
    "evidence_span_hash",
    "suggested_action",
    "your_decision",
    "your_note",
]


def _write_refined_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("5_refined_extraction")
    for j, h in enumerate(REFINED_HEADERS, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL
    for r in rows:
        ws.append(
            [
                str(r.get("research_id", "")),
                r.get("finding_date"),
                r.get("source_modality", ""),
                r.get("onset_class", ""),
                r.get("permanence_class", ""),
                r.get("evidence_strength", ""),
                r.get("confidence"),
                r.get("detection_date_inferred"),
                r.get("evidence_span_hash", ""),
                r.get("suggested_action", ""),
                "",
                "",
            ]
        )
    dec_col = get_column_letter(REFINED_HEADERS.index("your_decision") + 1)
    note_col = get_column_letter(REFINED_HEADERS.index("your_note") + 1)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in (dec_col, note_col):
                cell.fill = DECISION_FILL
    ws.freeze_panes = "A2"
    _autosize(ws, max_width=40.0, wrap_cols={note_col})


def _update_decisions_stub(
    bucket_counts: dict[str, int],
    bucket_row_counts: dict[str, int],
    phen_n: int,
    refined_n: int,
    run_ts: str,
) -> None:
    payload: dict[str, Any] = {}
    if DECISIONS_JSON.exists():
        try:
            payload = json.loads(DECISIONS_JSON.read_text(encoding="utf-8"))
        except Exception:
            payload = {}
    payload.setdefault("schema_version", 1)
    payload["last_updated_utc"] = run_ts
    payload.setdefault(
        "note",
        "Per-row decisions appended after Logan returns filled workbook via Cowork.",
    )
    payload.setdefault("complications", {})
    payload["complications"]["chyle_leak"] = {
        "mig_98b_workbook": str(OUT_XLSX.name),
        "bucket_pt_counts_at_generation": bucket_counts,
        "bucket_row_counts_at_generation": bucket_row_counts,
        "phenotype_rows_at_generation": phen_n,
        "refined_rows_at_generation": refined_n,
        "default_actions": DEFAULT_ACTIONS,
        "decisions": [],
    }
    DECISIONS_JSON.parent.mkdir(parents=True, exist_ok=True)
    DECISIONS_JSON.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.parse_args()
    run_ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    con = _connect_md()
    pt_rows = _fetch_per_patient(con)
    phen_rows = _fetch_phenotype_rows(con)
    ref_rows = _fetch_refined_rows(con)
    con.close()

    by_bucket: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in pt_rows:
        by_bucket[r["bucket"]].append(r)
    bucket_counts = {b: len(by_bucket[b]) for b in by_bucket}
    bucket_row_counts: dict[str, int] = defaultdict(int)
    for r in pt_rows:
        bucket_row_counts[r["bucket"]] += (
            int(r.get("n_op") or 0)
            + int(r.get("n_clinic") or 0)
            + int(r.get("n_disch_entity") or 0)
            + int(r.get("n_phen") or 0)
            + int(r.get("n_refined") or 0)
        )

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    _write_summary(
        wb,
        bucket_counts,
        dict(bucket_row_counts),
        len(phen_rows),
        len(ref_rows),
        run_ts,
    )
    _write_pt_sheet(wb, "1_opnote_only", by_bucket.get("A_OPNOTE_ONLY", []))
    _write_pt_sheet(wb, "2_opnote_with_postop", by_bucket.get("B_OPNOTE_WITH_POSTOP", []))
    _write_pt_sheet(
        wb, "3_no_opnote_backed_by_postop", by_bucket.get("C_NO_OPNOTE_BACKED_BY_POSTOP", [])
    )
    _write_phen_sheet(wb, phen_rows)
    _write_refined_sheet(wb, ref_rows)

    wb.properties.creator = "Logan Glosser <logan.glosser@gmail.com>"
    wb.save(OUT_XLSX)
    _update_decisions_stub(
        bucket_counts, dict(bucket_row_counts), len(phen_rows), len(ref_rows), run_ts
    )

    print("mig_98b chyle_leak review workbook")
    print(f"  wrote: {OUT_XLSX}")
    print(f"  decisions stub: {DECISIONS_JSON}")
    print(f"  cohort denom: {COHORT_DENOM}")
    print(f"  total present pts: {sum(bucket_counts.values())}")
    print(f"  pre-mig prevalence: {100.0*sum(bucket_counts.values())/COHORT_DENOM:.2f}%")
    print("  per-patient buckets:")
    for b in ("A_OPNOTE_ONLY", "B_OPNOTE_WITH_POSTOP", "C_NO_OPNOTE_BACKED_BY_POSTOP"):
        print(
            f"    {b}: {bucket_counts.get(b,0)} pts, "
            f"{bucket_row_counts.get(b,0)} rows  "
            f"(default: {DEFAULT_ACTIONS.get(b)})"
        )
    print(f"  D_PHENOTYPE_ROWS: {len(phen_rows)} rows (per-row review)")
    print(f"  E_REFINED_EXTRACTION: {len(ref_rows)} rows (per-row review)")


if __name__ == "__main__":
    main()
