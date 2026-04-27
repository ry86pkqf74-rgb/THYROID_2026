"""
qc_framework_v1/scripts/extract_fna_source_long.py
==================================================

Source extractor for the FNA pilot table (Protocol v2, Step B).

Reads:
    raw/FNAs 12_5_2025.xlsx, sheet 'FNA Bethesda'
    Layout: 11,701 rows x 63 cols. Cols 1-3 are Research_ID# + 2 misc;
    cols 4-63 are 12 FNA episode blocks of 5 fields each (Date,
    Specimen received, Path extended, History, Bethesda) keyed by
    column ORDINAL, not header text. Header text varies (whitespace,
    embedded newlines, capitalization, trailing-numbered styles).

Writes:
    verification_csvs/canonical_fna_events_v1/_source_long.parquet

Schema (one row per (research_id, fna_index) where any of the 5
episode fields is non-empty):

    research_id            VARCHAR
    fna_index              INTEGER  (1..12)
    source_workbook        VARCHAR
    source_sheet           VARCHAR
    source_row             INTEGER  (1-based Excel row, header is row 1, data starts row 2)
    source_col_date        INTEGER  (1-based Excel column ordinal for the Date field)
    source_col_specimen    INTEGER
    source_col_path        INTEGER
    source_col_history     INTEGER
    source_col_bethesda    INTEGER
    source_col_name_date   VARCHAR  (the literal header text, normalized)
    date_raw               VARCHAR  (string form of the Excel cell, untrimmed)
    specimen_raw           VARCHAR
    path_raw               VARCHAR
    history_raw            VARCHAR
    bethesda_raw           VARCHAR

This intermediate is reused by every column-level compare for the
FNA pilot. Each column verification CSV joins canonical_fna_events_v1
to this parquet on (research_id, fna_index) and emits the per-column
DB-vs-source compare rows.

Run:
    python3 qc_framework_v1/scripts/extract_fna_source_long.py
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parents[2]
SOURCE_XLSX = REPO_ROOT / "raw" / "FNAs 12_5_2025.xlsx"
SOURCE_SHEET = "FNA Bethesda"
OUT_DIR = REPO_ROOT / "verification_csvs" / "canonical_fna_events_v1"
OUT_PARQUET = OUT_DIR / "_source_long.parquet"

# Layout constants (1-based, matching openpyxl)
ID_COL = 1
N_PREFIX_COLS = 3   # Research_ID# + 2 misc cols (Misc_path_non-Thyroid, other thyroid+parathyroid)
N_FIELDS_PER_FNA = 5  # Date, Specimen, Path, History, Bethesda
N_MAX_FNA = 12

# Field offset within each 5-cell FNA block (1-based offset)
FIELD_OFFSET = {
    "date": 1,
    "specimen": 2,
    "path": 3,
    "history": 4,
    "bethesda": 5,
}


def col_for_field(fna_index: int, field: str) -> int:
    """1-based Excel column ordinal for (fna_index, field)."""
    return N_PREFIX_COLS + (fna_index - 1) * N_FIELDS_PER_FNA + FIELD_OFFSET[field]


def cell_to_string(value) -> str | None:
    """
    Stringify an Excel cell value the same way the DB build script
    would have. Preserves whitespace and newlines so we can trace
    where odd `\n` / `\t` artifacts originated. Returns None for
    truly empty cells.
    """
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        # Excel datetime: emit ISO with HH:MM:SS so it matches the
        # 2,535 DB rows already in YYYY-MM-DD HH:MM:SS format.
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d 00:00:00")
    if isinstance(value, dt.time):
        return value.strftime("%H:%M:%S")
    s = str(value)
    return s


def main() -> None:
    print(f"[extract_fna_source_long] reading {SOURCE_XLSX}")
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(SOURCE_XLSX)

    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True, read_only=True)
    if SOURCE_SHEET not in wb.sheetnames:
        raise KeyError(
            f"sheet {SOURCE_SHEET!r} not found in {SOURCE_XLSX}; "
            f"sheets: {wb.sheetnames}"
        )
    ws = wb[SOURCE_SHEET]

    # Snapshot header row to record source_col_name_date per FNA index
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if len(header_row) < N_PREFIX_COLS + N_MAX_FNA * N_FIELDS_PER_FNA:
        raise AssertionError(
            f"unexpected header width {len(header_row)} "
            f"(want >= {N_PREFIX_COLS + N_MAX_FNA * N_FIELDS_PER_FNA})"
        )
    date_header_per_fna: dict[int, str] = {}
    for n in range(1, N_MAX_FNA + 1):
        idx0 = col_for_field(n, "date") - 1
        date_header_per_fna[n] = (header_row[idx0] or "").replace("\n", " | ").strip()

    rows = []
    for excel_row_idx, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True), start=2
    ):
        if not row:
            continue
        rid_val = row[ID_COL - 1] if len(row) >= ID_COL else None
        if rid_val is None:
            continue
        # Normalize research_id to string (DB side is VARCHAR)
        if isinstance(rid_val, float) and rid_val.is_integer():
            rid = str(int(rid_val))
        else:
            rid = str(rid_val).strip()
        if not rid:
            continue

        for n in range(1, N_MAX_FNA + 1):
            c_date = col_for_field(n, "date")
            c_spec = col_for_field(n, "specimen")
            c_path = col_for_field(n, "path")
            c_hist = col_for_field(n, "history")
            c_beth = col_for_field(n, "bethesda")

            v_date = row[c_date - 1] if len(row) >= c_date else None
            v_spec = row[c_spec - 1] if len(row) >= c_spec else None
            v_path = row[c_path - 1] if len(row) >= c_path else None
            v_hist = row[c_hist - 1] if len(row) >= c_hist else None
            v_beth = row[c_beth - 1] if len(row) >= c_beth else None

            s_date = cell_to_string(v_date)
            s_spec = cell_to_string(v_spec)
            s_path = cell_to_string(v_path)
            s_hist = cell_to_string(v_hist)
            s_beth = cell_to_string(v_beth)

            # Skip episodes that are entirely empty (saves 90%+ of long-form rows)
            if not any(
                (x is not None and str(x).strip() != "")
                for x in (s_date, s_spec, s_path, s_hist, s_beth)
            ):
                continue

            rows.append(
                {
                    "research_id": rid,
                    "fna_index": n,
                    "source_workbook": "FNAs 12_5_2025.xlsx",
                    "source_sheet": SOURCE_SHEET,
                    "source_row": excel_row_idx,
                    "source_col_date": c_date,
                    "source_col_specimen": c_spec,
                    "source_col_path": c_path,
                    "source_col_history": c_hist,
                    "source_col_bethesda": c_beth,
                    "source_col_name_date": date_header_per_fna[n],
                    "date_raw": s_date,
                    "specimen_raw": s_spec,
                    "path_raw": s_path,
                    "history_raw": s_hist,
                    "bethesda_raw": s_beth,
                }
            )

    df = pd.DataFrame(rows)
    print(f"[extract_fna_source_long] long-form rows: {len(df):,}")
    print(
        f"[extract_fna_source_long] distinct patients: "
        f"{df['research_id'].nunique():,}"
    )
    print(
        f"[extract_fna_source_long] rows by fna_index:\n"
        f"{df['fna_index'].value_counts().sort_index().to_string()}"
    )
    print(
        f"[extract_fna_source_long] rows with non-null date_raw: "
        f"{df['date_raw'].notna().sum():,}"
    )

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df.to_parquet(OUT_PARQUET, index=False)
    print(f"[extract_fna_source_long] wrote {OUT_PARQUET}")


if __name__ == "__main__":
    main()
