#!/usr/bin/env python3
"""mig_98a — Build Logan review workbook for vocal_cord_paralysis in canonical_complications_events_v1.

Reads MotherDuck `thyroid_canonical_publication_v1_0` + pre-364 archives in
`Thyroid 2026 UPdated`.archive_pub_v1_0 for evidence_span and phenotype fields.

Does NOT mutate canonical tables or registry.

Author: Logan Glosser <logan.glosser@gmail.com>
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))
OUT_DIR = REPO_ROOT / "verification_csvs" / "canonical_complications_events_v1"
OUT_XLSX = OUT_DIR / "vocal_cord_paralysis_review__mig_98a.xlsx"
DECISIONS_JSON = OUT_DIR / "mig_98_decisions.json"

PUB_DB = "thyroid_canonical_publication_v1_0"
ARCH_DB = '"Thyroid 2026 UPdated".archive_pub_v1_0'
PHEN_PRE = f"{ARCH_DB}.complication_phenotype_v1_pre364_20260422_050902"
NEC_PRE = f"{ARCH_DB}.note_entities_complications_pre364_20260422_050902"

LARYNGO_RE = re.compile(
    r"laryngoscopy|fiberoptic|\bdl\b|direct\s+laryng|indirect\s+laryngoscopy",
    re.I,
)
PARALYSIS_RE = re.compile(
    r"vocal\s*cord\s*paralysis|true\s+vocal\s+cord|"
    r"(?:left|right)\s+vocal\s+cord|cord\s*paralysis",
    re.I,
)
NEGATION_CLEAR_RE = re.compile(
    r"vocal\s*cord[s]?\s*(?:are\s+)?mobile|no\s+vocal\s*cord\s*paralysis|"
    r"\bvcm\b|vocals?\s+mobile|normal\s+(?:motion|mobility)\s+of\s+(?:the\s+)?vocal",
    re.I,
)
HOARSE_RE = re.compile(
    r"hoarseness|voice\s+change|weak\s+voice|raspy|dysphonia",
    re.I,
)


def _connect_md() -> duckdb.DuckDBPyConnection:
    from motherduck_client import MotherDuckClient, MotherDuckConfig

    cfg = MotherDuckConfig(database=PUB_DB)
    con = MotherDuckClient(cfg).connect_rw()
    con.execute(f"USE {PUB_DB}")
    return con


def _event_id(
    research_id: str,
    source_table: str,
    source_row_id: str,
    finding_date: Any,
) -> str:
    fd = finding_date.isoformat() if hasattr(finding_date, "isoformat") else str(finding_date)
    raw = f"{research_id}|{source_table}|{source_row_id}|{fd}"
    return hashlib.md5(raw.encode("utf-8")).hexdigest()


def _archive_context_row(r: dict[str, Any]) -> str:
    parts: list[str] = []
    if r.get("source_table") == "complication_phenotype_v1":
        parts.append(f"final_complication_status={r.get('arch_final_status')}")
        parts.append(f"note_mention_flag={r.get('arch_note_mention_flag')}")
        parts.append(f"confirmed_flag={r.get('arch_confirmed_flag')}")
        parts.append(f"voice_resolution_noted={r.get('arch_voice_resolution_noted')}")
        parts.append(f"voice_permanence_noted={r.get('arch_voice_permanence_noted')}")
        parts.append(f"treatment_requiring_flag={r.get('arch_treatment_requiring_flag')}")
        parts.append(f"detection_date={r.get('arch_detection_date')}")
        parts.append(f"first_surgery_date={r.get('arch_first_surgery_date')}")
        parts.append(f"timing_window={r.get('arch_timing_window')}")
    elif r.get("source_table") == "note_entities_complications":
        parts.append(f"present_or_negated={r.get('arch_present_or_negated')}")
    return " | ".join(parts)


def assign_bucket(r: dict[str, Any]) -> tuple[str, str]:
    """Return (bucket, suggested_action)."""
    ev_raw = r.get("evidence_text") or ""
    ev = ev_raw.strip()
    fs = r.get("finding_status") or ""
    onset = r.get("onset_class") or ""
    es = r.get("evidence_strength") or ""
    pc = r.get("permanence_class") or ""
    st = r.get("source_table") or ""

    note_mention = r.get("arch_note_mention_flag")
    final_st = (r.get("arch_final_status") or "").strip()
    vres = r.get("arch_voice_resolution_noted")
    vperm = r.get("arch_voice_permanence_noted")
    pne = (r.get("arch_present_or_negated") or "").lower()

    # --- 1) Confirmed positive ---
    if fs == "present" and es == "definitive":
        return ("CONFIRMED_POSITIVE", "ACCEPT — canonical evidence_strength=definitive")
    if ev and PARALYSIS_RE.search(ev) and LARYNGO_RE.search(ev):
        return ("CONFIRMED_POSITIVE", "ACCEPT — explicit VCP + laryngoscopy in evidence span")

    # --- 2) Confirmed negative ---
    if st == "complication_phenotype_v1":
        if (
            fs == "absent"
            and final_st == "absent_or_unconfirmed"
            and note_mention is False
        ):
            return ("CONFIRMED_NEGATIVE", "ACCEPT — phenotype absent / no NLP mention")
    if ev and NEGATION_CLEAR_RE.search(ev):
        return ("CONFIRMED_NEGATIVE", "ACCEPT — documented mobile cords / explicit negation")
    if fs == "absent" and pne == "negated":
        return ("CONFIRMED_NEGATIVE", "ACCEPT — entity_legacy negated")

    # --- 3) Hoarseness-only (no paralysis wording, no laryngoscopy) ---
    if ev and HOARSE_RE.search(ev) and not LARYNGO_RE.search(ev) and not PARALYSIS_RE.search(ev):
        return (
            "HOARSENESS_ONLY",
            "FLAG — hoarseness / voice change without laryngoscopy or paralysis wording",
        )

    # --- 4) Negation / polarity risk ---
    if fs == "absent" and ev and PARALYSIS_RE.search(ev) and not NEGATION_CLEAR_RE.search(ev):
        return ("NEGATION_RISK", "FLAG — absent finding but positive paralysis language in span")
    if fs == "indeterminate":
        return ("NEGATION_RISK", "FLAG — indeterminate polarity; review entity + phenotype context")

    # --- 6) Onset ambiguous (before permanence) ---
    if fs == "present" and onset == "unspecified":
        return ("ONSET_AMBIGUOUS", "FLAG — present with onset_class=unspecified")

    # --- 5) Permanence unknown ---
    if fs == "present":
        if st == "complication_phenotype_v1":
            if vres is not True and vperm is not True:
                return (
                    "PERMANENCE_UNKNOWN",
                    "FLAG — present phenotype; voice resolution/permanence not documented",
                )
        if st == "note_entities_complications" and pc == "indeterminate":
            return (
                "PERMANENCE_UNKNOWN",
                "FLAG — entity row; permanence_class indeterminate",
            )
        return ("PERMANENCE_UNKNOWN", "FLAG — present; clarify transient vs permanent")

    return ("REVIEW_MISC", "FLAG — manual triage (unexpected pattern)")


def _fetch_rows(con: duckdb.DuckDBPyConnection) -> list[dict[str, Any]]:
    q = f"""
    SELECT
        c.research_id::VARCHAR AS research_id,
        c.source_table,
        c.source_row_id,
        c.source_kind,
        c.source_modality,
        c.finding_date,
        c.finding_status,
        c.evidence_strength,
        c.onset_class,
        c.permanence_class,
        c.source_evidence_type,
        phen.final_complication_status AS arch_final_status,
        phen.note_mention_flag AS arch_note_mention_flag,
        phen.confirmed_flag AS arch_confirmed_flag,
        phen.voice_resolution_noted AS arch_voice_resolution_noted,
        phen.voice_permanence_noted AS arch_voice_permanence_noted,
        phen.treatment_requiring_flag AS arch_treatment_requiring_flag,
        phen.detection_date AS arch_detection_date,
        phen.first_surgery_date AS arch_first_surgery_date,
        phen.timing_window AS arch_timing_window,
        nec.evidence_span AS arch_evidence_span,
        nec.present_or_negated AS arch_present_or_negated
    FROM main.canonical_complications_events_v1 c
    LEFT JOIN {PHEN_PRE} phen
      ON c.source_table = 'complication_phenotype_v1'
     AND c.source_row_id = CAST(hash(phen.research_id, phen.complication_entity,
                                      phen.detection_date) AS VARCHAR)
    LEFT JOIN {NEC_PRE} nec
      ON c.source_table = 'note_entities_complications'
     AND c.source_row_id = CAST(hash(nec.research_id, nec.note_row_id, nec.source_line,
                                      nec.entity_value_norm, nec.evidence_start) AS VARCHAR)
    WHERE c.complication_type = 'vocal_cord_paralysis'
    ORDER BY c.research_id, c.finding_date, c.source_table
    """
    res = con.execute(q)
    cols = [d[0] for d in res.description]
    out: list[dict[str, Any]] = []
    for tup in res.fetchall():
        rec = dict(zip(cols, tup))
        rec["evidence_text"] = rec.get("arch_evidence_span") or ""
        rec["archive_context"] = _archive_context_row(rec)
        rec["event_id"] = _event_id(
            str(rec["research_id"]),
            str(rec["source_table"]),
            str(rec["source_row_id"]),
            rec["finding_date"],
        )
        b, sug = assign_bucket(rec)
        rec["bucket"] = b
        rec["suggested_action"] = sug
        out.append(rec)
    return out


REVIEW_HEADERS = [
    "event_id",
    "research_id",
    "finding_date",
    "source_kind",
    "source_table",
    "archive_context",
    "evidence_text",
    "current_status",
    "current_evidence_strength",
    "current_permanence",
    "current_onset",
    "suggested_action",
    "your_decision",
    "your_note",
]


def _autosize_and_wrap(ws, wrap_col_letters: set[str], max_width: float = 60.0) -> None:
    for col in ws.columns:
        letter = col[0].column_letter
        mlen = 0
        for cell in col:
            if cell.value is None:
                continue
            mlen = max(mlen, min(len(str(cell.value)), 120))
        ws.column_dimensions[letter].width = min(max(10, mlen + 2), max_width)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in wrap_col_letters:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def _write_review_sheet(
    wb: Workbook,
    title: str,
    rows: list[dict[str, Any]],
    *,
    wrap_evidence: bool = True,
) -> None:
    ws = wb.create_sheet(title)
    ws.append(REVIEW_HEADERS)
    for r in rows:
        ws.append(
            [
                r["event_id"],
                r["research_id"],
                r["finding_date"],
                r["source_kind"],
                r["source_table"],
                r.get("archive_context", ""),
                r.get("evidence_text", ""),
                r.get("finding_status", ""),
                r.get("evidence_strength", ""),
                r.get("permanence_class", ""),
                r.get("onset_class", ""),
                r.get("suggested_action", ""),
                "",
                "",
            ]
        )
    wrap_cols = {get_column_letter(REVIEW_HEADERS.index("evidence_text") + 1)}
    if wrap_evidence:
        wrap_cols.add(get_column_letter(REVIEW_HEADERS.index("archive_context") + 1))
    _autosize_and_wrap(ws, wrap_cols)
    ws.freeze_panes = "A2"


def _cf91_rows() -> list[dict[str, Any]]:
    """Synthetic CF-91 absorption rows (not in canonical yet)."""
    return [
        {
            "event_id": "cf91_b79ddab49f77d446c6d232e4819ee2b0",
            "research_id": "5048",
            "finding_date": "2016-12-02",
            "source_kind": "invasion_mig91_deleted",
            "source_table": "(pending canonical_complications_events_v1)",
            "archive_context": "CF-91-VOCAL-CORD — deleted from canonical_invasion_events_v1",
            "evidence_text": "medialization of the right true vocal cord and arytenoid",
            "finding_status": "(none in canonical)",
            "evidence_strength": "",
            "permanence_class": "",
            "onset_class": "",
            "suggested_action": "ADD — no complication rows for rid; absorb as vocal_cord_paralysis?",
            "bucket": "CF91_ABSORPTION",
        },
        {
            "event_id": "cf91_8684d0f01e0ca1a16a90275f5314858b",
            "research_id": "11862",
            "finding_date": "2023-10-28",
            "source_kind": "invasion_mig91_deleted",
            "source_table": "(pending canonical_complications_events_v1)",
            "archive_context": "CF-91-VOCAL-CORD — deleted from canonical_invasion_events_v1; "
            "Cowork: rln_injury rows exist 2024-01-11, no vocal_cord_paralysis row",
            "evidence_text": "superior deviation towards the left",
            "finding_status": "(none in canonical)",
            "evidence_strength": "",
            "permanence_class": "",
            "onset_class": "",
            "suggested_action": "ADD vocal_cord_paralysis @ 2023-10-28 vs rely on rln_injury only — Logan",
            "bucket": "CF91_ABSORPTION",
        },
    ]


def _write_summary(
    wb: Workbook,
    bucket_counts: Counter,
    hoarse_patterns: list[tuple[str, int]],
    run_ts: str,
) -> None:
    ws = wb.create_sheet("summary", 0)
    ws["A1"] = "vocal_cord_paralysis review — mig_98a"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated (UTC): {run_ts}"
    ws["A3"] = "Author: Logan Glosser <logan.glosser@gmail.com>"
    ws["A5"] = "Methodology"
    ws["A5"].font = Font(bold=True)
    lines = [
        "Source: main.canonical_complications_events_v1 WHERE complication_type='vocal_cord_paralysis'.",
        "Archive joins: complication_phenotype_v1_pre364_20260422_050902, "
        "note_entities_complications_pre364_20260422_050902 (pre-364 Script-364 inputs).",
        "Buckets are rule-based triage hints; Logan fills your_decision / your_note.",
        "PHI: full evidence_text only in this workbook (verification_csvs/ is gitignored).",
        "Registry verification flags: deferred until mig_98a–h complete (per Cowork).",
        "",
        "Sign-off (Logan):",
        "  Date: _______________",
        "  Initials: _______________",
    ]
    for i, ln in enumerate(lines, start=6):
        ws.cell(row=i, column=1, value=ln)
    r0 = 6 + len(lines) + 1
    ws.cell(row=r0, column=1, value="Bucket counts (canonical rows + CF-91)").font = Font(bold=True)
    r0 += 1
    ordered = [
        "CONFIRMED_POSITIVE",
        "CONFIRMED_NEGATIVE",
        "HOARSENESS_ONLY",
        "NEGATION_RISK",
        "PERMANENCE_UNKNOWN",
        "ONSET_AMBIGUOUS",
        "REVIEW_MISC",
        "CF91_ABSORPTION",
        "TOTAL_CANONICAL",
    ]
    for k in ordered:
        if k not in bucket_counts:
            continue
        ws.cell(row=r0, column=1, value=k)
        ws.cell(row=r0, column=2, value=bucket_counts[k])
        r0 += 1
    r0 += 1
    ws.cell(row=r0, column=1, value="Top HOARSENESS_ONLY evidence patterns (first 60 chars)").font = Font(
        bold=True
    )
    r0 += 1
    for pat, n in hoarse_patterns:
        ws.cell(row=r0, column=1, value=pat)
        ws.cell(row=r0, column=2, value=n)
        r0 += 1


def _truncate_stdout(s: str, n: int = 80) -> str:
    s = re.sub(r"\s+", " ", (s or "").strip())
    if len(s) <= n:
        return s
    return s[: n - 3] + "..."


def _update_decisions_stub(bucket_counts: Counter, run_ts: str) -> None:
    """Append-ready stub for mig_98_decisions.json (decisions filled after Logan review)."""
    payload: dict[str, Any] = {
        "schema_version": 1,
        "last_updated_utc": run_ts,
        "note": "Per-row decisions appended after Logan returns filled workbook via Cowork.",
        "complications": {
            "vocal_cord_paralysis": {
                "mig_98a_workbook": str(OUT_XLSX.name),
                "canonical_row_count": int(bucket_counts.get("TOTAL_CANONICAL", 0)),
                "bucket_counts_at_generation": dict(bucket_counts),
                "decisions": [],
            }
        },
    }
    DECISIONS_JSON.parent.mkdir(parents=True, exist_ok=True)
    DECISIONS_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--stdout-evidence-chars",
        type=int,
        default=80,
        help="Truncate evidence snippets printed to stdout (PHI hygiene).",
    )
    args = ap.parse_args()
    run_ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    con = _connect_md()
    rows = _fetch_rows(con)
    con.close()

    bucket_counts: Counter = Counter(r["bucket"] for r in rows)
    bucket_counts["TOTAL_CANONICAL"] = len(rows)

    hoarse_rows = [r for r in rows if r["bucket"] == "HOARSENESS_ONLY"]
    pat_ctr: Counter = Counter()
    for r in hoarse_rows:
        ev = (r.get("evidence_text") or "").strip()
        key = ev[:60].replace("\n", " ") if ev else "(empty)"
        pat_ctr[key] += 1
    hoarse_top = pat_ctr.most_common(5)

    cf91 = _cf91_rows()
    bucket_counts["CF91_ABSORPTION"] = len(cf91)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    _write_summary(wb, bucket_counts, hoarse_top, run_ts)

    by_bucket: dict[str, list[dict[str, Any]]] = {b: [] for b in bucket_counts if b != "TOTAL_CANONICAL"}
    for r in rows:
        by_bucket.setdefault(r["bucket"], []).append(r)

    _write_review_sheet(wb, "1_confirmed_pos", by_bucket.get("CONFIRMED_POSITIVE", []))
    _write_review_sheet(wb, "2_confirmed_neg", by_bucket.get("CONFIRMED_NEGATIVE", []))
    _write_review_sheet(wb, "3_hoarseness_only", by_bucket.get("HOARSENESS_ONLY", []))
    _write_review_sheet(wb, "4_negation_risk", by_bucket.get("NEGATION_RISK", []))
    _write_review_sheet(wb, "5_permanence_unknown", by_bucket.get("PERMANENCE_UNKNOWN", []))
    _write_review_sheet(wb, "6_onset_ambiguous", by_bucket.get("ONSET_AMBIGUOUS", []))
    misc = by_bucket.get("REVIEW_MISC", [])
    if misc:
        _write_review_sheet(wb, "z_review_misc", misc)
    _write_review_sheet(wb, "7_cf91_absorption", cf91)

    wb.properties.creator = "Logan Glosser <logan.glosser@gmail.com>"
    wb.save(OUT_XLSX)

    _update_decisions_stub(bucket_counts, run_ts)

    # --- stdout summary (truncated evidence) ---
    print("mig_98a vocal_cord_paralysis review workbook")
    print(f"  wrote: {OUT_XLSX}")
    print(f"  decisions stub: {DECISIONS_JSON}")
    print("  bucket counts:")
    for name in (
        "CONFIRMED_POSITIVE",
        "CONFIRMED_NEGATIVE",
        "HOARSENESS_ONLY",
        "NEGATION_RISK",
        "PERMANENCE_UNKNOWN",
        "ONSET_AMBIGUOUS",
        "REVIEW_MISC",
        "CF91_ABSORPTION",
    ):
        if name == "CF91_ABSORPTION":
            print(f"    {name}: {len(cf91)}")
        else:
            print(f"    {name}: {bucket_counts.get(name, 0)}")
    print(f"  TOTAL_CANONICAL: {len(rows)}")
    print("  top HOARSENESS_ONLY patterns:")
    for pat, n in hoarse_top:
        print(f"    n={n}  {_truncate_stdout(pat, args.stdout_evidence_chars)!r}")
    print("  CF-91 absorption (full evidence in workbook; truncated here):")
    for r in cf91:
        print(
            f"    rid={r['research_id']} date={r['finding_date']} | "
            f"{_truncate_stdout(r['evidence_text'], args.stdout_evidence_chars)!r}"
        )


if __name__ == "__main__":
    main()
