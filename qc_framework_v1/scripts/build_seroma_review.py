#!/usr/bin/env python3
"""mig_98d — Build seroma candidates review workbook.

Reuses chyle/RLN methodology:
  - All 873 present pts get one row with full evidence
  - REAL/TEMPLATE note-text classifier (seroma-tuned)
  - claude_attribution from phenotype timing_days_post_surgery
  - Structured-signal flag (28 phen_confirmed/refined pts)
  - Logan reviews per row → KEEP/PMH/DELETE/NEEDS_CONTEXT

Decision vocabulary (your_decision):
  OPERATIVE       — drained/aspirated post-op seroma
  POSTOP_LATE     — operative but late detection (>30d)
  POSTOP_TX_REQ   — operative with treatment escalation
  PRIOR_OP        — from a prior (non-index) surgery
  PREEXISTING     — pre-existing or non-surgical fluid collection
  NOT_OPERATIVE   — not from index surgery (>5y)
  NEEDS_CONTEXT   — needs further chart review
  NO              — false positive / template noise

Output: verification_csvs/canonical_complications_events_v1/
        seroma_review__mig_98d.xlsx

Author: Logan Glosser <logan.glosser@gmail.com>
"""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

PUB_DB = "thyroid_canonical_publication_v1_0"
ARCH_DB = '"Thyroid 2026 UPdated".archive_pub_v1_0'
PHEN_PRE = f"{ARCH_DB}.complication_phenotype_v1_pre364_20260422_050902"

OUT_DIR = REPO_ROOT / "verification_csvs" / "canonical_complications_events_v1"
OUT_XLSX = OUT_DIR / "seroma_review__mig_98d.xlsx"

NOTE_TYPE_PRIORITY = [
    "DC_SUM", "HP", "OPNOTE", "ENDOCRINE_FM", "ED_NOTE",
    "OTHER_HISTORY", "OTHER_NOTES",
]

_ILLEGAL_CTRL_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
SEROMA_RE = re.compile(r"\bseroma", re.IGNORECASE)

TEMPLATE_PATTERNS = [
    # Consent risk lists ("infection, seroma, hematoma, ...")
    r"\brisks?\s+(?:include|of|involved)[^.]{0,300}seroma",
    r"\bconsent[^.]{0,200}seroma",
    r"\bcomplications?\s+(?:include|may include|of\s+(?:thyroid|surgery|the\s+procedure))[^.]{0,250}seroma",
    r"\bcounseled\s+(?:about|on|regarding)[^.]{0,80}seroma",
    r"\bdiscussed[^.]{0,150}\b(risks?|complications?|possibilit|sequelae)\b[^.]{0,150}seroma",
    # Prophylactic drain placement
    r"\bdrain\s+(?:was\s+)?placed[^.]{0,80}(?:to\s+)?prevent\s+seroma",
    r"\bto\s+prevent\s+(?:a\s+)?seroma",
    # Negation
    r"\bno\s+(?:evidence\s+of\s+)?seroma",
    r"\bwithout\s+(?:a\s+)?seroma",
    r"\bnegative\s+for\s+seroma",
    r"\bseroma\s+(?:was\s+)?not\s+(?:noted|seen|present|identified)",
    # Monitoring-only language
    r"\bmonitor(?:ed|ing)?\s+for[^.]{0,40}seroma",
    # Generic boilerplate listing
    r"\bseroma[, /][^.]{0,80}\bhematoma",
    r"\bhematoma[, /][^.]{0,80}\bseroma",
    r"\binfection[, /][^.]{0,40}seroma",
]

REAL_PATTERNS = [
    # Drainage / aspiration
    r"\b(?:drained|drainage|aspirat(?:ed|ion)|incision\s+and\s+drainage|i\s*&\s*d)\s+(?:of\s+)?(?:the\s+)?seroma",
    r"\bseroma[^.]{0,120}\b(?:drained|drainage|aspirat(?:ed|ion))",
    r"\bneedle\s+aspirat(?:ed|ion)[^.]{0,80}seroma",
    # Documented complication language
    r"\b(?:complicated\s+by|notable for|c\/b)\s+(?:a\s+|the\s+)?seroma",
    r"\bs/p[^.]{0,60}c\/b[^.]{0,60}seroma",
    r"\b(?:diagnosis|admit(?:ted|ting)?(?:\s+with)?)[^.]{0,80}seroma",
    # Active management / persistent
    r"\bseroma[^.]{0,120}\b(?:persistent|persisted|prolonged|increasing|recurrent)",
    r"\bseroma\s+(?:was\s+)?(?:treated|managed|addressed|noted|developed|present)",
    # Specific findings
    r"\bfluid\s+collection[^.]{0,80}seroma",
    r"\bseroma[^.]{0,120}(?:incision|wound|neck)",
    # Post-op clinic mentions
    r"\b(?:on |at |started\s+on\s+)POD\s*\d+[^.]{0,80}seroma",
    r"seroma[^.]{0,80}\bPOD\s*\d+",
]

TEMPLATE_RE = [re.compile(p, re.IGNORECASE) for p in TEMPLATE_PATTERNS]
REAL_RE = [re.compile(p, re.IGNORECASE) for p in REAL_PATTERNS]

HEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
DEC_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _connect():
    from motherduck_client import MotherDuckClient, MotherDuckConfig

    cfg = MotherDuckConfig(database=PUB_DB)
    con = MotherDuckClient(cfg).connect_rw()
    con.execute(f"USE {PUB_DB}")
    return con


def _squash(s, max_len=500):
    s = (s or "").replace("\r", " ").replace("\n", " ").strip()
    s = _ILLEGAL_CTRL_RE.sub(" ", s)
    while "  " in s:
        s = s.replace("  ", " ")
    if len(s) > max_len:
        s = s[: max_len - 3] + "..."
    return s


def _classify(ctx):
    is_template = any(r.search(ctx) for r in TEMPLATE_RE)
    is_real = any(r.search(ctx) for r in REAL_RE)
    return is_template, is_real


def _score_note(txt):
    found = []
    for m in SEROMA_RE.finditer(txt):
        pos = m.start()
        ctx = txt[max(0, pos - 200): pos + 300]
        is_t, is_r = _classify(ctx)
        found.append((pos, ctx, is_t, is_r))
    return found


def _classify_attribution(timing_days, timing_window):
    if timing_window == "pre_surgery":
        return "PREEXISTING", "phenotype timing_window=pre_surgery"
    if timing_days is None:
        return "UNKNOWN", "no timing_days_post_surgery"
    td = int(timing_days)
    if td < 0:
        return "PREEXISTING", f"finding precedes first surgery by {-td} days"
    if td <= 30:
        return "OPERATIVE", f"timing_days={td}"
    if td <= 180:
        return "POSTOP_LATE", f"timing_days={td}"
    if td <= 365:
        return "POSTOP_VERY_LATE", f"timing_days={td}"
    if td <= 365 * 5:
        return "POSSIBLY_PRIOR_OP", f"timing_days={td} ({td // 365}y)"
    return "NOT_OPERATIVE", f"timing_days={td} ({td // 365}y)"


def fetch(con):
    rids = [r[0] for r in con.execute(
        "SELECT DISTINCT research_id FROM main.canonical_complications_events_v1 "
        "WHERE complication_type='seroma' AND finding_status IN ('present','suspected')"
    ).fetchall()]

    rid_list = "','".join(rids)

    # Per-pt canonical aggregate
    canon_q = (
        f"SELECT c.research_id, "
        f"STRING_AGG(DISTINCT c.finding_status, ',' ORDER BY c.finding_status) AS statuses, "
        f"STRING_AGG(DISTINCT c.onset_class, ',' ORDER BY c.onset_class) AS onsets, "
        f"MIN(c.finding_date)::VARCHAR AS first_dt, "
        f"BOOL_OR(c.evidence_strength='definitive') AS any_definitive, "
        f"BOOL_OR(c.evidence_strength='probable') AS any_probable, "
        f"BOOL_OR(c.source_table='extracted_complications_refined_v5') AS has_refined_v5 "
        f"FROM main.canonical_complications_events_v1 c "
        f"WHERE c.complication_type='seroma' "
        f"AND c.research_id IN ('{rid_list}') "
        f"GROUP BY c.research_id"
    )
    canon = {r[0]: r for r in con.execute(canon_q).fetchall()}

    # First-surgery date
    fs_q = (
        f"SELECT research_id, MIN(surgery_date_native)::VARCHAR AS first_sx "
        f"FROM main.canonical_operative_events_v1 "
        f"WHERE research_id IN ('{rid_list}') "
        f"GROUP BY research_id"
    )
    first_sx = {r[0]: r[1] for r in con.execute(fs_q).fetchall()}

    # Phenotype data
    phen_q = (
        f"SELECT phen.research_id::VARCHAR, "
        f"BOOL_OR(phen.confirmed_flag) AS confirmed, "
        f"BOOL_OR(phen.suspected_flag) AS suspected, "
        f"BOOL_OR(phen.transient_flag) AS transient, "
        f"BOOL_OR(phen.permanent_flag) AS permanent, "
        f"BOOL_OR(phen.treatment_requiring_flag) AS treat_req, "
        f"MIN(phen.timing_days_post_surgery) AS min_td, "
        f"STRING_AGG(DISTINCT phen.source_tier_label, ',' ORDER BY phen.source_tier_label) AS tiers, "
        f"STRING_AGG(DISTINCT phen.timing_window, ',' ORDER BY phen.timing_window) AS windows "
        f"FROM {PHEN_PRE} phen "
        f"WHERE phen.research_id::VARCHAR IN ('{rid_list}') "
        f"AND phen.complication_entity ILIKE '%seroma%' "
        f"GROUP BY phen.research_id"
    )
    phen = {r[0]: r for r in con.execute(phen_q).fetchall()}

    # Notes
    con.execute("CREATE OR REPLACE TEMP TABLE tmp_pts (rid VARCHAR)")
    con.executemany("INSERT INTO tmp_pts VALUES (?)", [(r,) for r in rids])
    notes = con.execute(
        "SELECT cnl.research_id, cnl.note_type, cnl.note_index, cnl.note_text "
        "FROM main.clinical_notes_long cnl JOIN tmp_pts t ON cnl.research_id = t.rid "
        "WHERE POSITION('seroma' IN LOWER(cnl.note_text)) > 0"
    ).fetchall()

    return rids, canon, first_sx, phen, notes


def aggregate(notes):
    per_pt = defaultdict(lambda: {"real": [], "template": [], "unmarked": [], "all_count": 0})
    for rid, nt, idx, txt in notes:
        if not txt:
            continue
        for pos, ctx, is_t, is_r in _score_note(txt):
            per_pt[rid]["all_count"] += 1
            entry = {"note_type": nt, "note_index": idx, "pos": pos, "context": ctx}
            if is_r:
                per_pt[rid]["real"].append(entry)
            elif is_t:
                per_pt[rid]["template"].append(entry)
            else:
                per_pt[rid]["unmarked"].append(entry)
    return per_pt


HEADERS = [
    "research_id",
    "first_surgery_date",
    "claude_attribution",
    "claude_attribution_basis",
    "structured_signal",        # phen_confirmed OR refined_v5
    "canonical_statuses",
    "canonical_onsets",
    "canonical_first_finding_date",
    "any_definitive_evidence",
    "any_probable_evidence",
    "phen_confirmed",
    "phen_treat_req",
    "phen_permanent",
    "phen_transient",
    "phen_timing_days",
    "phen_timing_window",
    "phen_source_tier",
    "n_real_pattern_mentions",
    "n_template_mentions",
    "n_unmarked_mentions",
    "source1", "evidence1",
    "source2", "evidence2",
    "source3", "evidence3",
    "source4", "evidence4",
    "claude_suggested",
    "your_decision",
    "your_note",
]


def best_excerpts(info):
    """Pick best excerpt per note_type (REAL preferred, then TEMPLATE/UNMARKED)."""
    by_nt = defaultdict(list)
    for e in info["real"]:
        by_nt[e["note_type"]].append(_squash(e["context"]))
    for e in info["template"] + info["unmarked"]:
        nt = e["note_type"]
        if not by_nt.get(nt):
            by_nt[nt].append(_squash(e["context"]))
    ordered = [nt for nt in NOTE_TYPE_PRIORITY if nt in by_nt]
    sources = (ordered + [""] * 4)[:4]
    evidences = ["; ".join(by_nt.get(nt, [])[:2]) if nt else "" for nt in sources]
    return sources, evidences


def build_summary(wb, counts):
    ws = wb.create_sheet("summary", 0)
    ws["A1"] = "seroma review — mig_98d"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated (UTC): {datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')}"
    ws["A3"] = "Author: Logan Glosser <logan.glosser@gmail.com>"
    ws["A5"] = (
        "Method: same as mig_98b chyle / mig_98c voice/nerve. Pulled all 'seroma' "
        "mentions from clinical_notes_long for 873 canonical present pts; classified "
        "each 500-char window as REAL (drainage/aspiration/'complicated by'/treatment) "
        "vs TEMPLATE (consent risk lists, prophylactic drain placement, negation, "
        "monitoring-only language). claude_attribution from phenotype timing_days "
        "(PREEXISTING / OPERATIVE / POSTOP_LATE / POSSIBLY_PRIOR_OP / NOT_OPERATIVE)."
    )
    ws["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110
    ws.row_dimensions[5].height = 80

    ws["A7"] = "Counts"
    ws["A7"].font = Font(bold=True)
    rows = [
        f"Cohort: 10,871 pts",
        f"seroma rows in canonical: 1,407 present + 8 suspected + 844 absent",
        f"Distinct present|suspected pts: {counts['total']}",
        f"  -- with structured signal (phen_confirmed OR refined_v5): {counts['structured']}",
        f"  -- with REAL note-text mention(s):                        {counts['real']}",
        f"  -- TEMPLATE-only (likely FP):                             {counts['template_only']}",
        "",
        "Decision vocabulary:",
        "  OPERATIVE / POSTOP_LATE / POSTOP_TX_REQ — keep present",
        "  PRIOR_OP / PREEXISTING / NOT_OPERATIVE / POSSIBLY_PRIOR_OP — move to PMH",
        "  NEEDS_CONTEXT — defer",
        "  NO — DELETE (template FP)",
    ]
    for i, r in enumerate(rows, start=8):
        ws.cell(row=i, column=1, value=r)


def build_review_sheet(wb, rids, canon, first_sx, phen, per_pt):
    ws = wb.create_sheet("ALL_PATIENTS")
    for j, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL

    counts = {"total": len(rids), "structured": 0, "real": 0, "template_only": 0}

    rows_data = []
    for rid in rids:
        c = canon.get(rid)
        p = phen.get(rid)
        info = per_pt.get(rid, {"real": [], "template": [], "unmarked": [], "all_count": 0})
        is_struct = bool(c and c[6])  # has_refined_v5
        if not is_struct and p and p[1]:
            is_struct = True  # phen_confirmed
        if is_struct:
            counts["structured"] += 1

        td = p[6] if p else None
        tw = p[8] if p else None
        attr, basis = _classify_attribution(td, tw)

        sources, evidences = best_excerpts(info)
        if info["real"]:
            counts["real"] += 1
        elif info["template"] and not info["unmarked"]:
            counts["template_only"] += 1

        # claude_suggested rule
        if is_struct and attr == "OPERATIVE":
            sug = "REVIEW (likely OPERATIVE)"
        elif is_struct and attr in ("NOT_OPERATIVE", "POSSIBLY_PRIOR_OP"):
            sug = "REVIEW (likely PMH)"
        elif info["real"] and attr == "OPERATIVE":
            sug = "REVIEW (operative; real text)"
        elif not info["real"] and not info["template"] and not info["unmarked"]:
            sug = "NO (no seroma text)"
        elif not info["real"]:
            sug = "NO (template-only)"
        else:
            sug = "REVIEW"

        rows_data.append([
            rid, first_sx.get(rid), attr, basis, is_struct,
            c[1] if c else "",  # statuses
            c[2] if c else "",  # onsets
            c[3] if c else "",  # canonical first dt
            bool(c[4]) if c else None,
            bool(c[5]) if c else None,
            p[1] if p else None,
            p[5] if p else None,
            p[4] if p else None,
            p[3] if p else None,
            td,
            tw,
            p[7] if p else "",
            len(info["real"]),
            len(info["template"]),
            len(info["unmarked"]),
            sources[0], evidences[0],
            sources[1], evidences[1],
            sources[2], evidences[2],
            sources[3], evidences[3],
            sug,
            "",
            "",
        ])

    # Sort: structured first, then by # real mentions desc
    rows_data.sort(key=lambda r: (
        not r[4],  # structured first
        -r[17],    # n_real desc
        r[0]
    ))
    for d in rows_data:
        ws.append(d)

    dec_cols = {
        get_column_letter(HEADERS.index("your_decision") + 1),
        get_column_letter(HEADERS.index("your_note") + 1),
    }
    wrap_cols = {
        get_column_letter(HEADERS.index(h) + 1)
        for h in ("evidence1", "evidence2", "evidence3", "evidence4",
                  "claude_attribution_basis", "your_note")
    }
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in dec_cols:
                cell.fill = DEC_FILL
    ws.freeze_panes = "C2"
    for col in ws.columns:
        letter = col[0].column_letter
        mlen = 0
        for cell in col:
            if cell.value is None: continue
            mlen = max(mlen, min(len(str(cell.value)), 100))
        ws.column_dimensions[letter].width = min(max(10, mlen + 2), 60.0)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    return counts


def main():
    con = _connect()
    rids, canon, first_sx, phen, notes = fetch(con)
    con.close()

    per_pt = aggregate(notes)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    placeholder_counts = {"total": len(rids), "structured": 0, "real": 0, "template_only": 0}
    build_summary(wb, placeholder_counts)
    counts = build_review_sheet(wb, rids, canon, first_sx, phen, per_pt)
    wb.remove(wb["summary"])
    build_summary(wb, counts)
    wb.move_sheet("summary", offset=-1)

    wb.properties.creator = "Logan Glosser <logan.glosser@gmail.com>"
    wb.save(OUT_XLSX)
    print(f"wrote: {OUT_XLSX}")
    for k, v in counts.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
