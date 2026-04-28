"""
qc_framework_v1/scripts/build_esophageal_invasion_signoff.py
============================================================

ONE column to review for canonical_esophageal_invasion_events_v1:
the LLM's `present_or_negated` call per row.

The table is small (188 rows; raw note_entities-style extractions with no
staging column to derive). For each row the question is: given the
`evidence_text` quoted from the note, is the LLM's `present`/`negated`
call correct, AND is the row a real esophageal-invasion event vs a
contextual procedural mention?

Output: .xlsx (avoids the CSV comma-escape parsing issue that broke 2
of 47 rows in the t4b mig_91 review). evidence_text gets its own
distinctly-named column immediately before the decision column.

Decision vocabulary (column `your_decision`):
  ACCEPT  -- (default; blank == ACCEPT) keep present_or_negated as-is
  FLIP    -- LLM got the sense wrong; present <-> negated
  REJECT  -- not a real esophageal-invasion event (procedural-context-only
             mention; e.g. esophagus described as a landmark, not invaded)

Output:
  verification_csvs/canonical_esophageal_invasion_events_v1/
    esophageal_signoff__mig_93.xlsx
"""
from __future__ import annotations

from pathlib import Path

import duckdb
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

REPO_ROOT = Path(__file__).resolve().parents[2]
OUT = (
    REPO_ROOT
    / "verification_csvs"
    / "canonical_esophageal_invasion_events_v1"
    / "esophageal_signoff__mig_93.xlsx"
)
DB = "thyroid_canonical_publication_v1_0"

POSITIVE_ENTITY_TYPES = (
    "esophageal_invasion_present",
    "esophageal_invasion_extent",
    "esophageal_muscularis_invasion",
    "esophageal_mucosal_invasion",
    "esophageal_invasion_length_cm",
    "esophageal_repair_performed",
)


def main() -> None:
    con = duckdb.connect("md:")
    con.execute(f'USE "{DB}"')
    rows = con.execute(
        f"""
        SELECT
          research_id, note_type, note_date, source_column,
          entity_type, entity_value, present_or_negated,
          confidence, source_line, entity_date,
          evidence_text,
          CASE WHEN entity_type IN
               ({", ".join(f"'{t}'" for t in POSITIVE_ENTITY_TYPES)})
               THEN 1 ELSE 0 END AS is_invasion_entity
        FROM main.canonical_esophageal_invasion_events_v1
        ORDER BY is_invasion_entity DESC,
                 present_or_negated DESC,
                 entity_type,
                 research_id
        """
    ).fetchall()
    print(f"esophageal rows: {len(rows)}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "esophageal_signoff_mig_93"

    # Header rows: instructions then column headers.
    instructions = [
        "Esophageal invasion sign-off review -- canonical_esophageal_invasion_events_v1",
        "Protocol v2 LLM-output canonical (188 rows; full table review).",
        "",
        "ONE column to review per row: present_or_negated (the LLM's sense).",
        "",
        "your_decision vocabulary:",
        "  ACCEPT  -- (blank == ACCEPT) present_or_negated is correct",
        "  FLIP    -- LLM got the sense wrong; should be the opposite",
        "  REJECT  -- not a real esophageal-invasion event (procedural-context",
        "             only mention; e.g. esophagus described as a landmark)",
        "",
        "Notes:",
        " - is_invasion_entity=1 means entity_type is one of the 6 invasion",
        "   entity types (esophageal_invasion_present / _extent / _muscularis_",
        "   invasion / _mucosal_invasion / _invasion_length_cm / _repair_performed).",
        " - is_invasion_entity=0 rows are extractions for other esophageal-related",
        "   entities (procedural references etc.).",
        " - evidence_text is the quoted note span the LLM used to support the call.",
        "",
    ]
    for line in instructions:
        ws.append([line])
    instr_end_row = ws.max_row

    headers = [
        "research_id", "note_type", "note_date", "source_column",
        "entity_type", "entity_value", "present_or_negated",
        "confidence", "source_line", "entity_date",
        "evidence_text",  # dedicated column right before decision
        "is_invasion_entity",
        "your_decision", "your_note",
    ]
    ws.append(headers)
    header_row = ws.max_row

    # Bold + light-fill the header row.
    fill = PatternFill(start_color="FFE6E6E6",
                       end_color="FFE6E6E6", fill_type="solid")
    for col_ix in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_ix)
        cell.font = Font(bold=True)
        cell.fill = fill

    # Data rows.
    for r in rows:
        ws.append(list(r) + ["", ""])

    # Wrap evidence_text + reasonable column widths.
    col_widths = {
        "A": 12, "B": 16, "C": 12, "D": 14,        # rid, note_type, date, source_column
        "E": 32, "F": 22, "G": 16,                  # entity_type, entity_value, p_or_n
        "H": 10, "I": 10, "J": 12,                  # confidence, source_line, entity_date
        "K": 80,                                    # evidence_text (wide)
        "L": 16,                                    # is_invasion_entity
        "M": 14, "N": 30,                           # decision, note
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Wrap the evidence_text column in body rows.
    evidence_col_ix = headers.index("evidence_text") + 1
    decision_col_ix = headers.index("your_decision") + 1
    for row_ix in range(header_row + 1, ws.max_row + 1):
        ws.cell(row=row_ix, column=evidence_col_ix).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        # Yellow-tint the decision column so it's visually obvious where to type.
        ws.cell(row=row_ix, column=decision_col_ix).fill = PatternFill(
            start_color="FFFFFFCC", end_color="FFFFFFCC", fill_type="solid"
        )

    # Freeze the instructions + header so scrolling keeps the column titles in view.
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"-> wrote {OUT}")


if __name__ == "__main__":
    main()
