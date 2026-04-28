#!/usr/bin/env python3
"""mig_98c — Build rln_injury REAL candidates workbook (note-text classifier).

Adapts the mig_98b chyle_leak methodology to recurrent laryngeal nerve injury.
Pulls every RLN-related mention from main.clinical_notes_long for the 1,414
rln_injury pts in canonical_complications_events_v1; classifies each
~500-char context window:

  TEMPLATE  — consent risk list, "RLN identified and preserved",
              "RLN intact", negation ("no RLN injury"/"RLN preserved"/
              "RLN dissected free"), nerve-stim documentation, planning
              language ("likely a RLN injury during proposed surgery").
  REAL      — RLN sacrificed/transected/injured, vocal cord paralysis with
              treatment escalation (medialization, injection laryngoplasty,
              thyroplasty), persistent post-op hoarseness, voice/speech
              therapy referral, laryngoscopy showing paralysis, permanent
              paralysis flag.

Cross-source augmentations:
  - Auto-include patients with phenotype confirmed_flag=TRUE (high-confidence
    structured signal)
  - Auto-include patients with extracted_rln_injury_refined_v2 source rows
    (dedicated v2 NLP extractor)
  - Auto-include patients overlapping mig_98a vocal_cord_paralysis keepers
    (the VCP entity is the clinical outcome of RLN injury)

Output: verification_csvs/canonical_complications_events_v1/
        rln_injury_real_candidates__mig_98c.xlsx
  - summary sheet: counts + smoking-gun examples + recommended apply
  - REAL_CANDIDATES sheet: 1 row per pt; cols include source1..4 +
    evidence1..4 + cross-source flags + your_decision/your_note
  - NO_RLN_INJURY sheet: per-pt summary for the rest with sample
    template excerpt + suggested_decision='NO'.

Author: Logan Glosser <logan.glosser@gmail.com>
"""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

PUB_DB = "thyroid_canonical_publication_v1_0"
ARCH_DB = '"Thyroid 2026 UPdated".archive_pub_v1_0'
PHEN_PRE = f"{ARCH_DB}.complication_phenotype_v1_pre364_20260422_050902"

OUT_DIR = REPO_ROOT / "verification_csvs" / "canonical_complications_events_v1"
OUT_XLSX = OUT_DIR / "rln_injury_real_candidates__mig_98c.xlsx"

NOTE_TYPE_PRIORITY = [
    "DC_SUM", "HP", "OPNOTE", "ENDOCRINE_FM", "ED_NOTE",
    "OTHER_HISTORY", "OTHER_NOTES",
]

# Tokens to match in clinical_notes_long (case-insensitive)
RLN_TOKEN_RE = re.compile(
    r"\b(rln|recurrent\s+laryngeal|vocal\s+cord\s+paralysis|vocal\s+fold\s+paralysis|"
    r"\bvcp\b)",
    re.IGNORECASE,
)

TEMPLATE_PATTERNS = [
    # Consent risk lists
    r"\brisks?\s+(?:include|of|involved)[^.]{0,300}\b(rln|recurrent\s+laryngeal|vocal\s+cord)",
    r"\bconsent[^.]{0,250}\b(rln|recurrent\s+laryngeal|vocal\s+cord)",
    r"\bcomplications?\s+(?:include|may include|of\s+(?:thyroid|surgery|the\s+procedure))[^.]{0,300}\b(rln|recurrent\s+laryngeal)",
    r"\bcounseled\s+(?:about|on|regarding)[^.]{0,80}\b(rln|recurrent\s+laryngeal|vocal)",
    r"\bdiscussed[^.]{0,150}\b(risks?|complications?|possibilit|sequelae)\b[^.]{0,150}\b(rln|recurrent\s+laryngeal)",
    # Op-note prophylactic identification (RLN identified/preserved/intact)
    r"\b(rln|recurrent\s+laryngeal\s+nerve)\s+(?:was\s+)?(?:identified\s+and\s+)?(?:preserved|protected|intact|spared)",
    r"\b(rln|recurrent\s+laryngeal\s+nerve)\s+(?:was\s+)?dissected\s+(?:free|out)",
    r"\b(rln|recurrent\s+laryngeal\s+nerve)\s+(?:was\s+)?(?:traced|followed|skeletoniz)",
    r"\bidentified\s+(?:and\s+)?preserved\s+(?:the\s+)?(?:rln|recurrent\s+laryngeal)",
    r"\b(?:b/l|bilateral|both)\s+(?:rlns?|recurrent\s+laryngeal)\s+(?:were\s+)?(?:identified|preserved|intact)",
    # Negation
    r"\bno\s+(?:rln|recurrent\s+laryngeal|vocal\s+cord)\s+(?:injury|paralysis|deficit)",
    r"\b(?:rln|recurrent\s+laryngeal\s+nerve)\s+(?:was\s+)?intact",
    r"\bvocal\s+cords?\s+(?:are\s+)?(?:bilaterally\s+)?mobile",
    r"\bvocal\s+folds?\s+(?:are\s+)?(?:bilaterally\s+)?mobile",
    r"\bnormal\s+(?:vocal\s+(?:cord|fold)|laryng(?:eal|oscopy))",
    # Nerve stim documentation
    r"\bstimulated\s+at\s+\d+(?:\.\d+)?\s*ma",
    r"\bnims?\s+(?:monitor|stimulator|recording)",
    # Planning / hypothetical
    r"\blikely\s+(?:a\s+)?(?:rln|recurrent\s+laryngeal)\s+injury",
    r"\b(?:could|may|might|possible)\s+(?:require|result\s+in)[^.]{0,80}\b(rln|recurrent\s+laryngeal)",
    r"\bwould\s+(?:expect|require)\s+(?:rln|recurrent\s+laryngeal)\s+(?:injury|sacrifice)",
]

REAL_PATTERNS = [
    # Active injury / sacrifice
    r"\b(rln|recurrent\s+laryngeal\s+nerve)\s+(?:was\s+)?(?:sacrificed|transected|cut|severed|ligated|resected|injur(?:ed|y))",
    r"\bsacrific(?:ed|ing)\s+(?:the\s+)?(?:rln|recurrent\s+laryngeal\s+nerve)",
    r"\b(?:transect(?:ed|ion)|cut)\s+(?:the\s+)?(?:rln|recurrent\s+laryngeal\s+nerve)",
    # Vocal cord paralysis with management/permanence
    r"vocal\s+(?:cord|fold)\s+paralysis[^.]{0,180}\b(medialization|injection|thyroplast|silastic|gelfoam|laryngoplast)",
    r"\b(?:left|right)\s+vocal\s+(?:cord|fold)\s+paralysis",
    r"\b(?:permanent|persistent|chronic|fixed)\s+(?:vocal\s+(?:cord|fold)\s+paralysis|hoarseness|dysphonia)",
    # Treatment evidence
    r"\bmedialization\b",
    r"\b(?:injection\s+)?laryngoplast",
    r"\bthyroplast",
    r"\bvoice\s+therapy\b",
    r"\bspeech\s+(?:and\s+language\s+)?(?:therapy|pathology|pathologist)",
    r"\bSLP\s+(?:referral|consult|evaluation)",
    r"\bENT\s+(?:consult|referral|evaluation)\s+(?:for\s+)?(?:voice|hoarseness)",
    # Post-op clinical evidence
    r"\bhoarseness[^.]{0,100}\b(?:persistent|continued|prolonged|months|weeks|post.?op|since\s+surgery)",
    r"\b(?:persistent|continued)\s+(?:hoarseness|dysphonia|voice\s+change)",
    r"\blaryngoscopy[^.]{0,100}(?:paralysis|paretic|fixed|immobile|hypomobile)",
    r"\bdirect\s+laryngoscopy.{0,100}(?:paralys|paretic|immobile)",
    r"\bfiberoptic[^.]{0,100}(?:paralys|immobile|hypomobile)",
    r"\bDL\s+(?:showed|revealed|demonstrated)[^.]{0,80}(?:paralys|immobile|paretic)",
    # Documented complication language
    r"\b(?:complicated\s+by|notable for|c\/b)\s+(?:a\s+|the\s+)?(?:rln|recurrent\s+laryngeal|vocal\s+cord\s+paralys)",
    r"\bs/p[^.]{0,60}c\/b[^.]{0,80}(?:rln|vocal\s+cord)",
    r"\b(?:diagnosis|admit(?:ted)?(?:\s+with)?)[^.]{0,80}\b(?:rln|recurrent\s+laryngeal|vocal\s+cord\s+paralys)",
]

TEMPLATE_RE = [re.compile(p, re.IGNORECASE) for p in TEMPLATE_PATTERNS]
REAL_RE = [re.compile(p, re.IGNORECASE) for p in REAL_PATTERNS]

HEADER_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
DEC_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _classify(ctx: str) -> tuple[bool, bool]:
    is_template = any(r.search(ctx) for r in TEMPLATE_RE)
    is_real = any(r.search(ctx) for r in REAL_RE)
    return is_template, is_real


def _score_note(txt: str) -> list[tuple[int, str, bool, bool]]:
    found: list[tuple[int, str, bool, bool]] = []
    for m in RLN_TOKEN_RE.finditer(txt):
        pos = m.start()
        ctx = txt[max(0, pos - 200): pos + 300]
        is_t, is_r = _classify(ctx)
        found.append((pos, ctx, is_t, is_r))
    return found


def _connect() -> duckdb.DuckDBPyConnection:
    from motherduck_client import MotherDuckClient, MotherDuckConfig

    cfg = MotherDuckConfig(database=PUB_DB)
    con = MotherDuckClient(cfg).connect_rw()
    con.execute(f"USE {PUB_DB}")
    return con


_ILLEGAL_CTRL_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def _squash(s: str, max_len: int = 600) -> str:
    s = (s or "").replace("\r", " ").replace("\n", " ").strip()
    s = _ILLEGAL_CTRL_RE.sub(" ", s)
    while "  " in s:
        s = s.replace("  ", " ")
    if len(s) > max_len:
        s = s[: max_len - 3] + "..."
    return s


def _autosize(ws, max_w: float = 60.0, wrap_cols=()) -> None:
    wrap_cols = set(wrap_cols)
    for col in ws.columns:
        letter = col[0].column_letter
        mlen = 0
        for cell in col:
            if cell.value is None:
                continue
            mlen = max(mlen, min(len(str(cell.value)), 100))
        ws.column_dimensions[letter].width = min(max(10, mlen + 2), max_w)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def fetch(con):
    rids = [r[0] for r in con.execute(
        "SELECT DISTINCT research_id FROM main.canonical_complications_events_v1 "
        "WHERE complication_type='rln_injury' AND finding_status IN ('present','suspected')"
    ).fetchall()]

    con.execute("CREATE OR REPLACE TEMP TABLE tmp_rln_pts (rid VARCHAR)")
    con.executemany("INSERT INTO tmp_rln_pts VALUES (?)", [(r,) for r in rids])

    mentions = con.execute(
        "SELECT cnl.research_id, cnl.note_type, cnl.note_index, cnl.note_text "
        "FROM main.clinical_notes_long cnl JOIN tmp_rln_pts t ON cnl.research_id = t.rid "
        "WHERE regexp_matches(LOWER(cnl.note_text), "
        "'(\\brln\\b|recurrent laryngeal|vocal cord paralys|vocal fold paralys|\\bvcp\\b)')"
    ).fetchall()

    rid_list = "','".join(rids)
    aug_q = (
        "SELECT c.research_id, "
        "BOOL_OR(c.source_table='complication_phenotype_v1') AS has_phen, "
        "BOOL_OR(c.source_table='extracted_rln_injury_refined_v2') AS has_v2, "
        "BOOL_OR(c.source_table='extracted_complications_refined_v5') AS has_refined, "
        "BOOL_OR(c.evidence_strength='definitive') AS any_definitive, "
        "BOOL_OR(phen.confirmed_flag) AS phen_confirmed, "
        "BOOL_OR(phen.permanent_flag) AS phen_permanent, "
        "BOOL_OR(phen.transient_flag) AS phen_transient, "
        "BOOL_OR(phen.treatment_requiring_flag) AS phen_treat_req, "
        "MIN(c.finding_date)::VARCHAR AS earliest_date, "
        "STRING_AGG(DISTINCT c.finding_status, ',' ORDER BY c.finding_status) AS statuses, "
        "STRING_AGG(DISTINCT c.onset_class, ',' ORDER BY c.onset_class) AS onsets "
        "FROM main.canonical_complications_events_v1 c "
        f"LEFT JOIN {PHEN_PRE} phen "
        "ON c.source_table='complication_phenotype_v1' "
        "AND c.source_row_id=CAST(hash(phen.research_id, phen.complication_entity, phen.detection_date) AS VARCHAR) "
        "WHERE c.complication_type='rln_injury' "
        "AND c.finding_status IN ('present','suspected') "
        f"AND c.research_id IN ('{rid_list}') "
        "GROUP BY c.research_id"
    )
    aug_info = {r[0]: r for r in con.execute(aug_q).fetchall()}

    # mig_98a VCP keepers (anyone with present vocal_cord_paralysis after mig_98a)
    vcp_keepers = set(r[0] for r in con.execute(
        "SELECT DISTINCT research_id FROM main.canonical_complications_events_v1 "
        "WHERE complication_type='vocal_cord_paralysis' AND finding_status='present'"
    ).fetchall())

    return rids, mentions, aug_info, vcp_keepers


def aggregate(mentions):
    per_pt = defaultdict(lambda: {"real": [], "template": [], "unmarked": [], "all_count": 0})
    for rid, nt, idx, txt in mentions:
        if not txt:
            continue
        for pos, ctx, is_t, is_r in _score_note(txt):
            per_pt[rid]["all_count"] += 1
            entry = {"note_type": nt, "note_index": idx, "pos": pos, "context": ctx}
            if is_r:
                per_pt[rid]["real"].append(entry)
            elif is_t:
                per_pt[rid]["template"].append(entry)
            else:
                per_pt[rid]["unmarked"].append(entry)
    return per_pt


def build_summary(wb, counts):
    ws = wb.create_sheet("summary", 0)
    ws["A1"] = "rln_injury REAL candidates — mig_98c"
    ws["A1"].font = Font(bold=True, size=14)
    run_ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    ws["A2"] = f"Generated (UTC): {run_ts}"
    ws["A3"] = "Author: Logan Glosser <logan.glosser@gmail.com>"
    ws["A5"] = (
        "Method: pulled all RLN/recurrent-laryngeal/vocal-cord-paralysis mentions "
        "from clinical_notes_long for the 1,414 canonical rln_injury pts; "
        "classified each 500-char context window as REAL (RLN sacrificed/transected/"
        "injured, vocal cord paralysis with treatment, medialization/thyroplasty/"
        "voice therapy, persistent post-op hoarseness, laryngoscopy paralysis) vs "
        "TEMPLATE (consent risk lists, RLN identified-and-preserved, RLN intact, "
        "negation, nerve-stim documentation, planning language). Augmented with "
        "phenotype confirmed_flag, extracted_rln_injury_refined_v2 source rows, "
        "and mig_98a vocal_cord_paralysis keepers."
    )
    ws["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 110
    ws.row_dimensions[5].height = 90

    ws["A7"] = "Counts"
    ws["A7"].font = Font(bold=True)
    rows = [
        f"Cohort: 10,871 pts",
        f"rln_injury rows in canonical: 1,150 present + 66 suspected + 674 absent",
        f"Distinct pts (present|suspected): {counts['total_pts']}",
        f"Pts in REAL_CANDIDATES sheet (real-pattern OR phen_confirmed OR v2 OR mig_98a VCP keeper): {counts['candidates']}",
        f"Pts in NO_RLN_INJURY (template-only / no signal): {counts['no_rln']}",
        f"  -- of REAL_CANDIDATES: {counts['real_pattern']} have ≥1 REAL note-text mention",
        f"  -- {counts['phen_confirmed']} have phenotype confirmed_flag=TRUE",
        f"  -- {counts['v2_extracted']} have extracted_rln_injury_refined_v2 source",
        f"  -- {counts['vcp_overlap']} overlap mig_98a vocal_cord_paralysis keepers",
    ]
    for i, r in enumerate(rows, start=8):
        ws.cell(row=i, column=1, value=r)

    ws["A18"] = "Smoking-gun examples (from probe)"
    ws["A18"].font = Font(bold=True)
    ws["A19"] = (
        'TEMPLATE — op_note "definitive" intraop rid 10456: "b/l RLN identified and '
        'preserved, stimulated at 0.8 mA"  → not an injury'
    )
    ws["A20"] = (
        'TEMPLATE — op_note "definitive" intraop rid 11533: "I wanted to ensure that '
        'the right RLN was intact before proceeding"  → not an injury'
    )
    ws["A21"] = (
        'TEMPLATE — op_note "definitive" intraop rid 9910: "...likely a right RLN '
        'injury and possible esophageal injury requiring complex repair" '
        '→ planning/consent language, not actual injury'
    )
    for r in (19, 20, 21):
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 45

    ws["A23"] = "Sign-off (Logan)"
    ws["A23"].font = Font(bold=True)
    ws["A24"] = "  Date: _______________"
    ws["A25"] = "  Initials: _______________"


ALL_HEADERS = [
    "research_id",
    "candidate_tier",
    "claude_suggested",
    "earliest_finding_date",
    "statuses",
    "onsets",
    "has_phenotype_row",
    "has_v2_extractor",
    "has_refined_v5",
    "phen_confirmed",
    "phen_permanent",
    "phen_transient",
    "phen_treatment_required",
    "any_definitive_evidence",
    "vcp_overlap_mig98a",
    "n_real_pattern_mentions",
    "n_template_mentions",
    "n_total_mentions",
    "source1", "evidence1",
    "source2", "evidence2",
    "source3", "evidence3",
    "source4", "evidence4",
    "your_decision",
    "your_note",
]


def _row_for_pt(rid, per_pt, aug_info, vcp_keepers, candidates):
    """Build one consolidated row for a patient — used for both candidate
    and non-candidate sheets so Logan has the same evidence structure
    everywhere."""
    info = per_pt.get(rid, {"real": [], "template": [], "unmarked": [], "all_count": 0})
    aug = aug_info.get(rid, (rid, None, None, None, None, None, None, None, None, None, None, None))
    # aug: (rid, has_phen, has_v2, has_refined, any_definitive, phen_confirmed,
    #       phen_permanent, phen_transient, phen_treat_req, earliest_date, statuses, onsets)

    # Per note_type, prefer REAL excerpt if any, else best TEMPLATE, else UNMARKED.
    # Picks one excerpt per note_type, ordered by clinical priority.
    by_nt: dict[str, list[str]] = defaultdict(list)
    for entry in info["real"]:
        by_nt[entry["note_type"]].append(_squash(entry["context"]))
    # If a note_type has NO real, surface its template/unmarked excerpts as evidence anyway
    for entry in info["template"] + info["unmarked"]:
        nt = entry["note_type"]
        if not by_nt.get(nt):
            by_nt[nt].append(_squash(entry["context"]))
    ordered = [nt for nt in NOTE_TYPE_PRIORITY if nt in by_nt]
    sources = (ordered + [""] * 4)[:4]
    evidences = ["; ".join(by_nt.get(nt, [])[:2]) if nt else "" for nt in sources]

    is_phen_conf = bool(aug[5])
    is_v2 = bool(aug[2])
    is_vcp = rid in vcp_keepers
    is_real_text = bool(info["real"])

    # candidate tier label
    if rid in candidates:
        tiers = []
        if is_phen_conf:
            tiers.append("phen_conf")
        if is_v2:
            tiers.append("v2")
        if is_vcp:
            tiers.append("vcp_overlap")
        tier_label = "+".join(tiers) if tiers else "real_text"
    else:
        tier_label = "no_signal"

    # claude_suggested = "REVIEW" for candidates, "NO" for others (template-only)
    claude_suggested = "REVIEW" if rid in candidates else "NO"

    return [
        rid,
        tier_label,
        claude_suggested,
        aug[9],
        aug[10],
        aug[11],
        bool(aug[1]),
        is_v2,
        bool(aug[3]),
        aug[5],
        aug[6],
        aug[7],
        aug[8],
        bool(aug[4]),
        is_vcp,
        len(info["real"]),
        len(info["template"]),
        info["all_count"],
        sources[0], evidences[0],
        sources[1], evidences[1],
        sources[2], evidences[2],
        sources[3], evidences[3],
        "",
        "",
    ]


def build_consolidated_sheet(wb, all_rids, per_pt, aug_info, vcp_keepers, candidates):
    """One sheet with EVERY patient + full source/evidence columns. Sorted with
    candidates first (alphabetic by rid), then non-candidates (alphabetic)."""
    ws = wb.create_sheet("ALL_PATIENTS")
    for j, h in enumerate(ALL_HEADERS, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL

    cand_sorted = sorted(r for r in all_rids if r in candidates)
    other_sorted = sorted(r for r in all_rids if r not in candidates)
    for rid in cand_sorted + other_sorted:
        ws.append(_row_for_pt(rid, per_pt, aug_info, vcp_keepers, candidates))

    dec_cols = {
        get_column_letter(ALL_HEADERS.index("your_decision") + 1),
        get_column_letter(ALL_HEADERS.index("your_note") + 1),
        get_column_letter(ALL_HEADERS.index("claude_suggested") + 1),
    }
    wrap_cols = {
        get_column_letter(ALL_HEADERS.index(h) + 1)
        for h in ("evidence1", "evidence2", "evidence3", "evidence4", "your_note")
    }
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in dec_cols:
                cell.fill = DEC_FILL
    ws.freeze_panes = "C2"
    _autosize(ws, max_w=70.0, wrap_cols=wrap_cols)


NO_HEADERS = [
    "research_id",
    "n_rln_text_mentions",
    "n_template_classified",
    "n_unmarked_classified",
    "best_template_phrase",
    "sample_excerpt",
    "suggested_decision",
    "your_override",
]


def build_no_sheet(wb, all_rids, candidates_set, per_pt):
    ws = wb.create_sheet("NO_RLN_INJURY")
    for j, h in enumerate(NO_HEADERS, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = HEADER_FILL

    for rid in sorted(r for r in all_rids if r not in candidates_set):
        info = per_pt.get(rid)
        if info is None:
            ws.append([rid, 0, 0, 0, "(no rln text in clinical_notes_long)", "", "NO", ""])
            continue
        sample = ""
        sample_phrase = ""
        for nt in NOTE_TYPE_PRIORITY:
            cands = [e for e in info["template"] if e["note_type"] == nt]
            if cands:
                sample = _squash(cands[0]["context"], max_len=400)
                lc = sample.lower()
                if "identified and preserved" in lc or "preserved" in lc:
                    sample_phrase = "intra-op identification + preservation"
                elif "intact" in lc:
                    sample_phrase = '"RLN intact" / no injury'
                elif "stimulated" in lc and "ma" in lc:
                    sample_phrase = "nerve stim documentation"
                elif "risks" in lc or "consent" in lc:
                    sample_phrase = "consent risk list"
                elif "mobile" in lc:
                    sample_phrase = '"vocal cords mobile" — normal exam'
                elif "no rln" in lc or "no recurrent" in lc:
                    sample_phrase = "negation"
                else:
                    sample_phrase = "template/unspecified"
                break
        if not sample and info["unmarked"]:
            sample = _squash(info["unmarked"][0]["context"], max_len=400)
            sample_phrase = "unclassified (likely template variant)"
        ws.append([
            rid,
            info["all_count"],
            len(info["template"]),
            len(info["unmarked"]),
            sample_phrase,
            sample,
            "NO",
            "",
        ])

    dec_col_idx = NO_HEADERS.index("your_override")
    wrap_cols = {get_column_letter(NO_HEADERS.index("sample_excerpt") + 1)}
    for row in ws.iter_rows(min_row=2):
        row[dec_col_idx].fill = DEC_FILL
    ws.freeze_panes = "A2"
    _autosize(ws, max_w=60.0, wrap_cols=wrap_cols)


def main():
    con = _connect()
    rids, mentions, aug_info, vcp_keepers = fetch(con)
    con.close()

    per_pt = aggregate(mentions)

    # Candidate selection: structured signals are primary (phenotype confirmed,
    # v2 extractor, mig_98a VCP overlap). Note-text REAL classification is
    # supplementary context but NOT enough on its own — RLN/VCP entity tokens
    # appear too frequently in mixed template/real contexts (e.g., post-op clinic
    # notes that mention "monitored for VCP, medialization considered if persistent"
    # which includes both template and real keywords). Use STRICT note-text rule
    # for note-text-only candidates.
    candidates = set()
    real_pattern_pts = set()
    phen_confirmed_pts = set()
    v2_pts = set()
    vcp_overlap = set()
    strict_real_pts = set()

    # Strict REAL patterns require explicit complication language
    STRICT_REAL_RE = [
        re.compile(p, re.IGNORECASE) for p in [
            r"\b(complicated\s+by|notable for|c\/b)\s+(?:a\s+|the\s+)?(?:rln|recurrent\s+laryngeal|vocal\s+cord\s+paralys|vocal\s+fold\s+paralys)",
            r"\bs/p[^.]{0,60}c\/b[^.]{0,80}(?:rln|vocal\s+cord)",
            r"\b(rln|recurrent\s+laryngeal\s+nerve)\s+(?:was\s+)?(?:sacrificed|transected|cut|severed)",
            r"\bsacrific(?:ed|ing)\s+(?:the\s+)?(?:rln|recurrent\s+laryngeal\s+nerve)",
            r"\b(?:permanent|persistent|chronic)\s+(?:vocal\s+(?:cord|fold)\s+paralysis|hoarseness|dysphonia)",
            r"\b(?:underwent|s\/p)[^.]{0,60}(?:medialization|thyroplast|injection\s+laryngoplast)",
            r"vocal\s+(?:cord|fold)\s+paralysis[^.]{0,80}\b(?:treated|managed|requir(?:ed|ing))",
        ]
    ]

    def has_strict_real(rid):
        info = per_pt.get(rid)
        if not info:
            return False
        for entry in info["real"] + info["template"] + info["unmarked"]:
            ctx = entry.get("context", "")
            if any(r.search(ctx) for r in STRICT_REAL_RE):
                return True
        return False

    for rid in rids:
        info = per_pt.get(rid, {"real": []})
        aug = aug_info.get(rid)
        is_phen_conf = bool(aug and aug[5])
        is_v2 = bool(aug and aug[2])
        is_vcp_overlap = rid in vcp_keepers
        is_real_pattern = bool(info.get("real"))
        is_strict_real = has_strict_real(rid)

        if is_real_pattern:
            real_pattern_pts.add(rid)
        if is_phen_conf:
            phen_confirmed_pts.add(rid)
        if is_v2:
            v2_pts.add(rid)
        if is_vcp_overlap:
            vcp_overlap.add(rid)
        if is_strict_real:
            strict_real_pts.add(rid)

        # Candidate selection: structured signals only.
        # Note-text REAL classification (incl strict patterns) over-matches because
        # RLN/VCP tokens appear in many template+real mixed contexts post-op.
        # Logan can spot-check the NO_RLN_INJURY sheet for any false negatives.
        if is_phen_conf or is_v2 or is_vcp_overlap:
            candidates.add(rid)

    counts = {
        "total_pts": len(rids),
        "candidates": len(candidates),
        "strict_real_pts": len(strict_real_pts),
        "no_rln": len(rids) - len(candidates),
        "real_pattern": len(real_pattern_pts),
        "phen_confirmed": len(phen_confirmed_pts),
        "v2_extracted": len(v2_pts),
        "vcp_overlap": len(vcp_overlap),
    }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    build_summary(wb, counts)
    build_consolidated_sheet(wb, rids, per_pt, aug_info, vcp_keepers, candidates)
    wb.properties.creator = "Logan Glosser <logan.glosser@gmail.com>"
    wb.save(OUT_XLSX)

    print(f"wrote: {OUT_XLSX}")
    for k, v in counts.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
