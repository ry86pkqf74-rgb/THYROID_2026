#!/usr/bin/env python3
"""Lightweight workbook smoke check — does NOT run Excel calculation engine."""

from __future__ import annotations

import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook  # type: ignore[import-untyped]
except ImportError as e:
    raise SystemExit(
        "openpyxl required: pip install openpyxl (part of repo requirements)."
    ) from e

_ERR_PAT = re.compile(r"#(DIV/0!|REF!|VALUE!|NUM!|NULL!|NAME?|NA)")


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("usage: recalc.py <workbook.xlsx>")
    path = Path(sys.argv[1])
    if not path.exists():
        raise SystemExit(f"not found: {path}")

    bad: list[tuple[str, str, object]] = []
    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and (_ERR_PAT.search(v) or v.startswith("#")):
                    bad.append((ws.title, c.coordinate, v))

    if bad:
        for w, coord, raw in bad:
            print(f"ERROR_CELL\t{path}\t{w}\t{coord}\t{raw}")
        raise SystemExit(f"found {len(bad)} error-like cached formula values")

    print(f"[recalc-scan] OK: {path}")


if __name__ == "__main__":
    main()
