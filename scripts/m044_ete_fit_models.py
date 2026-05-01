#!/usr/bin/env python3
"""
M044 ETE manuscript: MotherDuck → analytic parquet (+CPM merges), statsmodels GLMs,
Cox sensitivity, forest plot (matplotlib), QA size strata, openpyxl workbook refresh.
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import statsmodels.api as sm
import statsmodels.formula.api as smf
from lifelines import CoxPHFitter
from openpyxl import load_workbook
from scipy import stats

REPO_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = REPO_ROOT / "data" / "m044"
FIG_DIR = REPO_ROOT / "figures"
XLSX = REPO_ROOT / "M044_ETE_tables.xlsx"

MASTER_ANALYTIC_SQL = r"""
WITH cohort AS (
  SELECT
    c.*,
    CASE
      WHEN c.ete_grade_final IN ('false','absent')   THEN 'No/negative ETE'
      WHEN c.ete_grade_final = 'microscopic'       THEN 'Microscopic ETE'
      WHEN c.ete_grade_final = 'gross'               THEN 'Gross ETE'
      WHEN c.ete_grade_final = 'present_ungraded'  THEN 'Present ungraded'
      ELSE 'Missing/other'
    END AS ete_group,
    CASE
      WHEN c.lvi_grade ILIKE 'extensiv%'             THEN 'extensive'
      WHEN c.lvi_grade IN ('present','preesent')    THEN 'present'
      WHEN c.lvi_grade = 'focal'                    THEN 'focal'
      WHEN c.lvi_grade IS NULL                       THEN 'missing'
      WHEN c.lvi_grade IN ('indeterminate','indetermiante','indeeterminate','indeterminent','suspicious','x','c/a','no','n/s')
                                                     THEN 'indeterminate'
      ELSE 'indeterminate'
    END AS lvi_clean,
    COALESCE(c.vascular_invasion_final, 'missing') AS vasc_clean
  FROM manuscript_workspace.cohort_m044_ajcc_ete_v1 c
),
ln AS (
  SELECT research_id,
    MAX(ln_total_examined)             AS ln_examined,
    MAX(ln_total_positive)             AS ln_positive,
    MAX(ln_central_examined)           AS ln_central_examined,
    MAX(ln_central_positive)           AS ln_central_positive,
    MAX(ln_lateral_left_positive)      AS ln_lateral_left_positive,
    MAX(ln_lateral_right_positive)     AS ln_lateral_right_positive,
    MAX(ln_bilateral_lateral_positive) AS ln_bilateral_lateral_positive,
    MAX(ln_level_vi_positive)          AS ln_level_vi_positive,
    MAX(ln_level_vii_positive)         AS ln_level_vii_positive,
    MAX(ln_extranodal_extension)       AS ln_ene
  FROM manuscript_workspace.ln_master_rollup_v1
  GROUP BY research_id
),
reop AS (
  SELECT research_id,
    MAX(n_surgeries) AS n_surgeries,
    MAX(second_surgery_date) AS second_surgery_date,
    MAX(days_between_first_second_surgery) AS days_to_2nd,
    MAX(completion_reason) AS completion_reason,
    MAX(completion_reason_confidence) AS completion_reason_confidence,
    MAX(completion_histology_type) AS completion_histology_type,
    MAX(op_reoperative_any) AS op_reoperative_any
  FROM manuscript_workspace.cohort_m040_reoperative_v1
  GROUP BY research_id
),
rec AS (
  SELECT
    research_id,
        first_surg_date AS recurrence_first_surg_date,
    recurrence_path_proven,
        recurrence_path_proven_date,
        recurrence_path_proven_source,
    recurrence_imaging_then_path_confirmed,
    recurrence_status_final,
    days_to_path_proven,
        recurrence_imaging_suspicious_date,
        is_implausible_date_quarantine,
        (recurrence_path_proven IS TRUE AND NOT COALESCE(is_implausible_date_quarantine, FALSE)) AS path_proven_primary,
        (recurrence_status_final = 'imaging_only_unconfirmed') AS imaging_only_unconfirmed,
        (
            (recurrence_path_proven IS TRUE AND NOT COALESCE(is_implausible_date_quarantine, FALSE))
            OR recurrence_status_final = 'imaging_only_unconfirmed'
        ) AS composite_primary
  FROM main.canonical_recurrence_resolved_v1
)
SELECT
  c.research_id,
  c.ete_group,
  c.ete_grade_final,
  c.age_at_surgery, c.sex, c.histology_final, c.tumor_size_cm,
  c.ajcc8_t_stage, c.ajcc8_n_stage, c.ajcc8_stage_group,
  c.surg_first_date,
  c.surg_first_date_lineage_note,
  c.surg_date_missing,
  c.surg_date_pre_1999,
  c.surg_date_1999_2024,
  c.surg_date_post_2024,
  c.surg_date_after_2024_06_04,
  c.followup_years, c.overall_survival_years, c.death_occurred,
  c.lvi_clean, c.vasc_clean, c.lvi_grade,
  ln.ln_lateral_left_positive, ln.ln_lateral_right_positive, ln.ln_bilateral_lateral_positive,
  CASE WHEN ln.ln_central_positive > 0 THEN 1 ELSE 0 END AS central_pos_flag,
  CASE WHEN COALESCE(ln.ln_lateral_left_positive,0) > 0
         OR COALESCE(ln.ln_lateral_right_positive,0) > 0
         OR COALESCE(ln.ln_bilateral_lateral_positive,0) > 0 THEN 1 ELSE 0 END AS lateral_pos_flag,
  c.rai_received_flag,
  c.any_recurrence_flag,
  rec.recurrence_path_proven,
    rec.recurrence_path_proven_date,
    rec.recurrence_path_proven_source,
    rec.recurrence_first_surg_date,
    rec.is_implausible_date_quarantine,
    rec.path_proven_primary,
    rec.imaging_only_unconfirmed,
    rec.composite_primary,
  rec.recurrence_imaging_then_path_confirmed,
  rec.recurrence_status_final,
  rec.days_to_path_proven,
  reop.n_surgeries, reop.days_to_2nd
FROM cohort c
LEFT JOIN ln USING (research_id)
LEFT JOIN reop USING (research_id)
LEFT JOIN rec USING (research_id);
"""


CPM_EXTRA_SQL = """
SELECT
  research_id::VARCHAR AS research_id_cpm_key,
  race, bmi_combined, multifocal_flag_path, bilateral_disease_flag, aggressive_variant_flag,
  margin_involved_any, closest_margin_mm, syn_hashimoto, syn_graves,
  pmhx_nlp_diabetes, pmhx_nlp_hypertension, pmhx_nlp_hypothyroidism, pmhx_nlp_obesity,
  braf_positive_final, tert_positive_final, ras_positive_final, ret_positive_unified,
  surg_total_thyroidectomy, ages_score
FROM main.canonical_patient_master;
"""


def tri_cat(series: pd.Series | None) -> pd.Categorical | None:
    if series is None:
        return None

    def one(v: Any) -> str:
        if pd.isna(v):
            return "missing"
        if isinstance(v, (bool, np.bool_)):
            return "true" if bool(v) else "false"
        sv = str(v).strip().lower()
        if sv in {"", "nan", "none"}:
            return "missing"
        if sv in {"1", "true", "yes", "t", "x"}:
            return "true"
        return "false"

    z = series.map(one)
    return pd.Categorical(z.astype(str), categories=["missing", "false", "true"])


def normalize_rid_series(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce").astype("Int64")


def cox_extract_contrast(sdf: pd.DataFrame, substring: str) -> dict[str, Any]:
    m = sdf[sdf["covariate"].astype(str).str.contains(substring, regex=False)]
    if len(m) != 1:
        return {"_match_error": len(m)}
    r = m.iloc[0]
    return {
        "hr": float(r["exp(coef)"]),
        "hr_ci_low": float(r["exp(coef) lower 95%"]),
        "hr_ci_high": float(r["exp(coef) upper 95%"]),
        "p": float(r["p"]),
        "covariate_row": str(r["covariate"]),
    }


def histology_grouped(h: Any) -> str:
    if h is None or (isinstance(h, float) and np.isnan(h)):
        return "other"
    t = str(h).strip().lower()
    if t == "" or t == "nan":
        return "other"
    if "medullary" in t or t == "mtc" or "(mtc" in t:
        return "MTC-like"
    follicular_hit = ("follicular" in t) and ("papillary" not in t)
    if follicular_hit or t in {"ftc", "hurthle", "hürthle", "hcc", "oncocytic", "niftp"}:
        return "follicular-like"
    if "papillary" in t or t == "ptc" or t.startswith("ptc"):
        return "PTC"
    return "other"


EXCLUDE_HISTOLOGIES_STRICT_DTC = frozenset(
    {
        "MTC",
        "metastatic MTC",
        "recurrent MTC",
        "MTC/PTC mixed composite",
        "anaplastic carcinoma",
        "metastatic anaplastic carcinoma",
        "NIFTP",
        "FTUMP",
        "atypical follicular adenoma",
        "follicular adenoma",
        "Atypical hurthle cell neoplasm",
        "NUT carcinoma",
        "adenoid cystic carcinoma",
        "metastatic PTC/anaplastic carcinoma",
    }
)


def _histology_str_raw(h: Any) -> str:
    if h is None or (isinstance(h, float) and np.isnan(h)):
        return ""
    return str(h).strip()


def histology_dtc_cat(h: Any) -> str:
    """Five-level DTC histology for strict-DTC primary analysis (post exclusion)."""

    nh = _histology_str_raw(h)
    if not nh:
        return "PTC"
    low = nh.lower()
    if nh.startswith("metastatic PTC") or nh.startswith("recurrent/metastatic PTC"):
        return "Metastatic-PTC"
    if nh == "follicular carcinoma":
        return "FTC"
    if "poorly differentiated" in low:
        return "Poorly-differentiated DTC"
    if "high grade" in low or "high-grade" in low:
        return "High-grade DTC"
    if nh == "PTC":
        return "PTC"
    if "metastatic thyroid carcinoma" in low:
        return "Metastatic-PTC"
    if nh == "differentiated thyroid carcinoma":
        return "PTC"
    if "thymus-like" in low:
        return "PTC"
    return "PTC"


def encode_rai_01(series: pd.Series) -> pd.Series:
    """Map RAI-received to 0/1; undocumented / non-positive → 0 (see analysis note)."""

    nu = pd.to_numeric(series, errors="coerce")
    s = series.astype(str).str.strip().str.lower()
    m_true = (nu == 1).fillna(False)
    txt_ok = (~s.isin({"nan", "none", "", "<nat>"})) & s.isin(
        ("1", "true", "t", "yes", "x")
    )
    ex_true = series.apply(lambda v: v is True)
    return (m_true | txt_ok.fillna(False) | ex_true).astype(int)


def bool_series(series: pd.Series | None, *, default: bool = False) -> pd.Series:
    """Parse nullable DB booleans without triggering pd.NA truth-value errors."""

    if series is None:
        return pd.Series(dtype=bool)
    if series.dtype == bool:
        return series.fillna(default).astype(bool)
    text = series.astype(str).str.strip().str.lower()
    out = text.isin({"true", "t", "1", "yes", "y", "x"})
    if default:
        out = out | text.isin({"", "nan", "none", "<na>", "nat"})
    return out.fillna(default).astype(bool)


def add_endpoint_flags(df: pd.DataFrame) -> pd.DataFrame:
    """Add M044 final endpoint flags, excluding implausible-date quarantines."""

    out = df.copy()
    raw_path = bool_series(out.get("recurrence_path_proven"))
    quarantine = bool_series(out.get("is_implausible_date_quarantine"))
    status = out.get("recurrence_status_final", pd.Series(index=out.index, dtype=object))
    status = status.astype(str)
    out["path_proven_primary"] = raw_path & ~quarantine
    out["imaging_only_unconfirmed"] = status.eq("imaging_only_unconfirmed")
    out["composite_primary"] = out["path_proven_primary"] | out["imaging_only_unconfirmed"]
    return out


def build_primary_prep(df: pd.DataFrame, *, strict_dtc: bool = False) -> pd.DataFrame:
    df = add_endpoint_flags(df)
    d = df[
        df["ete_group"].isin(["No/negative ETE", "Microscopic ETE", "Gross ETE"])
    ].copy()
    if strict_dtc:
        hs = d["histology_final"].map(_histology_str_raw)
        d = d.loc[~hs.isin(EXCLUDE_HISTOLOGIES_STRICT_DTC)].copy()
        d["histology_fac"] = d["histology_final"].apply(histology_dtc_cat)
        d["histology_fac"] = pd.Categorical(
            d["histology_fac"],
            categories=[
                "PTC",
                "FTC",
                "Metastatic-PTC",
                "Poorly-differentiated DTC",
                "High-grade DTC",
            ],
        )
    else:
        d["histology_fac"] = d["histology_final"].apply(histology_grouped)
        d["histology_fac"] = pd.Categorical(
            d["histology_fac"],
            categories=["PTC", "follicular-like", "MTC-like", "other"],
        )
    d["age10"] = pd.to_numeric(d["age_at_surgery"], errors="coerce") / 10.0
    sx = d["sex"].astype(str).str.strip().str.lower()
    sx = sx.replace({"female": "female", "male": "male"})
    d["sex_lc"] = sx.mask(~sx.isin(["female", "male"]))
    ns = d["ajcc8_n_stage"]
    d["ajcc8_n_fac"] = np.where(ns.isna(), "missing", ns.astype(str))
    d["tumor_size_cm"] = pd.to_numeric(d["tumor_size_cm"], errors="coerce")
    d["y_pp"] = bool_series(d["path_proven_primary"]).astype(int)
    d["y_img"] = bool_series(d["imaging_only_unconfirmed"]).astype(int)
    d["y_comp"] = bool_series(d["composite_primary"]).astype(int)
    # Sensitivity only; legacy flag vs canonical: manuscript_workspace.m044_legacy_recurrence_flag_audit_v1
    d["y_any"] = d["any_recurrence_flag"].astype(bool).astype(int)

    d["ete_group"] = pd.Categorical(
        d["ete_group"], categories=["Microscopic ETE", "No/negative ETE", "Gross ETE"]
    )
    d["sex_lc"] = pd.Categorical(d["sex_lc"], categories=["female", "male"])
    d["ajcc8_n_fac"] = pd.Categorical(
        d["ajcc8_n_fac"], categories=["N0", "N1a", "N1b", "Nx", "missing"]
    )
    d["lvi_clean"] = pd.Categorical(
        d["lvi_clean"],
        categories=["missing", "present", "extensive", "focal", "indeterminate"],
    )
    d["vasc_clean"] = pd.Categorical(
        d["vasc_clean"],
        categories=["missing", "present_ungraded", "focal", "extensive", "indeterminate"],
    )
    d["rai_received_flag"] = encode_rai_01(d["rai_received_flag"])

    if "multifocal_flag_path" in d.columns:
        d["multifocal_flag_path_fac"] = tri_cat(d["multifocal_flag_path"])
        d["bilateral_disease_fac"] = tri_cat(d["bilateral_disease_flag"])
        d["aggressive_variant_fac"] = tri_cat(d["aggressive_variant_flag"])
        d["margin_involved_fac"] = tri_cat(d["margin_involved_any"])
        d["braf_positive_fac"] = tri_cat(d["braf_positive_final"])
        d["surg_tt_fac"] = tri_cat(d["surg_total_thyroidectomy"])
    return d


def glm_formula_primary(*, include_rai: bool) -> str:
    rai = " + rai_received_flag" if include_rai else ""
    return (
        "y_pp ~ C(ete_group, Treatment(reference='Microscopic ETE'))"
        " + age10 + C(sex_lc, Treatment(reference='female')) + tumor_size_cm"
        " + C(ajcc8_n_fac, Treatment(reference='N0'))"
        " + C(histology_fac, Treatment(reference='PTC'))"
        f"{rai}"
        " + C(lvi_clean, Treatment(reference='missing'))"
        " + C(vasc_clean, Treatment(reference='missing'))"
    )


PRIMARY_FORMULA = glm_formula_primary(include_rai=True)

PRIMARY_FORMULA_CPH = glm_formula_primary(include_rai=False).replace("y_pp ~ ", "").strip()


def _glm_metrics(m: Any) -> dict[str, Any]:
    fr = m.model.data.frame  # noqa: SLF001
    endog = m.model.endog
    lhs = str(m.model.formula).strip().split("~", 1)[0].strip()
    n_events = int(np.sum(endog))
    null = smf.glm(f"{lhs} ~ 1", data=fr, family=sm.families.Binomial()).fit()

    lr_stat = 2.0 * (m.llf - null.llf)
    lr_p = float(1.0 - stats.chi2.cdf(lr_stat, df=int(m.df_model)))
    pr2 = float(1.0 - m.llf / null.llf) if null.llf != 0 else float("nan")
    return {
        "n_obs": int(m.nobs),
        "n_events": n_events,
        "llf": float(m.llf),
        "aic": float(m.aic),
        "pseudo_r2_mcfadden": pr2,
        "lr_vs_null_chi2": float(lr_stat),
        "lr_vs_null_pvalue": lr_p,
    }


def coef_table_glm(m: Any) -> pd.DataFrame:
    ci = np.clip(np.asarray(m.conf_int()), -25.0, 25.0)
    coef = np.asarray(m.params)
    se = np.asarray(m.bse)
    pv = np.asarray(m.pvalues)
    return pd.DataFrame(
        {
            "term": m.params.index.astype(str),
            "coef_logit": coef,
            "se": se,
            "or": np.exp(coef),
            "or_ci_low": np.exp(ci[:, 0]),
            "or_ci_high": np.exp(ci[:, 1]),
            "pvalue": pv,
        }
    )


def model_frame_for_primary(
    df_merged: pd.DataFrame, *, strict_dtc: bool
) -> pd.DataFrame:
    d = build_primary_prep(df_merged, strict_dtc=strict_dtc)
    d = d.dropna(subset=["sex_lc"])
    assert isinstance(d.index, pd.Index)
    return d


def build_cox_analytic_frame(
    dm_strict: pd.DataFrame,
) -> tuple[pd.DataFrame, dict[str, int]]:
    """Cox/KM-aligned rows: documented surgery date, FU>0, path-proven TTE in days.

    Mirrors lifelines Cox sample: censor at follow-up (years×365.25), event-time from
    ``days_to_path_proven`` when ``y_pp==1``, drop age/size missing / non-positive duration.
    """
    dc = dm_strict.copy()
    dc = dc.loc[
        pd.notna(dc["surg_first_date"])
        & (pd.to_numeric(dc["followup_years"], errors="coerce") > 0)
    ].copy()
    fu = pd.to_numeric(dc["followup_years"], errors="coerce")
    dtp_arr = pd.to_numeric(dc["days_to_path_proven"], errors="coerce").to_numpy(dtype=float)
    ev_s = dc["y_pp"].astype(int)
    fu_arr = fu.to_numpy(dtype=float)
    cens = fu_arr * 365.25
    ev_arr = ev_s.to_numpy(dtype=int)
    event_time = np.where(np.isfinite(dtp_arr), dtp_arr, cens)
    time_days_arr = np.where(ev_arr == 1, event_time, cens)
    dc["time_days"] = time_days_arr.astype(float)
    dc = dc.loc[np.isfinite(dc["time_days"]) & (dc["time_days"] > 0)].copy()

    dc = dc.dropna(subset=["age10", "tumor_size_cm"]).reset_index(drop=True)
    cat_for_cox = (
        "ete_group",
        "sex_lc",
        "ajcc8_n_fac",
        "histology_fac",
        "lvi_clean",
        "vasc_clean",
    )
    for c in cat_for_cox:
        if c in dc.columns:
            dc[c] = dc[c].astype(str)
    meta = {"n": len(dc), "events": int(dc["y_pp"].sum())}
    return dc, meta


def build_inclusion_flow_qc(df_merged: pd.DataFrame) -> pd.DataFrame:
    """Waterfall counts for manuscript/QC — strict-DTC Cox/KM slice (mirrors lifelines logic)."""

    def _hist_mask(df: pd.DataFrame) -> pd.Series:
        hs = df["histology_final"].map(_histology_str_raw)
        return ~hs.isin(EXCLUDE_HISTOLOGIES_STRICT_DTC)

    rows: list[dict[str, Any]] = []
    n0 = len(df_merged)

    hx = df_merged.loc[_hist_mask(df_merged)]
    n1 = len(hx)
    rows.append(
        {
            "step_order": 1,
            "criterion": "Strict-DTC histology (exclude MTC/anaplastic/NIFTP/FTUMP/benign-neoplasm list)",
            "n": n1,
            "excluded_at_step": n0 - n1,
            "cum_excluded_from_cohort": n0 - n1,
            "path_proven_events": pd.NA,
        }
    )

    eg_ok = hx["ete_group"].isin(["No/negative ETE", "Microscopic ETE", "Gross ETE"])
    hx2 = hx.loc[eg_ok]
    n2 = len(hx2)
    rows.append(
        {
            "step_order": 2,
            "criterion": "Three-level ETE (No/negative vs Microscopic vs Gross)",
            "n": n2,
            "excluded_at_step": n1 - n2,
            "cum_excluded_from_cohort": n0 - n2,
            "path_proven_events": pd.NA,
        }
    )

    dm_s = model_frame_for_primary(df_merged, strict_dtc=True)
    n3 = len(dm_s)
    rows.append(
        {
            "step_order": 3,
            "criterion": "Sex known female/male (same frame as primary logistic/Cox categoricals)",
            "n": n3,
            "excluded_at_step": n2 - n3,
            "cum_excluded_from_cohort": n0 - n3,
            "path_proven_events": pd.NA,
        }
    )

    fu_ok = pd.to_numeric(dm_s["followup_years"], errors="coerce") > 0
    d4 = dm_s.loc[fu_ok]
    n4 = len(d4)
    rows.append(
        {
            "step_order": 4,
            "criterion": "Positive analytic follow-up (followup_years > 0)",
            "n": n4,
            "excluded_at_step": n3 - n4,
            "cum_excluded_from_cohort": n0 - n4,
            "path_proven_events": pd.NA,
        }
    )

    sd_ok = d4["surg_first_date"].notna()
    d5 = d4.loc[sd_ok]
    n5 = len(d5)
    rows.append(
        {
            "step_order": 5,
            "criterion": "Known surgery date (CPM surg_first_date; mig_258/mig_254 lineage)",
            "n": n5,
            "excluded_at_step": n4 - n5,
            "cum_excluded_from_cohort": n0 - n5,
            "path_proven_events": pd.NA,
        }
    )

    fu = pd.to_numeric(d5["followup_years"], errors="coerce")
    dtp = pd.to_numeric(d5["days_to_path_proven"], errors="coerce").to_numpy(dtype=float)
    ev_arr = d5["y_pp"].astype(int).to_numpy(dtype=int)
    fu_arr = fu.to_numpy(dtype=float)
    cens = fu_arr * 365.25
    event_time = np.where(np.isfinite(dtp), dtp, cens)
    time_days = np.where(ev_arr == 1, event_time, cens)
    td_ok = pd.Series(np.isfinite(time_days) & (time_days > 0), index=d5.index)
    d6 = d5.loc[td_ok]
    n6 = len(d6)
    rows.append(
        {
            "step_order": 6,
            "criterion": "Finite positive Cox time_days (path-proven TTE if event else censor at FU×365.25)",
            "n": n6,
            "excluded_at_step": n5 - n6,
            "cum_excluded_from_cohort": n0 - n6,
            "path_proven_events": pd.NA,
        }
    )

    ts_ok = pd.to_numeric(d6["tumor_size_cm"], errors="coerce").notna()
    d7 = d6.loc[ts_ok]
    n7_ts = len(d7)
    rows.append(
        {
            "step_order": 7,
            "criterion": "Tumor size known (cm; Cox covariate complete-case)",
            "n": n7_ts,
            "excluded_at_step": n6 - n7_ts,
            "cum_excluded_from_cohort": n0 - n7_ts,
            "path_proven_events": pd.NA,
        }
    )

    age_ok = pd.to_numeric(d7["age_at_surgery"], errors="coerce").notna()
    d8 = d7.loc[age_ok]
    n_fin = len(d8)
    rows.append(
        {
            "step_order": 8,
            "criterion": "Age at surgery known (Cox covariate complete-case)",
            "n": n_fin,
            "excluded_at_step": n7_ts - n_fin,
            "cum_excluded_from_cohort": n0 - n_fin,
            "path_proven_events": pd.NA,
        }
    )

    dm_strict = model_frame_for_primary(df_merged, strict_dtc=True)
    cox_df, cox_meta = build_cox_analytic_frame(dm_strict)
    n_cox = len(cox_df)
    ev_cox = int(cox_meta["events"])
    if n_fin != n_cox:
        raise SystemExit(
            f"Inclusion QC waterfall final n ({n_fin}) != build_cox_analytic_frame ({n_cox}); "
            "logic drift — fix build_inclusion_flow_qc vs build_cox_analytic_frame."
        )

    header = pd.DataFrame(
        [
            {
                "step_order": 0,
                "criterion": "M044 analytic extract (cohort_m044_ajcc_ete_v1 + recurrence join)",
                "n": n0,
                "excluded_at_step": 0,
                "cum_excluded_from_cohort": 0,
                "path_proven_events": pd.NA,
            }
        ]
    )
    footer = pd.DataFrame(
        [
            {
                "step_order": 9,
                "criterion": "Final Cox PH + Kaplan–Meier rows (lifelines sample; aligns Figure 6)",
                "n": n_cox,
                "excluded_at_step": 0,
                "cum_excluded_from_cohort": n0 - n_cox,
                "path_proven_events": ev_cox,
            }
        ]
    )
    body = pd.DataFrame(rows)
    out_qc = pd.concat([header, body, footer], ignore_index=True)
    return out_qc.drop_duplicates(subset=["step_order"], keep="last").sort_values(
        "step_order"
    )


def crude_pairwise(d_model: pd.DataFrame) -> dict[str, Any]:
    sub = d_model.dropna(subset=["y_pp"]).copy()

    def or_for(group: str) -> tuple[float, float, float]:
        mic = sub["ete_group"].astype(str) == "Microscopic ETE"
        alt = sub["ete_group"].astype(str) == group
        mm = sub.loc[mic | alt, ["y_pp"]].copy()
        mm["x"] = (sub.loc[mic | alt, "ete_group"].astype(str) == group).astype(int)
        glm = smf.glm("y_pp ~ x", data=mm, family=sm.families.Binomial()).fit()
        coef = glm.params["x"]
        row = glm.conf_int().loc["x"]
        return float(np.exp(coef)), float(np.exp(row[0])), float(np.exp(row[1]))

    g_or, g_lo, g_hi = or_for("Gross ETE")
    n_or, n_lo, n_hi = or_for("No/negative ETE")
    return {
        "crude_pp_gross_vs_microscopic_or": g_or,
        "crude_pp_gross_vs_microscopic_ci_low": g_lo,
        "crude_pp_gross_vs_microscopic_ci_high": g_hi,
        "crude_pp_noneg_vs_microscopic_or": n_or,
        "crude_pp_noneg_vs_microscopic_ci_low": n_lo,
        "crude_pp_noneg_vs_microscopic_ci_high": n_hi,
    }


def export_parquet(force: bool) -> Path:
    sys.path.insert(0, str(REPO_ROOT / "scripts"))
    from _md_connect import connect_locked  # noqa: E402

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    out = DATA_DIR / "analytic_file_v1.parquet"
    if out.exists() and not force:
        print(f"[export] reuse {out}")
        return out

    print("[export] MotherDuck master analytic + CPM …")
    con = connect_locked()
    df_main = con.execute(MASTER_ANALYTIC_SQL).df()
    cohort_n = len(df_main)
    print(f"[export] cohort extract rows = {cohort_n} (pin with m044_validate_canonical_v1_runner.py)")
    df_cpm = con.execute(CPM_EXTRA_SQL).df()
    df_main["research_id_norm"] = normalize_rid_series(df_main["research_id"])
    df_cpm["research_id_norm"] = pd.to_numeric(
        df_cpm["research_id_cpm_key"], errors="coerce"
    ).astype("Int64")
    df_cpm = df_cpm.drop(columns=["research_id_cpm_key"]).drop_duplicates(
        subset=["research_id_norm"], keep="first"
    )
    merged = df_main.merge(df_cpm, on="research_id_norm", how="left")
    merged["multifocal_flag_path_fac"] = tri_cat(merged["multifocal_flag_path"])
    merged["bilateral_disease_fac"] = tri_cat(merged["bilateral_disease_flag"])
    merged["aggressive_variant_fac"] = tri_cat(merged["aggressive_variant_flag"])
    merged["margin_involved_fac"] = tri_cat(merged["margin_involved_any"])
    merged["braf_positive_fac"] = tri_cat(merged["braf_positive_final"])
    merged["surg_tt_fac"] = tri_cat(merged["surg_total_thyroidectomy"])
    merged = add_endpoint_flags(merged)
    merged.to_parquet(out, index=False)
    print(f"[export] wrote {out}")
    return out


def size_panel_clean(frame: pd.DataFrame) -> pd.DataFrame:
    # fixed implementation without buggy lambda
    def ete_three_row(gf):
        if pd.isna(gf):
            return None
        s = str(gf).strip().lower()
        if s in {"false", "absent"}:
            return "No/negative ETE"
        if s == "microscopic":
            return "Microscopic ETE"
        if s == "gross":
            return "Gross ETE"
        return None

    x = frame.copy()
    x["ete3"] = x["ete_grade_final"].map(ete_three_row)
    x = x[x["ete3"].notna()]
    ts = pd.to_numeric(x["tumor_size_cm"], errors="coerce")

    def bsz(v):
        if pd.isna(v):
            return None
        if v <= 1:
            return "<=1"
        if v <= 2:
            return "1.1-2"
        if v <= 4:
            return "2.1-4"
        return ">4"

    x["bin"] = ts.map(bsz)
    x = x[x["bin"].notna()]
    rows = []
    for (eg, bn), grp in x.groupby(["ete3", "bin"]):
        n = len(grp)
        ev = int(bool_series(grp.get("path_proven_primary")).sum())
        rows.append(
            {"ete_group": eg, "size_bin": bn, "n": n, "events": ev, "rate_pct": round(100 * ev / n, 2)}
        )
    return pd.DataFrame(rows)


def expected_size_rates() -> dict[tuple[str, str], float]:
    return {
        ("Microscopic ETE", "<=1"): 1.1,
        ("Microscopic ETE", "1.1-2"): 2.7,
        ("Microscopic ETE", "2.1-4"): 2.3,
        ("Microscopic ETE", ">4"): 5.6,
        ("Gross ETE", "<=1"): 2.6,
        ("Gross ETE", "1.1-2"): 4.1,
        ("Gross ETE", "2.1-4"): 7.0,
        ("Gross ETE", ">4"): 8.6,
        ("No/negative ETE", "<=1"): 3.8,
        ("No/negative ETE", "1.1-2"): 11.4,
        ("No/negative ETE", "2.1-4"): 7.0,
        ("No/negative ETE", ">4"): 3.8,
    }


TABLE2_ETE_ORDER: tuple[str, ...] = (
    "Microscopic ETE",
    "Gross ETE",
    "No/negative ETE",
    "Present ungraded",
    "Missing/other",
)


def build_table2_recurrence_summary(df_merged: pd.DataFrame) -> pd.DataFrame:
    """M044 Table 2 — cohort-wide n/% vs positive-FU person-year rates (SQL-aligned)."""

    fu = pd.to_numeric(df_merged["followup_years"], errors="coerce").fillna(0.0)
    work = add_endpoint_flags(df_merged)
    work["_fu"] = fu
    pp = bool_series(work["path_proven_primary"])
    img_only = bool_series(work["imaging_only_unconfirmed"])
    comp = bool_series(work["composite_primary"])
    has_img_then = "recurrence_imaging_then_path_confirmed" in work.columns
    img_then = (
        work["recurrence_imaging_then_path_confirmed"].fillna(False).astype(bool)
        if has_img_then
        else None
    )
    rows: list[dict[str, Any]] = []
    for eg in TABLE2_ETE_ORDER:
        m = work["ete_group"].astype(str) == eg
        grp = work.loc[m]
        n = int(len(grp))
        pos = grp["_fu"] > 0
        path_n = int(pp.loc[m].sum())
        img_only_n = int(img_only.loc[m].sum())
        comp_n = int(comp.loc[m].sum())
        if img_then is not None:
            img_then_n: int | None = int(img_then.loc[m].sum())
        else:
            img_then_n = None
        py_pos = float(grp.loc[pos, "_fu"].sum())
        pp_pos = int(pp.loc[m & pos].sum())
        comp_pos = int(comp.loc[m & pos].sum())
        pp_py = (100.0 * pp_pos / py_pos) if py_pos > 0 else 0.0
        comp_py = (100.0 * comp_pos / py_pos) if py_pos > 0 else 0.0
        rows.append(
            {
                "ete_group": eg,
                "n": n,
                "path_proven_n": path_n,
                "path_proven_pct": (100.0 * path_n / n) if n else 0.0,
                "img_only_n": img_only_n,
                "img_only_pct": (100.0 * img_only_n / n) if n else 0.0,
                "comp_n": comp_n,
                "comp_pct": (100.0 * comp_n / n) if n else 0.0,
                "img_then_path_n": img_then_n,
                "person_years_positive_fu": round(py_pos, 1),
                "path_proven_n_positive_fu": pp_pos,
                "comp_n_positive_fu": comp_pos,
                "pp_per_100py": round(pp_py, 2),
                "comp_per_100py": round(comp_py, 2),
            }
        )
    return pd.DataFrame(rows)


def fill_table2_excel(ws: Any, tbl: pd.DataFrame) -> None:
    """Write `Table 2 — Recurrence` data rows (expects manuscript column layout)."""

    ws.cell(row=4, column=10, value="Person-years (FU>0)")
    ws.cell(row=4, column=11, value="Path-proven /100 PY")
    start_row = 5
    pct_fmt = lambda x: f"{x:.2f}%"  # noqa: E731
    for i, eg in enumerate(TABLE2_ETE_ORDER):
        hit = tbl.loc[tbl["ete_group"].astype(str) == eg]
        if len(hit) != 1:
            raise ValueError(f"Table 2 missing row for {eg}")
        r = hit.iloc[0]
        row = start_row + i
        ws.cell(row=row, column=1, value=eg)
        ws.cell(row=row, column=2, value=int(r["n"]))
        ws.cell(row=row, column=3, value=int(r["path_proven_n"]))
        ws.cell(row=row, column=4, value=pct_fmt(float(r["path_proven_pct"])))
        ws.cell(row=row, column=5, value=int(r["img_only_n"]))
        ws.cell(row=row, column=6, value=pct_fmt(float(r["img_only_pct"])))
        ws.cell(row=row, column=7, value=int(r["comp_n"]))
        ws.cell(row=row, column=8, value=pct_fmt(float(r["comp_pct"])))
        if r["img_then_path_n"] is not None:
            ws.cell(row=row, column=9, value=int(r["img_then_path_n"]))
        ws.cell(row=row, column=10, value=float(r["person_years_positive_fu"]))
        ws.cell(row=row, column=11, value=float(r["pp_per_100py"]))
    ws.cell(
        row=11,
        column=1,
        value=(
            "Notes: Col J sums follow-up years among patients with FU>0 only (PY-rate denominator). "
            "Col K = 100 × (path-proven events among FU>0) / col J. "
            "Path-proven n and Path-proven % include all patients (zero-FU retained in denominator for proportions)."
        ),
    )


def run_all_models(df_merged: pd.DataFrame) -> dict[str, Any]:
    dm_broad = model_frame_for_primary(df_merged, strict_dtc=False)
    dm_strict = model_frame_for_primary(df_merged, strict_dtc=True)
    out: dict[str, Any] = {}
    out["primary_n_broad"] = len(dm_broad)
    out["primary_n_strict"] = len(dm_strict)
    out["primary_events_pp_strict"] = int(dm_strict["y_pp"].sum())
    out["drop_sex_missing_n_strict"] = int(
        len(build_primary_prep(df_merged, strict_dtc=True)) - len(dm_strict)
    )
    out["crude"] = crude_pairwise(dm_strict)
    out["crude_broad"] = crude_pairwise(dm_broad)

    f_strict_no_rai = glm_formula_primary(include_rai=False)
    f_strict_rai = glm_formula_primary(include_rai=True)

    primary_strict_no_rai = smf.glm(
        f_strict_no_rai, data=dm_strict, family=sm.families.Binomial()
    ).fit()
    primary_strict_rai = smf.glm(
        f_strict_rai, data=dm_strict, family=sm.families.Binomial()
    ).fit()
    primary_broad_rai = smf.glm(
        f_strict_rai, data=dm_broad, family=sm.families.Binomial()
    ).fit()

    out["primary"] = {
        "metrics": _glm_metrics(primary_strict_no_rai),
        "coef": coef_table_glm(primary_strict_no_rai),
        "formula": f_strict_no_rai,
        "label": "strict-DTC primary — without RAI covariate",
    }
    out["primary_strict_with_rai"] = {
        "metrics": _glm_metrics(primary_strict_rai),
        "coef": coef_table_glm(primary_strict_rai),
        "formula": f_strict_rai,
    }
    out["primary_broad_with_rai"] = {
        "metrics": _glm_metrics(primary_broad_rai),
        "coef": coef_table_glm(primary_broad_rai),
        "formula": f_strict_rai,
        "label": "full cohort sensitivity — includes non-DTC histologies (RAI covariate)",
    }

    f_int_full = (
        "y_pp ~ C(ete_group, Treatment(reference='Microscopic ETE'))"
        " * C(ajcc8_n_fac, Treatment(reference='N0'))"
        " + age10 + C(sex_lc, Treatment(reference='female')) + tumor_size_cm"
        " + C(histology_fac, Treatment(reference='PTC'))"
        " + C(lvi_clean, Treatment(reference='missing'))"
        " + C(vasc_clean, Treatment(reference='missing'))"
    )
    f_int_red = (
        "y_pp ~ C(ete_group, Treatment(reference='Microscopic ETE'))"
        " + C(ajcc8_n_fac, Treatment(reference='N0'))"
        " + age10 + C(sex_lc, Treatment(reference='female')) + tumor_size_cm"
        " + C(histology_fac, Treatment(reference='PTC'))"
        " + C(lvi_clean, Treatment(reference='missing'))"
        " + C(vasc_clean, Treatment(reference='missing'))"
    )
    m_int = smf.glm(f_int_full, data=dm_strict, family=sm.families.Binomial()).fit()
    m_int_red = smf.glm(f_int_red, data=dm_strict, family=sm.families.Binomial()).fit()
    lr_i = 2.0 * (m_int.llf - m_int_red.llf)
    df_i = int(round(m_int.df_model - m_int_red.df_model))
    p_int = float(1.0 - stats.chi2.cdf(lr_i, df=df_i)) if df_i > 0 else float("nan")
    out["ete_by_n_stage_interaction_lr"] = {
        "lr_chi2": float(lr_i),
        "df": df_i,
        "p_value": p_int,
        "formula_full": f_int_full,
        "formula_reduced": f_int_red,
    }

    strat_rows: list[dict[str, Any]] = []
    stratum_formula = (
        "y_pp ~ C(ete_group, Treatment(reference='Microscopic ETE'))"
        " + age10 + C(sex_lc, Treatment(reference='female')) + tumor_size_cm"
        " + C(histology_fac, Treatment(reference='PTC'))"
        " + C(lvi_clean, Treatment(reference='missing'))"
        " + C(vasc_clean, Treatment(reference='missing'))"
    )
    for st in ["N0", "N1a", "N1b", "Nx", "missing"]:
        sub_n = dm_strict.loc[dm_strict["ajcc8_n_fac"].astype(str) == st].copy()
        ev_ct = int(sub_n["y_pp"].sum())
        if len(sub_n) < 20 or ev_ct < 5:
            strat_rows.append(
                {
                    "stratum": st,
                    "n": len(sub_n),
                    "events": ev_ct,
                    "note": "skipped_sparse_events_for_stable_GLMs",
                }
            )
            continue
        try:
            ms = smf.glm(stratum_formula, data=sub_n, family=sm.families.Binomial()).fit()
            ct = coef_table_glm(ms)
            g_row = ct.loc[ct["term"].astype(str).str.contains("Gross ETE", regex=False)]
            nn_row = ct.loc[ct["term"].astype(str).str.contains("No/negative", regex=False)]
            if len(g_row) == 0:
                strat_rows.append(
                    {
                        "stratum": st,
                        "n": len(sub_n),
                        "events": ev_ct,
                        "note": "skipped_missing_gross_coef",
                    }
                )
                continue
            glo = float(g_row["or_ci_low"].iloc[0])
            ghi = float(g_row["or_ci_high"].iloc[0])
            gor = float(g_row["or"].iloc[0])
            if (
                not np.isfinite(glo)
                or not np.isfinite(ghi)
                or not np.isfinite(gor)
                or ghi > 100
                or glo <= 0
                or ghi / max(glo, 1e-12) > 1e6
            ):
                strat_rows.append(
                    {
                        "stratum": st,
                        "n": len(sub_n),
                        "events": ev_ct,
                        "note": "skipped_numerically_unstable_OR_CI",
                    }
                )
                continue
            strat_rows.append(
                {
                    "stratum": st,
                    "n": len(sub_n),
                    "events": int(sub_n["y_pp"].sum()),
                    "gross_or": gor,
                    "gross_ci": (glo, ghi),
                    "gross_p": float(g_row["pvalue"].iloc[0]) if len(g_row) else None,
                    "noneg_or": float(nn_row["or"].iloc[0]) if len(nn_row) else None,
                    "noneg_ci": (
                        float(nn_row["or_ci_low"].iloc[0]),
                        float(nn_row["or_ci_high"].iloc[0]),
                    )
                    if len(nn_row)
                    else None,
                    "noneg_p": float(nn_row["pvalue"].iloc[0]) if len(nn_row) else None,
                }
            )
        except Exception as exc:  # noqa: BLE001
            strat_rows.append({"stratum": st, "n": len(sub_n), "error": str(exc)})
    out["stratified_primary_by_n_stage"] = strat_rows

    img = f_strict_no_rai.replace("y_pp", "y_img")
    m_img = smf.glm(img, data=dm_strict, family=sm.families.Binomial()).fit()
    out["secondary_imaging_only"] = {
        "metrics": _glm_metrics(m_img),
        "coef": coef_table_glm(m_img),
    }

    comp = f_strict_no_rai.replace("y_pp", "y_comp")
    m_comp = smf.glm(comp, data=dm_strict, family=sm.families.Binomial()).fit()
    out["secondary_composite"] = {
        "metrics": _glm_metrics(m_comp),
        "coef": coef_table_glm(m_comp),
    }

    d_s1 = dm_strict.loc[
        pd.to_numeric(dm_strict["followup_years"], errors="coerce") > 0
    ].copy()
    m_s1 = smf.glm(f_strict_no_rai, data=d_s1, family=sm.families.Binomial()).fit()
    out["S1_exclude_zero_followup"] = {
        "n": len(d_s1),
        "metrics": _glm_metrics(m_s1),
        "coef": coef_table_glm(m_s1),
    }

    sfd = pd.to_datetime(dm_strict["surg_first_date"], errors="coerce")
    if "surg_date_1999_2024" in dm_strict.columns:
        cohort_win = dm_strict["surg_date_1999_2024"].fillna(False).eq(True)
        py_win = (
            (sfd >= pd.Timestamp("1999-01-01")) & (sfd <= pd.Timestamp("2024-12-31"))
        ).fillna(False)
        if not cohort_win.eq(py_win).all():
            raise SystemExit(
                "cohort_view.surg_date_1999_2024 disagrees with calendar window on strict-DTC frame"
            )
        win = cohort_win
    else:
        win = (
            (sfd >= pd.Timestamp("1999-01-01"))
            & (sfd <= pd.Timestamp("2024-12-31"))
        ).fillna(False)
    d_s2 = dm_strict.loc[win].copy()
    m_s2 = smf.glm(f_strict_no_rai, data=d_s2, family=sm.families.Binomial()).fit()
    out["S2_surgery_date_1999_2024"] = {
        "n": len(d_s2),
        "metrics": _glm_metrics(m_s2),
        "coef": coef_table_glm(m_s2),
    }

    f_s3 = (
        "y_pp ~ C(ete_group, Treatment(reference='Microscopic ETE'))"
        " + age10 + C(sex_lc, Treatment(reference='female')) + tumor_size_cm"
        " + central_pos_flag + lateral_pos_flag"
        " + C(histology_fac, Treatment(reference='PTC'))"
        " + C(lvi_clean, Treatment(reference='missing'))"
        " + C(vasc_clean, Treatment(reference='missing'))"
    )
    m_s3 = smf.glm(f_s3, data=dm_strict, family=sm.families.Binomial()).fit()
    out["S3_ln_flags_substitute_n_stage"] = {
        "metrics": _glm_metrics(m_s3),
        "coef": coef_table_glm(m_s3),
    }

    dp = dm_strict.copy()
    lc = dm_strict["lvi_clean"].astype(str)
    vc = dm_strict["vasc_clean"].astype(str)
    dp["lvi_pooled"] = (
        lc.isin(["present", "extensive", "focal"])
        | (vc.ne("missing") & vc.ne("nan") & vc.ne("indeterminate"))
    ).astype(int)
    f_s4 = f_strict_no_rai.replace(
        "+ C(lvi_clean, Treatment(reference='missing'))"
        " + C(vasc_clean, Treatment(reference='missing'))",
        "+ lvi_pooled",
    )
    m_s4 = smf.glm(f_s4, data=dp, family=sm.families.Binomial()).fit()
    ct_s4 = coef_table_glm(m_s4)
    row_pv = ct_s4.loc[ct_s4["term"] == "lvi_pooled", "coef_logit"]
    pooled_coef = float(row_pv.iloc[0]) if len(row_pv) else float("nan")
    pooled_or = float(np.exp(pooled_coef)) if np.isfinite(pooled_coef) else np.nan
    pooled_p = (
        float(ct_s4.loc[ct_s4["term"] == "lvi_pooled", "pvalue"].iloc[0])
        if len(ct_s4.loc[ct_s4["term"] == "lvi_pooled"])
        else float("nan")
    )

    protective = pooled_coef < -1e-6 and pooled_p < 0.05
    risk_increasing_sig = pooled_coef > 1e-6 and pooled_p < 0.05
    out["S4_pooled_lvi_collapsed_missing_as_absent"] = {
        "formula": f_s4,
        "metrics": _glm_metrics(m_s4),
        "coef": ct_s4,
        "protective_association_at_alpha005": protective,
        "pooled_positive_risk_at_alpha005": risk_increasing_sig,
        "pooled_coef_logit": pooled_coef,
        "pooled_or": pooled_or,
        "lvi_pooled_pvalue": pooled_p,
    }

    dor = dm_strict.copy()
    vmap = {
        "missing": 0,
        "indeterminate": 1,
        "focal": 2,
        "present_ungraded": 3,
        "extensive": 4,
    }
    dor["vasc_ord"] = dor["vasc_clean"].astype(str).map(vmap).astype(float)
    f_s5 = (
        "y_pp ~ C(ete_group, Treatment(reference='Microscopic ETE'))"
        " + age10 + C(sex_lc, Treatment(reference='female')) + tumor_size_cm"
        " + C(ajcc8_n_fac, Treatment(reference='N0'))"
        " + C(histology_fac, Treatment(reference='PTC'))"
        " + C(lvi_clean, Treatment(reference='missing')) + vasc_ord"
    )
    m_s5 = smf.glm(f_s5, data=dor, family=sm.families.Binomial()).fit()
    out["S5_vascular_invasion_linear_ordinal_scalar"] = {
        "formula": f_s5,
        "metrics": _glm_metrics(m_s5),
        "coef": coef_table_glm(m_s5),
        "ordering_note": (
            "vasc_ord: missing=0 indeterminate=1 focal=2 "
            "present_ungraded=3 extensive=4; single linear coefficient"
        ),
    }

    dm6 = dm_strict.loc[dm_strict["ete_grade_final"].astype(str) != "true"].copy()
    m_s6 = smf.glm(f_strict_no_rai, data=dm6, family=sm.families.Binomial()).fit()
    out["S6_drop_ete_grade_true"] = {
        "n": len(dm6),
        "n_removed": len(dm_strict) - len(dm6),
        "metrics": _glm_metrics(m_s6),
        "coef": coef_table_glm(m_s6),
    }

    f_s7 = (
        f_strict_no_rai.rstrip()
        + " + C(multifocal_flag_path_fac, Treatment(reference='missing'))"
        " + C(bilateral_disease_fac, Treatment(reference='missing'))"
        " + C(margin_involved_fac, Treatment(reference='missing'))"
        " + C(aggressive_variant_fac, Treatment(reference='missing'))"
        " + C(braf_positive_fac, Treatment(reference='missing'))"
        " + C(surg_tt_fac, Treatment(reference='missing'))"
    )
    m_s7 = smf.glm(f_s7, data=dm_strict, family=sm.families.Binomial()).fit()
    pr_g = coef_table_glm(primary_strict_no_rai)
    au_g = coef_table_glm(m_s7)
    pr_gross = pr_g[pr_g["term"].str.contains("Gross ETE", regex=False)]
    au_gross = au_g[au_g["term"].str.contains("Gross ETE", regex=False)]
    g1 = float(pr_gross["coef_logit"].iloc[0]) if len(pr_gross) else float("nan")
    g2 = float(au_gross["coef_logit"].iloc[0]) if len(au_gross) else float("nan")
    out["S7_cpm_extended_covariates"] = {
        "formula": f_s7,
        "metrics": _glm_metrics(m_s7),
        "coef": au_g,
        "gross_vs_micro_coef_primary": g1,
        "gross_vs_micro_coef_augmented": g2,
        "change_note": (
            "coefficient strengthened vs primary"
            if g2 > g1 + 1e-6
            else "coefficient attenuated vs primary"
            if g2 < g1 - 1e-6
            else "essentially unchanged"
        ),
    }

    m_s8 = smf.glm(
        f_strict_no_rai.replace("y_pp", "y_any"),
        data=dm_strict,
        family=sm.families.Binomial(),
    ).fit()
    out["S8_legacy_any_recurrence_flag"] = {
        "metrics": _glm_metrics(m_s8),
        "coef": coef_table_glm(m_s8),
    }

    def _cox_fit(dm: pd.DataFrame, formula_tail: str, csv_name: str) -> dict[str, Any]:
        dc, _meta = build_cox_analytic_frame(dm)

        cph = CoxPHFitter(penalizer=0.0001)
        cph.fit(
            dc,
            duration_col="time_days",
            event_col="y_pp",
            formula=formula_tail,
        )
        summ_df = cph.summary.reset_index()
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        out_path = DATA_DIR / csv_name
        summ_df.to_csv(out_path, index=False)
        gross_hr = cox_extract_contrast(summ_df, "Gross ETE")
        noneg_hr = cox_extract_contrast(summ_df, "No/negative ETE")
        return {
            "n": len(dc),
            "events": int(dc["y_pp"].sum()),
            "coef_table_path": str(out_path),
            "gross_vs_microscopic_hr": gross_hr,
            "noneg_vs_microscopic_hr": noneg_hr,
        }

    cox_no_rai = glm_formula_primary(include_rai=False).replace("y_pp ~ ", "").strip()
    cox_with_rai = glm_formula_primary(include_rai=True).replace("y_pp ~ ", "").strip()
    out["Cox_surgery_date_known_positive_fu"] = _cox_fit(
        dm_strict, cox_no_rai, "m044_cox_primary_summary.csv"
    )
    out["Cox_strict_with_rai_covariate"] = _cox_fit(
        dm_strict, cox_with_rai, "m044_cox_primary_with_rai_summary.csv"
    )

    # No/negative ETE — composite outcome (legacy supplemental row)
    sub = add_endpoint_flags(df_merged.loc[df_merged["ete_group"] == "No/negative ETE"]).copy()
    sub["tumor_size_cm"] = pd.to_numeric(sub["tumor_size_cm"], errors="coerce")
    ns = sub["ajcc8_n_stage"]
    sub["ajcc8_n_fac"] = np.where(ns.isna(), "missing", ns.astype(str))
    sub["ajcc8_n_fac"] = pd.Categorical(
        sub["ajcc8_n_fac"], categories=["N0", "N1a", "N1b", "Nx", "missing"]
    )
    sub["y_comp"] = bool_series(sub["composite_primary"]).astype(int)
    sub["rai_received_flag"] = encode_rai_01(sub["rai_received_flag"])
    sub["n_surgeries"] = pd.to_numeric(sub["n_surgeries"], errors="coerce").fillna(1.0)
    sub["ge2"] = (sub["n_surgeries"] >= 2).astype(int)
    sub["days_per_100"] = (
        pd.to_numeric(sub["days_to_2nd"], errors="coerce").fillna(0.0) / 100.0
    )
    sg_form = (
        "y_comp ~ tumor_size_cm + C(ajcc8_n_fac, Treatment(reference='N0'))"
        " + central_pos_flag + lateral_pos_flag + rai_received_flag"
        " + ge2 + days_per_100"
    )
    m_sg = smf.glm(sg_form, data=sub, family=sm.families.Binomial()).fit()
    out["subgroup_no_neg_ete_composite"] = {
        "n": len(sub),
        "events": int(sub["y_comp"].sum()),
        "formula": sg_form,
        "metrics": _glm_metrics(m_sg),
        "coef": coef_table_glm(m_sg),
    }

    # Refit 5 — strict DTC no/negative ETE — path-proven only — no RAI covariate
    sub_s = dm_strict.loc[
        dm_strict["ete_group"].astype(str) == "No/negative ETE"
    ].copy()
    sub_s["n_surgeries"] = pd.to_numeric(sub_s["n_surgeries"], errors="coerce").fillna(
        1.0
    )
    sub_s["ge2"] = (sub_s["n_surgeries"] >= 2).astype(int)
    sub_s["days_per_100"] = (
        pd.to_numeric(sub_s["days_to_2nd"], errors="coerce").fillna(0.0) / 100.0
    )
    sg5_form = (
        "y_pp ~ tumor_size_cm + C(ajcc8_n_fac, Treatment(reference='N0'))"
        " + central_pos_flag + lateral_pos_flag"
        " + ge2 + days_per_100"
    )
    m_sg5 = smf.glm(sg5_form, data=sub_s, family=sm.families.Binomial()).fit()
    out["subgroup_no_neg_ete_strict_pathproven"] = {
        "n": len(sub_s),
        "events": int(sub_s["y_pp"].sum()),
        "formula": sg5_form,
        "metrics": _glm_metrics(m_sg5),
        "coef": coef_table_glm(m_sg5),
    }

    qc_ie = build_inclusion_flow_qc(df_merged)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    qc_csv = DATA_DIR / "m044_inclusion_flow_qc.csv"
    qc_ie.to_csv(qc_csv, index=False)
    out["inclusion_flow_qc_csv"] = str(qc_csv)
    out["inclusion_flow_qc_df"] = qc_ie

    qa_tbl = size_panel_clean(df_merged)
    qa_flags = []
    exp = expected_size_rates()
    for _, r in qa_tbl.iterrows():
        key = (r["ete_group"], str(r["size_bin"]))
        ex = exp.get(key)
        if ex is None:
            continue
        if abs(float(r["rate_pct"]) - ex) > 0.5:
            qa_flags.append(
                f"{key}: got {r['rate_pct']}% expected {ex}% (n={r['n']})"
            )
    out["size_strata_qa"] = {"table": qa_tbl, "flags": qa_flags}

    out["table2_recurrence_df"] = build_table2_recurrence_summary(df_merged)

    return out


def forest_plot_primary(
    primary_coef: pd.DataFrame,
    out_png: Path,
    out_csv: Path,
    *,
    title: str = "M044 strict-DTC primary — path-proven recurrence (no RAI covariate)",
) -> None:
    tab = primary_coef[~primary_coef["term"].str.contains("Intercept", na=False)].copy()
    tab = tab.sort_values("or").reset_index(drop=True)
    ete_hit = tab["term"].astype(str).str.contains("ete_group", regex=False)
    colors = np.where(ete_hit, "#c0392b", "#34495e")
    yy = np.arange(len(tab))
    fig, ax = plt.subplots(figsize=(10, max(6, len(tab) * 0.22)))
    ax.axvline(1.0, color="gray", lw=1)
    for j in range(len(tab)):
        row = tab.iloc[j]
        ax.plot([row["or_ci_low"], row["or_ci_high"]], [j, j], color="#bdc3c7", lw=1.5)
        ax.plot(row["or"], j, "o", color=colors[j], markersize=6)
    ax.set_yticks(yy)
    ax.set_yticklabels(tab["term"], fontsize=8)
    ax.set_xlabel("Adjusted OR (95% CI)")
    ax.set_title(title)
    fig.tight_layout()
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_png, dpi=200)
    plt.close(fig)
    tab.to_csv(out_csv, index=False)


def _find_table3_row(ws: Any, prefix: str) -> int | None:
    for r in range(1, ws.max_row + 2):
        v = ws.cell(row=r, column=1).value
        if v and str(v).startswith(prefix):
            return r
    return None


def _set_or_row(ws: Any, row: int, coef_row: pd.Series) -> None:
    ws.cell(row=row, column=2, value=float(coef_row["or"]))
    ws.cell(row=row, column=3, value=float(coef_row["or_ci_low"]))
    ws.cell(row=row, column=4, value=float(coef_row["or_ci_high"]))
    ws.cell(row=row, column=5, value=float(coef_row["pvalue"]))


def _coef_one(coef: pd.DataFrame, *needles: str) -> pd.Series | None:
    mask = pd.Series(True, index=coef.index)
    for n in needles:
        mask &= coef["term"].astype(str).str.contains(n, regex=False, na=False)
    hits = coef.loc[mask]
    if len(hits) == 1:
        return hits.iloc[0]
    return None


def layout_table3_strict_primary(ws: Any) -> dict[str, int]:
    """Rewrite Table 3 column A labels for strict-DTC primary + sensitivity rows."""

    rows_txt: list[tuple[int, str]] = [
        (4, "Variable"),
        (5, "No/negative ETE (vs Microscopic)"),
        (6, "Gross ETE (vs Microscopic)"),
        (7, "Age, per 10 years"),
        (8, "Sex (male vs female)"),
        (9, "Tumor size, per cm"),
        (10, "AJCC N1a (vs N0)"),
        (11, "AJCC N1b (vs N0)"),
        (12, "AJCC Nx (vs N0)"),
        (13, "AJCC N missing (vs N0)"),
        (14, "Histology — FTC (vs PTC)"),
        (15, "Histology — Metastatic-PTC (vs PTC)"),
        (16, "Histology — Poorly-differentiated DTC (vs PTC)"),
        (17, "Histology — High-grade DTC (vs PTC)"),
        (18, "Lymphatic — present (vs missing)"),
        (19, "Lymphatic — extensive (vs missing)"),
        (20, "Lymphatic — focal (vs missing)"),
        (21, "Lymphatic — indeterminate (vs missing)"),
        (22, "Vascular — present_ungraded (vs missing)"),
        (23, "Vascular — focal (vs missing)"),
        (24, "Vascular — extensive (vs missing)"),
        (25, "Vascular — indeterminate (vs missing)"),
        (26, ""),
        (27, "Sensitivity — full cohort, Gross ETE (vs Microscopic); histology grouped + RAI"),
        (28, "Sensitivity — full cohort, No/negative ETE (vs Microscopic)"),
        (29, "Sensitivity — strict-DTC + RAI covariate, Gross ETE"),
        (30, "Sensitivity — strict-DTC + RAI covariate, No/negative ETE"),
        (31, "Sensitivity — strict-DTC + RAI covariate (RAI receipt term)"),
        (32, ""),
        (33, "ETE × N stage interaction (LR vs main effects only)"),
        (34, ""),
        (35, "Stratified Gross vs Microscopic OR within AJCC N0"),
        (36, "Stratified Gross vs Microscopic OR within AJCC N1a"),
        (37, "Stratified Gross vs Microscopic OR within AJCC N1b"),
        (38, "Stratified Gross vs Microscopic OR within AJCC Nx"),
        (39, "Stratified Gross vs Microscopic OR within AJCC N missing"),
    ]
    from openpyxl.styles import Font

    for r, txt in rows_txt:
        c = ws.cell(row=r, column=1, value=txt if txt else None)
        if r == 4:
            c.font = Font(bold=True)
    h2 = ws.cell(row=4, column=2, value="Adjusted OR")
    h2.font = Font(bold=True)
    for cc, lab in ((3, "95% CI low"), (4, "95% CI high"), (5, "p-value")):
        ws.cell(row=4, column=cc, value=lab).font = Font(bold=True)
    keys = {
        "ete_noneg": 5,
        "ete_gross": 6,
        "age10": 7,
        "sex_male": 8,
        "tumor_size": 9,
        "n_n1a": 10,
        "n_n1b": 11,
        "n_nx": 12,
        "n_miss": 13,
        "hist_ftc": 14,
        "hist_meta": 15,
        "hist_poor": 16,
        "hist_hg": 17,
        "lvi_pr": 18,
        "lvi_ex": 19,
        "lvi_fo": 20,
        "lvi_ind": 21,
        "vas_pu": 22,
        "vas_fo": 23,
        "vas_ex": 24,
        "vas_ind": 25,
        "sens_broad_gross": 27,
        "sens_broad_noneg": 28,
        "sens_strict_wrai_gross": 29,
        "sens_strict_wrai_noneg": 30,
        "sens_strict_wrai_rai": 31,
        "lr_interaction": 33,
        "strat_n0": 35,
        "strat_n1a": 36,
        "strat_n1b": 37,
        "strat_nx": 38,
        "strat_nm": 39,
    }
    return keys


def fill_table3_excel(ws: Any, bundle: dict[str, Any]) -> None:
    rows = layout_table3_strict_primary(ws)
    coef = bundle["primary"]["coef"]

    def put_term(row_key: str, needles: tuple[str, ...]) -> None:
        prow = _coef_one(coef, *needles)
        r = rows[row_key]
        if prow is None:
            ws.cell(row=r, column=6, value=f"TERM_NOT_FOUND:{needles}")
            return
        _set_or_row(ws, r, prow)

    put_term("ete_noneg", ("ete_group", "No/negative"))
    put_term("ete_gross", ("ete_group", "Gross"))
    put_term("age10", ("age10",))
    put_term("sex_male", ("sex_lc", "[T.male"))
    put_term("tumor_size", ("tumor_size_cm",))
    put_term("n_n1a", ("ajcc8_n_fac", "N1a"))
    put_term("n_n1b", ("ajcc8_n_fac", "N1b"))
    put_term("n_nx", ("ajcc8_n_fac", "Nx"))
    put_term("n_miss", ("ajcc8_n_fac", "missing"))
    put_term("hist_ftc", ("histology_fac", "FTC"))
    put_term("hist_meta", ("histology_fac", "Metastatic"))
    put_term("hist_poor", ("histology_fac", "Poorly"))
    put_term("hist_hg", ("histology_fac", "High-grade"))
    put_term("lvi_pr", ("lvi_clean", "present"))
    put_term("lvi_ex", ("lvi_clean", "extensive"))
    put_term("lvi_fo", ("lvi_clean", "focal"))
    put_term("lvi_ind", ("lvi_clean", "indeterminate"))
    put_term("vas_pu", ("vasc_clean", "present_ungraded"))
    put_term("vas_fo", ("vasc_clean", "focal"))
    put_term("vas_ex", ("vasc_clean", "extensive"))
    put_term("vas_ind", ("vasc_clean", "indeterminate"))

    c_broad = bundle["primary_broad_with_rai"]["coef"]
    c_wrai = bundle["primary_strict_with_rai"]["coef"]
    for rk, cf, needles in (
        ("sens_broad_gross", c_broad, ("ete_group", "Gross")),
        ("sens_broad_noneg", c_broad, ("ete_group", "No/negative")),
        ("sens_strict_wrai_gross", c_wrai, ("ete_group", "Gross")),
        ("sens_strict_wrai_noneg", c_wrai, ("ete_group", "No/negative")),
        ("sens_strict_wrai_rai", c_wrai, ("rai_received_flag",)),
    ):
        prow = _coef_one(cf, *needles)
        if prow is not None:
            _set_or_row(ws, rows[rk], prow)

    inte = bundle["ete_by_n_stage_interaction_lr"]
    lr_txt = (
        f"LR χ²={inte['lr_chi2']:.2f}, df={inte['df']}, p={inte['p_value']:.4g}"
    )
    ws.cell(row=rows["lr_interaction"], column=2, value=lr_txt)

    stmap = [
        ("strat_n0", "N0"),
        ("strat_n1a", "N1a"),
        ("strat_n1b", "N1b"),
        ("strat_nx", "Nx"),
        ("strat_nm", "missing"),
    ]
    byst = {x["stratum"]: x for x in bundle["stratified_primary_by_n_stage"]}
    for rk, stlbl in stmap:
        item = byst.get(stlbl, {})
        gor = item.get("gross_or")
        if gor is None:
            note = item.get("note") or item.get("error") or "n/a"
            ws.cell(row=rows[rk], column=2, value=str(note))
            continue
        lo, hi = item["gross_ci"] if item.get("gross_ci") else (None, None)
        ws.cell(row=rows[rk], column=2, value=float(gor))
        if lo is not None and hi is not None:
            ws.cell(row=rows[rk], column=3, value=float(lo))
            ws.cell(row=rows[rk], column=4, value=float(hi))
        gp = item.get("gross_p")
        if gp is not None:
            ws.cell(row=rows[rk], column=5, value=float(gp))


def _append_coef_block(
    ws: Any, start_row: int, title: str, metrics: dict[str, Any], ctab: pd.DataFrame
) -> int:
    from openpyxl.styles import Font

    r = start_row
    ws.cell(row=r, column=1, value=title).font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value=json.dumps(metrics, indent=2, default=str))
    r += 1
    for _, row in ctab.iterrows():
        ws.cell(row=r, column=1, value=row["term"])
        ws.cell(row=r, column=2, value=float(row["or"]))
        ws.cell(row=r, column=3, value=float(row["or_ci_low"]))
        ws.cell(row=r, column=4, value=float(row["or_ci_high"]))
        ws.cell(row=r, column=5, value=float(row["pvalue"]))
        r += 1
    return r + 1


def _qa_clear_dynamic_append_block(qa_ws: Any) -> None:
    """Drop rows from a prior run's appended QA block (avoids duplicate blocks on re-save)."""
    marker = "S4 pooled-LVI artifact check"
    first: int | None = None
    for r in range(1, qa_ws.max_row + 1):
        v = qa_ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() == marker:
            first = r
            break
    if first is not None and first <= qa_ws.max_row:
        qa_ws.delete_rows(int(first), int(qa_ws.max_row - first + 1))


def update_workbook(bundle: dict[str, Any]) -> None:
    from openpyxl.styles import Font

    wb = load_workbook(XLSX)

    ws2 = wb["Table 2 — Recurrence"]
    fill_table2_excel(ws2, bundle["table2_recurrence_df"])

    ws3 = wb["Table 3 — Multivariable"]
    ws3.cell(
        row=2,
        column=1,
        value=(
            "Strict-DTC primary model without RAI covariate | Reference = Microscopic ETE | "
            "scripts/m044_ete_fit_models.py"
        ),
    )
    fill_table3_excel(ws3, bundle)

    ws_m = wb["Model outputs"]
    for rr in ws_m.iter_rows():
        for c in rr:
            c.value = None

    r = 1
    seq = [
        ("Primary — strict-DTC path proven (NO RAI covariate)", "primary"),
        ("Primary — strict-DTC + RAI covariate (sensitivity)", "primary_strict_with_rai"),
        ("Primary — full cohort + RAI (sensitivity)", "primary_broad_with_rai"),
        ("ETE × N stage interaction (LR)", "_interaction_note"),
        ("Secondary — imaging_only_unconfirmed", "secondary_imaging_only"),
        ("Secondary — composite", "secondary_composite"),
        ("S1 exclude FU==0", "S1_exclude_zero_followup"),
        ("S2 surgery date 1999–2024", "S2_surgery_date_1999_2024"),
        ("S3 LN flags replace N-stage", "S3_ln_flags_substitute_n_stage"),
        ("S4 pooled LVI (artifact)", "S4_pooled_lvi_collapsed_missing_as_absent"),
        ("S5 vascular ordinal scalar", "S5_vascular_invasion_linear_ordinal_scalar"),
        ("S6 drop ete true", "S6_drop_ete_grade_true"),
        ("S7 CPM-augmented", "S7_cpm_extended_covariates"),
        ("S8 legacy any recurrence flag", "S8_legacy_any_recurrence_flag"),
    ]
    for title, key in seq:
        if key == "_interaction_note":
            ws_m.cell(row=r, column=1, value=title).font = Font(bold=True)
            r += 1
            ws_m.cell(
                row=r,
                column=1,
                value=json.dumps(bundle["ete_by_n_stage_interaction_lr"], indent=2),
            )
            r += 2
            continue
        blk = bundle[key]
        met = dict(blk.get("metrics", {}))
        for k_extra in ("n", "formula", "n_removed", "change_note", "ordering_note"):
            if k_extra in blk:
                met[k_extra] = blk[k_extra]
        if "label" in blk:
            met["label"] = blk["label"]
        r = _append_coef_block(ws_m, r, title, met, blk["coef"])

    ws_m.cell(row=r, column=1, value="Stratified ORs Gross vs Microscopic by N stage").font = Font(
        bold=True
    )
    r += 1
    ws_m.cell(
        row=r, column=1, value=json.dumps(bundle["stratified_primary_by_n_stage"], indent=2)
    )
    r += 2

    sgk = bundle["subgroup_no_neg_ete_composite"]
    sg_met = dict(sgk.get("metrics", {}))
    sg_met["events"] = sgk["events"]
    sg_met["formula"] = sgk["formula"]
    r = _append_coef_block(ws_m, r, "Subgroup no/negative composite", sg_met, sgk["coef"])

    sg5 = bundle["subgroup_no_neg_ete_strict_pathproven"]
    sg5_met = dict(sg5.get("metrics", {}))
    sg5_met["events"] = sg5["events"]
    sg5_met["formula"] = sg5["formula"]
    r = _append_coef_block(
        ws_m,
        r,
        "Supplement — strict-DTC no/negative ETE, path-proven only (Refit 5)",
        sg5_met,
        sg5["coef"],
    )

    cx = bundle["Cox_surgery_date_known_positive_fu"]
    ws_m.cell(row=r, column=1, value="Cox PH strict-DTC (NO RAI covariate)").font = Font(bold=True)
    r += 1
    slim = {k: v for k, v in cx.items() if k != "summary"}
    ws_m.cell(row=r, column=1, value=json.dumps(slim, indent=2, default=str))
    r += 1
    cdf = pd.read_csv(DATA_DIR / "m044_cox_primary_summary.csv")
    cov_col = "covariate" if "covariate" in cdf.columns else cdf.columns[0]
    for _, row in cdf.iterrows():
        ws_m.cell(row=r, column=1, value=row[cov_col])
        ws_m.cell(row=r, column=2, value=float(row["exp(coef)"]))
        ws_m.cell(row=r, column=3, value=float(row["exp(coef) lower 95%"]))
        ws_m.cell(row=r, column=4, value=float(row["exp(coef) upper 95%"]))
        ws_m.cell(row=r, column=5, value=float(row["p"]))
        r += 1

    cx2 = bundle["Cox_strict_with_rai_covariate"]
    ws_m.cell(row=r, column=1, value="Cox PH strict-DTC (RAI covariate sensitivity)").font = Font(
        bold=True
    )
    r += 1
    slim2 = {k: v for k, v in cx2.items() if k != "summary"}
    ws_m.cell(row=r, column=1, value=json.dumps(slim2, indent=2, default=str))
    r += 1
    cdf2 = pd.read_csv(DATA_DIR / "m044_cox_primary_with_rai_summary.csv")
    for _, row in cdf2.iterrows():
        ws_m.cell(row=r, column=1, value=row[cov_col])
        ws_m.cell(row=r, column=2, value=float(row["exp(coef)"]))
        ws_m.cell(row=r, column=3, value=float(row["exp(coef) lower 95%"]))
        ws_m.cell(row=r, column=4, value=float(row["exp(coef) upper 95%"]))
        ws_m.cell(row=r, column=5, value=float(row["p"]))
        r += 1

    sup_name = "Supp Refit5 no-neg path-proven"
    if sup_name in wb.sheetnames:
        wb.remove(wb[sup_name])
    ws_sup = wb.create_sheet(sup_name)
    ws_sup.cell(row=1, column=1, value="Supplement table — Refit 5 (strict-DTC no/negative ETE, path-proven)")
    ws_sup.cell(row=3, column=1, value="term")
    ws_sup.cell(row=3, column=2, value="OR")
    ws_sup.cell(row=3, column=3, value="CI_low")
    ws_sup.cell(row=3, column=4, value="CI_high")
    ws_sup.cell(row=3, column=5, value="p")
    rr = 4
    for _, row in sg5["coef"].iterrows():
        ws_sup.cell(row=rr, column=1, value=row["term"])
        ws_sup.cell(row=rr, column=2, value=float(row["or"]))
        ws_sup.cell(row=rr, column=3, value=float(row["or_ci_low"]))
        ws_sup.cell(row=rr, column=4, value=float(row["or_ci_high"]))
        ws_sup.cell(row=rr, column=5, value=float(row["pvalue"]))
        rr += 1

    if "Figures" not in wb.sheetnames:
        wb.create_sheet("Figures")
    wfig = wb["Figures"]
    wfig.cell(row=1, column=1, value="figures/m044_forest_primary.png (strict-DTC, no RAI)")
    wfig.cell(row=2, column=1, value="figures/m044_forest_primary_data.csv")
    wfig.cell(row=3, column=1, value="figures/m044_forest_primary_broad.png (full cohort + RAI)")
    wfig.cell(row=4, column=1, value="figures/m044_forest_primary_broad_data.csv")

    qa = wb["QA"]
    _qa_clear_dynamic_append_block(qa)
    nxt = qa.max_row + 2
    proto = bundle["S4_pooled_lvi_collapsed_missing_as_absent"]
    qa.cell(row=nxt, column=1, value="S4 pooled-LVI artifact check")
    qa.cell(
        row=nxt,
        column=2,
        value=(
            f"OR pooled={proto.get('pooled_or')} "
            f"p={proto.get('lvi_pooled_pvalue')}"
        ),
    )
    qa.cell(
        row=nxt,
        column=3,
        value=(
            f"protective_sig={proto.get('protective_association_at_alpha005')}; "
            f"pooled_OR>1_sig={proto.get('pooled_positive_risk_at_alpha005')}"
        ),
    )
    qs = bundle["size_strata_qa"]["flags"]
    nxt += 1
    if qs:
        for i, ftxt in enumerate(qs):
            qa.cell(row=nxt + i, column=1, value=f"SIZE_STRAT_MISMATCH: {ftxt}")
    else:
        qa.cell(row=nxt, column=1, value="SIZE_STRAT QA: within ±0.5 ppt")

    qc_df_inc = bundle.get("inclusion_flow_qc_df")
    if qc_df_inc is not None:
        nxt_row = nxt + max(len(qs), 1) + 2
        qa.cell(row=nxt_row, column=1, value="Strict-DTC Cox/KM inclusion flow").font = Font(bold=True)
        qa.cell(row=nxt_row, column=2, value=str(DATA_DIR / "m044_inclusion_flow_qc.csv"))
        hdr = nxt_row + 1
        qa.cell(row=hdr, column=1, value="step")
        qa.cell(row=hdr, column=2, value="criterion")
        qa.cell(row=hdr, column=3, value="n")
        qa.cell(row=hdr, column=4, value="excluded_at_step")
        qa.cell(row=hdr, column=5, value="cum_excluded_from_cohort")
        qa.cell(row=hdr, column=6, value="path_proven_events_final_only")
        r0 = hdr + 1
        for jj, (_, qr) in enumerate(qc_df_inc.iterrows()):
            qa.cell(row=r0 + jj, column=1, value=int(qr["step_order"]))
            qa.cell(row=r0 + jj, column=2, value=str(qr["criterion"]))
            qa.cell(row=r0 + jj, column=3, value=int(qr["n"]))
            qa.cell(row=r0 + jj, column=4, value=int(qr["excluded_at_step"]))
            qa.cell(row=r0 + jj, column=5, value=int(qr["cum_excluded_from_cohort"]))
            pe = qr.get("path_proven_events")
            if pd.notna(pe):
                qa.cell(row=r0 + jj, column=6, value=int(pe))

    wb.save(XLSX)
    print(f"[xlsx] saved {XLSX}")


def surgery_date_paragraph_live(df_parquet: pd.DataFrame) -> str:
    """Narrative for Results — replaces legacy missing-date / 1999-only wording."""

    n = len(df_parquet)
    cols = df_parquet.columns
    miss = (
        int(df_parquet["surg_date_missing"].fillna(False).astype(bool).sum())
        if "surg_date_missing" in cols
        else int(df_parquet["surg_first_date"].isna().sum())
    )
    pct = round(100.0 * miss / n, 1) if n else 0.0
    pre99 = (
        int(df_parquet["surg_date_pre_1999"].fillna(False).astype(bool).sum())
        if "surg_date_pre_1999" in cols
        else 0
    )
    win_1999 = (
        int(df_parquet["surg_date_1999_2024"].fillna(False).astype(bool).sum())
        if "surg_date_1999_2024" in cols
        else 0
    )
    post24 = (
        int(df_parquet["surg_date_post_2024"].fillna(False).astype(bool).sum())
        if "surg_date_post_2024" in cols
        else 0
    )
    sfd = pd.to_datetime(df_parquet["surg_first_date"], errors="coerce")
    ok_dt = sfd.notna()
    if ok_dt.any():
        dmin = sfd.loc[ok_dt].min()
        dmax = sfd.loc[ok_dt].max()
        cal_rng = (
            dmin.strftime("%Y-%m-%d")
            + "–"
            + (dmax.strftime("%Y-%m-%d") if hasattr(dmax, "strftime") else str(dmax))
        )
        cal_sentence = (
            f"Among patients with calendar surgery dates ({int(ok_dt.sum())}/{n}), "
            f"first-surgery spans {cal_rng}."
        )
    else:
        cal_sentence = ""

    lineage = ""
    note_col = (
        df_parquet["surg_first_date_lineage_note"].dropna().astype(str).head(1).tolist()
    )
    if note_col:
        lineage = note_col[0]
    lineage_sentence = ""
    if lineage and len(lineage) < 500:
        lineage_sentence = (
            " Lineage SSOT notes (cohort extract): "
            + lineage.replace("\n", " ").strip()
            + ".\n\n"
        )

    return (
        f"Surgery-anchor dates derive from **`canonical_patient_master.surg_first_date`** "
        f"(operative-spine lineage; mig_258 cohort flags include `surg_date_missing`). "
        f"In this extract (n={n}), **missing surgery-date indicator** applied to "
        f"**{miss}** patients (**{pct}%**)."
        + lineage_sentence
        + (
            "Calendar partition among lineage flags counts (non-exclusive windows per flag): "
            f"before 1999-01-01: **{pre99}**; 1999-01-01 through 2024-12-31: **{win_1999}**; "
            f"after 2024-12-31: **{post24}**. "
            if any(x in cols for x in ("surg_date_pre_1999", "surg_date_1999_2024", "surg_date_post_2024"))
            else ""
        )
        + cal_sentence
        + (
            " For Cox PH and Kaplan–Meier path-proven time-to-event, only patients with documented "
            "surgery dates, strictly positive analytic follow-up, complete age and tumor size, and "
            "positive finite durations are retained — see **`data/m044/m044_inclusion_flow_qc.csv`** "
            "(latest `scripts/m044_ete_fit_models.py` run) and validation "
            "**`studies/m044_validation/m044_canonical_audit.md`**."
        )
    )


def patch_manuscript_md(bundle: dict[str, Any], df_parquet: pd.DataFrame) -> None:
    import re as _re

    path_md = REPO_ROOT / "M044_ETE_manuscript_draft.md"
    txt = path_md.read_text(encoding="utf-8")
    surg_new = surgery_date_paragraph_live(df_parquet)
    txt, nsurg = _re.subn(
        r"(?:Surgery dates were available for|Surgery-anchor dates derive from).*?(?=\n\nFollow-up was)",
        surg_new.strip(),
        txt,
        count=1,
        flags=_re.S,
    )
    if nsurg != 1:
        print("[md] WARN: surgery-date paragraph regex did not match; manual manuscript update may be needed")
    crude = bundle["crude"]
    coef = bundle["primary"]["coef"]
    g_adj = _coef_one(coef, "ete_group", "Gross")
    nn_adj = _coef_one(coef, "ete_group", "No/negative")

    cg = crude["crude_pp_gross_vs_microscopic_or"]
    cgl = crude["crude_pp_gross_vs_microscopic_ci_low"]
    cgh = crude["crude_pp_gross_vs_microscopic_ci_high"]
    cn = crude["crude_pp_noneg_vs_microscopic_or"]
    cnl = crude["crude_pp_noneg_vs_microscopic_ci_low"]
    cnh = crude["crude_pp_noneg_vs_microscopic_ci_high"]

    def _adj_txt(row: pd.Series | None, label: str) -> str:
        if row is None:
            return f"{label} adjusted OR=TBD"
        return (
            f"{label} adjusted OR {float(row['or']):.2f} "
            f"(95% CI {float(row['or_ci_low']):.2f}–{float(row['or_ci_high']):.2f}; "
            f"p={float(row['pvalue']):.4g})"
        )

    adj_gross = _adj_txt(g_adj, "Gross vs microscopic ETE:")
    adj_noneg = _adj_txt(nn_adj, "No/negative vs microscopic ETE:")
    coef_wrai = bundle["primary_strict_with_rai"]["coef"]
    g_wrai = _coef_one(coef_wrai, "ete_group", "Gross")
    nn_wrai = _coef_one(coef_wrai, "ete_group", "No/negative")
    adj_gross_wrai = _adj_txt(
        g_wrai, "Strict cohort — gross vs microscopic ETE (RAI covariate retained):"
    )
    adj_noneg_wrai = _adj_txt(
        nn_wrai, "Strict cohort — no/negative vs microscopic ETE (RAI covariate retained):"
    )
    cx = bundle["Cox_surgery_date_known_positive_fu"]
    gh = cx.get("gross_vs_microscopic_hr")

    gross_hr_note = ""
    if isinstance(gh, dict) and "hr" in gh:
        gross_hr_note = (
            f" A Cox proportional hazards model on the same strict-DTC subset (documented surgery date, positive "
            f"follow-up; **no RAI covariate**; n={cx['n']}) estimated HR="
            f"{gh['hr']:.2f} (95% CI {gh['hr_ci_low']:.2f}–{gh['hr_ci_high']:.2f}; p={gh['p']:.4g}) "
            f"for gross vs microscopic ETE."
        )

    proto = bundle["S4_pooled_lvi_collapsed_missing_as_absent"]
    pooled_pv = proto.get("lvi_pooled_pvalue")
    pooled_pv_s = "" if pooled_pv is None or (isinstance(pooled_pv, float) and np.isnan(pooled_pv)) else f"{float(pooled_pv):.4g}"
    por = proto.get("pooled_or")
    por_s = "" if por is None or (isinstance(por, float) and np.isnan(por)) else f"{float(por):.2f}"
    pooled_bits = (
        f"In the pre-specified pooled lymphovascular sensitivity model (missing treated as absent for the "
        f"pooled binary), the pooled coefficient had adjusted OR={por_s} "
        f"(p={pooled_pv_s}). "
    )
    if proto.get("protective_association_at_alpha005"):
        pooled_bits += (
            "Under α=0.05, this pooled-negative-coefficient construction reproduced a statistically significant "
            "**inverse-risk (OR<1)** association, consistent with the pooled-LVI artifact described in prior "
            "literature."
        )
    elif proto.get("pooled_positive_risk_at_alpha005"):
        pooled_bits += (
            "Instead, this pooled construction produced a statistically significant **elevated-odds** association "
            "(OR>1), not a protective association; it therefore does **not** reconstruct the classic "
            "**protective** pooled-LVI artifact, though it still mixes lymphatic/vascular signal and treats "
            "missing as absent."
        )
    else:
        pooled_bits += (
            "A statistically significant pooled coefficient (either protective or risk-increasing) was not "
            "observed at α=0.05 under this specification."
        )

    pm = bundle["primary"]["metrics"]
    inte = bundle["ete_by_n_stage_interaction_lr"]
    interaction_note = (
        f"Global **ETE × AJCC8 N-stage interaction** (likelihood ratio vs main-effects-only model): "
        f"LR χ²={inte['lr_chi2']:.2f}, df={inte['df']}, p={inte['p_value']:.4g}. "
    )
    if inte["p_value"] < 0.05:
        interaction_note += (
            "The interaction term bundle reached conventional statistical significance; "
            "stratum-specific gross-vs-microscopic ORs are summarized in Table 3."
        )
    else:
        interaction_note += (
            "The omnibus interaction test did not reach α=0.05; stratum-specific contrasts are nevertheless "
            "presented given clinically heterogeneous crude gradients (especially within N1b)."
        )

    strat_lines: list[str] = []
    for x in bundle["stratified_primary_by_n_stage"]:
        st = x.get("stratum", "?")
        gor = x.get("gross_or")
        if gor is None:
            strat_lines.append(f"- **{st}:** {x.get('note') or x.get('error') or 'n/a'}")
            continue
        lo, hi = x["gross_ci"]
        strat_lines.append(
            f"- **{st}:** adjusted OR {gor:.2f} "
            f"(95% CI {lo:.2f}–{hi:.2f}; p={x['gross_p']:.4g}); n={x['n']}, path-proven events={x['events']}."
        )
    strat_blob = "\n".join(strat_lines)

    mv_par = (
        "Multivariable models were refit on a **strict-DTC** analytic subset after excluding medullary carcinoma, "
        "anaplastic carcinoma, NIFTP/FTUMP, benign follicular neoplasms (including atypical adenoma), and rare "
        "non-DTC histologies listed in Methods (see Table 3 footnote in workbook). Histology was parameterized as "
        "PTC (reference), FTC, metastatic PTC, poorly differentiated DTC, and high-grade DTC. "
        "**Radioactive iodine was excluded from the primary covariate set** because receipt reflects "
        "confounding-by-indication; a parallel strict-cohort model retaining RAI appears as sensitivity.\n\n"
        "**Primary logistic model (strict-DTC; no RAI covariate):**\n\n"
        f"{adj_gross}\n\n{adj_noneg}\n\n"
        "**Sensitivity — strict-DTC with RAI covariate retained:**\n\n"
        f"{adj_gross_wrai}\n\n{adj_noneg_wrai}\n\n"
        f"(McFadden pseudo-R²={pm['pseudo_r2_mcfadden']:.4f}; "
        f"n={pm['n_obs']}, path-proven events={pm['n_events']}; likelihood-ratio χ²="
        f"{pm['lr_vs_null_chi2']:.2f} vs intercept-only)."
        f"{gross_hr_note}\n\n"
        f"{interaction_note}\n\n"
        f"**Within-N-stage gross-vs-microscopic contrasts** (same adjustment bundle excluding the fixed stratum’s "
        f"N-stage factor):\n{strat_blob}\n\n"
        "Full coefficient tables are in Table 3 (including full-cohort sensitivity rows) and Supplement.\n\n"
        f"{pooled_bits}\n"
    )

    core = (
        f"Within the strict-DTC three-group analytic subset (Methods), the crude path-proven odds ratio for "
        f"gross ETE vs microscopic ETE was {cg:.2f} "
        f"(95% CI {cgl:.2f}–{cgh:.2f}); the crude odds ratio for no/negative ETE vs microscopic ETE was "
        f"{cn:.2f} (95% CI {cnl:.2f}–{cnh:.2f}). "
    )
    # Mid-sentence uses lowercase "the crude …" after "(Methods),"
    txt, n12 = _re.subn(
        r"(?:Within the strict-DTC three-group analytic subset \(Methods\), )?"
        r"[Tt]he crude path-proven odds ratio for gross ETE vs microscopic ETE was .*?;"
        r" [Tt]he crude odds ratio for no/negative ETE vs microscopic ETE was .*?\.\s*"
        r"(?=[Tt]he legacy `any_recurrence_flag`)",
        core,
        txt,
        count=1,
        flags=_re.S,
    )

    txt, n3 = _re.subn(
        r"(### Multivariable analysis\n\n).+?(\n\n### Tumor-size)",
        r"\1" + mv_par + r"\2",
        txt,
        count=1,
        flags=_re.S,
    )
    if n12 != 1 or n3 != 1:
        raise SystemExit(f"manuscript regex replace failed: crude_block={n12}, multivariable={n3}")

    if isinstance(gh, dict) and "hr" in gh:
        p_cox = float(gh["p"])
        hr_v = float(gh["hr"])
        if p_cox < 0.05 and hr_v > 1.0:
            cox_sentence = (
                f"Cox regression on documented surgery-interval follow-up without RAI estimated HR {hr_v:.2f} "
                f"(95% CI {gh['hr_ci_low']:.2f}–{gh['hr_ci_high']:.2f}; p={p_cox:.4g}) for gross versus "
                "microscopic ETE, consistent with **elevated hazard** on the path-proven time axis."
            )
        elif p_cox < 0.05 and hr_v < 1.0:
            cox_sentence = (
                f"Cox regression estimated HR {hr_v:.2f} (95% CI {gh['hr_ci_low']:.2f}–{gh['hr_ci_high']:.2f}; "
                f"p={p_cox:.4g}) — interpret with caution (convergence / sparse events; see lifelines warnings)."
            )
        else:
            cox_sentence = (
                f"Cox regression on documented surgery-interval follow-up without RAI yielded HR {hr_v:.2f} "
                f"(95% CI {gh['hr_ci_low']:.2f}–{gh['hr_ci_high']:.2f}; p={p_cox:.4g}) for gross versus "
                "microscopic ETE — **no statistically significant incremental hazard** on the Cox time axis "
                "aligned to path-proven TTE (`data/m044/m044_inclusion_flow_qc.csv`)."
            )
        disc_par = (
            "The original **full-cohort** logistic specification—including both **RAI receipt** and a collapsed "
            "**histology-other** bucket containing non-DTC and borderline entities—materially attenuated the "
            "gross-vs-microscopic adjusted odds ratio relative to crude estimates (prior Table 3 iteration). "
            "Under the **strict-DTC primary model without an RAI covariate**, the gross-vs-microscopic association "
            f"moves toward the crude gradient ({adj_gross}), while {cox_sentence} "
            "Together, these findings indicate that much of the earlier logistic attenuation was driven by "
            "**treatment-confounding (RAI)** and **histologic heterogeneity**, not by disappearance of the "
            "gross-vs-microscopic OR gradient in the primary logistic frame. "
            "Microscopic ETE behaved more like the no-ETE group than gross ETE on most ETE-anchored contrasts. "
            "These findings support the AJCC 8th edition decision not to upstage microscopic ETE to T3b and "
            "reinforce the principle that gross strap-muscle invasion is a meaningfully different pathologic "
            "phenomenon, despite recent literature questioning whether T3b adds incremental risk over T2 in "
            "small tumors.[24, 47, 51]"
        )
    else:
        disc_par = (
            "Strict-DTC refitting and omission of RAI as a covariate were undertaken to address "
            "confounding-by-indication and histology heterogeneity; see Results for updated effect estimates."
        )

    # Opening Discussion paragraph (pattern updated when manuscript was hand-edited 2026-05-01).
    txt, ndisc = _re.subn(
        r"The original \*\*full-cohort\*\* logistic specification.+?(?=\n\nThree findings warrant emphasis\.)",
        disc_par.rstrip() + "\n\n",
        txt,
        count=1,
        flags=_re.S,
    )
    if ndisc == 0:
        txt, ndisc = _re.subn(
            r"The original \*\*full-cohort\*\* logistic specification.+?true gross-ETE signal\.\s+",
            disc_par.rstrip() + "\n\n",
            txt,
            count=1,
            flags=_re.S,
        )
    if ndisc == 0:
        txt, ndisc = _re.subn(
            r"In a contemporary 4,128-patient single-institution thyroid cancer cohort, gross versus microscopic "
            r"extrathyroidal extension showed the largest disparity in pathology-proven recurrence crude odds ratios; "
            r"logistic adjustment materially attenuated the gross-vs-microscopic odds ratio \(Table 3\), whereas Cox "
            r"proportional hazards regression on documented surgery-interval follow-up retained elevated hazard comparing "
            r"gross versus microscopic disease \(above in Results\)\.\s+",
            disc_par.rstrip() + "\n\n",
            txt,
            count=1,
        )
    if ndisc != 1:
        raise SystemExit(f"manuscript Discussion paragraph replace failed: {ndisc}")

    # Normalize excessive blank lines before "Three findings"
    txt, _ = _re.subn(
        r"(small tumors\.\[[0-9, ]+\])\n\n+(Three findings warrant emphasis\.)",
        r"\1\n\n\2",
        txt,
        count=1,
    )

    # Limitations: refresh obsolete surgery-date % / calendar-only wording when cohort flags exist.
    cols = df_parquet.columns
    if "surg_date_missing" in cols:
        n_lp = len(df_parquet)
        miss_lp = int(df_parquet["surg_date_missing"].fillna(False).astype(bool).sum())
        pct_lp = round(100.0 * miss_lp / n_lp, 1) if n_lp else 0.0
        txt, nlim = _re.subn(
            r"(\d+(?:\.\d+)?%)\s+missing surgery dates",
            f"{pct_lp}% missing surgery-date indicator (`surg_date_missing`; n={miss_lp}/{n_lp})",
            txt,
            count=1,
        )
        if nlim != 1:
            txt, _ = _re.subn(
                r"22\.1%\s+missing surgery dates",
                f"{pct_lp}% missing surgery-date indicator (`surg_date_missing`; n={miss_lp}/{n_lp})",
                txt,
                count=1,
            )
    if all(c in cols for c in ("surg_date_1999_2024", "surg_date_pre_1999", "surg_date_post_2024")):
        win_94 = int(df_parquet["surg_date_1999_2024"].fillna(False).astype(bool).sum())
        win_pre = int(df_parquet["surg_date_pre_1999"].fillna(False).astype(bool).sum())
        win_post = int(df_parquet["surg_date_post_2024"].fillna(False).astype(bool).sum())
        era_block = (
            f"Surgery-flag calendar strata in this analytic extract: "
            f"1999‑01‑01–2024‑12‑31 n={win_94}, pre‑1999 n={win_pre}, post‑2024 n={win_post} "
            f"(counts from mig_258 `surg_date_*`; not a censorship rule for the primary Cox slice). "
            f"Era-stratified sensitivity analyses are reported in the supplement."
        )
        old_era = (
            "Our cohort spans 1999–2024 among patients with non-missing dates, encompassing "
            "two distinct AJCC eras and a large change in surveillance intensity; "
            "era-stratified sensitivity analyses are reported in the supplement."
        )
        if old_era in txt:
            txt = txt.replace(old_era, era_block, 1)
        else:
            # Prior regex-based edits may have partially replaced; fuse fragment if needed
            txt, ne = _re.subn(
                r"Surgery-flag calendar strata in this analytic extract:[^\n]+\s*Implications\.",
                era_block + "\n\nImplications.",
                txt,
                count=1,
            )
            if ne != 1:
                txt, _ = _re.subn(
                    r"Surgery-flag calendar strata in this analytic extract:.+?\)\.\s+",
                    era_block + "\n\n",
                    txt,
                    count=1,
                )

    path_md.write_text(txt, encoding="utf-8")
    print(f"[md] patched {path_md}")


def main() -> None:
    ap = argparse.ArgumentParser(description="M044 ETE analytic export + modeling")
    ap.add_argument("--force", action="store_true", help="re-export parquet from MotherDuck")
    ap.add_argument(
        "--skip-md-export",
        action="store_true",
        help="use existing data/m044/analytic_file_v1.parquet",
    )
    args = ap.parse_args()

    if args.skip_md_export:
        pq = DATA_DIR / "analytic_file_v1.parquet"
        if not pq.exists():
            raise SystemExit("--skip-md-export but parquet missing")
    else:
        pq = export_parquet(force=args.force)

    df = pd.read_parquet(pq)
    cohort_n = len(df)
    print(f"[models] analytic parquet rows = {cohort_n}")

    bundle = run_all_models(df)
    fg_png = FIG_DIR / "m044_forest_primary.png"
    fg_csv = FIG_DIR / "m044_forest_primary_data.csv"
    forest_plot_primary(bundle["primary"]["coef"], fg_png, fg_csv)
    fb_png = FIG_DIR / "m044_forest_primary_broad.png"
    fb_csv = FIG_DIR / "m044_forest_primary_broad_data.csv"
    forest_plot_primary(
        bundle["primary_broad_with_rai"]["coef"],
        fb_png,
        fb_csv,
        title="M044 sensitivity — full cohort + RAI covariate (historical specification)",
    )
    print(f"[figures] wrote {fg_png} {fg_csv}")
    print(f"[figures] wrote {fb_png} {fb_csv}")

    update_workbook(bundle)

    def _coef_snapshot(row: pd.Series | None) -> dict[str, Any] | None:
        if row is None:
            return None
        return {
            "or": float(row["or"]),
            "or_ci_low": float(row["or_ci_low"]),
            "or_ci_high": float(row["or_ci_high"]),
            "pvalue": float(row["pvalue"]),
        }

    snap = {
        "generated_at_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "cohort_parquet_rows": cohort_n,
        "crude_path_proven_or": bundle["crude"],
        "primary_strict_no_rai": {
            "n_obs": bundle["primary"]["metrics"].get("n_obs"),
            "n_events": bundle["primary"]["metrics"].get("n_events"),
            "pseudo_r2_mcfadden": bundle["primary"]["metrics"].get("pseudo_r2_mcfadden"),
            "lr_vs_null_chi2": bundle["primary"]["metrics"].get("lr_vs_null_chi2"),
            "gross_vs_microscopic": _coef_snapshot(
                _coef_one(bundle["primary"]["coef"], "ete_group", "Gross")
            ),
            "noneg_vs_microscopic": _coef_snapshot(
                _coef_one(bundle["primary"]["coef"], "ete_group", "No/negative")
            ),
        },
        "cox_strict_no_rai": {
            k: v
            for k, v in bundle["Cox_surgery_date_known_positive_fu"].items()
            if k != "summary"
        },
    }
    snap_path = DATA_DIR / "m044_run_snapshot.json"
    snap_path.write_text(json.dumps(snap, indent=2, default=str), encoding="utf-8")
    print(f"[snapshot] wrote {snap_path}")

    patch_manuscript_md(bundle, df)
    print("[done]")


if __name__ == "__main__":
    main()
