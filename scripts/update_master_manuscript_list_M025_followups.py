"""
Update MASTER_MANUSCRIPT_LIST_DATA_READINESS_20260504.xlsx
to incorporate M025-derived expansions and the M048-M054 follow-up portfolio.

Run from THYROID_2026 root:
    python3 scripts/update_master_manuscript_list_M025_followups.py

Source of truth for the M048-M054 ideas: M025 v2 manuscript draft + Cowork chat
2026-05-05.
"""

from copy import copy
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill

PATH = "/sessions/eloquent-serene-johnson/mnt/THYROID_2026/MASTER_MANUSCRIPT_LIST_DATA_READINESS_20260504.xlsx"


def main() -> None:
    wb = openpyxl.load_workbook(PATH)
    ws = wb["Master Manuscript List"]

    # ------------------------------------------------------------------
    # 1) Re-scope existing rows that M025 directly informs
    # ------------------------------------------------------------------
    # Row 22 = ID 45 (Multimodal Risk: TI-RADS + Bethesda + Final Path)
    ws.cell(row=22, column=2).value = (
        "Multimodal Risk: TI-RADS + Bethesda + Final Path "
        "(per-nodule grain, post-M025)"
    )
    ws.cell(row=22, column=4).value = "In Progress"
    ws.cell(row=22, column=7).value = "READY"
    ws.cell(row=22, column=9).value = (
        "M025 expansion: re-scope to per-nodule grain framework "
        "(use cohort_m025_nodule_level_v1 strict subset n=3,687 + Bethesda "
        "bridge from imaging_fna_linkage_v3). Adds joint TI-RADS x Bethesda "
        "ROM table at nodule grain; no new data work."
    )

    # Row 23 = ID 46 (Bethesda ROM with NIFTP-Era Stratification)
    ws.cell(row=23, column=2).value = (
        "Bethesda + TI-RADS ROM with NIFTP/FT-UMP-Era Stratification"
    )
    ws.cell(row=23, column=9).value = (
        "M025 expansion: extend to TI-RADS calibration under NIFTP "
        "reclassification (n=58 NIFTP, n=22 FT-UMP in M025 cohort). "
        "Single paper instead of two. Needs NIFTP reclass logic vs WHO 2022."
    )

    # ------------------------------------------------------------------
    # 2) Append M025 follow-up portfolio (IDs 48-54)
    # ------------------------------------------------------------------
    new_rows = [
        (
            48,
            "Racial Disparities in TI-RADS Performance "
            "(45.5% Black operative cohort)",
            "M025 Follow-Up Portfolio",
            "Proposed",
            "High",
            "Priority Active",
            "READY",
            "canonical_patient_master (race), cohort_m025_tirads_performance_v1, "
            "cohort_m025_nodule_level_v1",
            "Same data as M025; pre-specified per-race AUC, per-TR ROM at "
            "patient + nodule grain. Target: Thyroid or JAMA Otolaryngology.",
        ),
        (
            49,
            "Multinodular Attribution Error: Generalizable Framework for "
            "RSS Validation (methods)",
            "M025 Follow-Up Portfolio",
            "Proposed",
            "High",
            "Priority Active",
            "NEAR READY",
            "imaging_nodule_master_v1 features + reapply EU-TIRADS / K-TIRADS / "
            "ATA pattern / AI-TIRADS scoring algorithms",
            "Methods extension of M025: re-score same cohort under 4 alternative "
            "RSS systems; closed-form attribution-error formula. Target: "
            "Stat Med / DPR / J Clin Epi or Thyroid commentary companion.",
        ),
        (
            50,
            "Imaging-Pathology Correlation in Multifocal Thyroid Cancer",
            "M025 Follow-Up Portfolio",
            "Proposed",
            "High",
            "Priority Active",
            "READY",
            "canonical_path_malignant_events_v1 (4,022 pts; mean 1.61 tumors/pt), "
            "canonical_us_nodule_v2",
            "Index-nodule concordance (TR-max = path-dominant cancer Y/N); "
            "occult cancer rate; size-discordance. Target: "
            "Annals of Surgical Oncology, Surgery, or Thyroid.",
        ),
        (
            51,
            "TI-RADS Performance in Hashimoto's / CLT Background Thyroid",
            "M025 Follow-Up Portfolio",
            "Proposed",
            "Medium",
            "Priority Active",
            "READY",
            "CLT cohort 1,967 pts (per EERC reconciliation), Graves 574 pts; "
            "M025 analytic tables",
            "Per-TR ROM stratified by CLT status; Graves comparator arm; "
            "feature-distribution shift analysis. Target: "
            "Thyroid or Endocrine Practice.",
        ),
        (
            52,
            "LLM + Structured-Feature Pipeline for Retrospective TI-RADS "
            "Harmonization (methods)",
            "M025 Follow-Up Portfolio",
            "Proposed",
            "Medium",
            "Priority Active",
            "READY",
            "Qwen2.5-32B extraction logs, imaging_nodule_master_v1, "
            "us_nodules_tirads_vs_inm_v1_discordance_v1",
            "Documents M025 re-scoring pipeline (99.3% structured, 0.7% LLM-aug). "
            "Target: Radiology: AI, JAMIA, NPJ Digital Medicine.",
        ),
        (
            53,
            "ACR FNA-Eligibility Decision-Curve Analysis "
            "(1,553 unnecessary / 472 missed)",
            "M025 Follow-Up Portfolio",
            "Proposed",
            "High",
            "Priority Active",
            "READY",
            "M025 analytic tables + size_cm from imaging_nodule_master_v1",
            "Net-benefit curves at default ACR thresholds vs alternatives; "
            "subgroup decision curves. Target: JACR or Endocrine Practice.",
        ),
        (
            54,
            "Pre-/Post-2017 Era Comparison: Natural Experiment in TI-RADS "
            "Guideline Adoption",
            "M025 Follow-Up Portfolio",
            "Proposed",
            "Medium",
            "Priority Active",
            "NEAR READY",
            "M025 era-split tables (pre n=422 / post n=2,953 patients; "
            "pre n=381 / post n=3,306 nodules) + canonical_us_exam temporal cols",
            "Interrupted-time-series at 2017-05-01 boundary; documentation "
            "completeness, FNA yield, malignancy yield. Distinct from #36 "
            "(staging-focused). Target: JACR, Radiology, or Thyroid.",
        ),
    ]

    # Capture template formatting from the last existing row (24) so the
    # new rows pick up the same look.
    template_row = 24
    template_cells = [ws.cell(row=template_row, column=c) for c in range(1, ws.max_column + 1)]

    next_row = ws.max_row + 1
    for offset, payload in enumerate(new_rows):
        target_row = next_row + offset
        for col_idx, value in enumerate(payload, start=1):
            tc = template_cells[col_idx - 1]
            cell = ws.cell(row=target_row, column=col_idx, value=value)
            if tc.has_style:
                cell.font = copy(tc.font)
                cell.fill = copy(tc.fill)
                cell.border = copy(tc.border)
                cell.alignment = copy(tc.alignment)
                cell.number_format = tc.number_format
                cell.protection = copy(tc.protection)

    # ------------------------------------------------------------------
    # 3) Stamp a provenance note on the cover sheet if one exists, else
    #    add a small note row at the bottom of the master sheet.
    # ------------------------------------------------------------------
    ws.cell(
        row=ws.max_row + 2,
        column=1,
        value=(
            "Updated 2026-05-05 — M025 expansions to #45, #46 + new "
            "M048-M054 follow-up portfolio appended. See "
            "M025_v2_manuscript_DRAFT_v1_0.md for source thesis."
        ),
    ).font = Font(italic=True, color="666666")

    wb.save(PATH)
    print(f"Saved {PATH}")
    print(f"Rows now: 1-{ws.max_row}")


if __name__ == "__main__":
    main()
