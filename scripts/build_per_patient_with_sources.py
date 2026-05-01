#!/usr/bin/env python3
"""
M044 — Build comprehensive 1-row-per-research_id analytic table with explicit
column-level source map.

Output: /sessions/wonderful-trusting-babbage/mnt/THyroid 2026/M044_per_patient_with_sources.xlsx

Sheets:
  1. Per-patient master  (n=4128 × ~150 cols)
  2. Source map          (every col → source DB.schema.table.column → rule)
  3. Cohort view raw     (29 cols, pass-through)
  4. CPM key cols raw    (~75 cols pulled fresh)
  5. Recurrence raw      (20 cols, full canonical_recurrence_resolved_v1)
  6. LN agg raw
  7. Reop agg raw
  8. README
"""
from __future__ import annotations
import os, sys, json
from datetime import datetime
from pathlib import Path
import duckdb, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DB = "thyroid_canonical_publication_v1_0"
OUT = Path("/sessions/wonderful-trusting-babbage/mnt/THyroid 2026/M044_per_patient_with_sources.xlsx")
OUT.parent.mkdir(parents=True, exist_ok=True)

HF = PatternFill("solid", fgColor="1F4E78"); HFONT = Font(bold=True, color="FFFFFF", size=11)
TF = Font(bold=True, size=14, color="1F4E78"); SF = Font(bold=True, size=11, color="404040")
BD = Border(left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"))


# ---------------------------------------------------------------------------
# SQL — one-row-per-research_id pull (everything we use anywhere)
# ---------------------------------------------------------------------------

ANALYTIC_SQL = r"""
WITH cohort AS (
  SELECT
    c.*,
    -- Derived from cohort_m044 + histology rules
    CASE
      WHEN c.ete_grade_final IN ('false','absent')  THEN 'No/negative ETE'
      WHEN c.ete_grade_final = 'microscopic'        THEN 'Microscopic ETE'
      WHEN c.ete_grade_final = 'gross'              THEN 'Gross ETE'
      WHEN c.ete_grade_final = 'present_ungraded'   THEN 'Present ungraded'
      ELSE 'Missing/other'
    END AS ete_group,
    CASE
      WHEN c.lvi_grade ILIKE 'extensiv%'            THEN 'extensive'
      WHEN c.lvi_grade IN ('present','preesent')    THEN 'present'
      WHEN c.lvi_grade = 'focal'                    THEN 'focal'
      WHEN c.lvi_grade IS NULL                      THEN 'missing'
      WHEN c.lvi_grade IN ('indeterminate','indetermiante','indeeterminate','indeterminent','suspicious','x','c/a','no','n/s')
                                                    THEN 'indeterminate'
      ELSE 'indeterminate'
    END AS lvi_clean,
    COALESCE(c.vascular_invasion_final, 'missing')  AS vasc_clean,
    CASE
      WHEN c.histology_final IN ('PTC','differentiated thyroid carcinoma','high-grade PTC with thymic-like features',
                                 'infiltrating carcinoma with thymus-like differentiation','poorly differentiated PTC')                THEN 'PTC'
      WHEN c.histology_final = 'follicular carcinoma'                                                                                    THEN 'FTC'
      WHEN c.histology_final ILIKE 'metastatic PTC%' OR c.histology_final ILIKE 'recurrent/metastatic PTC%'
        OR c.histology_final ILIKE 'metastatic thyroid carcinoma%' OR c.histology_final ILIKE 'metastatic follicular carcinoma%'
        OR c.histology_final = 'recurrent/metastatic follicular carcinoma'                                                               THEN 'Metastatic-PTC'
      WHEN c.histology_final ILIKE '%poorly differentiated%' OR c.histology_final ILIKE '%poorly differntiated%'                         THEN 'Poorly-differentiated DTC'
      WHEN c.histology_final ILIKE '%high grade%' OR c.histology_final ILIKE '%high-grade%'                                              THEN 'High-grade DTC'
      WHEN c.histology_final IN ('MTC','metastatic MTC','recurrent MTC','MTC/PTC mixed composite')                                       THEN 'NON-DTC: MTC'
      WHEN c.histology_final IN ('anaplastic carcinoma','metastatic anaplastic carcinoma','metastatic PTC/anaplastic carcinoma')         THEN 'NON-DTC: Anaplastic'
      WHEN c.histology_final IN ('NIFTP','FTUMP','follicular adenoma','atypical follicular adenoma','Atypical hurthle cell neoplasm')    THEN 'NON-DTC: NIFTP/borderline'
      WHEN c.histology_final IN ('NUT carcinoma','adenoid cystic carcinoma','high grade carcinoma with focal squamous features')        THEN 'NON-DTC: Other rare'
      ELSE 'Unclassified'
    END AS histology_dtc_5level,
    CASE
      WHEN c.histology_final IN ('MTC','metastatic MTC','recurrent MTC','MTC/PTC mixed composite',
                                 'anaplastic carcinoma','metastatic anaplastic carcinoma','metastatic PTC/anaplastic carcinoma',
                                 'NIFTP','FTUMP','follicular adenoma','atypical follicular adenoma','Atypical hurthle cell neoplasm',
                                 'NUT carcinoma','adenoid cystic carcinoma') THEN 0
      ELSE 1
    END AS strict_dtc_include
  FROM manuscript_workspace.cohort_m044_ajcc_ete_v1 c
),
ln AS (
  SELECT research_id,
    MAX(ln_total_examined) AS ln_examined, MAX(ln_total_positive) AS ln_positive,
    MAX(ln_central_examined) AS ln_central_examined, MAX(ln_central_positive) AS ln_central_positive,
    MAX(ln_lateral_left_positive) AS ln_lateral_left_positive,
    MAX(ln_lateral_right_positive) AS ln_lateral_right_positive,
    MAX(ln_bilateral_lateral_positive) AS ln_bilateral_lateral_positive,
    MAX(ln_level_vi_positive) AS ln_level_vi_positive,
    MAX(ln_level_vii_positive) AS ln_level_vii_positive,
    MAX(ln_extranodal_extension) AS ln_ene
  FROM manuscript_workspace.ln_master_rollup_v1 GROUP BY research_id
),
reop AS (
  SELECT research_id,
    MAX(n_surgeries) AS n_surgeries, MAX(second_surgery_date) AS second_surgery_date,
    MAX(days_between_first_second_surgery) AS days_to_2nd,
    MAX(completion_reason) AS completion_reason,
    MAX(completion_reason_confidence) AS completion_reason_confidence,
    MAX(completion_histology_type) AS completion_histology_type,
    MAX(op_reoperative_any) AS op_reoperative_any
  FROM manuscript_workspace.cohort_m040_reoperative_v1 GROUP BY research_id
),
rec AS (
  SELECT research_id,
    recurrence_path_proven, recurrence_path_proven_date, recurrence_path_proven_source, recurrence_path_proven_evidence,
    days_to_path_proven, recurrence_imaging_suspicious, recurrence_imaging_suspicious_date,
    recurrence_imaging_modality, recurrence_imaging_modality_summary, recurrence_imaging_source,
    recurrence_imaging_finding_text, recurrence_imaging_n_events, days_to_imaging_suspicious,
    recurrence_imaging_then_path_confirmed, recurrence_status_final, is_implausible_date_quarantine
  FROM main.canonical_recurrence_resolved_v1
),
cpm AS (
  SELECT CAST(research_id AS VARCHAR) AS research_id,
    -- Demographics
    race, bmi_combined,
    -- Tumor / surgical
    multifocal_flag_path, bilateral_disease_flag, aggressive_variant_flag,
    margin_involved_any, closest_margin_mm, capsular_invasion_v6, capsular_ordinal_worst,
    syn_hashimoto, syn_graves, syn_chronic_thyroiditis,
    -- PMHx (NLP)
    pmhx_nlp_diabetes, pmhx_nlp_hypertension, pmhx_nlp_hypothyroidism, pmhx_nlp_hyperthyroidism,
    pmhx_nlp_obesity, pmhx_nlp_smoking_status,
    pmhx_nlp_family_hx_thyroid, pmhx_nlp_family_hx_cancer,
    pmhx_nlp_radiation_exposure, pmhx_nlp_men_syndrome,
    pmhx_nlp_prior_cancer_hx, pmhx_nlp_breast_cancer,
    pmhx_nlp_cad, pmhx_nlp_ckd, pmhx_nlp_copd, pmhx_nlp_depression,
    pmhx_nlp_autoimmune_thyroid_hx,
    -- Molecular (resolved + raw counts)
    braf_positive_final, tert_positive_final, ras_positive_final, ret_positive_unified,
    molecular_tested_confirmed,
    mol_first_test_date, mol_first_test_days_from_surg, mol_n_tests, mol_genes_list,
    mol_has_thyroseq, mol_has_afirma, mol_has_snv, mol_has_fusion,
    mol_n_snvs, mol_n_fusions, mol_n_distinct_genes, mol_variant_classes, mol_platform,
    -- Surgical extent
    surg_total_thyroidectomy,
    -- AGES + ATA
    ages_score, ages_calculable_flag, ata_initial_risk, ata_risk_category AS ata_risk_category_cpm,
    ata_response_category, ata_response_calculable_flag, ata_response_is_provisional,
    -- RAI (full panel)
    rai_received_flag AS rai_received_flag_cpm,
    rai_max_dose_mci, rai_min_dose_mci, rai_dose_v9, rai_intent_v9,
    rai_first_episode_date, rai_first_episode_days_from_surg,
    rai_last_episode_date, rai_last_episode_days_from_surg,
    rai_n_episodes_with_dose, rai_n_distinct_intents, rai_episode_date_span_days,
    rai_avid_flag, rai_avidity, rai_eligible_flag,
    rai_dose_data_available, rai_has_completion_status, rai_has_adjudication,
    -- Operative-note prior history
    ops_prior_neck_irradiation, ops_prior_neck_operation,
    -- Dates / first surgery
    first_surgery_date AS first_surgery_date_cpm,
    -- Follow-up + death
    last_contact_date, last_contact_days_from_surg,
    death_date, death_days_from_surg, followup_or_death_date, followup_or_death_days_from_surg,
    -- Recurrence diagnostics from CPM
    first_recurrence_date AS first_recurrence_date_cpm, first_recurrence_days_from_surg,
    -- Lab dates (TSH / Tg / Calcium / PTH / VitD)
    lab_tsh_first_date, lab_tsh_first_days_from_surg, lab_tsh_last_date, lab_tsh_last_days_from_surg,
    lab_tsh_most_recent_date,
    first_tg_date, first_tg_days_from_surg, last_tg_date, last_tg_days_from_surg,
    max_stimulated_tg_date,
    lab_calcium_first_date, lab_calcium_first_days_from_surg, lab_calcium_last_date,
    lab_pth_first_date, lab_pth_first_days_from_surg, lab_pth_last_date,
    lab_vitd_first_date, lab_vitd_first_days_from_surg,
    -- Imaging dates (US / CT)
    lnus_first_date, lnus_first_days_from_surg, lnus_last_date,
    ct_first_date, ct_first_days_from_surg, ct_last_date,
    -- Cervical LN (ultrasound + surg)
    cnln_earliest_date, cnln_latest_date,
    cnln_img_first_date, cnln_img_last_date,
    cnln_surg_first_date, cnln_surg_last_date,
    -- Med dates
    med_nlp_levothyroxine_date, med_nlp_levothyroxine_days_from_surg,
    med_nlp_calcium_supplement_date, med_nlp_calcium_supplement_days_from_surg,
    med_nlp_calcitriol_date, med_nlp_calcitriol_days_from_surg
  FROM main.canonical_patient_master
)
SELECT
  -- Identity
  c.research_id,

  -- Cohort flags (DERIVED)
  c.strict_dtc_include, c.histology_dtc_5level, c.ete_group, c.lvi_clean, c.vasc_clean,

  -- ETE columns (cohort_m044)
  c.ete_grade_final, c.ete_grade, c.ete_grade_source,
  c.gross_ete_flag, c.path_gross_ete_flag, c.ete_op_note_grade, c.ete_original_grade,

  -- Demographics
  c.age_at_surgery, c.sex, cpm.race, cpm.bmi_combined,

  -- Tumor characteristics
  c.histology_final, c.tumor_size_cm,
  cpm.multifocal_flag_path, cpm.bilateral_disease_flag, cpm.aggressive_variant_flag,
  cpm.margin_involved_any, cpm.closest_margin_mm,
  cpm.capsular_invasion_v6, cpm.capsular_ordinal_worst,
  cpm.syn_hashimoto, cpm.syn_graves, cpm.syn_chronic_thyroiditis,

  -- AJCC staging
  c.ajcc8_t_stage, c.ajcc8_n_stage, c.ajcc8_m_stage, c.ajcc8_stage_group,
  c.ata_risk_category, cpm.ata_initial_risk, cpm.ata_response_category,

  -- Lymph nodes (LN rollup MAX per rid)
  ln.ln_examined, ln.ln_positive, ln.ln_central_examined, ln.ln_central_positive,
  ln.ln_lateral_left_positive, ln.ln_lateral_right_positive, ln.ln_bilateral_lateral_positive,
  ln.ln_level_vi_positive, ln.ln_level_vii_positive, ln.ln_ene,
  CASE WHEN ln.ln_central_positive > 0 THEN 1 ELSE 0 END AS central_pos_flag,
  CASE WHEN COALESCE(ln.ln_lateral_left_positive,0) > 0
         OR COALESCE(ln.ln_lateral_right_positive,0) > 0
         OR COALESCE(ln.ln_bilateral_lateral_positive,0) > 0 THEN 1 ELSE 0 END AS lateral_pos_flag,
  c.ln_positive_flag AS ln_positive_flag_cohort,
  c.ln_total_positive AS ln_total_positive_cohort,

  -- Lymphatic / vascular invasion (raw + cleaned)
  c.lvi_grade, c.vascular_invasion_final,

  -- Molecular
  cpm.braf_positive_final, cpm.tert_positive_final, cpm.ras_positive_final, cpm.ret_positive_unified,
  cpm.molecular_tested_confirmed,
  cpm.mol_first_test_date, cpm.mol_first_test_days_from_surg, cpm.mol_n_tests,
  cpm.mol_genes_list, cpm.mol_platform,
  cpm.mol_has_thyroseq, cpm.mol_has_afirma, cpm.mol_has_snv, cpm.mol_has_fusion,
  cpm.mol_n_snvs, cpm.mol_n_fusions, cpm.mol_n_distinct_genes, cpm.mol_variant_classes,

  -- Comorbidities (PMHx NLP)
  cpm.pmhx_nlp_diabetes, cpm.pmhx_nlp_hypertension, cpm.pmhx_nlp_hypothyroidism, cpm.pmhx_nlp_hyperthyroidism,
  cpm.pmhx_nlp_obesity, cpm.pmhx_nlp_smoking_status,
  cpm.pmhx_nlp_family_hx_thyroid, cpm.pmhx_nlp_family_hx_cancer,
  cpm.pmhx_nlp_radiation_exposure, cpm.pmhx_nlp_men_syndrome,
  cpm.pmhx_nlp_prior_cancer_hx, cpm.pmhx_nlp_breast_cancer,
  cpm.pmhx_nlp_cad, cpm.pmhx_nlp_ckd, cpm.pmhx_nlp_copd, cpm.pmhx_nlp_depression,
  cpm.pmhx_nlp_autoimmune_thyroid_hx,

  -- Surgical extent
  c.surg_procedure_type, cpm.surg_total_thyroidectomy,

  -- AGES
  cpm.ages_score, cpm.ages_calculable_flag,

  -- RAI (full panel)
  c.rai_received_flag,
  cpm.rai_max_dose_mci, cpm.rai_min_dose_mci, cpm.rai_dose_v9, cpm.rai_intent_v9,
  cpm.rai_first_episode_date, cpm.rai_first_episode_days_from_surg,
  cpm.rai_last_episode_date, cpm.rai_last_episode_days_from_surg,
  cpm.rai_n_episodes_with_dose, cpm.rai_n_distinct_intents, cpm.rai_episode_date_span_days,
  cpm.rai_avid_flag, cpm.rai_avidity, cpm.rai_eligible_flag,

  -- Operative-note prior history
  cpm.ops_prior_neck_irradiation, cpm.ops_prior_neck_operation,

  -- Surgery dates (cohort + CPM + reop)
  c.surg_first_date AS surg_first_date_cohort,
  cpm.first_surgery_date_cpm,
  reop.second_surgery_date,
  reop.days_to_2nd,
  reop.n_surgeries, reop.completion_reason, reop.completion_reason_confidence,
  reop.completion_histology_type, reop.op_reoperative_any,

  -- Follow-up
  c.followup_years, c.overall_survival_years, c.death_occurred,
  cpm.last_contact_date, cpm.last_contact_days_from_surg,
  cpm.death_date, cpm.death_days_from_surg,
  cpm.followup_or_death_date, cpm.followup_or_death_days_from_surg,

  -- Recurrence (PRIMARY ENDPOINT)
  rec.recurrence_path_proven, rec.recurrence_path_proven_date, rec.recurrence_path_proven_source,
  rec.recurrence_path_proven_evidence, rec.days_to_path_proven,
  rec.recurrence_imaging_suspicious, rec.recurrence_imaging_suspicious_date,
  rec.recurrence_imaging_modality, rec.recurrence_imaging_modality_summary,
  rec.recurrence_imaging_source, rec.recurrence_imaging_finding_text,
  rec.recurrence_imaging_n_events, rec.days_to_imaging_suspicious,
  rec.recurrence_imaging_then_path_confirmed, rec.recurrence_status_final,
  rec.is_implausible_date_quarantine,
  CASE WHEN rec.recurrence_status_final IN ('path_proven','imaging_only_unconfirmed') THEN TRUE ELSE FALSE END AS recurrence_composite,
  cpm.first_recurrence_date_cpm, cpm.first_recurrence_days_from_surg,
  c.any_recurrence_flag AS legacy_any_recurrence_flag,
  c.structural_recurrence_flag AS legacy_structural_recurrence_flag,

  -- Lab first/last dates
  cpm.lab_tsh_first_date, cpm.lab_tsh_first_days_from_surg, cpm.lab_tsh_last_date, cpm.lab_tsh_most_recent_date,
  cpm.first_tg_date, cpm.first_tg_days_from_surg, cpm.last_tg_date, cpm.max_stimulated_tg_date,
  cpm.lab_calcium_first_date, cpm.lab_calcium_first_days_from_surg, cpm.lab_calcium_last_date,
  cpm.lab_pth_first_date, cpm.lab_pth_first_days_from_surg, cpm.lab_pth_last_date,
  cpm.lab_vitd_first_date, cpm.lab_vitd_first_days_from_surg,

  -- Imaging first/last dates
  cpm.lnus_first_date, cpm.lnus_first_days_from_surg, cpm.lnus_last_date,
  cpm.ct_first_date, cpm.ct_first_days_from_surg, cpm.ct_last_date,
  cpm.cnln_earliest_date, cpm.cnln_latest_date,
  cpm.cnln_img_first_date, cpm.cnln_img_last_date,
  cpm.cnln_surg_first_date, cpm.cnln_surg_last_date,

  -- Med dates
  cpm.med_nlp_levothyroxine_date, cpm.med_nlp_levothyroxine_days_from_surg,
  cpm.med_nlp_calcium_supplement_date, cpm.med_nlp_calcium_supplement_days_from_surg,
  cpm.med_nlp_calcitriol_date, cpm.med_nlp_calcitriol_days_from_surg

FROM cohort c
LEFT JOIN ln   USING (research_id)
LEFT JOIN reop USING (research_id)
LEFT JOIN rec  USING (research_id)
LEFT JOIN cpm  USING (research_id)
ORDER BY c.research_id;
"""


# Source map: derived from the SQL SELECT list above
SOURCE_MAP = [
    # (column, source_database, source_object, source_column, aggregation_rule, type, used_in_models, notes)
    ("research_id", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "research_id", "PASS", "VARCHAR", "ALL (key)", "Patient identifier; joins all sources."),

    # DERIVED flags
    ("strict_dtc_include", DB, "DERIVED", "histology_final (CASE)", "DERIVED", "INT", "ALL strict-DTC fits", "1 if histology_final NOT IN exclusion list (MTC, anaplastic, NIFTP, FTUMP, follicular adenoma, atypical Hurthle, NUT, adenoid cystic). Per Standing Rule #5."),
    ("histology_dtc_5level", DB, "DERIVED", "histology_final (CASE)", "DERIVED", "VARCHAR", "Refit #3 histology factor", "5-level: PTC / FTC / Metastatic-PTC / Poorly-differentiated DTC / High-grade DTC; non-DTC labelled separately."),
    ("ete_group", DB, "DERIVED", "ete_grade_final (CASE)", "DERIVED", "VARCHAR", "PRIMARY EXPOSURE", "5-level: Gross / Microscopic / No-negative / Present-ungraded / Missing-other. Reference = Microscopic."),
    ("lvi_clean", DB, "DERIVED", "lvi_grade (CASE / regex)", "DERIVED", "VARCHAR", "Primary logistic + Cox", "extensive / present / focal / missing / indeterminate (collapses 'preesent', 'extensivre', etc.)."),
    ("vasc_clean", DB, "DERIVED", "vascular_invasion_final (COALESCE)", "DERIVED", "VARCHAR", "Primary logistic + Cox", "COALESCE(., 'missing'); preserves present_ungraded/focal/extensive/indeterminate."),

    # ETE
    ("ete_grade_final", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "ete_grade_final", "PASS", "VARCHAR", "PRIMARY EXPOSURE source", "Resolved ETE grade (extraction_audit_engine_v7 + script_390_rule_a_20260422 + tumor_episode_master_v2)."),
    ("ete_grade", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "ete_grade", "PASS", "VARCHAR", "ETE secondary tie-break", "Secondary ETE grade column."),
    ("ete_grade_source", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "ete_grade_source", "PASS", "VARCHAR", "Provenance", "Source/engine that produced ete_grade_final."),
    ("gross_ete_flag", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "gross_ete_flag", "PASS", "BOOLEAN", "Gross-ETE indicator", "Boolean any-gross-ETE."),
    ("path_gross_ete_flag", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "path_gross_ete_flag", "PASS", "BIGINT", "Path gross-ETE 0/1", ""),
    ("ete_op_note_grade", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "ete_op_note_grade", "PASS", "VARCHAR", "OPnote ETE provenance", ""),
    ("ete_original_grade", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "ete_original_grade", "PASS", "VARCHAR", "Pre-resolution ETE", ""),

    # Demographics
    ("age_at_surgery", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "age_at_surgery", "PASS", "BIGINT", "PRIMARY covariate", "Age at first surgery, years."),
    ("sex", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "sex", "PASS", "VARCHAR", "PRIMARY covariate", "female / male."),
    ("race", DB, "main.canonical_patient_master", "race", "PASS", "VARCHAR", "Table 1 only", "Race (4,124/4,128 populated)."),
    ("bmi_combined", DB, "main.canonical_patient_master", "bmi_combined", "PASS", "DOUBLE", "Sensitivity (80% missing)", "BMI combined source; not primary."),

    # Tumor
    ("histology_final", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "histology_final", "PASS", "VARCHAR", "Strict-DTC filter + factor", "Raw resolved histology."),
    ("tumor_size_cm", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "tumor_size_cm", "PASS", "DOUBLE", "PRIMARY covariate", "Largest tumor dimension cm."),
    ("multifocal_flag_path", DB, "main.canonical_patient_master", "multifocal_flag_path", "PASS", "BOOLEAN", "Table 1 + sensitivity", "Multifocal disease (pathology)."),
    ("bilateral_disease_flag", DB, "main.canonical_patient_master", "bilateral_disease_flag", "PASS", "BOOLEAN", "Table 1 + sensitivity", "Bilateral thyroid involvement."),
    ("aggressive_variant_flag", DB, "main.canonical_patient_master", "aggressive_variant_flag", "PASS", "BOOLEAN", "Table 1", "Aggressive histologic variant."),
    ("margin_involved_any", DB, "main.canonical_patient_master", "margin_involved_any", "PASS", "BOOLEAN", "Table 1 + Discussion", "Any positive surgical margin."),
    ("closest_margin_mm", DB, "main.canonical_patient_master", "closest_margin_mm", "PASS", "DOUBLE", "Table 1 + Discussion", "Closest margin distance, mm."),
    ("capsular_invasion_v6", DB, "main.canonical_patient_master", "capsular_invasion_v6", "PASS", "VARCHAR", "Reference / sensitivity", "Capsular invasion v6 categorical."),
    ("capsular_ordinal_worst", DB, "main.canonical_patient_master", "capsular_ordinal_worst", "PASS", "INT", "Reference / sensitivity", "Worst-ordinal capsular invasion."),
    ("syn_hashimoto", DB, "main.canonical_patient_master", "syn_hashimoto", "PASS", "BOOLEAN", "Table 1", "Hashimoto thyroiditis (synoptic)."),
    ("syn_graves", DB, "main.canonical_patient_master", "syn_graves", "PASS", "BOOLEAN", "Table 1", "Graves disease (synoptic)."),
    ("syn_chronic_thyroiditis", DB, "main.canonical_patient_master", "syn_chronic_thyroiditis", "PASS", "BOOLEAN", "Table 1", "Chronic thyroiditis (synoptic)."),

    # Stage
    ("ajcc8_t_stage", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "ajcc8_t_stage", "PASS", "VARCHAR", "Table 1 + stratified", "AJCC 8 T category."),
    ("ajcc8_n_stage", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "ajcc8_n_stage", "PASS", "VARCHAR", "PRIMARY covariate (4-level)", "AJCC 8 N category. Levels N0/N1a/N1b/Nx/missing in models."),
    ("ajcc8_m_stage", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "ajcc8_m_stage", "PASS", "VARCHAR", "Table 1", "AJCC 8 M category."),
    ("ajcc8_stage_group", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "ajcc8_stage_group", "PASS", "VARCHAR", "Table 1", "AJCC 8 stage group."),
    ("ata_risk_category", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "ata_risk_category", "PASS", "VARCHAR", "Table 1", "ATA risk (cohort view)."),
    ("ata_initial_risk", DB, "main.canonical_patient_master", "ata_initial_risk", "PASS", "VARCHAR", "Table 1 + sensitivity", "ATA initial risk (CPM)."),
    ("ata_response_category", DB, "main.canonical_patient_master", "ata_response_category", "PASS", "VARCHAR", "Reference", "ATA dynamic response."),

    # LN
    ("ln_examined", DB, "manuscript_workspace.ln_master_rollup_v1", "ln_total_examined", "MAX per research_id", "INT", "Table 1", "Total LN examined."),
    ("ln_positive", DB, "manuscript_workspace.ln_master_rollup_v1", "ln_total_positive", "MAX per research_id", "INT", "Table 1", "Total LN positive."),
    ("ln_central_examined", DB, "manuscript_workspace.ln_master_rollup_v1", "ln_central_examined", "MAX per research_id", "INT", "Reference", ""),
    ("ln_central_positive", DB, "manuscript_workspace.ln_master_rollup_v1", "ln_central_positive", "MAX per research_id", "INT", "Reference (drives central_pos_flag)", ""),
    ("ln_lateral_left_positive", DB, "manuscript_workspace.ln_master_rollup_v1", "ln_lateral_left_positive", "MAX per research_id", "INT", "Reference (drives lateral_pos_flag)", ""),
    ("ln_lateral_right_positive", DB, "manuscript_workspace.ln_master_rollup_v1", "ln_lateral_right_positive", "MAX per research_id", "INT", "Reference (drives lateral_pos_flag)", ""),
    ("ln_bilateral_lateral_positive", DB, "manuscript_workspace.ln_master_rollup_v1", "ln_bilateral_lateral_positive", "MAX per research_id", "INT", "Reference (drives lateral_pos_flag)", ""),
    ("ln_level_vi_positive", DB, "manuscript_workspace.ln_master_rollup_v1", "ln_level_vi_positive", "MAX per research_id", "INT", "Reference", ""),
    ("ln_level_vii_positive", DB, "manuscript_workspace.ln_master_rollup_v1", "ln_level_vii_positive", "MAX per research_id", "INT", "Reference", ""),
    ("ln_ene", DB, "manuscript_workspace.ln_master_rollup_v1", "ln_extranodal_extension", "MAX per research_id", "INT", "Reference", "Extranodal extension flag."),
    ("central_pos_flag", DB, "DERIVED", "ln_central_positive", "DERIVED", "INT", "Refit #5 no-neg subgroup", "1 if MAX(ln_central_positive) > 0."),
    ("lateral_pos_flag", DB, "DERIVED", "ln_lateral_*_positive", "DERIVED", "INT", "Refit #5 no-neg subgroup", "1 if any of MAX(left/right/bilateral) > 0."),
    ("ln_positive_flag_cohort", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "ln_positive_flag", "PASS", "INT", "Reference", "Legacy LN flag from cohort view."),
    ("ln_total_positive_cohort", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "ln_total_positive", "PASS", "INT", "Reference", "Legacy LN count from cohort view."),

    # LVI / vascular (raw)
    ("lvi_grade", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "lvi_grade", "PASS", "VARCHAR", "Reference (cleaned in lvi_clean)", "Raw lymphatic invasion (free-text)."),
    ("vascular_invasion_final", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "vascular_invasion_final", "PASS", "VARCHAR", "Reference (cleaned in vasc_clean)", "Raw resolved vascular invasion."),

    # Molecular
    ("braf_positive_final", DB, "main.canonical_patient_master", "braf_positive_final", "PASS", "BOOLEAN", "Table 1", "Resolved BRAF V600E."),
    ("tert_positive_final", DB, "main.canonical_patient_master", "tert_positive_final", "PASS", "BOOLEAN", "Table 1", "Resolved TERT promoter."),
    ("ras_positive_final", DB, "main.canonical_patient_master", "ras_positive_final", "PASS", "BOOLEAN", "Table 1", "Resolved RAS."),
    ("ret_positive_unified", DB, "main.canonical_patient_master", "ret_positive_unified", "PASS", "BOOLEAN", "Table 1", "Unified RET fusion/mutation."),
    ("molecular_tested_confirmed", DB, "main.canonical_patient_master", "molecular_tested_confirmed", "PASS", "BOOLEAN", "Denominator note", "Molecular testing confirmed."),
    ("mol_first_test_date", DB, "main.canonical_patient_master", "mol_first_test_date", "PASS", "DATE", "Reference", "Date of first molecular test."),
    ("mol_first_test_days_from_surg", DB, "main.canonical_patient_master", "mol_first_test_days_from_surg", "PASS", "INT", "Reference", "Days first surg → first mol test."),
    ("mol_n_tests", DB, "main.canonical_patient_master", "mol_n_tests", "PASS", "BIGINT", "Reference", ""),
    ("mol_genes_list", DB, "main.canonical_patient_master", "mol_genes_list", "PASS", "VARCHAR", "Reference", "Comma-separated genes tested/found."),
    ("mol_platform", DB, "main.canonical_patient_master", "mol_platform", "PASS", "VARCHAR", "Reference", ""),
    ("mol_has_thyroseq", DB, "main.canonical_patient_master", "mol_has_thyroseq", "PASS", "BOOLEAN", "Reference", ""),
    ("mol_has_afirma", DB, "main.canonical_patient_master", "mol_has_afirma", "PASS", "BOOLEAN", "Reference", ""),
    ("mol_has_snv", DB, "main.canonical_patient_master", "mol_has_snv", "PASS", "BOOLEAN", "Reference", ""),
    ("mol_has_fusion", DB, "main.canonical_patient_master", "mol_has_fusion", "PASS", "BOOLEAN", "Reference", ""),
    ("mol_n_snvs", DB, "main.canonical_patient_master", "mol_n_snvs", "PASS", "BIGINT", "Reference", ""),
    ("mol_n_fusions", DB, "main.canonical_patient_master", "mol_n_fusions", "PASS", "BIGINT", "Reference", ""),
    ("mol_n_distinct_genes", DB, "main.canonical_patient_master", "mol_n_distinct_genes", "PASS", "BIGINT", "Reference", ""),
    ("mol_variant_classes", DB, "main.canonical_patient_master", "mol_variant_classes", "PASS", "VARCHAR", "Reference", ""),

    # Comorbidities
    ("pmhx_nlp_diabetes", DB, "main.canonical_patient_master", "pmhx_nlp_diabetes", "PASS", "BOOLEAN", "Table 1", ""),
    ("pmhx_nlp_hypertension", DB, "main.canonical_patient_master", "pmhx_nlp_hypertension", "PASS", "BOOLEAN", "Table 1", ""),
    ("pmhx_nlp_hypothyroidism", DB, "main.canonical_patient_master", "pmhx_nlp_hypothyroidism", "PASS", "BOOLEAN", "Table 1", ""),
    ("pmhx_nlp_hyperthyroidism", DB, "main.canonical_patient_master", "pmhx_nlp_hyperthyroidism", "PASS", "BOOLEAN", "Table 1", ""),
    ("pmhx_nlp_obesity", DB, "main.canonical_patient_master", "pmhx_nlp_obesity", "PASS", "BOOLEAN", "Table 1", ""),
    ("pmhx_nlp_smoking_status", DB, "main.canonical_patient_master", "pmhx_nlp_smoking_status", "PASS", "VARCHAR", "Limitation (99.7% NULL)", ""),
    ("pmhx_nlp_family_hx_thyroid", DB, "main.canonical_patient_master", "pmhx_nlp_family_hx_thyroid", "PASS", "BOOLEAN", "Limitation (under-extracted)", ""),
    ("pmhx_nlp_family_hx_cancer", DB, "main.canonical_patient_master", "pmhx_nlp_family_hx_cancer", "PASS", "BOOLEAN", "Limitation (under-extracted)", ""),
    ("pmhx_nlp_radiation_exposure", DB, "main.canonical_patient_master", "pmhx_nlp_radiation_exposure", "PASS", "BOOLEAN", "Limitation (under-extracted)", ""),
    ("pmhx_nlp_men_syndrome", DB, "main.canonical_patient_master", "pmhx_nlp_men_syndrome", "PASS", "BOOLEAN", "Reference", ""),
    ("pmhx_nlp_prior_cancer_hx", DB, "main.canonical_patient_master", "pmhx_nlp_prior_cancer_hx", "PASS", "BOOLEAN", "Reference", ""),
    ("pmhx_nlp_breast_cancer", DB, "main.canonical_patient_master", "pmhx_nlp_breast_cancer", "PASS", "BOOLEAN", "Reference", ""),
    ("pmhx_nlp_cad", DB, "main.canonical_patient_master", "pmhx_nlp_cad", "PASS", "BOOLEAN", "Reference", ""),
    ("pmhx_nlp_ckd", DB, "main.canonical_patient_master", "pmhx_nlp_ckd", "PASS", "BOOLEAN", "Reference", ""),
    ("pmhx_nlp_copd", DB, "main.canonical_patient_master", "pmhx_nlp_copd", "PASS", "BOOLEAN", "Reference", ""),
    ("pmhx_nlp_depression", DB, "main.canonical_patient_master", "pmhx_nlp_depression", "PASS", "BOOLEAN", "Reference", ""),
    ("pmhx_nlp_autoimmune_thyroid_hx", DB, "main.canonical_patient_master", "pmhx_nlp_autoimmune_thyroid_hx", "PASS", "BOOLEAN", "Reference", ""),

    # Surgical
    ("surg_procedure_type", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "surg_procedure_type", "PASS", "VARCHAR", "Table 1", ""),
    ("surg_total_thyroidectomy", DB, "main.canonical_patient_master", "surg_total_thyroidectomy", "PASS", "BOOLEAN", "Table 1 + Discussion", "62.0% gross / 47.9% micro / 21.9% no-neg."),

    # AGES
    ("ages_score", DB, "main.canonical_patient_master", "ages_score", "PASS", "DOUBLE", "Table 1", "Mayo AGES."),
    ("ages_calculable_flag", DB, "main.canonical_patient_master", "ages_calculable_flag", "PASS", "BOOLEAN", "Reference", ""),

    # RAI
    ("rai_received_flag", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "rai_received_flag", "PASS", "BOOLEAN", "Sensitivity ONLY (with-RAI)", "Confounded by indication."),
    ("rai_max_dose_mci", DB, "main.canonical_patient_master", "rai_max_dose_mci", "PASS", "DOUBLE", "Reference", ""),
    ("rai_min_dose_mci", DB, "main.canonical_patient_master", "rai_min_dose_mci", "PASS", "DOUBLE", "Reference", ""),
    ("rai_dose_v9", DB, "main.canonical_patient_master", "rai_dose_v9", "PASS", "VARCHAR", "Reference", ""),
    ("rai_intent_v9", DB, "main.canonical_patient_master", "rai_intent_v9", "PASS", "VARCHAR", "Reference", "adjuvant / therapy / remnant."),
    ("rai_first_episode_date", DB, "main.canonical_patient_master", "rai_first_episode_date", "PASS", "DATE", "Reference", ""),
    ("rai_first_episode_days_from_surg", DB, "main.canonical_patient_master", "rai_first_episode_days_from_surg", "PASS", "BIGINT", "Reference", ""),
    ("rai_last_episode_date", DB, "main.canonical_patient_master", "rai_last_episode_date", "PASS", "DATE", "Reference", ""),
    ("rai_last_episode_days_from_surg", DB, "main.canonical_patient_master", "rai_last_episode_days_from_surg", "PASS", "BIGINT", "Reference", ""),
    ("rai_n_episodes_with_dose", DB, "main.canonical_patient_master", "rai_n_episodes_with_dose", "PASS", "BIGINT", "Reference", ""),
    ("rai_n_distinct_intents", DB, "main.canonical_patient_master", "rai_n_distinct_intents", "PASS", "BIGINT", "Reference", ""),
    ("rai_episode_date_span_days", DB, "main.canonical_patient_master", "rai_episode_date_span_days", "PASS", "BIGINT", "Reference", ""),
    ("rai_avid_flag", DB, "main.canonical_patient_master", "rai_avid_flag", "PASS", "BOOLEAN", "Reference", ""),
    ("rai_avidity", DB, "main.canonical_patient_master", "rai_avidity", "PASS", "VARCHAR", "Reference", ""),
    ("rai_eligible_flag", DB, "main.canonical_patient_master", "rai_eligible_flag", "PASS", "BOOLEAN", "Reference", ""),

    # Operative-note prior history
    ("ops_prior_neck_irradiation", DB, "main.canonical_patient_master", "ops_prior_neck_irradiation", "PASS", "BOOLEAN", "Reference", ""),
    ("ops_prior_neck_operation", DB, "main.canonical_patient_master", "ops_prior_neck_operation", "PASS", "BOOLEAN", "Reference", ""),

    # Surgery dates
    ("surg_first_date_cohort", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "surg_first_date", "PASS", "DATE", "Cox subset filter", "Date of first surgery (cohort)."),
    ("first_surgery_date_cpm", DB, "main.canonical_patient_master", "first_surgery_date", "PASS", "DATE", "Reference (compare to cohort)", "Date of first surgery (CPM)."),
    ("second_surgery_date", DB, "manuscript_workspace.cohort_m040_reoperative_v1", "second_surgery_date", "MAX per research_id", "DATE", "Refit #5 covariate", "Second surgery date."),
    ("days_to_2nd", DB, "manuscript_workspace.cohort_m040_reoperative_v1", "days_between_first_second_surgery", "MAX per research_id", "INT", "Refit #5 covariate", "Days first → second surgery."),
    ("n_surgeries", DB, "manuscript_workspace.cohort_m040_reoperative_v1", "n_surgeries", "MAX per research_id", "INT", "Refit #5 covariate", "Number of surgeries."),
    ("completion_reason", DB, "manuscript_workspace.cohort_m040_reoperative_v1", "completion_reason", "MAX per research_id", "VARCHAR", "Reference", ""),
    ("completion_reason_confidence", DB, "manuscript_workspace.cohort_m040_reoperative_v1", "completion_reason_confidence", "MAX per research_id", "VARCHAR", "Reference", ""),
    ("completion_histology_type", DB, "manuscript_workspace.cohort_m040_reoperative_v1", "completion_histology_type", "MAX per research_id", "VARCHAR", "Reference", ""),
    ("op_reoperative_any", DB, "manuscript_workspace.cohort_m040_reoperative_v1", "op_reoperative_any", "MAX per research_id", "BOOLEAN", "Reference", ""),

    # Follow-up
    ("followup_years", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "followup_years", "PASS", "DOUBLE", "Cox time variable", "Follow-up time, years."),
    ("overall_survival_years", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "overall_survival_years", "PASS", "DOUBLE", "Reference", ""),
    ("death_occurred", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "death_occurred", "PASS", "BOOLEAN", "Reference", ""),
    ("last_contact_date", DB, "main.canonical_patient_master", "last_contact_date", "PASS", "DATE", "Reference", ""),
    ("last_contact_days_from_surg", DB, "main.canonical_patient_master", "last_contact_days_from_surg", "PASS", "INT", "Reference", ""),
    ("death_date", DB, "main.canonical_patient_master", "death_date", "PASS", "DATE", "Reference", ""),
    ("death_days_from_surg", DB, "main.canonical_patient_master", "death_days_from_surg", "PASS", "INT", "Reference", ""),
    ("followup_or_death_date", DB, "main.canonical_patient_master", "followup_or_death_date", "PASS", "DATE", "Reference", ""),
    ("followup_or_death_days_from_surg", DB, "main.canonical_patient_master", "followup_or_death_days_from_surg", "PASS", "INT", "Reference", ""),

    # Recurrence
    ("recurrence_path_proven", DB, "main.canonical_recurrence_resolved_v1", "recurrence_path_proven", "PASS", "BOOLEAN", "PRIMARY ENDPOINT", "Path-proven recurrence (PRIMARY)."),
    ("recurrence_path_proven_date", DB, "main.canonical_recurrence_resolved_v1", "recurrence_path_proven_date", "PASS", "DATE", "Cox event date", "Date of path-proven event."),
    ("recurrence_path_proven_source", DB, "main.canonical_recurrence_resolved_v1", "recurrence_path_proven_source", "PASS", "VARCHAR", "Provenance", ""),
    ("recurrence_path_proven_evidence", DB, "main.canonical_recurrence_resolved_v1", "recurrence_path_proven_evidence", "PASS", "VARCHAR", "Provenance", "Evidence text."),
    ("days_to_path_proven", DB, "main.canonical_recurrence_resolved_v1", "days_to_path_proven", "PASS", "BIGINT", "Cox time-to-event (alt)", ""),
    ("recurrence_imaging_suspicious", DB, "main.canonical_recurrence_resolved_v1", "recurrence_imaging_suspicious", "PASS", "BOOLEAN", "Secondary endpoint", "Imaging-suspicious flag."),
    ("recurrence_imaging_suspicious_date", DB, "main.canonical_recurrence_resolved_v1", "recurrence_imaging_suspicious_date", "PASS", "DATE", "Reference", ""),
    ("recurrence_imaging_modality", DB, "main.canonical_recurrence_resolved_v1", "recurrence_imaging_modality", "PASS", "VARCHAR", "Reference", ""),
    ("recurrence_imaging_modality_summary", DB, "main.canonical_recurrence_resolved_v1", "recurrence_imaging_modality_summary", "PASS", "VARCHAR", "Reference", ""),
    ("recurrence_imaging_source", DB, "main.canonical_recurrence_resolved_v1", "recurrence_imaging_source", "PASS", "VARCHAR", "Provenance", ""),
    ("recurrence_imaging_finding_text", DB, "main.canonical_recurrence_resolved_v1", "recurrence_imaging_finding_text", "PASS", "VARCHAR", "Provenance", ""),
    ("recurrence_imaging_n_events", DB, "main.canonical_recurrence_resolved_v1", "recurrence_imaging_n_events", "PASS", "BIGINT", "Reference", ""),
    ("days_to_imaging_suspicious", DB, "main.canonical_recurrence_resolved_v1", "days_to_imaging_suspicious", "PASS", "BIGINT", "Reference", ""),
    ("recurrence_imaging_then_path_confirmed", DB, "main.canonical_recurrence_resolved_v1", "recurrence_imaging_then_path_confirmed", "PASS", "BOOLEAN", "Reference", "Imaging finding subsequently confirmed by pathology."),
    ("recurrence_status_final", DB, "main.canonical_recurrence_resolved_v1", "recurrence_status_final", "PASS", "VARCHAR", "Composite endpoint source", "none / imaging_only_unconfirmed / path_proven."),
    ("is_implausible_date_quarantine", DB, "main.canonical_recurrence_resolved_v1", "is_implausible_date_quarantine", "PASS", "BOOLEAN", "QA filter", "True if implausible date detected (132 quarantined)."),
    ("recurrence_composite", DB, "DERIVED", "recurrence_status_final (CASE)", "DERIVED", "BOOLEAN", "Pre-specified secondary", "TRUE if status_final IN ('path_proven','imaging_only_unconfirmed')."),
    ("first_recurrence_date_cpm", DB, "main.canonical_patient_master", "first_recurrence_date", "PASS", "DATE", "Reference (compare to canonical)", ""),
    ("first_recurrence_days_from_surg", DB, "main.canonical_patient_master", "first_recurrence_days_from_surg", "PASS", "INT", "Reference", ""),
    ("legacy_any_recurrence_flag", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "any_recurrence_flag", "PASS", "BOOLEAN", "SENSITIVITY ONLY", "LEGACY; 318/503 inconsistent with canonical resolved."),
    ("legacy_structural_recurrence_flag", DB, "manuscript_workspace.cohort_m044_ajcc_ete_v1", "structural_recurrence_flag", "PASS", "BOOLEAN", "SENSITIVITY ONLY", "LEGACY; 1467/1819 inconsistent with canonical resolved."),

    # Lab dates
    ("lab_tsh_first_date", DB, "main.canonical_patient_master", "lab_tsh_first_date", "PASS", "DATE", "Reference", "First TSH lab date."),
    ("lab_tsh_first_days_from_surg", DB, "main.canonical_patient_master", "lab_tsh_first_days_from_surg", "PASS", "INT", "Reference", ""),
    ("lab_tsh_last_date", DB, "main.canonical_patient_master", "lab_tsh_last_date", "PASS", "DATE", "Reference", ""),
    ("lab_tsh_most_recent_date", DB, "main.canonical_patient_master", "lab_tsh_most_recent_date", "PASS", "DATE", "Reference", ""),
    ("first_tg_date", DB, "main.canonical_patient_master", "first_tg_date", "PASS", "DATE", "Reference", "First Tg date."),
    ("first_tg_days_from_surg", DB, "main.canonical_patient_master", "first_tg_days_from_surg", "PASS", "INT", "Reference", ""),
    ("last_tg_date", DB, "main.canonical_patient_master", "last_tg_date", "PASS", "DATE", "Reference", ""),
    ("max_stimulated_tg_date", DB, "main.canonical_patient_master", "max_stimulated_tg_date", "PASS", "DATE", "Reference", ""),
    ("lab_calcium_first_date", DB, "main.canonical_patient_master", "lab_calcium_first_date", "PASS", "DATE", "Reference", ""),
    ("lab_calcium_first_days_from_surg", DB, "main.canonical_patient_master", "lab_calcium_first_days_from_surg", "PASS", "INT", "Reference", ""),
    ("lab_calcium_last_date", DB, "main.canonical_patient_master", "lab_calcium_last_date", "PASS", "DATE", "Reference", ""),
    ("lab_pth_first_date", DB, "main.canonical_patient_master", "lab_pth_first_date", "PASS", "DATE", "Reference", ""),
    ("lab_pth_first_days_from_surg", DB, "main.canonical_patient_master", "lab_pth_first_days_from_surg", "PASS", "INT", "Reference", ""),
    ("lab_pth_last_date", DB, "main.canonical_patient_master", "lab_pth_last_date", "PASS", "DATE", "Reference", ""),
    ("lab_vitd_first_date", DB, "main.canonical_patient_master", "lab_vitd_first_date", "PASS", "DATE", "Reference", ""),
    ("lab_vitd_first_days_from_surg", DB, "main.canonical_patient_master", "lab_vitd_first_days_from_surg", "PASS", "INT", "Reference", ""),

    # Imaging dates
    ("lnus_first_date", DB, "main.canonical_patient_master", "lnus_first_date", "PASS", "DATE", "Reference", "First LN ultrasound."),
    ("lnus_first_days_from_surg", DB, "main.canonical_patient_master", "lnus_first_days_from_surg", "PASS", "INT", "Reference", ""),
    ("lnus_last_date", DB, "main.canonical_patient_master", "lnus_last_date", "PASS", "DATE", "Reference", ""),
    ("ct_first_date", DB, "main.canonical_patient_master", "ct_first_date", "PASS", "DATE", "Reference", "First CT scan."),
    ("ct_first_days_from_surg", DB, "main.canonical_patient_master", "ct_first_days_from_surg", "PASS", "INT", "Reference", ""),
    ("ct_last_date", DB, "main.canonical_patient_master", "ct_last_date", "PASS", "DATE", "Reference", ""),
    ("cnln_earliest_date", DB, "main.canonical_patient_master", "cnln_earliest_date", "PASS", "DATE", "Reference", "Earliest cervical-LN documentation date."),
    ("cnln_latest_date", DB, "main.canonical_patient_master", "cnln_latest_date", "PASS", "DATE", "Reference", ""),
    ("cnln_img_first_date", DB, "main.canonical_patient_master", "cnln_img_first_date", "PASS", "DATE", "Reference", ""),
    ("cnln_img_last_date", DB, "main.canonical_patient_master", "cnln_img_last_date", "PASS", "DATE", "Reference", ""),
    ("cnln_surg_first_date", DB, "main.canonical_patient_master", "cnln_surg_first_date", "PASS", "DATE", "Reference", ""),
    ("cnln_surg_last_date", DB, "main.canonical_patient_master", "cnln_surg_last_date", "PASS", "DATE", "Reference", ""),

    # Med dates
    ("med_nlp_levothyroxine_date", DB, "main.canonical_patient_master", "med_nlp_levothyroxine_date", "PASS", "DATE", "Reference", ""),
    ("med_nlp_levothyroxine_days_from_surg", DB, "main.canonical_patient_master", "med_nlp_levothyroxine_days_from_surg", "PASS", "INT", "Reference", ""),
    ("med_nlp_calcium_supplement_date", DB, "main.canonical_patient_master", "med_nlp_calcium_supplement_date", "PASS", "DATE", "Reference", ""),
    ("med_nlp_calcium_supplement_days_from_surg", DB, "main.canonical_patient_master", "med_nlp_calcium_supplement_days_from_surg", "PASS", "INT", "Reference", ""),
    ("med_nlp_calcitriol_date", DB, "main.canonical_patient_master", "med_nlp_calcitriol_date", "PASS", "DATE", "Reference", ""),
    ("med_nlp_calcitriol_days_from_surg", DB, "main.canonical_patient_master", "med_nlp_calcitriol_days_from_surg", "PASS", "INT", "Reference", ""),
]


# ---------------------------------------------------------------------------
# openpyxl helpers
# ---------------------------------------------------------------------------

def _safe(v):
    try:
        if v is None: return None
        if isinstance(v, (str, int, bool, bytes)): return v
        if isinstance(v, float):
            import math
            if math.isnan(v): return None
            return v
        from datetime import datetime as dt, date as dt_date
        if isinstance(v, (dt, dt_date)): return v
        try:
            if pd.isna(v): return None
        except Exception: pass
        return str(v)
    except Exception:
        return str(v)


def style_hdr(ws, n):
    for c in range(1, n+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HF; cell.font = HFONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BD


def autosize(ws, df, cap=48):
    for i, c in enumerate(df.columns, start=1):
        try:
            sample = df[c].astype(str).head(200)
            mx = max([len(str(c))] + sample.map(len).tolist())
        except Exception:
            mx = max(len(str(c)), 18)
        ws.column_dimensions[get_column_letter(i)].width = min(max(mx + 2, 12), cap)


def write_df(ws, df):
    df = df.copy()
    for c in df.columns:
        if df[c].dtype.kind == "O":
            df[c] = df[c].where(df[c].notna(), None)
        elif "datetime" in str(df[c].dtype):
            try: df[c] = df[c].dt.tz_localize(None)
            except (TypeError, AttributeError): pass
            df[c] = df[c].where(df[c].notna(), None)
    for c_idx, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=c_idx, value=str(col))
    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_safe(value))
    style_hdr(ws, len(df.columns))
    autosize(ws, df)
    ws.freeze_panes = ws.cell(row=2, column=2)


# ---------------------------------------------------------------------------
def main():
    tok = os.environ['MDT']
    print(f"-> Connecting to {DB}", flush=True)
    con = duckdb.connect(f"md:{DB}?motherduck_token={tok}")

    print("-> Pulling per-patient master ...", flush=True)
    pm = con.execute(ANALYTIC_SQL).fetchdf()
    print(f"   shape = {pm.shape}", flush=True)
    assert len(pm) == 4128, f"got {len(pm)} rows, expected 4128"

    smap = pd.DataFrame(SOURCE_MAP, columns=[
        "column_name", "source_database", "source_object", "source_column",
        "aggregation_rule", "data_type", "used_in_models", "notes"
    ])

    # Cross-check that source map covers the per-patient master
    in_pm  = set(pm.columns)
    in_map = set(smap['column_name'])
    missing_in_map = sorted(in_pm - in_map)
    extra_in_map   = sorted(in_map - in_pm)
    print(f"   columns missing from source map: {len(missing_in_map)}")
    if missing_in_map: print("     " + ", ".join(missing_in_map[:25]))
    print(f"   columns in source map not in pm: {len(extra_in_map)}")
    if extra_in_map: print("     " + ", ".join(extra_in_map[:25]))

    # Source raw dumps already exist in M044_ETE_master_data.xlsx; do not duplicate.
    print("-> Building workbook ...", flush=True)
    wb = Workbook()
    ws = wb.active; ws.title = "README"
    ws["A1"] = "M044 - Per-patient analytic table with column source map"; ws["A1"].font = TF
    ws["A2"] = ("One row per research_id (n=4,128). Sheet 1 = patient analytic. "
                "Sheet 2 = column source map. Sheet 3 = audit (columns missing from map / extra).")
    ws["A2"].font = SF
    notes = [
        ("",""),
        ("Date prepared", datetime.utcnow().strftime("%Y-%m-%d")),
        ("Database", DB),
        ("Cohort source", "manuscript_workspace.cohort_m044_ajcc_ete_v1 (n = 4,128)"),
        ("Sheet 1 columns", str(len(pm.columns))),
        ("Source-map rows", str(len(smap))),
        ("Pre-aggregation rule", "ln_master_rollup_v1 and cohort_m040_reoperative_v1 are MAX(...) per research_id."),
        ("Strict-DTC denominator", "n = 3,789 (drop MTC, anaplastic, NIFTP, FTUMP, follicular adenoma, atypical Hurthle, NUT, adenoid cystic)."),
        ("Primary 3-level analytic", "n = 3,756 (Gross/Microscopic/No-negative ETE only); path-proven events = 139."),
        ("Cox subset", "n = 2,025 (strict-DTC + surgery date known + FU > 0)."),
        ("Standing rule", "(Logan, 2026-05-01): every manuscript final must include this kind of per-patient + source-map workbook."),
        ("Companion file", "M044_ETE_master_data.xlsx contains the raw-source tab dumps; this file replaces the patient-analytic tab with explicit per-column provenance."),
    ]
    for i,(k,v) in enumerate(notes, start=4):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=str(v)).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 32; ws.column_dimensions["B"].width = 110
    for r in range(4, 4+len(notes)): ws.row_dimensions[r].height = 30

    # Sheet: per-patient analytic
    ws = wb.create_sheet("Per-patient analytic")
    write_df(ws, pm)

    # Sheet: source map
    ws = wb.create_sheet("Source map")
    write_df(ws, smap)

    # Sheet: audit (columns not covered)
    audit_df = pd.DataFrame({
        "column_name": missing_in_map + extra_in_map,
        "status": (["IN_PM_NOT_IN_MAP"]*len(missing_in_map) + ["IN_MAP_NOT_IN_PM"]*len(extra_in_map)),
    })
    ws = wb.create_sheet("Audit (gaps)")
    write_df(ws, audit_df if len(audit_df) else pd.DataFrame({"column_name":["(none)"], "status":["clean"]}))

    print(f"-> Save -> {OUT}")
    wb.save(OUT)
    print(f"OK {OUT.stat().st_size:,} bytes")

if __name__ == "__main__":
    main()
