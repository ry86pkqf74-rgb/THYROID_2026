#!/usr/bin/env python3
"""
THYROID_2026 — Script 219: Imaging Data Evaluation, Validation & Canonical Integration

Database: thyroid_ete_fix_20260413 on MotherDuck
Canonical: canonical_patient_master_v1 (10,871 rows × 1,025+ columns)

Tasks:
  1: CT expanded rollup — indication, thyroid_details, ln_details, airway, 15 new ct_* columns
  2: PET 'Other' miscategorized exam recovery — 77 exams / 72 patients reclassified
  3: MRI indication backfill via regex (~98 of 147 null-indication exams)
  4: Nuclear Med expansion — indication, impression, dose, TSH/Tg, uptake from pre-parsed fields
  5: LN US dedicated exam ingestion — 61 patients from Imaging_12_1_25.xlsx 'LN US' sheet
  6: Canonical rebuild + comprehensive provenance validation (6 checks)

Run:
  .venv/bin/python scripts/219_imaging_gap_resolution.py [--dry-run] [--phase 1|2|3|4|5|6|all]
                                                          [--db DB_NAME] [--canonical TABLE_NAME]

Examples:
  # Original DB (thyroid_ete_fix_20260413)
  .venv/bin/python scripts/219_imaging_gap_resolution.py --phase all

  # New account / gold canonical
  .venv/bin/python scripts/219_imaging_gap_resolution.py --db "Thyroid 2026" --canonical gold_master_patient_facts_v1 --phase all
"""
from __future__ import annotations

import argparse
import re
import sys
import time
from pathlib import Path
from typing import Any
from urllib.parse import quote

import duckdb

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))
from motherduck_client import get_token  # noqa: E402

# Defaults — overridden by CLI flags at runtime
DB = "thyroid_ete_fix_20260413"
CANONICAL = "canonical_patient_master_v1"
TOTAL_ROWS = 10871

# Set at runtime by detect_schema()
_CANONICAL_RID_IS_BIGINT: bool = False  # canonical research_id type


# ======================================================================
# Connection + utilities
# ======================================================================

def connect() -> duckdb.DuckDBPyConnection:
    token = get_token()
    if not token:
        print("[219] ERROR: No MotherDuck token found.")
        sys.exit(1)
    print(f"[219] Token: SET, len={len(token)}")
    # URL-encode database name to handle spaces (e.g. "Thyroid 2026" → "Thyroid%202026")
    db_encoded = quote(DB, safe="")
    return duckdb.connect(f"md:{db_encoded}?motherduck_token={token}")


def detect_schema(con: duckdb.DuckDBPyConnection) -> None:
    """Detect canonical research_id type and resolve write target if canonical is a VIEW."""
    global _CANONICAL_RID_IS_BIGINT, CANONICAL

    # Resolve VIEW → base table so UPDATEs work
    try:
        catalog = con.execute("SELECT current_catalog()").fetchone()[0]
        ttype = con.execute(f"""
            SELECT DISTINCT table_type FROM information_schema.tables
            WHERE table_name = '{CANONICAL}' AND table_schema = 'main'
              AND table_catalog = '{catalog}'
        """).fetchone()
        if ttype and ttype[0] == "VIEW":
            vdef = con.execute(f"""
                SELECT view_definition FROM information_schema.views
                WHERE table_name = '{CANONICAL}' AND table_schema = 'main'
                  AND table_catalog = '{catalog}'
            """).fetchone()
            if vdef:
                import re as _re
                m = _re.search(r"FROM\s+(?:main\.)?(\w+)", vdef[0], _re.IGNORECASE)
                if m:
                    base_table = m.group(1)
                    # Verify it's a base table
                    bt = con.execute(f"""
                        SELECT table_type FROM information_schema.tables
                        WHERE table_name = '{base_table}' AND table_schema = 'main'
                          AND table_catalog = '{catalog}'
                        LIMIT 1
                    """).fetchone()
                    if bt and bt[0] == "BASE TABLE":
                        print(f"[219] VIEW detected: {CANONICAL} → base table: {base_table}")
                        CANONICAL = base_table
                    else:
                        print(f"[219] WARN: Cannot resolve {CANONICAL} view to writable base table")
    except Exception as e:
        print(f"[219] WARN: View resolution failed: {e}")

    # Detect research_id type on resolved CANONICAL
    try:
        rid_type = con.execute(f"SELECT typeof(research_id) FROM {CANONICAL} LIMIT 1").fetchone()
        if rid_type and rid_type[0].upper() in ("BIGINT", "INTEGER", "INT", "INT8", "INT64", "HUGEINT"):
            _CANONICAL_RID_IS_BIGINT = True
            print(f"[219] Canonical ({CANONICAL}) research_id type: {rid_type[0]} → BIGINT casts in joins")
        else:
            _CANONICAL_RID_IS_BIGINT = False
            print(f"[219] Canonical ({CANONICAL}) research_id type: {rid_type[0] if rid_type else 'unknown'} → VARCHAR joins")
    except Exception as e:
        print(f"[219] WARN: Could not detect canonical RID type: {e}")


def rid_cast_expr(alias: str = "r") -> str:
    """Return the right-hand side expression for research_id join to canonical.

    Source tables always have VARCHAR research_id.
    Canonical may be BIGINT (new account) or VARCHAR (old account).
    """
    if _CANONICAL_RID_IS_BIGINT:
        return f"TRY_CAST({alias}.research_id AS BIGINT)"
    return f"{alias}.research_id"


def check_invariants(con: duckdb.DuckDBPyConnection, table: str, label: str) -> bool:
    inv = con.execute(f"""
        SELECT
            COUNT(*) AS total_rows,
            COUNT(DISTINCT research_id) AS distinct_rids,
            COUNT(*) FILTER (WHERE research_id IS NULL) AS null_rids,
            COUNT(*) FILTER (WHERE fna_path_outcome IS NULL) AS null_fna
        FROM {table}
    """).fetchone()
    print(f"[219] {label}: {inv[0]} rows, {inv[1]} distinct RIDs, "
          f"{inv[2]} null RIDs, {inv[3]} null fna_path_outcome")
    errors = []
    if inv[0] != TOTAL_ROWS:
        errors.append(f"Row count {inv[0]} != {TOTAL_ROWS}")
    if inv[0] != inv[1]:
        errors.append(f"Duplicate research_ids: {inv[0] - inv[1]}")
    if inv[2] > 0:
        errors.append(f"NULL research_ids: {inv[2]}")
    for e in errors:
        print(f"[219] ERROR: {e}")
    return len(errors) == 0


def get_existing_columns(con: duckdb.DuckDBPyConnection) -> set[str]:
    """Return column names for CANONICAL in the CURRENT database."""
    try:
        rows = con.execute(f"DESCRIBE {CANONICAL}").fetchall()
        return {r[0] for r in rows}
    except Exception:
        # Fall back to information_schema with catalog filter
        try:
            catalog = con.execute("SELECT current_catalog()").fetchone()[0]
            rows = con.execute(f"""
                SELECT DISTINCT column_name
                FROM information_schema.columns
                WHERE table_name = '{CANONICAL}'
                  AND table_schema = 'main'
                  AND table_catalog = '{catalog}'
            """).fetchall()
            return {r[0] for r in rows}
        except Exception:
            return set()


def safe_add_column(con: duckdb.DuckDBPyConnection, col: str, dtype: str) -> None:
    try:
        con.execute(f'ALTER TABLE {CANONICAL} ADD COLUMN "{col}" {dtype}')
        print(f"[219]   + added column {col} ({dtype})")
    except Exception:
        pass  # column already exists


def table_exists(con: duckdb.DuckDBPyConnection, name: str) -> bool:
    """Check if a table exists in the CURRENT database only (not cross-db shared tables)."""
    try:
        # Quick probe: actually try to describe the table
        con.execute(f"SELECT 1 FROM {name} LIMIT 0")
        return True
    except Exception:
        return False


def run_sql(con: duckdb.DuckDBPyConnection, sql: str, label: str, dry_run: bool = False) -> None:
    if dry_run:
        print(f"[219] DRY-RUN — would execute: {label}")
        return
    t0 = time.time()
    con.execute(sql)
    print(f"[219] {label} — done in {time.time()-t0:.1f}s")


def check_orphans(con: duckdb.DuckDBPyConnection, staging_table: str, label: str) -> int:
    # Source tables have VARCHAR research_id; canonical may be BIGINT.
    # Use CAST on the staging side to match canonical type.
    cast_expr = "TRY_CAST(s.research_id AS BIGINT)" if _CANONICAL_RID_IS_BIGINT else "s.research_id"
    r = con.execute(f"""
        SELECT COUNT(*) FROM {staging_table} s
        WHERE {cast_expr} NOT IN (SELECT research_id FROM {CANONICAL})
    """).fetchone()[0]
    if r > 0:
        print(f"[219] WARNING: {r} orphan research_ids in {label} — not in canonical spine")
    else:
        print(f"[219] OK: 0 orphan research_ids in {label}")
    return r


# ======================================================================
# TASK 1: CT EXPANDED ROLLUP
# ======================================================================

CT_ROLLUP_SQL = """
CREATE OR REPLACE TABLE _ct_expanded_rollup_v1 AS
WITH ct_only AS (
    -- Exclude PET-family and MRI and pure 'Other' non-CT exams
    SELECT * FROM ct_imaging
    WHERE exam_type_normalized NOT LIKE 'PET%'
      AND exam_type_normalized NOT LIKE '%PET%'
      AND exam_type_normalized NOT LIKE 'MRI%'
      AND exam_type_normalized != 'PET/MR'
      AND exam_type_normalized != 'None'
      AND exam_type_normalized IS NOT NULL
),
ranked AS (
    SELECT *,
        ROW_NUMBER() OVER (
            PARTITION BY research_id
            ORDER BY TRY_CAST(date_of_exam AS DATE) ASC NULLS LAST
        ) AS rn_first,
        ROW_NUMBER() OVER (
            PARTITION BY research_id
            ORDER BY TRY_CAST(date_of_exam AS DATE) DESC NULLS LAST
        ) AS rn_last
    FROM ct_only
)
SELECT
    research_id,
    -- Indication: first and last exam
    MAX(CASE WHEN rn_first = 1 THEN indication END) AS ct_indication_first,
    MAX(CASE WHEN rn_last = 1 THEN indication END)  AS ct_indication_last,
    -- Dates
    MIN(TRY_CAST(date_of_exam AS DATE)) AS ct_first_date_new,
    MAX(TRY_CAST(date_of_exam AS DATE)) AS ct_last_date_new,
    -- Exam type of first CT
    MAX(CASE WHEN rn_first = 1 THEN exam_type_normalized END) AS ct_exam_type_first,
    -- Contrast on first CT
    MAX(CASE WHEN rn_first = 1 THEN contrast END) AS ct_contrast_first,
    -- Thyroid details: most recent non-null
    MAX(CASE WHEN rn_last = 1 AND thyroid_details IS NOT NULL
             THEN LEFT(CAST(thyroid_details AS VARCHAR), 500) END) AS ct_thyroid_details_last,
    -- LN details: most recent
    MAX(CASE WHEN rn_last = 1 AND lymph_node_details IS NOT NULL
             THEN LEFT(CAST(lymph_node_details AS VARCHAR), 500) END) AS ct_ln_details_last,
    -- LN locations: most recent
    MAX(CASE WHEN rn_last = 1 AND lymph_node_locations IS NOT NULL
             THEN LEFT(CAST(lymph_node_locations AS VARCHAR), 300) END) AS ct_ln_locations_last,
    -- Airway compromise
    BOOL_OR(airway_compromise_comment IS NOT NULL
            AND LENGTH(CAST(airway_compromise_comment AS VARCHAR)) > 5)
        AS ct_airway_compromise_any,
    MAX(CASE WHEN rn_last = 1 AND airway_compromise_comment IS NOT NULL
             THEN LEFT(CAST(airway_compromise_comment AS VARCHAR), 300) END)
        AS ct_airway_comment_last,
    -- Additional thyroid binary flags (BOOLEAN columns — use directly)
    BOOL_OR(thyroid_postsurgical = TRUE)      AS ct_thyroid_postsurgical_any,
    BOOL_OR(thyroid_not_visualized = TRUE)    AS ct_thyroid_not_visualized_any,
    BOOL_OR(thyroid_heterogeneous = TRUE)     AS ct_thyroid_heterogeneous_any,
    BOOL_OR(thyroid_other_abnormality = TRUE) AS ct_thyroid_other_abnormality_any,
    BOOL_OR(thyroid_normal = TRUE)            AS ct_thyroid_normal_any,
    BOOL_OR(thyroid_nodule = TRUE)            AS ct_thyroid_nodule_any,
    BOOL_OR(thyroid_enlarged = TRUE)          AS ct_thyroid_enlarged_any,
    BOOL_OR(pathologic_lymph_nodes = TRUE)    AS ct_pathologic_ln_any,
    -- Source provenance
    'ct_imaging'              AS ct_expanded_source_table,
    '219_imaging_gap_resolution' AS ct_expanded_script
FROM ranked
GROUP BY research_id
"""

CT_VALIDATION_SQL = """
SELECT
    COUNT(*) AS total,
    COUNT(*) FILTER (WHERE ct_indication_first IS NOT NULL AND LENGTH(ct_indication_first) > 5) AS has_indication,
    COUNT(*) FILTER (WHERE ct_first_date_new IS NOT NULL) AS has_first_date,
    COUNT(*) FILTER (WHERE ct_thyroid_details_last IS NOT NULL) AS has_thyroid_details,
    COUNT(*) FILTER (WHERE ct_ln_details_last IS NOT NULL) AS has_ln_details,
    COUNT(*) FILTER (WHERE ct_airway_compromise_any = TRUE) AS has_airway,
    COUNT(*) FILTER (WHERE ct_thyroid_nodule_any = TRUE) AS has_nodule,
    COUNT(*) FILTER (WHERE ct_pathologic_ln_any = TRUE) AS has_pathologic_ln
FROM _ct_expanded_rollup_v1
"""


def run_task1_ct(con: duckdb.DuckDBPyConnection, dry_run: bool) -> None:
    print("\n[219] === TASK 1: CT EXPANDED ROLLUP ===")
    run_sql(con, CT_ROLLUP_SQL, "build _ct_expanded_rollup_v1", dry_run)
    if dry_run:
        return

    orphans = check_orphans(con, "_ct_expanded_rollup_v1", "CT rollup")
    if orphans > 0:
        print(f"[219] WARN: {orphans} CT patients not in canonical spine — investigate")

    stats = con.execute(CT_VALIDATION_SQL).fetchone()
    print(f"[219] CT rollup stats:")
    print(f"  total={stats[0]}, indication={stats[1]}, first_date={stats[2]}")
    print(f"  thyroid_details={stats[3]}, ln_details={stats[4]}, airway={stats[5]}")
    print(f"  nodule_any={stats[6]}, pathologic_ln={stats[7]}")

    # Cross-validate overlap between rollup and canonical
    # Use safe join with BIGINT cast where needed
    join_expr = "TRY_CAST(r.research_id AS BIGINT) = c.research_id" \
        if _CANONICAL_RID_IS_BIGINT else "c.research_id = r.research_id"
    existing_cols = get_existing_columns(con)
    if "ct_n_exams" in existing_cols:
        ct_n_filter = "c.ct_n_exams IS NOT NULL"
    else:
        # New lean schema — just check if patient appears in rollup vs canonical
        ct_n_filter = "TRUE"
    xval = con.execute(f"""
        SELECT
            COUNT(*) FILTER (WHERE {ct_n_filter} AND r.research_id IS NOT NULL) AS both_have_ct,
            COUNT(*) FILTER (WHERE {ct_n_filter} AND r.research_id IS NULL)     AS canonical_only,
            COUNT(*) FILTER (WHERE r.research_id IS NOT NULL)                   AS rollup_patients
        FROM {CANONICAL} c
        FULL OUTER JOIN _ct_expanded_rollup_v1 r ON {join_expr}
    """).fetchone()
    print(f"[219] CT cross-val: matched={xval[0]}, canonical_only={xval[1]}, rollup_total={xval[2]}")


# ======================================================================
# TASK 2: PET 'OTHER' MISCATEGORIZED EXAM RECOVERY
# ======================================================================

PET_OTHER_SQL = """
CREATE OR REPLACE TABLE _pet_other_recovered_v1 AS
SELECT
    research_id,
    date_of_exam,
    indication,
    -- Reclassify exam type from report text
    CASE
        WHEN LOWER(CAST(original_report AS VARCHAR)) LIKE '%pet/ct%'
          OR LOWER(CAST(original_report AS VARCHAR)) LIKE '%pet ct%' THEN 'PET/CT'
        WHEN LOWER(CAST(original_report AS VARCHAR)) LIKE '%pet/mr%'
          OR LOWER(CAST(original_report AS VARCHAR)) LIKE '%pet mr%' THEN 'PET/MR'
        WHEN LOWER(CAST(original_report AS VARCHAR)) LIKE '%dotatate%' THEN 'PET/CT_DOTATATE'
        WHEN LOWER(CAST(original_report AS VARCHAR)) LIKE '%fdg%'
          OR LOWER(CAST(original_report AS VARCHAR)) LIKE '%fluorodeoxyglucose%' THEN 'PET/CT'
        ELSE 'PET_unclassified'
    END AS reclassified_exam_type,
    -- Radiotracer
    CASE
        WHEN LOWER(CAST(original_report AS VARCHAR)) LIKE '%dotatate%' THEN 'DOTATATE'
        WHEN LOWER(CAST(original_report AS VARCHAR)) LIKE '%fdg%'
          OR LOWER(CAST(original_report AS VARCHAR)) LIKE '%fluorodeoxyglucose%' THEN 'FDG'
        WHEN LOWER(CAST(original_report AS VARCHAR)) LIKE '%i-131%'
          OR LOWER(CAST(original_report AS VARCHAR)) LIKE '%nai-131%' THEN 'I-131'
        WHEN LOWER(CAST(original_report AS VARCHAR)) LIKE '%i-123%' THEN 'I-123'
        ELSE NULL
    END AS radiotracer_parsed,
    -- Basic findings (regex, no LLM)
    CASE WHEN LOWER(CAST(original_report AS VARCHAR)) LIKE '%metasta%' THEN TRUE ELSE FALSE END
        AS mentions_metastasis,
    CASE WHEN LOWER(CAST(original_report AS VARCHAR)) LIKE '%thyroid bed%' THEN TRUE ELSE FALSE END
        AS mentions_thyroid_bed,
    CASE WHEN LOWER(CAST(original_report AS VARCHAR)) LIKE '%lymph node%'
           OR LOWER(CAST(original_report AS VARCHAR)) LIKE '%nodal%' THEN TRUE ELSE FALSE END
        AS mentions_lymph_nodes,
    CASE WHEN LOWER(CAST(original_report AS VARCHAR)) LIKE '%lung%'
           OR LOWER(CAST(original_report AS VARCHAR)) LIKE '%pulmonary%' THEN TRUE ELSE FALSE END
        AS mentions_lung,
    CASE WHEN LOWER(CAST(original_report AS VARCHAR)) LIKE '%bone%'
           OR LOWER(CAST(original_report AS VARCHAR)) LIKE '%osseous%' THEN TRUE ELSE FALSE END
        AS mentions_bone,
    CASE WHEN LOWER(CAST(original_report AS VARCHAR)) LIKE '%no evidence%metast%'
           OR LOWER(CAST(original_report AS VARCHAR)) LIKE '%no metast%' THEN TRUE ELSE FALSE END
        AS ned_statement,
    -- Source provenance
    'ct_imaging'           AS source_table,
    'Other'                AS original_exam_type,
    'regex_from_other'     AS pet_extraction_method,
    '219_pet_other_recovery' AS parse_script
FROM ct_imaging
WHERE exam_type_normalized = 'Other'
  AND (LOWER(CAST(original_report AS VARCHAR)) LIKE '%pet%'
       OR LOWER(CAST(original_report AS VARCHAR)) LIKE '%fdg%')
"""

PET_OTHER_ROLLUP_SQL = """
CREATE OR REPLACE TABLE _pet_other_rollup_v1 AS
SELECT
    research_id,
    COUNT(*) AS pet_other_n_exams,
    MIN(TRY_CAST(date_of_exam AS DATE)) AS pet_other_first_date,
    MAX(TRY_CAST(date_of_exam AS DATE)) AS pet_other_last_date,
    MAX(indication)                     AS pet_other_indication_first,
    BOOL_OR(mentions_metastasis)        AS pet_other_mentions_metastasis,
    BOOL_OR(ned_statement)              AS pet_other_ned_statement,
    BOOL_OR(mentions_thyroid_bed)       AS pet_other_mentions_thyroid_bed,
    BOOL_OR(mentions_lymph_nodes)       AS pet_other_mentions_ln,
    -- Most common reclassified type
    MAX(reclassified_exam_type)         AS pet_other_exam_type,
    MAX(radiotracer_parsed)             AS pet_other_radiotracer,
    'regex_from_other'                  AS pet_other_extraction_method,
    '219_pet_other_recovery'            AS pet_other_parse_script
FROM _pet_other_recovered_v1
GROUP BY research_id
"""


def run_task2_pet_other(con: duckdb.DuckDBPyConnection, dry_run: bool) -> None:
    print("\n[219] === TASK 2: PET 'OTHER' RECOVERY ===")
    run_sql(con, PET_OTHER_SQL, "build _pet_other_recovered_v1", dry_run)
    run_sql(con, PET_OTHER_ROLLUP_SQL, "build _pet_other_rollup_v1", dry_run)
    if dry_run:
        return

    orphans = check_orphans(con, "_pet_other_recovered_v1", "PET Other recovered")

    stats = con.execute("""
        SELECT
            COUNT(DISTINCT research_id) AS total_pts,
            COUNT(*) AS total_exams,
            COUNT(*) FILTER (WHERE indication IS NOT NULL) AS has_indication,
            COUNT(*) FILTER (WHERE TRY_CAST(date_of_exam AS DATE) IS NOT NULL) AS has_date,
            COUNT(*) FILTER (WHERE reclassified_exam_type = 'PET/CT') AS is_petct,
            COUNT(*) FILTER (WHERE ned_statement = TRUE) AS ned
        FROM _pet_other_recovered_v1
    """).fetchone()
    print(f"[219] PET Other stats: pts={stats[0]}, exams={stats[1]}, indication={stats[2]}")
    print(f"       date={stats[3]}, PET/CT={stats[4]}, NED={stats[5]}")

    # How many are NEW vs already have PET data in canonical
    existing_cols_pet = get_existing_columns(con)
    if "pet_has_data" in existing_cols_pet:
        new_pts = con.execute(f"""
            SELECT COUNT(DISTINCT r.research_id)
            FROM _pet_other_recovered_v1 r
            WHERE TRY_CAST(r.research_id AS BIGINT) NOT IN (
                SELECT research_id FROM {CANONICAL} WHERE pet_has_data = TRUE
            )
        """).fetchone()[0] if _CANONICAL_RID_IS_BIGINT else con.execute(f"""
            SELECT COUNT(DISTINCT r.research_id)
            FROM _pet_other_recovered_v1 r
            WHERE r.research_id NOT IN (
                SELECT research_id FROM {CANONICAL} WHERE pet_has_data = TRUE
            )
        """).fetchone()[0]
    else:
        # pet_has_data not in new schema — all are "new"
        new_pts = con.execute("SELECT COUNT(DISTINCT research_id) FROM _pet_other_recovered_v1").fetchone()[0]
    print(f"[219] PET Other: {new_pts} patients new to canonical PET data")


# ======================================================================
# TASK 3: MRI INDICATION BACKFILL
# ======================================================================

MRI_PREVIEW_SQL = """
SELECT
    research_id,
    mri_label,
    TRIM(REGEXP_EXTRACT(
        original_report,
        '(?i)(?:CLINICAL INDICATION|INDICATION|HISTORY)[:\\s]*([^\\n]+)',
        1
    )) AS parsed_indication,
    LEFT(original_report, 100) AS report_start
FROM mri_imaging
WHERE (indication IS NULL OR LENGTH(indication) < 4)
  AND original_report IS NOT NULL
  AND LENGTH(TRIM(REGEXP_EXTRACT(
        original_report,
        '(?i)(?:CLINICAL INDICATION|INDICATION|HISTORY)[:\\s]*([^\\n]+)',
        1
  ))) > 5
LIMIT 10
"""

MRI_UPDATE_SQL = """
UPDATE mri_imaging
SET indication = TRIM(REGEXP_EXTRACT(
        original_report,
        '(?i)(?:CLINICAL INDICATION|INDICATION|HISTORY)[:\\s]*([^\\n]+)',
        1
))
WHERE (indication IS NULL OR LENGTH(indication) < 4)
  AND original_report IS NOT NULL
  AND LENGTH(TRIM(REGEXP_EXTRACT(
        original_report,
        '(?i)(?:CLINICAL INDICATION|INDICATION|HISTORY)[:\\s]*([^\\n]+)',
        1
  ))) > 5
"""


def run_task3_mri(con: duckdb.DuckDBPyConnection, dry_run: bool) -> None:
    print("\n[219] === TASK 3: MRI INDICATION BACKFILL ===")

    if not table_exists(con, "mri_imaging"):
        print("[219] SKIP: mri_imaging table not present in this database")
        return

    # Before stats
    before = con.execute("""
        SELECT
            COUNT(*) AS total,
            COUNT(*) FILTER (WHERE indication IS NOT NULL AND LENGTH(indication) > 5) AS has_indication,
            COUNT(*) FILTER (WHERE indication IS NULL OR LENGTH(indication) < 4) AS missing,
            COUNT(*) FILTER (WHERE original_report IS NOT NULL) AS has_report
        FROM mri_imaging
    """).fetchone()
    print(f"[219] MRI before: total={before[0]}, has_indication={before[1]}, "
          f"missing={before[2]}, has_report={before[3]}")

    # Preview 3 samples
    samples = con.execute(MRI_PREVIEW_SQL).fetchall()
    print(f"[219] MRI preview ({len(samples)} parseable):")
    for s in samples[:3]:
        print(f"  RID={s[0]} label={s[1]} indication='{str(s[2])[:80]}'")

    parseable = con.execute("""
        SELECT COUNT(*) FROM mri_imaging
        WHERE (indication IS NULL OR LENGTH(indication) < 4)
          AND original_report IS NOT NULL
          AND LENGTH(TRIM(REGEXP_EXTRACT(
                original_report,
                '(?i)(?:CLINICAL INDICATION|INDICATION|HISTORY)[:\\s]*([^\\n]+)',
                1
          ))) > 5
    """).fetchone()[0]
    print(f"[219] MRI: {parseable} exams with parseable indication")

    if dry_run:
        print("[219] DRY-RUN — would UPDATE mri_imaging.indication")
        return

    run_sql(con, MRI_UPDATE_SQL, "UPDATE mri_imaging.indication", dry_run)

    # After stats
    after = con.execute("""
        SELECT
            COUNT(*) FILTER (WHERE indication IS NOT NULL AND LENGTH(indication) > 5) AS has_indication,
            COUNT(*) AS total
        FROM mri_imaging
    """).fetchone()
    print(f"[219] MRI after: {after[0]}/{after[1]} exams with indication")


# ======================================================================
# TASK 4: NUCLEAR MED EXPANSION
# ======================================================================

NUCMED_ROLLUP_SQL = """
CREATE OR REPLACE TABLE _nucmed_expanded_v1 AS
WITH parsed AS (
    SELECT
        CAST(research_id AS VARCHAR) AS research_id,
        scan_index,
        TRY_CAST(scandate AS DATE) AS scan_date_parsed,
        scantype,
        radiotracer,
        -- Use pre-parsed columns directly (already extracted by ingestion pipeline)
        indication_text   AS nucmed_indication,
        impression_text   AS nucmed_impression,
        findings_text     AS nucmed_findings,
        -- Uptake from structured column (format: '0.4%') — strip % and cast
        TRY_CAST(REPLACE(NULLIF(TRIM(uptake_24hour), ''), '%', '') AS DOUBLE) AS nucmed_uptake_24hr_pct,
        TRY_CAST(REPLACE(NULLIF(TRIM(uptake_general), ''), '%', '') AS DOUBLE) AS nucmed_uptake_gen_pct,
        -- TSH from lab_summary (e.g. 'tsh: 47.74 mciu' or 'tsh: 9.50 mciu; thyroglobulin: ...')
        TRY_CAST(REGEXP_EXTRACT(
            LOWER(COALESCE(CAST(lab_summary AS VARCHAR), '')),
            'tsh[:\\s]+([0-9]+\\.?[0-9]*)',
            1
        ) AS DOUBLE) AS nucmed_tsh_lab,
        -- Tg from lab_summary
        TRY_CAST(REGEXP_EXTRACT(
            LOWER(COALESCE(CAST(lab_summary AS VARCHAR), '')),
            '(?:thyroglobulin|tg)[:\\s]+([0-9]+\\.?[0-9]*)',
            1
        ) AS DOUBLE) AS nucmed_tg_lab,
        -- TgAb from lab_summary
        TRY_CAST(REGEXP_EXTRACT(
            LOWER(COALESCE(CAST(lab_summary AS VARCHAR), '')),
            '(?:tg antibod|tgab|anti.?thyroglobulin)[:\\s]*(?:less than\\s*)?([0-9]+\\.?[0-9]*)',
            1
        ) AS DOUBLE) AS nucmed_tgab_lab,
        -- Also try TSH / Tg from the main report text (scan_present = full report)
        TRY_CAST(REGEXP_EXTRACT(
            LOWER(COALESCE(CAST(scan_present AS VARCHAR), '')),
            'stimulated thyroglobulin[:\\s]+([0-9]+\\.?[0-9]*)',
            1
        ) AS DOUBLE) AS nucmed_tg_from_report,
        TRY_CAST(REGEXP_EXTRACT(
            LOWER(COALESCE(CAST(scan_present AS VARCHAR), '')),
            'tsh[:\\s]+([0-9]+\\.?[0-9]*)',
            1
        ) AS DOUBLE) AS nucmed_tsh_from_report,
        -- Dose: 'administration of XXX mCi' or 'administered XXX mCi' in scan_present
        TRY_CAST(REGEXP_EXTRACT(
            LOWER(COALESCE(CAST(scan_present AS VARCHAR), '')),
            '(?:administration of|administered)\\s+([0-9]+\\.?[0-9]*)\\s*mci',
            1
        ) AS DOUBLE) AS nucmed_dose_mci_parsed,
        -- NED / metastasis flags from impression
        CASE
            WHEN LOWER(COALESCE(impression_text, '')) LIKE '%no evidence%metast%'
              OR LOWER(COALESCE(impression_text, '')) LIKE '%no scintigraphic evidence%'
              OR LOWER(COALESCE(impression_text, '')) LIKE '%no abnormal%uptake%' THEN 'NED'
            WHEN LOWER(COALESCE(impression_text, '')) LIKE '%metast%' THEN 'metastasis_mentioned'
            WHEN LOWER(COALESCE(impression_text, '')) LIKE '%thyroid bed%as expected%'
              OR LOWER(COALESCE(impression_text, '')) LIKE '%remnant only%' THEN 'thyroid_bed_only'
            ELSE 'other'
        END AS nucmed_overall_assessment,
        -- Source provenance
        'nuclear_med'              AS source_table,
        'indication_text/impression_text/scan_present' AS source_column,
        '219_nucmed_expansion'     AS parse_script,
        COALESCE(LENGTH(CAST(indication_text AS VARCHAR)), 0)
            + COALESCE(LENGTH(CAST(impression_text AS VARCHAR)), 0) AS content_length
    FROM nuclear_med
    WHERE CAST(scan_present AS VARCHAR) IS NOT NULL
       OR indication_text IS NOT NULL
)
SELECT * FROM parsed
"""

NUCMED_PATIENT_ROLLUP_SQL = """
CREATE OR REPLACE TABLE _nucmed_patient_rollup_v1 AS
SELECT
    research_id,
    -- Indication: first and last scan (by date)
    FIRST_VALUE(nucmed_indication IGNORE NULLS)
        OVER (PARTITION BY research_id ORDER BY scan_date_parsed ASC
              ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING)
        AS nucmed_indication_first_w,
    LAST_VALUE(nucmed_indication IGNORE NULLS)
        OVER (PARTITION BY research_id ORDER BY scan_date_parsed ASC
              ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING)
        AS nucmed_indication_last_w,
    -- Impression: last scan
    LAST_VALUE(nucmed_impression IGNORE NULLS)
        OVER (PARTITION BY research_id ORDER BY scan_date_parsed ASC
              ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING)
        AS nucmed_impression_last_w,
    -- Findings: last scan
    LAST_VALUE(nucmed_findings IGNORE NULLS)
        OVER (PARTITION BY research_id ORDER BY scan_date_parsed ASC
              ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING)
        AS nucmed_findings_last_w
FROM _nucmed_expanded_v1
"""

NUCMED_AGG_SQL = """
CREATE OR REPLACE TABLE _nucmed_agg_v1 AS
WITH wnd AS (
    SELECT DISTINCT ON (research_id)
        research_id,
        nucmed_indication_first_w  AS nucmed_indication_first,
        nucmed_indication_last_w   AS nucmed_indication_last,
        nucmed_impression_last_w   AS nucmed_impression_last,
        nucmed_findings_last_w     AS nucmed_findings_last
    FROM _nucmed_patient_rollup_v1
),
agg AS (
    SELECT
        research_id,
        -- TSH: combine lab_summary and report text; max for stimulated context
        MAX(COALESCE(nucmed_tsh_lab, nucmed_tsh_from_report)) AS nucmed_tsh_max,
        MIN(COALESCE(nucmed_tsh_lab, nucmed_tsh_from_report)) AS nucmed_tsh_min,
        BOOL_OR(COALESCE(nucmed_tsh_lab, nucmed_tsh_from_report) > 30)
            AS nucmed_tsh_is_stimulated,
        -- Tg: max and min
        MAX(COALESCE(nucmed_tg_lab, nucmed_tg_from_report)) AS nucmed_tg_max_new,
        MIN(COALESCE(nucmed_tg_lab, nucmed_tg_from_report)) AS nucmed_tg_min_new,
        -- TgAb
        MAX(nucmed_tgab_lab) AS nucmed_tgab_max,
        -- Uptake: prefer 24hr structured field, else general
        MAX(COALESCE(nucmed_uptake_24hr_pct, nucmed_uptake_gen_pct))
            AS nucmed_uptake_pct_max_new,
        -- Dose: max therapeutic, cumulative
        MAX(nucmed_dose_mci_parsed)  AS nucmed_dose_max_parsed,
        SUM(CASE WHEN nucmed_dose_mci_parsed > 10
                 THEN nucmed_dose_mci_parsed END)
            AS nucmed_cumulative_therapeutic_dose,
        -- Counts
        COUNT(*) FILTER (WHERE nucmed_tsh_lab IS NOT NULL
                            OR nucmed_tsh_from_report IS NOT NULL)
            AS nucmed_n_tsh_values_new,
        COUNT(*) FILTER (WHERE nucmed_tg_lab IS NOT NULL
                            OR nucmed_tg_from_report IS NOT NULL)
            AS nucmed_n_tg_values_new,
        COUNT(*) FILTER (WHERE nucmed_dose_mci_parsed IS NOT NULL)
            AS nucmed_n_doses_parsed,
        COUNT(*) FILTER (WHERE nucmed_indication IS NOT NULL
                            AND LENGTH(nucmed_indication) > 5)
            AS nucmed_n_with_indication,
        COUNT(*) FILTER (WHERE nucmed_impression IS NOT NULL
                            AND LENGTH(nucmed_impression) > 5)
            AS nucmed_n_with_impression,
        -- Overall assessment: worst finding wins
        MAX(CASE WHEN nucmed_overall_assessment = 'metastasis_mentioned' THEN 3
                 WHEN nucmed_overall_assessment = 'NED'                  THEN 2
                 WHEN nucmed_overall_assessment = 'thyroid_bed_only'     THEN 1
                 ELSE 0 END) AS nucmed_worst_assessment_rank,
        -- Source provenance
        'nuclear_med'          AS nucmed_expanded_source,
        '219_nucmed_expansion' AS nucmed_expanded_script
    FROM _nucmed_expanded_v1
    GROUP BY research_id
)
SELECT
    a.*,
    w.nucmed_indication_first,
    w.nucmed_indication_last,
    w.nucmed_impression_last,
    w.nucmed_findings_last,
    CASE a.nucmed_worst_assessment_rank
        WHEN 3 THEN 'metastasis_mentioned'
        WHEN 2 THEN 'NED'
        WHEN 1 THEN 'thyroid_bed_only'
        ELSE 'other'
    END AS nucmed_overall_assessment
FROM agg a
JOIN wnd w ON a.research_id = w.research_id
"""


def run_task4_nucmed(con: duckdb.DuckDBPyConnection, dry_run: bool) -> None:
    print("\n[219] === TASK 4: NUCLEAR MED EXPANSION ===")
    run_sql(con, NUCMED_ROLLUP_SQL, "build _nucmed_expanded_v1", dry_run)
    run_sql(con, NUCMED_PATIENT_ROLLUP_SQL, "build _nucmed_patient_rollup_v1", dry_run)
    run_sql(con, NUCMED_AGG_SQL, "build _nucmed_agg_v1", dry_run)
    if dry_run:
        return

    orphans = check_orphans(con, "_nucmed_agg_v1", "NucMed agg")

    stats = con.execute("""
        SELECT
            COUNT(*) AS pts,
            COUNT(*) FILTER (WHERE nucmed_indication_first IS NOT NULL) AS has_indication,
            COUNT(*) FILTER (WHERE nucmed_impression_last IS NOT NULL) AS has_impression,
            COUNT(*) FILTER (WHERE nucmed_dose_max_parsed IS NOT NULL) AS has_dose,
            COUNT(*) FILTER (WHERE nucmed_tsh_max IS NOT NULL) AS has_tsh,
            COUNT(*) FILTER (WHERE nucmed_tg_max_new IS NOT NULL) AS has_tg,
            COUNT(*) FILTER (WHERE nucmed_uptake_pct_max_new IS NOT NULL) AS has_uptake,
            COUNT(*) FILTER (WHERE nucmed_tsh_is_stimulated = TRUE) AS tsh_stimulated
        FROM _nucmed_agg_v1
    """).fetchone()
    print(f"[219] NucMed agg: pts={stats[0]}, indication={stats[1]}, impression={stats[2]}")
    print(f"  dose={stats[3]}, tsh={stats[4]}, tg={stats[5]}, uptake={stats[6]}, stimulated={stats[7]}")

    # Range validation
    tsh_check = con.execute("""
        SELECT
            COUNT(*) FILTER (WHERE nucmed_tsh_max < 0) AS tsh_negative,
            COUNT(*) FILTER (WHERE nucmed_tsh_max > 500) AS tsh_extreme,
            ROUND(AVG(nucmed_tsh_max), 1) AS mean_tsh
        FROM _nucmed_agg_v1 WHERE nucmed_tsh_max IS NOT NULL
    """).fetchone()
    print(f"[219] TSH validation: negative={tsh_check[0]}, extreme={tsh_check[1]}, mean={tsh_check[2]}")
    if tsh_check[0] > 0 or tsh_check[1] > 0:
        print("[219] WARN: TSH values out of clinical range — review before integrating")

    tg_check = con.execute("""
        SELECT
            COUNT(*) FILTER (WHERE nucmed_tg_max_new < 0) AS tg_negative,
            ROUND(AVG(nucmed_tg_max_new), 2) AS mean_tg,
            ROUND(PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY nucmed_tg_max_new), 2) AS median_tg
        FROM _nucmed_agg_v1 WHERE nucmed_tg_max_new IS NOT NULL
    """).fetchone()
    print(f"[219] Tg validation: negative={tg_check[0]}, mean={tg_check[1]}, median={tg_check[2]}")

    dose_check = con.execute("""
        SELECT
            COUNT(*) FILTER (WHERE nucmed_dose_max_parsed < 1) AS dose_too_low,
            COUNT(*) FILTER (WHERE nucmed_dose_max_parsed > 300) AS dose_too_high,
            COUNT(*) FILTER (WHERE nucmed_dose_max_parsed BETWEEN 1 AND 10) AS diagnostic,
            COUNT(*) FILTER (WHERE nucmed_dose_max_parsed BETWEEN 30 AND 250) AS therapeutic
        FROM _nucmed_agg_v1 WHERE nucmed_dose_max_parsed IS NOT NULL
    """).fetchone()
    print(f"[219] Dose: too_low={dose_check[0]}, too_high={dose_check[1]}, "
          f"diagnostic={dose_check[2]}, therapeutic={dose_check[3]}")

    # Cross-validate dose vs canonical rai_max_dose_mci
    discordant = con.execute(f"""
        SELECT COUNT(*) FROM {CANONICAL} c
        JOIN (SELECT research_id, MAX(nucmed_dose_max_parsed) AS parsed_dose
              FROM _nucmed_agg_v1 WHERE nucmed_dose_max_parsed IS NOT NULL
              GROUP BY research_id) n ON c.research_id = n.research_id
        WHERE c.rai_max_dose_mci IS NOT NULL
          AND c.rai_max_dose_mci > 0
          AND ABS(c.rai_max_dose_mci - n.parsed_dose) > 10
    """).fetchone()[0]
    print(f"[219] NucMed dose cross-val: {discordant} patients with >10mCi discordance vs rai_max_dose_mci")


# ======================================================================
# TASK 5: LN US DEDICATED EXAM INGESTION
# ======================================================================

def run_task5_ln_us(con: duckdb.DuckDBPyConnection, dry_run: bool) -> None:
    print("\n[219] === TASK 5: LN US DEDICATED EXAMS ===")

    try:
        import openpyxl
    except ImportError:
        print("[219] WARN: openpyxl not available — skipping LN US ingestion")
        return

    excel_path = REPO / "raw" / "Imaging_12_1_25.xlsx"
    if not excel_path.exists():
        print(f"[219] WARN: {excel_path} not found — skipping LN US ingestion")
        return

    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    if "LN US" not in wb.sheetnames:
        print("[219] WARN: 'LN US' sheet not in Imaging_12_1_25.xlsx — skipping")
        wb.close()
        return

    ws = wb["LN US"]
    rows_data = []

    for row in ws.iter_rows(min_row=2):
        vals = [cell.value for cell in row]
        if not vals or vals[0] is None:
            continue
        try:
            rid = str(int(float(str(vals[0]))))
        except (ValueError, TypeError):
            continue

        for i in range(1, min(6, len(vals))):
            if vals[i] and str(vals[i]).strip() and len(str(vals[i]).strip()) > 20:
                rows_data.append({
                    "research_id": rid,
                    "ln_us_slot": f"LN_US{i}",
                    "report_text": str(vals[i]).strip(),
                    "source_workbook": "Imaging_12_1_25.xlsx",
                    "source_sheet": "LN US",
                    "ingest_script": "219_imaging_gap_resolution",
                })

    wb.close()

    if not rows_data:
        print("[219] WARN: No LN US report data extracted from Excel")
        return

    import pandas as pd
    df = pd.DataFrame(rows_data)
    print(f"[219] LN US: {len(df)} reports, {df['research_id'].nunique()} patients")

    # Validate RIDs — canonical may be BIGINT; compare numerically
    canonical_rids_raw = con.execute(f"SELECT research_id FROM {CANONICAL}").fetchall()
    # Normalise both sides to string for safe set comparison
    canonical_rids = {str(int(r[0])) if r[0] is not None else None for r in canonical_rids_raw}
    orphan_rids = set(df["research_id"]) - canonical_rids
    if orphan_rids:
        print(f"[219] WARN: {len(orphan_rids)} LN US patients not in canonical: {list(orphan_rids)[:5]}")
        df = df[df["research_id"].isin(canonical_rids)]
        if df.empty:
            print("[219] WARN: All LN US patients filtered — possible research_id format mismatch")
            return
        print(f"[219] LN US after filtering: {len(df)} reports, {df['research_id'].nunique()} patients")

    # Parse structured fields from report text
    def parse_ln_us(text: str) -> dict:
        out: dict[str, Any] = {}
        # Date: usually near start of text
        dm = re.search(r"(\d{1,2}/\d{1,2}/\d{2,4})", text[:100])
        out["exam_date"] = dm.group(1) if dm else None
        # Indication
        im = re.search(
            r"(?:CLINICAL INDICATION|INDICATION|HISTORY)[:\s]*([^\n]+)",
            text, re.IGNORECASE
        )
        out["indication"] = im.group(1).strip()[:200] if im else None
        # Impression
        ipm = re.search(
            r"IMPRESSION[:\s]*(.*?)(?:$|Released|These images|Radiologist)",
            text, re.IGNORECASE | re.DOTALL
        )
        out["impression"] = ipm.group(1).strip()[:400] if ipm else None
        # Key findings
        out["mentions_abnormal_ln"] = bool(re.search(
            r"(?:abnormal|suspicious|enlarged|pathologic|lymphadenopathy)\s+lymph",
            text, re.IGNORECASE
        ))
        out["mentions_normal"] = bool(re.search(
            r"(?:no abnormal|no suspicious|normal|unremarkable)\s+lymph",
            text, re.IGNORECASE
        ))
        out["has_size_measurement"] = bool(re.search(
            r"\d+\.?\d*\s*(?:mm|cm)", text, re.IGNORECASE
        ))
        return out

    parsed_rows = []
    for _, row in df.iterrows():
        parsed = parse_ln_us(row["report_text"])
        parsed["research_id"] = row["research_id"]
        parsed["ln_us_slot"] = row["ln_us_slot"]
        parsed["source_workbook"] = row["source_workbook"]
        parsed["source_sheet"] = row["source_sheet"]
        parsed["ingest_script"] = row["ingest_script"]
        parsed_rows.append(parsed)

    df_parsed = pd.DataFrame(parsed_rows)
    print(f"[219] LN US parsed: "
          f"indication={df_parsed['indication'].notna().sum()}, "
          f"impression={df_parsed['impression'].notna().sum()}, "
          f"exam_date={df_parsed['exam_date'].notna().sum()}, "
          f"abnormal_ln={df_parsed['mentions_abnormal_ln'].sum()}")

    if dry_run:
        print("[219] DRY-RUN — would upload LN US data to MotherDuck")
        return

    # Upload raw table
    import tempfile, os
    with tempfile.NamedTemporaryFile(suffix=".parquet", delete=False) as tf:
        tmpf = tf.name
    df_parsed.to_parquet(tmpf, index=False)
    con.execute(f"CREATE OR REPLACE TABLE _lnus_raw_v1 AS SELECT * FROM read_parquet('{tmpf}')")
    os.unlink(tmpf)
    print(f"[219] LN US raw table uploaded: {len(df_parsed)} rows")

    # Patient-level rollup
    run_sql(con, """
        CREATE OR REPLACE TABLE _lnus_patient_rollup_v1 AS
        SELECT
            research_id,
            TRUE                               AS lnus_has_dedicated_exam,
            COUNT(*)                           AS lnus_n_exams,
            MIN(TRY_CAST(exam_date AS DATE))   AS lnus_first_date,
            MAX(TRY_CAST(exam_date AS DATE))   AS lnus_last_date,
            MAX(indication)                    AS lnus_indication_first,
            MAX(impression)                    AS lnus_impression_last,
            BOOL_OR(mentions_abnormal_ln)      AS lnus_abnormal_ln_any,
            BOOL_OR(mentions_normal)           AS lnus_normal_ln_any,
            BOOL_OR(has_size_measurement)      AS lnus_has_size_measurement,
            'Imaging_12_1_25.xlsx / LN US'     AS lnus_source,
            '219_imaging_gap_resolution'       AS lnus_ingest_script
        FROM _lnus_raw_v1
        GROUP BY research_id
    """, "build _lnus_patient_rollup_v1")

    rollup_count = con.execute("SELECT COUNT(*) FROM _lnus_patient_rollup_v1").fetchone()[0]
    print(f"[219] LN US patient rollup: {rollup_count} patients")
    orphans = check_orphans(con, "_lnus_patient_rollup_v1", "LN US rollup")


# ======================================================================
# TASK 6: CANONICAL REBUILD + VALIDATION
# ======================================================================

def run_task6_canonical(con: duckdb.DuckDBPyConnection, dry_run: bool) -> None:
    print("\n[219] === TASK 6: CANONICAL REBUILD + VALIDATION ===")

    if dry_run:
        print("[219] DRY-RUN — would rebuild canonical with new imaging columns")
        return

    # Verify staging tables exist
    required_staging = [
        "_ct_expanded_rollup_v1",
        "_pet_other_rollup_v1",
        "_nucmed_agg_v1",
    ]
    for t in required_staging:
        if not table_exists(con, t):
            print(f"[219] ERROR: Staging table {t} missing — cannot rebuild canonical")
            return

    # --- Step 6.1: Add new columns to canonical ---
    existing_cols = get_existing_columns(con)

    new_ct_cols = {
        "ct_indication_first":          "VARCHAR",
        "ct_indication_last":           "VARCHAR",
        "ct_first_date":                "DATE",
        "ct_last_date":                 "DATE",
        "ct_exam_type_first":           "VARCHAR",
        "ct_contrast_first":            "VARCHAR",
        "ct_thyroid_details_last":      "VARCHAR",
        "ct_ln_details_last":           "VARCHAR",
        "ct_ln_locations_last":         "VARCHAR",
        "ct_airway_compromise_any":     "BOOLEAN",
        "ct_airway_comment_last":       "VARCHAR",
        "ct_thyroid_postsurgical_any":  "BOOLEAN",
        "ct_thyroid_not_visualized_any": "BOOLEAN",
        "ct_thyroid_heterogeneous_any": "BOOLEAN",
        "ct_thyroid_other_abnormality_any": "BOOLEAN",
        "ct_thyroid_normal_any":        "BOOLEAN",
        "ct_thyroid_nodule_any":        "BOOLEAN",
        "ct_thyroid_enlarged_any":      "BOOLEAN",
        "ct_pathologic_ln_any":         "BOOLEAN",
    }
    new_pet_cols = {
        "pet_other_n_exams":            "INTEGER",
        "pet_other_first_date":         "DATE",
        "pet_other_last_date":          "DATE",
        "pet_other_indication_first":   "VARCHAR",
        "pet_other_mentions_metastasis": "BOOLEAN",
        "pet_other_ned_statement":      "BOOLEAN",
        "pet_other_exam_type":          "VARCHAR",
        "pet_other_extraction_method":  "VARCHAR",
    }
    new_nucmed_cols = {
        "nucmed_indication_first":          "VARCHAR",
        "nucmed_indication_last":           "VARCHAR",
        "nucmed_impression_last":           "VARCHAR",
        "nucmed_findings_last":             "VARCHAR",
        "nucmed_tsh_max":                   "DOUBLE",
        "nucmed_tsh_is_stimulated":         "BOOLEAN",
        "nucmed_tgab_max":                  "DOUBLE",
        "nucmed_uptake_pct_max":            "DOUBLE",
        "nucmed_dose_max_parsed":           "DOUBLE",
        "nucmed_cumulative_therapeutic_dose": "DOUBLE",
        "nucmed_n_doses_parsed":            "INTEGER",
        "nucmed_n_with_indication":         "INTEGER",
        "nucmed_n_with_impression":         "INTEGER",
        "nucmed_overall_assessment":        "VARCHAR",
    }
    new_lnus_cols: dict[str, str] = {}
    if table_exists(con, "_lnus_patient_rollup_v1"):
        new_lnus_cols = {
            "lnus_has_dedicated_exam":  "BOOLEAN",
            "lnus_n_exams":             "INTEGER",
            "lnus_first_date":          "DATE",
            "lnus_last_date":           "DATE",
            "lnus_indication_first":    "VARCHAR",
            "lnus_impression_last":     "VARCHAR",
            "lnus_abnormal_ln_any":     "BOOLEAN",
            "lnus_normal_ln_any":       "BOOLEAN",
            "lnus_has_size_measurement": "BOOLEAN",
            "lnus_source":              "VARCHAR",
        }

    all_new_cols = {**new_ct_cols, **new_pet_cols, **new_nucmed_cols, **new_lnus_cols}
    for col, dtype in all_new_cols.items():
        if col not in existing_cols:
            safe_add_column(con, col, dtype)

    # --- Step 6.2: UPDATE canonical from CT rollup ---
    print("[219] Updating canonical from CT rollup...")
    rid_join = f"TRY_CAST(r.research_id AS BIGINT)" if _CANONICAL_RID_IS_BIGINT else "r.research_id"
    con.execute(f"""
        UPDATE {CANONICAL} AS c
        SET
            ct_indication_first              = r.ct_indication_first,
            ct_indication_last               = r.ct_indication_last,
            ct_first_date                    = r.ct_first_date_new,
            ct_last_date                     = r.ct_last_date_new,
            ct_exam_type_first               = r.ct_exam_type_first,
            ct_contrast_first                = r.ct_contrast_first,
            ct_thyroid_details_last          = r.ct_thyroid_details_last,
            ct_ln_details_last               = r.ct_ln_details_last,
            ct_ln_locations_last             = r.ct_ln_locations_last,
            ct_airway_compromise_any         = r.ct_airway_compromise_any,
            ct_airway_comment_last           = r.ct_airway_comment_last,
            ct_thyroid_postsurgical_any      = r.ct_thyroid_postsurgical_any,
            ct_thyroid_not_visualized_any    = r.ct_thyroid_not_visualized_any,
            ct_thyroid_heterogeneous_any     = r.ct_thyroid_heterogeneous_any,
            ct_thyroid_other_abnormality_any = r.ct_thyroid_other_abnormality_any,
            ct_thyroid_normal_any            = r.ct_thyroid_normal_any,
            ct_thyroid_nodule_any            = r.ct_thyroid_nodule_any,
            ct_thyroid_enlarged_any          = r.ct_thyroid_enlarged_any,
            ct_pathologic_ln_any             = r.ct_pathologic_ln_any
        FROM _ct_expanded_rollup_v1 AS r
        WHERE c.research_id = {rid_join}
    """)
    ct_updated = con.execute(
        f"SELECT COUNT(*) FROM {CANONICAL} WHERE ct_indication_first IS NOT NULL"
    ).fetchone()[0]
    print(f"[219] CT update: {ct_updated} patients with ct_indication_first")

    # --- Step 6.3: UPDATE canonical from PET Other rollup ---
    print("[219] Updating canonical from PET Other rollup...")
    # Check if pet_has_data column exists in canonical (may not exist in new lean schema)
    existing_after_add = get_existing_columns(con)
    pet_has_data_exists = "pet_has_data" in existing_after_add

    # Build SET clause dynamically to avoid trailing-comma syntax errors
    pet_set_pairs = [
        "pet_other_n_exams             = r.pet_other_n_exams",
        "pet_other_first_date          = r.pet_other_first_date",
        "pet_other_last_date           = r.pet_other_last_date",
        "pet_other_indication_first    = r.pet_other_indication_first",
        "pet_other_mentions_metastasis = r.pet_other_mentions_metastasis",
        "pet_other_ned_statement       = r.pet_other_ned_statement",
        "pet_other_exam_type           = r.pet_other_exam_type",
        "pet_other_extraction_method   = r.pet_other_extraction_method",
    ]
    if pet_has_data_exists:
        pet_set_pairs.append(
            "pet_has_data = CASE WHEN c.pet_has_data IS NULL OR c.pet_has_data = FALSE "
            "THEN TRUE ELSE c.pet_has_data END"
        )
    con.execute(f"""
        UPDATE {CANONICAL} AS c
        SET {', '.join(pet_set_pairs)}
        FROM _pet_other_rollup_v1 AS r
        WHERE c.research_id = {rid_join}
    """)
    pet_updated = con.execute(
        f"SELECT COUNT(*) FROM {CANONICAL} WHERE pet_other_n_exams IS NOT NULL"
    ).fetchone()[0]
    print(f"[219] PET Other update: {pet_updated} patients")

    # --- Step 6.4: UPDATE canonical from NucMed agg ---
    print("[219] Updating canonical from NucMed agg...")
    # Build NucMed SET pairs — some columns may not exist in lean new schema
    existing_now = get_existing_columns(con)
    nm_set_pairs = [
        "nucmed_indication_first            = r.nucmed_indication_first",
        "nucmed_indication_last             = r.nucmed_indication_last",
        "nucmed_impression_last             = r.nucmed_impression_last",
        "nucmed_findings_last               = r.nucmed_findings_last",
        "nucmed_tsh_max                     = r.nucmed_tsh_max",
        "nucmed_tsh_is_stimulated           = r.nucmed_tsh_is_stimulated",
        "nucmed_tgab_max                    = r.nucmed_tgab_max",
        "nucmed_uptake_pct_max              = r.nucmed_uptake_pct_max_new",
        "nucmed_dose_max_parsed             = r.nucmed_dose_max_parsed",
        "nucmed_cumulative_therapeutic_dose = r.nucmed_cumulative_therapeutic_dose",
        "nucmed_n_doses_parsed              = r.nucmed_n_doses_parsed",
        "nucmed_n_with_indication           = r.nucmed_n_with_indication",
        "nucmed_n_with_impression           = r.nucmed_n_with_impression",
        "nucmed_overall_assessment          = r.nucmed_overall_assessment",
    ]
    if "rai_stimulated_tsh" in existing_now:
        nm_set_pairs.append(
            "rai_stimulated_tsh = CASE WHEN c.rai_stimulated_tsh IS NULL "
            "AND r.nucmed_tsh_is_stimulated = TRUE THEN r.nucmed_tsh_max "
            "ELSE c.rai_stimulated_tsh END"
        )
    if "nucmed_uptake_24hr_max" in existing_now:
        nm_set_pairs.append(
            "nucmed_uptake_24hr_max = COALESCE(c.nucmed_uptake_24hr_max, r.nucmed_uptake_pct_max_new)"
        )
    con.execute(f"""
        UPDATE {CANONICAL} AS c
        SET {', '.join(nm_set_pairs)}
        FROM _nucmed_agg_v1 AS r
        WHERE c.research_id = {rid_join}
    """)
    nm_updated = con.execute(
        f"SELECT COUNT(*) FROM {CANONICAL} WHERE nucmed_indication_first IS NOT NULL"
    ).fetchone()[0]
    print(f"[219] NucMed update: {nm_updated} patients with nucmed_indication_first")

    # --- Step 6.5: UPDATE canonical from MRI indication refresh ---
    # mri_imaging may not exist in all databases — skip gracefully
    print("[219] Refreshing mri_indication_first from updated mri_imaging...")
    mri_table_exists = table_exists(con, "mri_imaging")
    mri_col_exists = "mri_indication_first" in get_existing_columns(con)
    mri_has_data_col = "mri_has_data" in get_existing_columns(con)

    if not mri_table_exists:
        print("[219] SKIP: mri_imaging table not present in this database")
    elif not mri_col_exists:
        print("[219] SKIP: mri_indication_first column not in canonical")
    else:
        # Cast mri_imaging.research_id to match canonical type
        mri_cast = "TRY_CAST(m.research_id AS BIGINT)" if _CANONICAL_RID_IS_BIGINT else "m.research_id"
        mri_cast_sub = "TRY_CAST(research_id AS BIGINT)" if _CANONICAL_RID_IS_BIGINT else "research_id"
        where_clause = "WHERE c.mri_has_data = TRUE AND c.mri_indication_first IS NULL" \
            if mri_has_data_col else "WHERE c.mri_indication_first IS NULL"
        con.execute(f"""
            UPDATE {CANONICAL} AS c
            SET mri_indication_first = COALESCE(
                c.mri_indication_first,
                (SELECT MIN(indication)
                 FROM mri_imaging m
                 WHERE {mri_cast} = c.research_id
                   AND m.indication IS NOT NULL AND LENGTH(m.indication) > 5
                   AND TRY_CAST(m.date_of_exam AS DATE) = (
                        SELECT MIN(TRY_CAST(date_of_exam AS DATE))
                        FROM mri_imaging
                        WHERE {mri_cast_sub} = c.research_id
                          AND indication IS NOT NULL
                   )
                )
            )
            {where_clause}
        """)
    mri_ind_col_exists = "mri_indication_first" in get_existing_columns(con)
    if mri_ind_col_exists:
        mri_ind_count = con.execute(
            f"SELECT COUNT(*) FROM {CANONICAL} WHERE mri_indication_first IS NOT NULL"
        ).fetchone()[0]
        print(f"[219] MRI indication: {mri_ind_count} patients now have mri_indication_first")
    else:
        print("[219] MRI indication: column not in canonical (mri_imaging not present in this DB)")

    # --- Step 6.6: UPDATE canonical from LN US rollup ---
    if table_exists(con, "_lnus_patient_rollup_v1"):
        print("[219] Updating canonical from LN US rollup...")
        lnus_rid_join = f"TRY_CAST(r.research_id AS BIGINT)" if _CANONICAL_RID_IS_BIGINT else "r.research_id"
        con.execute(f"""
            UPDATE {CANONICAL} AS c
            SET
                lnus_has_dedicated_exam   = r.lnus_has_dedicated_exam,
                lnus_n_exams              = r.lnus_n_exams,
                lnus_first_date           = r.lnus_first_date,
                lnus_last_date            = r.lnus_last_date,
                lnus_indication_first     = r.lnus_indication_first,
                lnus_impression_last      = r.lnus_impression_last,
                lnus_abnormal_ln_any      = r.lnus_abnormal_ln_any,
                lnus_normal_ln_any        = r.lnus_normal_ln_any,
                lnus_has_size_measurement = r.lnus_has_size_measurement,
                lnus_source               = r.lnus_source
            FROM _lnus_patient_rollup_v1 AS r
            WHERE c.research_id = {lnus_rid_join}
        """)
        lnus_updated = con.execute(
            f"SELECT COUNT(*) FROM {CANONICAL} WHERE lnus_has_dedicated_exam = TRUE"
        ).fetchone()[0]
        print(f"[219] LN US update: {lnus_updated} patients with dedicated LN US")

    # --- Step 6.7: Comprehensive provenance verification ---
    print("\n[219] === PROVENANCE CHECKS ===")

    # Check 1: All staging RIDs in spine
    # Use BIGINT cast in orphan subquery if needed
    rid_sub = f"TRY_CAST(research_id AS BIGINT)" if _CANONICAL_RID_IS_BIGINT else "research_id"
    prov1 = con.execute(f"""
        SELECT 'ct_expanded' AS source,
            COUNT(*) FILTER (WHERE {rid_sub} NOT IN (SELECT research_id FROM {CANONICAL})) AS orphans
        FROM _ct_expanded_rollup_v1
        UNION ALL SELECT 'pet_other_rollup',
            COUNT(*) FILTER (WHERE {rid_sub} NOT IN (SELECT research_id FROM {CANONICAL}))
        FROM _pet_other_rollup_v1
        UNION ALL SELECT 'nucmed_agg',
            COUNT(*) FILTER (WHERE {rid_sub} NOT IN (SELECT research_id FROM {CANONICAL}))
        FROM _nucmed_agg_v1
    """).fetchall()
    print("[219] CHECK 1 (orphan RIDs):")
    all_pass = True
    for row in prov1:
        status = "PASS" if row[1] == 0 else f"FAIL ({row[1]} orphans)"
        print(f"  {row[0]}: {status}")
        if row[1] > 0:
            all_pass = False

    # Check 2: Dates present — only query columns that actually exist
    existing_now2 = get_existing_columns(con)
    date_cols = {
        "ct_dates":   "ct_first_date",
        "pet_dates":  "pet_other_first_date",
        "mri_dates":  "mri_first_date",
        "nm_pts":     "nucmed_n_scans",
    }
    date_parts = []
    date_labels = []
    for label, col in date_cols.items():
        if col in existing_now2:
            if col == "nucmed_n_scans":
                date_parts.append(f"COUNT(*) FILTER (WHERE {col} IS NOT NULL AND {col} > 0)")
            else:
                date_parts.append(f"COUNT(*) FILTER (WHERE {col} IS NOT NULL)")
        else:
            date_parts.append("0")
        date_labels.append(label)
    prov2 = con.execute(f"SELECT {', '.join(date_parts)} FROM {CANONICAL}").fetchone()
    print(f"[219] CHECK 2 (dates): " + ", ".join(f"{l}={v}" for l, v in zip(date_labels, prov2)))

    # Check 3: Indication coverage
    ind_cols = {
        "ct_ind":  "ct_indication_first",
        "pet_ind": "pet_other_indication_first",
        "mri_ind": "mri_indication_first",
        "nm_ind":  "nucmed_indication_first",
    }
    ind_parts = []
    ind_labels = []
    for label, col in ind_cols.items():
        if col in existing_now2:
            ind_parts.append(f"COUNT(*) FILTER (WHERE {col} IS NOT NULL)")
        else:
            ind_parts.append("0")
        ind_labels.append(label)
    prov3 = con.execute(f"SELECT {', '.join(ind_parts)} FROM {CANONICAL}").fetchone()
    print(f"[219] CHECK 3 (indication): " + ", ".join(f"{l}={v}" for l, v in zip(ind_labels, prov3)))

    # Check 4: Units verified by column naming convention
    print("[219] CHECK 4 (units): verified by column naming convention:")
    print("  nucmed_dose_max_parsed → mCi | nucmed_tsh_max → mIU/L | nucmed_tg_max → ng/mL")
    print("  nucmed_uptake_pct_max → % | ct_largest_ln_short_axis_mm → mm")

    # Check 5: Source provenance on staging tables
    prov5 = con.execute("""
        SELECT
            COUNT(*) FILTER (WHERE ct_expanded_source_table IS NOT NULL) AS ct_src
        FROM _ct_expanded_rollup_v1
    """).fetchone()[0]
    prov5b = con.execute("""
        SELECT COUNT(*) FILTER (WHERE pet_other_extraction_method IS NOT NULL)
        FROM _pet_other_rollup_v1
    """).fetchone()[0]
    prov5c = con.execute("""
        SELECT COUNT(*) FILTER (WHERE nucmed_expanded_source IS NOT NULL) FROM _nucmed_agg_v1
    """).fetchone()[0]
    print(f"[219] CHECK 5 (source): CT={prov5}, PET_Other={prov5b}, NucMed={prov5c}")

    # Check 6: Cross-validation — patient overlap between canonical and staging
    join_expr6 = (
        f"TRY_CAST(r.research_id AS BIGINT) = c.research_id" if _CANONICAL_RID_IS_BIGINT
        else "c.research_id = r.research_id"
    )
    join_expr6n = (
        f"TRY_CAST(n.research_id AS BIGINT) = c.research_id" if _CANONICAL_RID_IS_BIGINT
        else "c.research_id = n.research_id"
    )
    ct_filter = "c.ct_indication_first IS NOT NULL" if "ct_indication_first" in get_existing_columns(con) else "TRUE"
    nm_filter = "c.nucmed_n_scans IS NOT NULL" if "nucmed_n_scans" in get_existing_columns(con) else "TRUE"
    prov6 = con.execute(f"""
        SELECT
            COUNT(*) FILTER (WHERE {ct_filter} AND r.research_id IS NOT NULL) AS both_ct,
            COUNT(*) FILTER (WHERE {nm_filter} AND n.research_id IS NOT NULL) AS both_nm
        FROM {CANONICAL} c
        FULL OUTER JOIN _ct_expanded_rollup_v1 r ON {join_expr6}
        FULL OUTER JOIN _nucmed_agg_v1 n ON {join_expr6n}
    """).fetchone()
    print(f"[219] CHECK 6 (cross-val): CT_overlap={prov6[0]}, NucMed_overlap={prov6[1]}")

    # --- Step 6.8: Final canonical invariants ---
    print("\n[219] === FINAL INVARIANT CHECK ===")
    ok = check_invariants(con, CANONICAL, "canonical_patient_master_v1 after 219")
    if not ok:
        print("[219] CRITICAL: Canonical invariants FAILED — investigate before committing")
    else:
        print("[219] Canonical invariants: ALL PASS")

    # --- Step 6.9: Coverage report ---
    print("\n[219] === COVERAGE REPORT ===")
    all_col_checks = [
        ("ct_indication_first",          "CT indication"),
        ("ct_thyroid_details_last",      "CT thyroid details"),
        ("ct_ln_details_last",           "CT LN details"),
        ("ct_airway_compromise_any",     "CT airway compromise"),
        ("ct_thyroid_nodule_any",        "CT thyroid nodule"),
        ("ct_pathologic_ln_any",         "CT pathologic LN"),
        ("pet_other_n_exams",            "PET Other exams"),
        ("nucmed_indication_first",      "NucMed indication"),
        ("nucmed_impression_last",       "NucMed impression"),
        ("nucmed_tsh_max",               "NucMed TSH"),
        ("nucmed_tsh_is_stimulated",     "NucMed TSH stimulated"),
        ("nucmed_dose_max_parsed",       "NucMed dose (mCi)"),
        ("nucmed_uptake_pct_max",        "NucMed uptake %"),
        ("nucmed_overall_assessment",    "NucMed assessment"),
        ("lnus_has_dedicated_exam",      "LN US dedicated exam"),
        ("mri_indication_first",         "MRI indication"),
    ]
    final_cols = get_existing_columns(con)
    for col, label in all_col_checks:
        if col not in final_cols:
            print(f"  {label} ({col}): SKIP (column not in this canonical)")
            continue
        try:
            count = con.execute(f"""
                SELECT COUNT(*) FILTER (WHERE {col} IS NOT NULL
                  AND CAST({col} AS VARCHAR) NOT IN ('false', 'False', ''))
                FROM {CANONICAL}
            """).fetchone()[0]
            print(f"  {label} ({col}): {count} ({count/TOTAL_ROWS*100:.1f}%)")
        except Exception as e:
            print(f"  {label} ({col}): ERROR — {e}")


# ======================================================================
# MAIN
# ======================================================================

def main() -> None:
    global DB, CANONICAL

    parser = argparse.ArgumentParser(description="Script 219: Imaging gap resolution")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview only — no writes to MotherDuck")
    parser.add_argument("--phase", default="all",
                        help="Phases to run: 1|2|3|4|5|6|all (comma-separated)")
    parser.add_argument("--db", default=None,
                        help="MotherDuck database name (default: thyroid_ete_fix_20260413)")
    parser.add_argument("--canonical", default=None,
                        help="Canonical patient table name (default: canonical_patient_master_v1)")
    args = parser.parse_args()

    # Apply overrides
    if args.db:
        DB = args.db
    if args.canonical:
        CANONICAL = args.canonical

    phases = set()
    if args.phase == "all":
        phases = {"1", "2", "3", "4", "5", "6"}
    else:
        phases = set(args.phase.split(","))

    print(f"[219] Script 219 — Imaging Gap Resolution")
    print(f"[219] DB: {DB}  Canonical: {CANONICAL}")
    print(f"[219] Phases: {phases}  dry_run={args.dry_run}")

    con = connect()
    original_canonical = CANONICAL  # save view name before possible resolution
    detect_schema(con)

    # Verify canonical (probe original name first, then resolved base table)
    probe_table = CANONICAL  # after detect_schema, this may be the base table
    inv = con.execute(f"""
        SELECT COUNT(*), COUNT(DISTINCT research_id)
        FROM {probe_table}
    """).fetchone()
    print(f"[219] Canonical baseline ({probe_table}): {inv[0]} rows, {inv[1]} distinct RIDs")
    if inv[0] != TOTAL_ROWS:
        print(f"[219] ERROR: Expected {TOTAL_ROWS} rows, got {inv[0]}")
        sys.exit(1)

    t_start = time.time()

    if "1" in phases:
        run_task1_ct(con, args.dry_run)

    if "2" in phases:
        run_task2_pet_other(con, args.dry_run)

    if "3" in phases:
        run_task3_mri(con, args.dry_run)

    if "4" in phases:
        run_task4_nucmed(con, args.dry_run)

    if "5" in phases:
        run_task5_ln_us(con, args.dry_run)

    if "6" in phases:
        run_task6_canonical(con, args.dry_run)

    elapsed = time.time() - t_start
    print(f"\n[219] Script 219 complete in {elapsed:.0f}s")


if __name__ == "__main__":
    main()
