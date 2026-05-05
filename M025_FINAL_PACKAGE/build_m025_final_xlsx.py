#!/usr/bin/env python3
"""
M025 v2.0 — final-package xlsx builder.

Pulls from MotherDuck thyroid_canonical_publication_v1_0:
  - manuscript_workspace.m025_analytic_master_patient_v1   (3,375 rows)
  - manuscript_workspace.m025_analytic_master_nodule_v1    (37,438 rows)
  - manuscript_workspace.m025_threshold_metrics_v1         (6 rows)
  - manuscript_workspace.m025_rom_by_tr_v1                 (9 rows)
  - manuscript_workspace.m025_bethesda_x_tr_v1             (25 rows)
  - manuscript_workspace.m025_sens_era_* / m025_sens_window_nodule_v1 (mig_307c)

Produces:
  1. M025_FINAL_PACKAGE/M025_master_data.xlsx
       Cover, patient_master, nodule_master_strict, nodule_master_all,
       threshold_metrics, rom_by_tr, bethesda_x_tr, run_snapshot,
       definitions, gates.
  2. M025_FINAL_PACKAGE/M025_tables_and_summary.xlsx
       Cover, Table 1–7, subgroups, sensitivity sheets (Wilson CIs via mig_307d), QA_gates.
  3. M025_FINAL_PACKAGE/06_figures_sensitivity/
       CSV sidecars + 300 DPI PNGs (forest + match-window line chart).

After rebuild, INSERT signoff (once): .venv/bin/python M025_FINAL_PACKAGE/m025_sensitivity_mig_307d.py

Usage:
  cd /Users/loganglosser/THYROID_2026
  .venv/bin/python M025_FINAL_PACKAGE/build_m025_final_xlsx.py
"""

from __future__ import annotations

import os
import sys
from datetime import datetime, timezone

import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Allow import of motherduck_client.py from repo root + local package modules
_PKG_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, _PKG_DIR)
sys.path.insert(0, os.path.dirname(_PKG_DIR))
from motherduck_client import get_token  # type: ignore

from m025_sensitivity_lib import (
    add_rom_ci_lo_hi,
    augment_window_table,
    export_sensitivity_csv_bundle,
    fetch_diagnostic_by_era,
    fetch_per_era_auc,
    render_forest_rom_by_era,
    render_linechart_match_window,
    wilson_ci,
)

DB = "thyroid_canonical_publication_v1_0"
OUTDIR = _PKG_DIR
MASTER_PATH = os.path.join(OUTDIR, "M025_master_data.xlsx")
TABLES_PATH = os.path.join(OUTDIR, "M025_tables_and_summary.xlsx")
SENS_FIG_DIR = os.path.join(OUTDIR, "06_figures_sensitivity")

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SUB_FONT = Font(italic=True, color="595959")
THIN = Side(border_style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def connect():
    token = get_token()
    return duckdb.connect(f"md:{DB}?motherduck_token={token}")


def _strip_tz(df: pd.DataFrame) -> pd.DataFrame:
    """Strip tzinfo from datetime columns so openpyxl can write them."""
    df = df.copy()
    for c in df.columns:
        s = df[c]
        if pd.api.types.is_datetime64_any_dtype(s):
            try:
                if getattr(s.dt, "tz", None) is not None:
                    df[c] = s.dt.tz_convert("UTC").dt.tz_localize(None)
            except Exception:
                df[c] = pd.to_datetime(s, errors="coerce").dt.tz_localize(None) if hasattr(s.dt, "tz_localize") else s
        elif s.dtype == object:
            # mixed-type object cols may contain tz-aware datetimes
            def _coerce(v):
                if hasattr(v, "tzinfo") and v.tzinfo is not None:
                    return v.replace(tzinfo=None)
                return v
            df[c] = s.map(_coerce)
    return df


def write_dataframe(ws, df: pd.DataFrame, start_row: int = 1, header: bool = True, freeze: bool = True):
    df = _strip_tz(df)
    if header:
        for j, col in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=j, value=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = BORDER
        start_row += 1
    for i, (_, row) in enumerate(df.iterrows()):
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=start_row + i, column=j, value=(None if pd.isna(v) else v))
            cell.border = BORDER
            if isinstance(v, float):
                cell.number_format = "0.00"
    # column widths
    for j, col in enumerate(df.columns, 1):
        try:
            maxlen = max(len(str(col)), df[col].astype(str).map(len).max() if len(df) else 0)
        except Exception:
            maxlen = max(len(str(col)), 12)
        ws.column_dimensions[get_column_letter(j)].width = min(max(12, int(maxlen) + 2), 60)
    if freeze and header:
        ws.freeze_panes = "A2"


def title_block(ws, title: str, subtitle: str | None = None, row: int = 1):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    if subtitle:
        cell2 = ws.cell(row=row + 1, column=1, value=subtitle)
        cell2.font = SUB_FONT
        return row + 3
    return row + 2


def main():
    con = connect()
    print(f"[{datetime.now(timezone.utc).isoformat()}] Connected to {DB}")

    # ---------------------------------------------------------------
    # Pull data
    # ---------------------------------------------------------------
    print("Pulling patient_master ...")
    df_p = con.execute("SELECT * FROM manuscript_workspace.m025_analytic_master_patient_v1 ORDER BY research_id").df()
    print(f"  patient: {len(df_p):,} rows × {df_p.shape[1]} cols")

    print("Pulling nodule_master_all ...")
    df_n = con.execute("SELECT * FROM manuscript_workspace.m025_analytic_master_nodule_v1 ORDER BY research_id, nodule_master_id").df()
    print(f"  nodule_all: {len(df_n):,} rows × {df_n.shape[1]} cols")

    df_n_strict = df_n[df_n["analytic_eligible_strict_acr_pernodule"] == True].copy()
    print(f"  nodule_strict: {len(df_n_strict):,} rows")

    print("Pulling threshold_metrics + ROM + Bethesda ...")
    df_thr = con.execute("SELECT * FROM manuscript_workspace.m025_threshold_metrics_v1 ORDER BY grain, threshold").df()
    df_rom = con.execute("SELECT * FROM manuscript_workspace.m025_rom_by_tr_v1 ORDER BY grain, tr_category").df()
    df_btr = con.execute("SELECT * FROM manuscript_workspace.m025_bethesda_x_tr_v1 ORDER BY bethesda_bucket, tr_category").df()

    # Add Wilson CIs to threshold metrics
    def add_ci(row):
        sens_lo, sens_hi = wilson_ci(int(row["tp"]), int(row["tp"] + row["fn"]))
        spec_lo, spec_hi = wilson_ci(int(row["tn"]), int(row["tn"] + row["fp"]))
        ppv_lo, ppv_hi = wilson_ci(int(row["tp"]), int(row["tp"] + row["fp"]))
        npv_lo, npv_hi = wilson_ci(int(row["tn"]), int(row["tn"] + row["fn"]))
        return pd.Series({
            "sens_lo_95": round(100 * sens_lo, 2), "sens_hi_95": round(100 * sens_hi, 2),
            "spec_lo_95": round(100 * spec_lo, 2), "spec_hi_95": round(100 * spec_hi, 2),
            "ppv_lo_95": round(100 * ppv_lo, 2), "ppv_hi_95": round(100 * ppv_hi, 2),
            "npv_lo_95": round(100 * npv_lo, 2), "npv_hi_95": round(100 * npv_hi, 2),
        })
    ci_cols = df_thr.apply(add_ci, axis=1)
    df_thr_full = pd.concat([df_thr.drop(columns=["built_at"]), ci_cols], axis=1)

    # Add Wilson CIs to ROM by TR
    def add_rom_ci(row):
        lo, hi = wilson_ci(int(row["n_malignant"]), int(row["n_total"]))
        return pd.Series({"rom_lo_95": round(100 * lo, 2), "rom_hi_95": round(100 * hi, 2)})
    rom_ci = df_rom.apply(add_rom_ci, axis=1)
    df_rom_full = pd.concat([df_rom, rom_ci], axis=1)

    # Bethesda x TR pivot for readability
    df_btr_pivot = df_btr.pivot_table(index="bethesda_bucket", columns="tr_category", values="n", fill_value=0).reset_index()
    df_btr_pivot.columns.name = None

    # ---------------------------------------------------------------
    # Run snapshot
    # ---------------------------------------------------------------
    snapshot = {
        "build_utc": datetime.now(timezone.utc).isoformat(),
        "database": DB,
        "patient_cohort_n": int(len(df_p)),
        "patient_cohort_n_malignant": int(df_p["is_malignant"].sum()),
        "patient_overall_malignancy_pct": round(100 * df_p["is_malignant"].mean(), 2),
        "nodule_total_rows": int(len(df_n)),
        "nodule_distinct_patients": int(df_n["research_id"].nunique()),
        "nodule_strict_eligible": int(len(df_n_strict)),
        "nodule_strict_distinct_patients": int(df_n_strict["research_id"].nunique()),
        "nodule_strict_path_malignant": int(df_n_strict["nodule_path_proven_malignant"].sum()),
        "patient_auc_locked": 0.6478,
        "patient_auc_lo_95": 0.6301,
        "patient_auc_hi_95": 0.6665,
        "nodule_auc_locked": 0.6399,
        "youden_optimal_threshold": "TR>=TR4",
        "youden_J": 0.271,
    }

    # ---------------------------------------------------------------
    # WORKBOOK 1 — master_data.xlsx
    # ---------------------------------------------------------------
    print("Building master_data.xlsx ...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Cover"
    title_block(ws, "M025 v2.0 — Master Data Workbook",
                f"Per-patient and per-nodule data points for manuscript writing. Built {snapshot['build_utc']}.")
    cover_rows = [
        ("Database", DB),
        ("Patient cohort view", "manuscript_workspace.cohort_m025_tirads_performance_v1"),
        ("Nodule cohort view", "manuscript_workspace.cohort_m025_nodule_level_v1 (mig_306)"),
        ("Patient analytic master", "manuscript_workspace.m025_analytic_master_patient_v1 (mig_307b)"),
        ("Nodule analytic master", "manuscript_workspace.m025_analytic_master_nodule_v1 (mig_307b)"),
        ("Build UTC", snapshot["build_utc"]),
        ("Patient cohort n", snapshot["patient_cohort_n"]),
        ("Patient malignant n", snapshot["patient_cohort_n_malignant"]),
        ("Overall malignancy rate %", snapshot["patient_overall_malignancy_pct"]),
        ("Nodule total rows", snapshot["nodule_total_rows"]),
        ("Nodule distinct patients", snapshot["nodule_distinct_patients"]),
        ("Nodule strict-eligible n", snapshot["nodule_strict_eligible"]),
        ("Nodule strict-eligible distinct pts", snapshot["nodule_strict_distinct_patients"]),
        ("Nodule strict path-malignant n", snapshot["nodule_strict_path_malignant"]),
        ("Patient AUC (locked)", snapshot["patient_auc_locked"]),
        ("Patient AUC 95% CI", f"[{snapshot['patient_auc_lo_95']}, {snapshot['patient_auc_hi_95']}]"),
        ("Nodule AUC (locked)", snapshot["nodule_auc_locked"]),
        ("Youden optimal threshold", snapshot["youden_optimal_threshold"]),
        ("Youden J", snapshot["youden_J"]),
        ("Sister patient package", "M025_submission_package_v1_0/ (frozen, mig_292)"),
        ("Manuscript framing", "Patient-level headline + nodule-level sister analysis"),
    ]
    for i, (k, v) in enumerate(cover_rows, start=4):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 64

    # patient_master sheet
    ws = wb.create_sheet("patient_master")
    title_block(ws, "Patient analytic master (n=3,375)",
                "1 row per patient. is_malignant = TRUE when histology proves thyroid malignancy.")
    write_dataframe(ws, df_p, start_row=4)

    # nodule_master_strict
    ws = wb.create_sheet("nodule_master_strict")
    title_block(ws, "Nodule analytic master — strict ACR-eligible (n=3,687)",
                "1 row per nodule passing analytic_eligible_strict_acr_pernodule = TRUE. Used for nodule-level sister analysis.")
    write_dataframe(ws, df_n_strict, start_row=4)

    # nodule_master_all
    ws = wb.create_sheet("nodule_master_all")
    title_block(ws, "Nodule analytic master — full spine (n=37,438)",
                "1 row per US nodule observation. Strict subset is `analytic_eligible_strict_acr_pernodule = TRUE`.")
    write_dataframe(ws, df_n, start_row=4)

    # threshold_metrics
    ws = wb.create_sheet("threshold_metrics")
    title_block(ws, "Threshold metrics — Sens/Spec/PPV/NPV at TR>=TR3, TR4, TR5",
                "Both grains. Wilson 95% CIs in *_lo_95 / *_hi_95 columns.")
    write_dataframe(ws, df_thr_full, start_row=4)

    # rom_by_tr
    ws = wb.create_sheet("rom_by_tr")
    title_block(ws, "Risk of malignancy by TI-RADS category",
                "patient grain = max_tirads_category_ever; nodule_strict = analytic_eligible_strict_acr_pernodule.")
    write_dataframe(ws, df_rom_full, start_row=4)

    # bethesda_x_tr
    ws = wb.create_sheet("bethesda_x_tr")
    title_block(ws, "Bethesda × TI-RADS — strict-eligible nodules", "n per cell.")
    write_dataframe(ws, df_btr_pivot, start_row=4)

    # run_snapshot
    ws = wb.create_sheet("run_snapshot")
    title_block(ws, "Run snapshot", "Key counts, AUCs, optimal threshold.")
    snap_df = pd.DataFrame([(k, v) for k, v in snapshot.items()], columns=["key", "value"])
    write_dataframe(ws, snap_df, start_row=4)

    # definitions
    ws = wb.create_sheet("definitions")
    title_block(ws, "Definitions and notes")
    defs = [
        ("max_tirads_category_ever", "Patient-level max ACR TI-RADS category across all imaging exams"),
        ("acr2017_tirads_category", "Per-nodule ACR 2017 TR category (strict ACR points complete)"),
        ("analytic_eligible_strict_acr_pernodule", "TRUE if the nodule has all ACR2017 features scored AND no size-quarantine"),
        ("nodule_path_proven_malignant", "Nodule-level: matched to an operative path tumor on same side ≤365 days post-US"),
        ("is_malignant", "Patient-level: any pathology-proven thyroid malignancy in the cohort"),
        ("bethesda_final / bethesda_final_num", "Patient/nodule Bethesda category (1-6)"),
        ("predicted_pos_TR3/TR4/TR5", "Treats TR>=TR3 / TR>=TR4 / TR>=TR5 as 'predict malignant' for diagnostic-performance metrics"),
        ("Wilson 95% CI", "Score-based binomial CI (Wilson, 1927); preferred over Wald for proportions near 0/100"),
        ("Strict-DTC subset", "Patient cohort restricted to PTC + follicular carcinoma; NOT applied here — full cohort retained"),
        ("Patient-level headline", "AUC 0.6478 [0.6301-0.6665], Youden J 0.271 at TR>=TR4"),
        ("Nodule-level sister", "AUC 0.6399; per-nodule TR4 ROM 18.7% / TR5 26.1% inside ACR bands"),
        ("Multinodular attribution", "Patient inflation = sum of ROM across all that-patient's nodules; ETE/multinodularity confounds patient grain"),
    ]
    defs_df = pd.DataFrame(defs, columns=["term", "definition"])
    write_dataframe(ws, defs_df, start_row=4)

    # gates
    ws = wb.create_sheet("gates")
    title_block(ws, "Build gates — should all PASS")
    gates = [
        ("Patient cohort n == 3,375", snapshot["patient_cohort_n"] == 3375),
        ("Patient malignant n == 1,479", snapshot["patient_cohort_n_malignant"] == 1479),
        ("Nodule total == 37,438", snapshot["nodule_total_rows"] == 37438),
        ("Nodule strict == 3,687", snapshot["nodule_strict_eligible"] == 3687),
        ("Nodule strict path-malig == 631", snapshot["nodule_strict_path_malignant"] == 631),
        ("Threshold metrics rows == 6", len(df_thr) == 6),
        ("ROM-by-TR rows == 9", len(df_rom) == 9),
        ("Bethesda × TR rows == 25", len(df_btr) == 25),
    ]
    gates_df = pd.DataFrame(gates, columns=["check", "pass"])
    write_dataframe(ws, gates_df, start_row=4)

    wb.save(MASTER_PATH)
    print(f"  → {MASTER_PATH}")

    # ---------------------------------------------------------------
    # WORKBOOK 2 — tables_and_summary.xlsx
    # ---------------------------------------------------------------
    print("Building tables_and_summary.xlsx ...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Cover"
    title_block(ws, "M025 v2.0 — Manuscript Tables & Statistical Summary",
                f"Built {snapshot['build_utc']}. Manuscript headline = patient-level. Nodule-level = sister analysis.")
    cover2 = [
        ("Tables included",
         "T1 Baseline by max-TR | T2 Threshold metrics | T3 Patient-vs-nodule ROM | T4 Bethesda × TR | T5 Histology | T6 FNA-path concordance | T7 Race/era | Subgroups | Sensitivity arms | QA"),
        ("Database", DB),
        ("Patient cohort n", snapshot["patient_cohort_n"]),
        ("Patient AUC", f"{snapshot['patient_auc_locked']} [{snapshot['patient_auc_lo_95']}-{snapshot['patient_auc_hi_95']}]"),
        ("Nodule strict n", snapshot["nodule_strict_eligible"]),
        ("Nodule AUC", snapshot["nodule_auc_locked"]),
        ("Build UTC", snapshot["build_utc"]),
        ("Sign-off", "mig_307b (analytic spine) + mig_307c (sensitivity tables) + mig_307d (sensitivity publication outputs)"),
    ]
    for i, (k, v) in enumerate(cover2, start=4):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90

    # ---------------- Table 1 — Baseline by max-TR ----------------
    print(" Table 1 ...")
    df_t1 = con.execute("""
        SELECT
          max_tirads_category_ever AS max_TR,
          COUNT(*) AS n,
          ROUND(100.0*COUNT(*)/SUM(COUNT(*)) OVER (),2) AS pct_of_cohort,
          ROUND(AVG(age_at_surgery),1) AS age_mean,
          ROUND(STDDEV(age_at_surgery),1) AS age_sd,
          MEDIAN(age_at_surgery) AS age_median,
          COUNT_IF(sex='female') AS n_female,
          ROUND(100.0*COUNT_IF(sex='female')/COUNT(*),1) AS pct_female,
          COUNT_IF(race='Black or African American') AS n_black,
          COUNT_IF(race='White') AS n_white,
          ROUND(MEDIAN(imaging_nodule_size_cm),2) AS img_size_med_cm,
          ROUND(MEDIAN(tumor_size_cm),2) AS path_size_med_cm,
          COUNT_IF(is_malignant) AS n_malignant,
          ROUND(100.0*COUNT_IF(is_malignant)/COUNT(*),2) AS rom_pct,
          COUNT_IF(bethesda_final IS NOT NULL) AS n_with_bethesda,
          COUNT_IF(histology_final IS NOT NULL) AS n_with_path
        FROM manuscript_workspace.m025_analytic_master_patient_v1
        GROUP BY max_tirads_category_ever
        ORDER BY max_tirads_category_ever
    """).df()
    ws = wb.create_sheet("Table_1_Baseline")
    title_block(ws, "Table 1 — Baseline characteristics by max TI-RADS category",
                f"Patient cohort, n={snapshot['patient_cohort_n']:,}.")
    write_dataframe(ws, df_t1, start_row=4)

    # ---------------- Table 2 — Threshold metrics ----------------
    ws = wb.create_sheet("Table_2_Thresholds")
    title_block(ws, "Table 2 — Diagnostic performance at TI-RADS thresholds",
                "Sens / Spec / PPV / NPV with Wilson 95% CIs. Patient grain = manuscript primary; nodule grain = sister analysis.")
    write_dataframe(ws, df_thr_full, start_row=4)

    # ---------------- Table 3 — Patient vs nodule ROM ----------------
    print(" Table 3 ...")
    df_t3 = pd.DataFrame()
    p = df_rom_full[df_rom_full["grain"] == "patient"].rename(columns={"n_total": "patient_n", "n_malignant": "patient_k", "rom_pct": "patient_rom_pct", "rom_lo_95": "patient_lo_95", "rom_hi_95": "patient_hi_95"})
    n = df_rom_full[df_rom_full["grain"] == "nodule_strict"].rename(columns={"n_total": "nodule_n", "n_malignant": "nodule_k", "rom_pct": "nodule_rom_pct", "rom_lo_95": "nodule_lo_95", "rom_hi_95": "nodule_hi_95"})
    df_t3 = p[["tr_category", "patient_n", "patient_k", "patient_rom_pct", "patient_lo_95", "patient_hi_95"]].merge(
        n[["tr_category", "nodule_n", "nodule_k", "nodule_rom_pct", "nodule_lo_95", "nodule_hi_95"]],
        on="tr_category", how="outer")
    acr_band = {"TR1": "<2%", "TR2": "<2%", "TR3": "<5%", "TR4": "5-20%", "TR5": ">20%"}
    df_t3["acr_expected_band"] = df_t3["tr_category"].map(acr_band)

    def in_band(row):
        if pd.isna(row.get("nodule_rom_pct")):
            return ""
        rom = row["nodule_rom_pct"]
        if row["tr_category"] in ("TR1", "TR2"):
            return "YES" if rom < 2 else "no"
        if row["tr_category"] == "TR3":
            return "YES" if rom < 5 else "no"
        if row["tr_category"] == "TR4":
            return "YES" if 5 <= rom <= 20 else "no"
        if row["tr_category"] == "TR5":
            return "YES" if rom > 20 else "no"
        return ""
    df_t3["nodule_in_acr_band"] = df_t3.apply(in_band, axis=1)
    df_t3["inflation_pp_patient_vs_nodule"] = df_t3.apply(
        lambda r: round(r["patient_rom_pct"] - r["nodule_rom_pct"], 2)
        if pd.notna(r.get("patient_rom_pct")) and pd.notna(r.get("nodule_rom_pct")) else None, axis=1)
    ws = wb.create_sheet("Table_3_Patient_vs_Nodule")
    title_block(ws, "Table 3 — Patient-level vs nodule-level ROM by TI-RADS category",
                "ACR expected bands per ACR TI-RADS 2017. Inflation = patient_rom - nodule_rom (percentage points).")
    write_dataframe(ws, df_t3, start_row=4)

    # ---------------- Table 4 — Bethesda × TR ----------------
    ws = wb.create_sheet("Table_4_Bethesda_x_TR")
    title_block(ws, "Table 4 — Bethesda × TI-RADS (strict-eligible nodules)",
                "n per cell. 'missing' = nodule has no FNA linkage in this version.")
    write_dataframe(ws, df_btr_pivot, start_row=4)

    # ---------------- Table 5 — Histology ----------------
    print(" Table 5 ...")
    df_t5 = con.execute("""
        SELECT histology_final,
          COUNT(*) AS n,
          ROUND(100.0*COUNT(*)/SUM(COUNT(*)) OVER (),2) AS pct,
          COUNT_IF(is_malignant) AS n_malignant
        FROM manuscript_workspace.m025_analytic_master_patient_v1
        GROUP BY histology_final ORDER BY n DESC
    """).df()
    ws = wb.create_sheet("Table_5_Histology")
    title_block(ws, "Table 5 — Histology distribution (patient cohort)",
                "NULL = no operative pathology available.")
    write_dataframe(ws, df_t5, start_row=4)

    # ---------------- Table 6 — FNA-path concordance ----------------
    print(" Table 6 ...")
    df_t6 = con.execute("""
        SELECT fna_path_concordance_category,
          COUNT(*) AS n,
          ROUND(100.0*COUNT(*)/SUM(COUNT(*)) OVER (),2) AS pct,
          COUNT_IF(is_malignant) AS n_malignant,
          ROUND(100.0*COUNT_IF(is_malignant)/COUNT(*),2) AS rom_pct
        FROM manuscript_workspace.m025_analytic_master_patient_v1
        GROUP BY fna_path_concordance_category ORDER BY n DESC
    """).df()
    ws = wb.create_sheet("Table_6_FNA_Path_Concordance")
    title_block(ws, "Table 6 — FNA cytology vs final pathology concordance categories",
                "Categories from canonical FNA → path bridge.")
    write_dataframe(ws, df_t6, start_row=4)

    # ---------------- Table 7 — Race / era ----------------
    print(" Table 7 ...")
    df_t7a = con.execute("""
        SELECT race, COUNT(*) AS n, ROUND(100.0*COUNT(*)/SUM(COUNT(*)) OVER (),2) AS pct,
               COUNT_IF(is_malignant) AS n_malignant
        FROM manuscript_workspace.m025_analytic_master_patient_v1
        GROUP BY race ORDER BY n DESC
    """).df()
    df_t7b = con.execute("""
        SELECT
          CASE
            WHEN surg_year < 2010 THEN '_pre_2010'
            WHEN surg_year < 2015 THEN 'a_2010-2014'
            WHEN surg_year < 2020 THEN 'b_2015-2019'
            WHEN surg_year < 2025 THEN 'c_2020-2024'
            ELSE 'd_2025+' END AS era,
          COUNT(*) AS n,
          COUNT_IF(is_malignant) AS n_malignant,
          ROUND(100.0*COUNT_IF(is_malignant)/COUNT(*),2) AS rom_pct
        FROM manuscript_workspace.m025_analytic_master_patient_v1
        GROUP BY era ORDER BY era
    """).df()
    ws = wb.create_sheet("Table_7_Race_and_Era")
    next_row = title_block(ws, "Table 7 — Race and surgery-era distribution",
                            "Race per CPM canonical. Era = surgery year buckets.")
    ws.cell(row=next_row, column=1, value="A) Race").font = Font(bold=True, size=12)
    write_dataframe(ws, df_t7a, start_row=next_row + 1)
    next_row2 = next_row + 2 + len(df_t7a) + 2
    ws.cell(row=next_row2, column=1, value="B) Surgery era").font = Font(bold=True, size=12)
    write_dataframe(ws, df_t7b, start_row=next_row2 + 1)

    # ---------------- Subgroup analyses ----------------
    print(" Subgroups ...")
    df_subg_age = con.execute("""
        SELECT
          CASE
            WHEN age_at_surgery < 40 THEN 'a_<40'
            WHEN age_at_surgery < 55 THEN 'b_40-54'
            WHEN age_at_surgery < 70 THEN 'c_55-69'
            ELSE 'd_>=70' END AS age_bin,
          max_tirads_category_ever AS max_TR,
          COUNT(*) AS n,
          COUNT_IF(is_malignant) AS n_malignant,
          ROUND(100.0*COUNT_IF(is_malignant)/COUNT(*),2) AS rom_pct
        FROM manuscript_workspace.m025_analytic_master_patient_v1
        GROUP BY age_bin, max_tirads_category_ever
        ORDER BY age_bin, max_tirads_category_ever
    """).df()
    ws = wb.create_sheet("Subgroup_Age_x_TR")
    title_block(ws, "Subgroup — age × max TR (patient grain)")
    write_dataframe(ws, df_subg_age, start_row=4)

    df_subg_sex = con.execute("""
        SELECT sex, max_tirads_category_ever AS max_TR, COUNT(*) AS n,
               COUNT_IF(is_malignant) AS n_malignant,
               ROUND(100.0*COUNT_IF(is_malignant)/COUNT(*),2) AS rom_pct
        FROM manuscript_workspace.m025_analytic_master_patient_v1
        GROUP BY sex, max_tirads_category_ever
        ORDER BY sex, max_tirads_category_ever
    """).df()
    ws = wb.create_sheet("Subgroup_Sex_x_TR")
    title_block(ws, "Subgroup — sex × max TR (patient grain)")
    write_dataframe(ws, df_subg_sex, start_row=4)

    df_subg_hist = con.execute("""
        SELECT histology_category AS hist_cat, max_tirads_category_ever AS max_TR, COUNT(*) AS n,
               COUNT_IF(is_malignant) AS n_malignant,
               ROUND(100.0*COUNT_IF(is_malignant)/COUNT(*),2) AS rom_pct
        FROM manuscript_workspace.m025_analytic_master_patient_v1
        GROUP BY histology_category, max_tirads_category_ever
        ORDER BY histology_category, max_tirads_category_ever
    """).df()
    ws = wb.create_sheet("Subgroup_Histology_x_TR")
    title_block(ws, "Subgroup — histology category × max TR (patient grain)")
    write_dataframe(ws, df_subg_hist, start_row=4)

    df_multi = con.execute("""
        SELECT
          CASE WHEN n_us_exams = 1 THEN '1_exam'
               WHEN n_us_exams BETWEEN 2 AND 3 THEN '2-3_exams'
               WHEN n_us_exams >= 4 THEN '>=4_exams' ELSE 'unknown' END AS exam_cat,
          COUNT(*) AS n,
          COUNT_IF(is_malignant) AS n_malignant,
          ROUND(100.0*COUNT_IF(is_malignant)/COUNT(*),2) AS rom_pct
        FROM manuscript_workspace.m025_analytic_master_patient_v1
        GROUP BY exam_cat ORDER BY exam_cat
    """).df()
    ws = wb.create_sheet("Subgroup_NumExams")
    title_block(ws, "Subgroup — number of US exams per patient",
                "Multi-exam patients drive part of the patient-level vs nodule-level ROM inflation.")
    write_dataframe(ws, df_multi, start_row=4)

    # ---------------- Sensitivity arm summary ----------------
    sens_rows = [
        ("Primary", "Patient cohort n=3,375; AUC 0.6478 [0.6301-0.6665]; Youden TR>=TR4 (J=0.271)"),
        ("Sister nodule analysis", "Strict ACR n=3,687; AUC 0.6399; TR4 ROM 18.7% [16.3-21.5]; TR5 26.1% [23.7-28.6]"),
        ("S1A relaxed nodule cohort (15,309)", "TR4 ROM ~23.0%; PPV 23.1% Sens 44.6% Spec 78.0%"),
        ("S1B first-US-only", "Identical to primary (nodule_master_id already deduplicates)"),
        ("S1C single-nodule patients (n=782)", "TR4 ROM 30.7%; TR5 ROM 34.9% (selection effect)"),
        ("S1D unilateral path-only", "TR4 ROM 8.5%; TR5 ROM 10.7% (conservative bilateral exclusion)"),
        ("S2 era split (patient grain)", "Pre-2017 n=422 ROM 40.0%; Post-2017 n=2,953 ROM 44.4%; per-TR ROM directionally similar (see Sensitivity_Era_Patient sheet)"),
        ("S2 era split (nodule strict)", "Pre-2017 n=381 ROM 25.2%; Post-2017 n=3,306 ROM 16.2%; post-2017-only TR4 18.0% / TR5 24.4% (both in ACR bands)"),
        ("S3 tighter US-to-surgery window (nodule strict)", "180d: TR4 15.7% / TR5 22.2% (still in ACR bands); 90d: TR4 11.3% / TR5 16.8%; 30d: tight bound"),
        ("Provenance check (strict subset)", "99.3% inm_v1 structured source (Script 246); 0.7% LLM-augmented; institutional independent verification of feature extractions"),
        ("ACR FNA compliance flagged", "1,553 unnecessary FNAs by ACR criteria; 472 cancers below ACR FNA threshold"),
    ]
    sens_df = pd.DataFrame(sens_rows, columns=["arm", "result"])
    ws = wb.create_sheet("Sensitivity_Arms")
    title_block(ws, "Sensitivity / supplementary analyses",
                "mig_307c sensitivity tables; mig_307d adds Wilson 95% CIs, diagnostics CSVs, and figures under 06_figures_sensitivity/.")
    write_dataframe(ws, sens_df, start_row=4)

    # ---------------- Era subset (patient grain) ----------------
    print(" Sensitivity — era × patient ...")
    df_era_p = con.execute("""
        SELECT * FROM manuscript_workspace.m025_sens_era_patient_v1
        ORDER BY era, tr_category NULLS LAST
    """).df()
    df_era_p = add_rom_ci_lo_hi(df_era_p)
    ws = wb.create_sheet("Sensitivity_Era_Patient")
    title_block(ws, "Sensitivity S2 — Patient cohort split by ACR-2017 era",
                "Era boundary = surg_first_date < 2017-05-01 (ACR TI-RADS 2017 publication date). "
                "lo_95 / hi_95 = Wilson 95% CI for ROM (n_malignant / n_total).")
    write_dataframe(ws, df_era_p, start_row=4)

    # ---------------- Era subset (nodule strict) ----------------
    print(" Sensitivity — era × nodule ...")
    df_era_n = con.execute("""
        SELECT * FROM manuscript_workspace.m025_sens_era_nodule_v1
        ORDER BY era, tr_category NULLS LAST
    """).df()
    df_era_n = add_rom_ci_lo_hi(df_era_n)
    ws = wb.create_sheet("Sensitivity_Era_Nodule")
    title_block(ws, "Sensitivity S2 — Nodule strict cohort split by ACR-2017 era",
                "Era boundary = exam_date < 2017-05-01. lo_95 / hi_95 = Wilson CI for ROM.")
    write_dataframe(ws, df_era_n, start_row=4)

    # ---------------- Time-window subset (nodule strict) ----------------
    print(" Sensitivity — time window × nodule ...")
    df_win = con.execute("""
        SELECT * FROM manuscript_workspace.m025_sens_window_nodule_v1
        ORDER BY tr_category
    """).df()
    df_win = augment_window_table(df_win)
    ws = wb.create_sheet("Sensitivity_Match_Window")
    title_block(ws, "Sensitivity S3 — Tighter US-to-surgery match-window (nodule strict)",
                "ROM at 365 / 180 / 90 / 30 d: Wilson 95% CIs in rom_w*_lo_95 / rom_w*_hi_95 vs n_total.")
    write_dataframe(ws, df_win, start_row=4)

    # ---------------- Sensitivity publication bundle (mig_307d) ----------------
    print(" Sensitivity — CSVs + figures (06_figures_sensitivity) ...")
    df_diag = fetch_diagnostic_by_era(con)
    df_auc = fetch_per_era_auc(con)
    export_sensitivity_csv_bundle(SENS_FIG_DIR, df_era_p, df_era_n, df_win, df_diag, df_auc)
    render_forest_rom_by_era(
        df_era_p,
        df_era_n,
        os.path.join(SENS_FIG_DIR, "M025_fig_sens_forest_tr4_tr5_rom_by_era.png"),
        dpi=300,
    )
    render_linechart_match_window(
        df_win,
        os.path.join(SENS_FIG_DIR, "M025_fig_sens_rom_by_match_window.png"),
        dpi=300,
    )

    # ---------------- QA gates ----------------
    qa_df = pd.DataFrame([
        ("patient_cohort_n", snapshot["patient_cohort_n"], 3375, snapshot["patient_cohort_n"] == 3375),
        ("patient_malignant_n", snapshot["patient_cohort_n_malignant"], 1479, snapshot["patient_cohort_n_malignant"] == 1479),
        ("nodule_total_n", snapshot["nodule_total_rows"], 37438, snapshot["nodule_total_rows"] == 37438),
        ("nodule_strict_n", snapshot["nodule_strict_eligible"], 3687, snapshot["nodule_strict_eligible"] == 3687),
        ("nodule_strict_path_malig", snapshot["nodule_strict_path_malignant"], 631, snapshot["nodule_strict_path_malignant"] == 631),
        ("threshold_metrics rows", len(df_thr), 6, len(df_thr) == 6),
        ("rom_by_tr rows", len(df_rom), 9, len(df_rom) == 9),
        ("bethesda_x_tr rows", len(df_btr), 25, len(df_btr) == 25),
    ], columns=["check", "observed", "expected", "pass"])
    ws = wb.create_sheet("QA_Gates")
    title_block(ws, "QA gates")
    write_dataframe(ws, qa_df, start_row=4)

    wb.save(TABLES_PATH)
    print(f"  → {TABLES_PATH}")

    print("\nDone.")
    print(f"  Master:  {MASTER_PATH}")
    print(f"  Tables:  {TABLES_PATH}")


if __name__ == "__main__":
    main()
