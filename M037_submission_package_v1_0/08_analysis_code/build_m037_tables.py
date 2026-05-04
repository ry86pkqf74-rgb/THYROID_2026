#!/usr/bin/env python3
"""
build_m037_tables.py
====================
Builds 04_tables.xlsx + 05_master_data.xlsx for M037 submission package v1.0.

Run from repo root:
    .venv/bin/python M037_submission_package_v1_0/08_analysis_code/build_m037_tables.py
"""
from __future__ import annotations

import json
import os
import sys
from datetime import datetime, timezone

import duckdb
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from scipy import stats
import statsmodels.api as sm

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
from motherduck_client import get_token

PKG_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
ALT_FILL = PatternFill("solid", fgColor="DEEAF1")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
BODY_FONT = Font(size=10, name="Calibri")
THIN = Side(style="thin", color="AAAAAA")


def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 50)


def write_df_to_sheet(ws, df: pd.DataFrame, title: str, sub: str = ""):
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=12, name="Calibri", color="1F4E79")
    if sub:
        ws.append([sub])
        ws["A2"].font = Font(italic=True, size=9, name="Calibri", color="444444")
        ws.append([])
        hdr_row = 4
    else:
        ws.append([])
        hdr_row = 3
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=hdr_row, column=ci, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border()
    for ri, row in enumerate(df.itertuples(index=False), hdr_row + 1):
        fill = ALT_FILL if ri % 2 == 0 else WHITE_FILL
        for ci, val in enumerate(row, 1):
            if val is None or (hasattr(val, "__class__") and val.__class__.__name__ in ("NAType", "NaT")):
                val = ""
            try:
                if pd.isna(val):
                    val = ""
            except (TypeError, ValueError):
                pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = thin_border()
    _autofit(ws)


def connect():
    tok = get_token()
    return duckdb.connect(f"md:thyroid_canonical_publication_v1_0?motherduck_token={tok}")


def pull_spine(con) -> pd.DataFrame:
    nb = """
    CASE
      WHEN UPPER(COALESCE(c.ajcc8_n_stage, '')) IN ('N0') THEN 'N0'
      WHEN UPPER(COALESCE(c.ajcc8_n_stage, '')) LIKE 'N1A%' THEN 'N1a'
      WHEN UPPER(COALESCE(c.ajcc8_n_stage, '')) LIKE 'N1B%' THEN 'N1b'
      WHEN UPPER(COALESCE(c.ajcc8_n_stage, '')) LIKE 'NX%' OR c.ajcc8_n_stage IS NULL THEN 'Nx'
      WHEN UPPER(COALESCE(c.ajcc8_n_stage, '')) LIKE 'N1%' THEN 'N1_other'
      ELSE 'Other'
    END
    """
    sql = f"""
    SELECT
      c.*,
      ({nb}) AS n_bucket,
      p.ln_status_source,
      p.ajcc8_t_stage,
      p.race
    FROM manuscript_workspace.cohort_m037_ln_metastasis_v1 c
    INNER JOIN main.canonical_patient_master p
      ON CAST(c.research_id AS VARCHAR) = CAST(p.research_id AS VARCHAR)
    """
    return con.execute(sql).fetchdf()


def add_derived(df: pd.DataFrame) -> pd.DataFrame:
    d = df.copy()
    stage = d["ajcc8_n_stage"].fillna("").astype(str).str.upper()
    d["y_ln_pos"] = stage.str.startswith("N1").astype(int)
    d["sex_male"] = d["sex"].astype(str).str.lower().isin(("male", "m")).astype(int)
    d["fhx"] = d["pmhx_nlp_family_hx_thyroid"].apply(lambda x: x is True).astype(int)
    d["tumor_size_cm"] = pd.to_numeric(d["tumor_size_cm"], errors="coerce")
    d["age_at_surgery"] = pd.to_numeric(d["age_at_surgery"], errors="coerce")
    ts = d["ajcc8_t_stage"].astype(str).str.upper()
    d["t_high"] = ts.isin(["T3A", "T3B", "T4A", "T4B"]).astype(int)
    ete = d["ete_grade_final"].astype(str).str.lower()
    d["ete_any"] = ete.isin(["microscopic", "gross", "present_ungraded"]).astype(int)
    hist = d["histology_final"].astype(str)
    d["hist_non_ptc"] = (~hist.str.contains("PTC|PAPILLARY", case=False, regex=True)).astype(int)
    d["braf_pos"] = d["braf_positive_final"].apply(lambda x: x is True or str(x).lower() in ("true", "t", "1")).astype(int)
    d["surg_total"] = d["surg_procedure_type"].astype(str).str.lower().eq("total_thyroidectomy").astype(int)
    d["ln_total_examined"] = pd.to_numeric(d["ln_total_examined"], errors="coerce")
    return d


def fit_primary_model(df: pd.DataFrame):
    sub = df[["age_at_surgery", "tumor_size_cm", "sex_male", "fhx", "y_ln_pos"]].dropna().copy()
    for c in ("age_at_surgery", "tumor_size_cm", "sex_male", "fhx", "y_ln_pos"):
        sub[c] = pd.to_numeric(sub[c], errors="coerce")
    sub = sub.dropna()
    X = sm.add_constant(sub[["sex_male", "fhx", "age_at_surgery", "tumor_size_cm"]].astype(float))
    y = sub["y_ln_pos"].astype(float)
    m = sm.Logit(y, X).fit(disp=False, maxiter=200)
    rows = []
    for name in m.params.index:
        coef = float(m.params[name])
        or_ = float(np.exp(coef))
        ci_lo = float(np.exp(coef - 1.96 * float(m.bse[name])))
        ci_hi = float(np.exp(coef + 1.96 * float(m.bse[name])))
        p = float(m.pvalues[name])
        rows.append(
            {
                "term": name,
                "OR": round(or_, 4),
                "CI_low": round(ci_lo, 4),
                "CI_high": round(ci_hi, 4),
                "p_value": p,
            }
        )
    return pd.DataFrame(rows), int(len(sub)), float(m.prsquared)


def fit_extended_model(df: pd.DataFrame):
    cols = [
        "age_at_surgery",
        "tumor_size_cm",
        "sex_male",
        "hist_non_ptc",
        "t_high",
        "ete_any",
        "braf_pos",
        "surg_total",
        "ln_total_examined",
        "y_ln_pos",
    ]
    sub = df[cols].dropna().copy()
    for c in cols:
        sub[c] = pd.to_numeric(sub[c], errors="coerce")
    sub = sub.dropna()
    pred = [
        "age_at_surgery",
        "tumor_size_cm",
        "sex_male",
        "hist_non_ptc",
        "t_high",
        "ete_any",
        "braf_pos",
        "surg_total",
        "ln_total_examined",
    ]
    X = sm.add_constant(sub[pred].astype(float))
    y = sub["y_ln_pos"].astype(float)
    m = sm.Logit(y, X).fit(disp=False, maxiter=300)
    rows = []
    for name in m.params.index:
        coef = float(m.params[name])
        or_ = float(np.exp(coef))
        ci_lo = float(np.exp(coef - 1.96 * float(m.bse[name])))
        ci_hi = float(np.exp(coef + 1.96 * float(m.bse[name])))
        p = float(m.pvalues[name])
        rows.append({"term": name, "OR": round(or_, 4), "CI_low": round(ci_lo, 4), "CI_high": round(ci_hi, 4), "p_value": p})
    return pd.DataFrame(rows), int(len(sub)), float(m.prsquared)


def table1_demo_by_n_bucket(df: pd.DataFrame) -> pd.DataFrame:
    buckets = ["N0", "N1a", "N1b", "Nx", "N1_other", "Other"]
    rows = []
    for b in buckets:
        sub = df[df["n_bucket"] == b]
        if sub.empty:
            continue
        rows.append(
            {
                "n_bucket": b,
                "n": len(sub),
                "age_mean": round(float(sub["age_at_surgery"].mean()), 1) if sub["age_at_surgery"].notna().any() else None,
                "age_sd": round(float(sub["age_at_surgery"].std()), 1) if sub["age_at_surgery"].notna().sum() > 1 else None,
                "pct_male": round(100.0 * float(sub["sex_male"].mean()), 1),
                "tumor_cm_median": round(float(sub["tumor_size_cm"].median()), 2) if sub["tumor_size_cm"].notna().any() else None,
                "pct_fhx_nlp": round(100.0 * float(sub["fhx"].mean()), 1),
                "pct_syndrome": round(
                    100.0 * float(sub["family_syndrome_flag"].apply(lambda x: x is True).mean()),
                    1,
                ),
                "ln_examined_median": round(float(sub["ln_total_examined"].median()), 1)
                if sub["ln_total_examined"].notna().any()
                else None,
            }
        )
    return pd.DataFrame(rows)


def table5_hist_ln_crosstab(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["hist_grp"] = np.where(
        df["histology_final"].astype(str).str.contains("(?i)papillary|ptc"),
        "PTC",
        np.where(df["histology_final"].astype(str).str.contains("(?i)follicular|ftc"), "FTC", "Other"),
    )
    ct = pd.crosstab(df["hist_grp"], df["n_bucket"], margins=True)
    return ct.reset_index().rename(columns={"hist_grp": "histology_group"})


def main():
    con = connect()
    raw = pull_spine(con)
    con.close()
    df = add_derived(raw)

    out_dir = os.path.join(PKG_DIR, "08_analysis_outputs")
    os.makedirs(out_dir, exist_ok=True)
    snap = {
        "generated_utc": datetime.now(timezone.utc).isoformat(),
        "n_cohort": int(len(df)),
        "n_ln_pos": int(df["y_ln_pos"].sum()),
        "pct_ln_pos": round(100.0 * float(df["y_ln_pos"].mean()), 2),
    }
    with open(os.path.join(out_dir, "m037_run_snapshot.json"), "w", encoding="utf-8") as f:
        json.dump(snap, f, indent=2)
    df.to_parquet(os.path.join(out_dir, "m037_analytic_spine.parquet"), index=False)

    tab1 = table1_demo_by_n_bucket(df)
    groups = [
        df.loc[df["n_bucket"] == b, "age_at_surgery"].dropna()
        for b in ("N0", "N1a", "N1b", "Nx")
        if (df["n_bucket"] == b).any() and df.loc[df["n_bucket"] == b, "age_at_surgery"].notna().any()
    ]
    if len(groups) >= 2:
        _, p_age = stats.kruskal(*groups)
        tab1_note = f"Kruskal-Wallis age p={p_age:.4g} across N0/N1a/N1b/Nx with data."
    else:
        tab1_note = "Kruskal-Wallis age: skipped (insufficient strata)."

    prim, n_cc, pr2 = fit_primary_model(df)
    prim.insert(0, "model", "primary_4var_family_hx")
    prim_meta = pd.DataFrame(
        [{"metric": "complete_case_n", "value": n_cc}, {"metric": "pseudo_r2_mcfadden", "value": round(pr2, 4)}]
    )

    ext, n_ext, pr2e = fit_extended_model(df)
    ext.insert(0, "model", "extended_9var")
    ext_meta = pd.DataFrame(
        [{"metric": "complete_case_n", "value": n_ext}, {"metric": "pseudo_r2_mcfadden", "value": round(pr2e, 4)}]
    )

    df_sens = df[df["ln_status_source"].isna() | (df["ln_status_source"].astype(str) != "staging")]
    prim_s, n_s, pr2s = fit_primary_model(df_sens)
    prim_s.insert(0, "model", "primary_4var_sensitivity_excl_staging_only_source")

    strat_rows = []
    for label, mask in [("female", df["sex_male"] == 0), ("male", df["sex_male"] == 1)]:
        subm = df.loc[mask]
        try:
            res, n_, _pr = fit_primary_model(subm)
            tr = res.set_index("term").loc["tumor_size_cm"]
            strat_rows.append(
                {
                    "stratum": label,
                    "n_complete_primary_model": n_,
                    "OR_tumor_per_cm": tr["OR"],
                    "CI_low": tr["CI_low"],
                    "CI_high": tr["CI_high"],
                    "p_tumor": tr["p_value"],
                }
            )
        except Exception as exc:
            strat_rows.append({"stratum": label, "n_complete_primary_model": len(subm), "error": str(exc)})
    tab3 = pd.DataFrame(strat_rows)

    tab5 = table5_hist_ln_crosstab(df)

    n1 = df[df["n_bucket"].isin(["N1a", "N1b"])]
    supp1 = (
        n1.groupby("n_bucket", observed=True)
        .agg(
            n=("research_id", "count"),
            age_mean=("age_at_surgery", "mean"),
            tumor_median=("tumor_size_cm", "median"),
            pct_male=("sex_male", "mean"),
            pct_fhx=("fhx", "mean"),
        )
        .reset_index()
    )
    for c in ("age_mean", "tumor_median", "pct_male", "pct_fhx"):
        if c in supp1.columns:
            supp1[c] = supp1[c].round(4)

    df_sy = df.copy()
    df_sy["syndrome_arm"] = np.where(df_sy["family_syndrome_flag"].apply(lambda x: x is True), "syndrome", "no_syndrome")
    supp2 = df_sy.groupby("syndrome_arm", observed=True).agg(n=("research_id", "count"), ln_pos_rate=("y_ln_pos", "mean")).reset_index()
    supp2["ln_pos_rate_pct"] = (supp2["ln_pos_rate"] * 100).round(2)
    supp2.drop(columns=["ln_pos_rate"], inplace=True)

    cover = pd.DataFrame(
        [
            ("Package", "M037 v1.0 — LN metastasis predictors"),
            ("Database", "thyroid_canonical_publication_v1_0"),
            ("Cohort view", "manuscript_workspace.cohort_m037_ln_metastasis_v1"),
            ("Build UTC", snap["generated_utc"]),
            ("Cohort n", snap["n_cohort"]),
            ("LN pos (N1+) n / %", f'{snap["n_ln_pos"]} / {snap["pct_ln_pos"]}%'),
            ("Table1 foot", tab1_note),
            ("Primary model complete n", n_cc),
            ("Cowork target complete-case n", 2147),
        ],
        columns=["key", "value"],
    )

    qa = pd.DataFrame(
        [
            {"check": "cohort_non_empty", "status": "PASS" if snap["n_cohort"] > 0 else "FAIL"},
            {"check": "primary_complete_n_approx_cowork", "status": "PASS" if 1800 <= n_cc <= 2400 else "REVIEW"},
            {"check": "parquet_spine_written", "status": "PASS"},
        ]
    )

    wb = Workbook()
    wb.remove(wb.active)
    ws0 = wb.create_sheet("Cover")
    write_df_to_sheet(ws0, cover, "M037 — Tables workbook cover", "")

    ws = wb.create_sheet("Table1_demographics_by_N")
    write_df_to_sheet(ws, tab1, "Table 1 — Demographics by AJCC N bucket", tab1_note)

    ws = wb.create_sheet("Table2a_primary_meta")
    write_df_to_sheet(ws, prim_meta, "Primary model fit summary", "")
    ws = wb.create_sheet("Table2b_primary_coef")
    write_df_to_sheet(
        ws,
        prim,
        "Table 2 — Primary logistic regression",
        "Outcome: N1+ vs N0/Nx. Predictors: male sex, NLP family hx thyroid (TRUE only), age, tumor size (cm).",
    )

    ws = wb.create_sheet("Table2c_extended_meta")
    write_df_to_sheet(ws, ext_meta, "Extended model fit summary", "")
    ws = wb.create_sheet("Table2d_extended_coef")
    write_df_to_sheet(ws, ext, "Extended multivariable (publication-style covariates)", "")

    ws = wb.create_sheet("Table3_stratified_tumor_OR")
    write_df_to_sheet(ws, tab3, "Table 3 — Tumor size OR (primary 4-var) by sex", "")

    ws = wb.create_sheet("Table4_sensitivity_ln_source")
    write_df_to_sheet(ws, prim_s, f"Sensitivity: exclude ln_status_source=staging only; n={len(df_sens)}; complete-case n={n_s}", "")

    ws = wb.create_sheet("Table5_histology_crosstab")
    write_df_to_sheet(ws, tab5, "Histology group × N bucket", "")

    ws = wb.create_sheet("SuppS1_N1a_vs_N1b")
    write_df_to_sheet(ws, supp1, "Supplement S1 — N1a vs N1b", "")

    ws = wb.create_sheet("SuppS2_syndrome_flag")
    write_df_to_sheet(ws, supp2, "Supplement S2 — family_syndrome_flag prevalence + LN+ rate", "")

    ws = wb.create_sheet("QA")
    write_df_to_sheet(ws, qa, "QA checks", "")

    out_xlsx = os.path.join(PKG_DIR, "04_tables.xlsx")
    wb.save(out_xlsx)
    print(f"Wrote {out_xlsx}")

    master_path = os.path.join(PKG_DIR, "05_master_data.xlsx")
    with pd.ExcelWriter(master_path, engine="openpyxl") as xw:
        df.to_excel(xw, sheet_name="analytic_spine", index=False)
        pd.DataFrame([snap]).to_excel(xw, sheet_name="run_snapshot", index=False)
    print(f"Wrote {master_path}")


if __name__ == "__main__":
    main()
