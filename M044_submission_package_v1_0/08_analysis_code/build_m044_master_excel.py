#!/usr/bin/env python3
"""
M044 ETE manuscript - Per-Patient Master Data Workbook builder.
Pulls live from MotherDuck via duckdb (.eras account); no local-write COPY needed.
"""
from __future__ import annotations
import os, sys
from datetime import datetime
from pathlib import Path
import duckdb, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

DB = "thyroid_canonical_publication_v1_0"
OUT = Path("/sessions/wonderful-trusting-babbage/mnt/THyroid 2026/M044_ETE_master_data.xlsx")
OUT.parent.mkdir(parents=True, exist_ok=True)

HF = PatternFill("solid", fgColor="1F4E78"); HFONT = Font(bold=True, color="FFFFFF", size=11)
TF = Font(bold=True, size=14, color="1F4E78"); SF = Font(bold=True, size=11, color="404040")
BD = Border(left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"))

ANALYTIC_SQL = open("/tmp/analytic.sql").read()
SRC_COHORT_SQL = "SELECT * FROM manuscript_workspace.cohort_m044_ajcc_ete_v1 ORDER BY research_id"
SRC_CPM_SQL = open("/tmp/cpm.sql").read()
SRC_REC_SQL = """SELECT r.* FROM main.canonical_recurrence_resolved_v1 r
                 JOIN manuscript_workspace.cohort_m044_ajcc_ete_v1 c USING (research_id) ORDER BY r.research_id"""
SRC_LN_SQL = """SELECT research_id,
  MAX(ln_total_examined) AS ln_examined, MAX(ln_total_positive) AS ln_positive,
  MAX(ln_central_examined) AS ln_central_examined, MAX(ln_central_positive) AS ln_central_positive,
  MAX(ln_lateral_left_positive) AS ln_lateral_left_positive,
  MAX(ln_lateral_right_positive) AS ln_lateral_right_positive,
  MAX(ln_bilateral_lateral_positive) AS ln_bilateral_lateral_positive,
  MAX(ln_level_vi_positive) AS ln_level_vi_positive,
  MAX(ln_level_vii_positive) AS ln_level_vii_positive,
  MAX(ln_extranodal_extension) AS ln_ene
FROM manuscript_workspace.ln_master_rollup_v1
WHERE research_id IN (SELECT research_id FROM manuscript_workspace.cohort_m044_ajcc_ete_v1)
GROUP BY research_id ORDER BY research_id"""
SRC_REOP_SQL = """SELECT research_id,
  MAX(n_surgeries) AS n_surgeries, MAX(second_surgery_date) AS second_surgery_date,
  MAX(days_between_first_second_surgery) AS days_to_2nd,
  MAX(completion_reason) AS completion_reason,
  MAX(completion_reason_confidence) AS completion_reason_confidence,
  MAX(completion_histology_type) AS completion_histology_type,
  MAX(op_reoperative_any) AS op_reoperative_any
FROM manuscript_workspace.cohort_m040_reoperative_v1
WHERE research_id IN (SELECT research_id FROM manuscript_workspace.cohort_m044_ajcc_ete_v1)
GROUP BY research_id ORDER BY research_id"""
SRC_PATH_SQL = """SELECT ps.research_id,
  ps.tumor_1_lymphatic_invasion, ps.tumor_1_angioinvasion,
  ps.tumor_1_angioinvasion_quantify, ps.tumor_1_extrathyroidal_extension
FROM main.path_synoptics ps
JOIN manuscript_workspace.cohort_m044_ajcc_ete_v1 c USING (research_id)
ORDER BY ps.research_id"""

# Crosswalk + dictionary - load from JSON
import json
CROSSWALK_ROWS = json.load(open("/tmp/crosswalk.json"))
DICTIONARY_ROWS = json.load(open("/tmp/dictionary.json"))

def style_hdr(ws, n):
    for c in range(1, n+1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HF; cell.font = HFONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BD

def autosize(ws, df):
    for i, c in enumerate(df.columns, start=1):
        try:
            sample = df[c].astype(str).head(200)
            mx = max([len(str(c))] + sample.map(len).tolist())
        except Exception:
            mx = max(len(str(c)), 18)
        ws.column_dimensions[get_column_letter(i)].width = min(max(mx + 2, 12), 48)

def _safe(v):
    """Coerce pandas NA/NaT/numpy NaN to None for openpyxl."""
    try:
        if v is None: return None
        if isinstance(v, (str, int, float, bool, bytes)):
            if isinstance(v, float):
                import math
                if math.isnan(v): return None
            return v
        # pandas Timestamp / datetime
        try:
            from datetime import datetime, date
            if isinstance(v, (datetime, date)): return v
        except Exception: pass
        # pandas NA / NaT
        try:
            if pd.isna(v): return None
        except Exception: pass
        return str(v)
    except Exception:
        return str(v)

def write_df(ws, df, start_row=1):
    # Pre-clean df: replace NaN/NaT with None
    df = df.copy()
    for c in df.columns:
        if df[c].dtype.kind == "O":
            df[c] = df[c].where(df[c].notna(), None)
        elif "datetime" in str(df[c].dtype):
            try:
                df[c] = df[c].dt.tz_localize(None)
            except (TypeError, AttributeError):
                pass
            df[c] = df[c].where(df[c].notna(), None)
    # Header
    for c_idx, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=c_idx, value=str(col))
    # Body
    for r_idx, (_, row) in enumerate(df.iterrows(), start=start_row+1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_safe(value))
    style_hdr(ws, len(df.columns))
    autosize(ws, df)
    ws.freeze_panes = ws.cell(row=start_row+1, column=2)

def main():
    tok = os.environ['MDT']
    print(f"-> Connect MD ({DB})", flush=True)
    con = duckdb.connect(f"md:{DB}?motherduck_token={tok}")

    print("-> Tab 2 analytic ...", flush=True)
    analytic = con.execute(ANALYTIC_SQL).fetchdf()
    print(f"   shape={analytic.shape}")
    assert len(analytic)==4128

    print("-> Tab 3 cohort raw ...", flush=True)
    src_cohort = con.execute(SRC_COHORT_SQL).fetchdf()
    print(f"   shape={src_cohort.shape}")

    print("-> Tab 4 CPM selected ...", flush=True)
    src_cpm = con.execute(SRC_CPM_SQL).fetchdf()
    print(f"   shape={src_cpm.shape}")

    print("-> Tab 5 recurrence ...", flush=True)
    src_rec = con.execute(SRC_REC_SQL).fetchdf()
    print(f"   shape={src_rec.shape}")

    print("-> Tab 6 LN agg ...", flush=True)
    src_ln = con.execute(SRC_LN_SQL).fetchdf()
    print(f"   shape={src_ln.shape}")

    print("-> Tab 7 reop agg ...", flush=True)
    src_reop = con.execute(SRC_REOP_SQL).fetchdf()
    print(f"   shape={src_reop.shape}")

    print("-> Tab 8 path LVI ...", flush=True)
    src_path = con.execute(SRC_PATH_SQL).fetchdf()
    print(f"   shape={src_path.shape}")

    # Save analytic to parquet for refits and reuse
    pq_path = Path("/sessions/wonderful-trusting-babbage/mnt/outputs/m044_workspace/data/m044_analytic_v2.parquet")
    pq_path.parent.mkdir(parents=True, exist_ok=True)
    analytic.to_parquet(pq_path, compression="zstd")
    print(f"-> Saved analytic parquet -> {pq_path}")

    # Build QA flags
    print("-> QA flags ...", flush=True)
    qa_rows = []
    for _, r in analytic.iterrows():
        f = []
        if pd.isna(r.get("surg_first_date")): f.append("surg_first_date_missing")
        if pd.isna(r.get("bmi_combined")): f.append("bmi_missing")
        if pd.isna(r.get("pmhx_nlp_smoking_status")): f.append("smoking_status_null")
        if r.get("followup_years",1)==0: f.append("zero_followup")
        if (r.get("any_recurrence_flag") is True) and (r.get("recurrence_status_final")=="none"):
            f.append("legacy_anyrec_TRUE_canonical_none")
        if (r.get("structural_recurrence_flag") is True) and (r.get("recurrence_status_final")=="none"):
            f.append("legacy_struc_TRUE_canonical_none")
        sd = r.get("surg_first_date")
        try:
            if sd is not None and pd.notna(sd) and pd.Timestamp(sd) < pd.Timestamp("1999-01-01"):
                f.append("surgery_pre_1999_outlier")
        except Exception: pass
        if r.get("ete_grade_final") == "true": f.append("ete_grade_final_ambiguous_true")
        if isinstance(r.get("lvi_grade"), str) and r.get("lvi_grade") in ("preesent","extensivre","extensiver","indetermiante","indeeterminate"):
            f.append("lvi_grade_spelling_variant")
        if (r.get("recurrence_path_proven") is True) and (pd.notna(r.get("n_surgeries"))) and (int(r.get("n_surgeries") or 0)>=2):
            if (r.get("days_to_2nd") or 0) >= 365:
                f.append("recur_via_completion_pathway_long_interval")
        if r.get("strict_dtc_include")==0: f.append("excluded_from_strict_DTC")
        if f:
            qa_rows.append({"research_id": r["research_id"], "ete_group": r["ete_group"],
                            "n_flags": len(f), "flags": "; ".join(f)})
    qa = pd.DataFrame(qa_rows)
    print(f"   QA flagged rows: {len(qa)}")

    crosswalk = pd.DataFrame(CROSSWALK_ROWS, columns=["derived_column","source_object","source_columns","cleaning_rule"])
    dictionary = pd.DataFrame(DICTIONARY_ROWS, columns=["column","type","definition","allowed_values_or_range","source_object"])

    print("-> Assembling workbook ...", flush=True)
    wb = Workbook()
    ws = wb.active; ws.title = "Cover"
    ws["A1"] = "M044 - Per-Patient Master Data Workbook"; ws["A1"].font = TF
    ws["A2"] = "Microscopic vs Gross Extrathyroidal Extension and Recurrence in Differentiated Thyroid Cancer"
    ws["A2"].font = SF
    cover = [
        ("",""),
        ("Date prepared", datetime.utcnow().strftime("%Y-%m-%d")),
        ("Cohort", "manuscript_workspace.cohort_m044_ajcc_ete_v1 (n = 4,128)"),
        ("Database", DB),
        ("Primary endpoint", "main.canonical_recurrence_resolved_v1.recurrence_path_proven (n = 145)"),
        ("Secondary endpoint A", "imaging_only_unconfirmed (n = 195)"),
        ("Secondary endpoint B", "Composite path-or-imaging-suspicious (n = 340)"),
        ("Source objects", "cohort_m044_ajcc_ete_v1; canonical_patient_master; canonical_recurrence_resolved_v1; "
                         "ln_master_rollup_v1; cohort_m040_reoperative_v1; path_synoptics"),
        ("Aggregation rule", "ln_master_rollup_v1 and cohort_m040_reoperative_v1 are pre-aggregated MAX(...) per research_id."),
        ("Strict-DTC denominator", "n = 3,787 patients. Excludes MTC, anaplastic, NIFTP, FTUMP, follicular adenoma, NUT, adenoid cystic."),
        ("Standing rules applied", "(1) Demographics + full canonical schema audited; "
                                 "(2) Recurrence dual-track preserved; (3) LVI/vascular separated with explicit missing; "
                                 "(4) RAI is sensitivity, not primary; (5) Strict-DTC inclusion."),
        ("Tabs", "1 Cover - 2 Patient analytic - 3-8 Source - 9 Crosswalk - 10 Data dictionary - 11 QA flags"),
        ("Reproducibility", "Build script: build_master_excel.py; SQL: M044_ETE_analysis.sql"),
        ("Contact", "Logan Glosser, Emory Surgery Research"),
    ]
    for i,(k,v) in enumerate(cover, start=4):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=str(v)).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 28; ws.column_dimensions["B"].width = 100
    for r in range(4, 4+len(cover)): ws.row_dimensions[r].height = 36

    layout = [
        ("Patient analytic", analytic),
        ("Source-cohort_m044", src_cohort),
        ("Source-canonical_patient", src_cpm),
        ("Source-recurrence_resolved", src_rec),
        ("Source-ln_master_rollup", src_ln),
        ("Source-cohort_m040_reop", src_reop),
        ("Source-path_synoptics_LVI", src_path),
        ("Crosswalk", crosswalk),
        ("Data dictionary", dictionary),
        ("QA flags", qa if len(qa) else pd.DataFrame({"research_id":[], "flags":[]})),
    ]
    for name, df in layout:
        ws = wb.create_sheet(name[:31])
        write_df(ws, df, 1)

    print(f"-> Save -> {OUT}")
    wb.save(OUT)
    print(f"OK {OUT.stat().st_size:,} bytes")

    summary = {
        "Patient analytic": str(analytic.shape),
        "src cohort": str(src_cohort.shape),
        "src CPM": str(src_cpm.shape),
        "src recurrence": str(src_rec.shape),
        "src LN agg": str(src_ln.shape),
        "src reop agg": str(src_reop.shape),
        "src path LVI": str(src_path.shape),
        "Crosswalk": str(crosswalk.shape),
        "Dictionary": str(dictionary.shape),
        "QA": str(qa.shape),
    }
    print("\nFinal tab summary:")
    for k,v in summary.items(): print(f"  {k}: {v}")

if __name__ == "__main__":
    main()
